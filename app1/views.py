import zipfile
from calendar import monthrange
from django.contrib.auth import get_user_model
from django.core.mail import EmailMultiAlternatives
from django.utils.timezone import now
from .utils import get_client_ip, is_last_week_of_month
from .utils import get_work_duration,  get_half_hour, timedelta_from_datetime_or_time
from .clock_out import clocked_out
from .clock_in import clocked_in
from num2words import num2words
from reportlab.platypus import Paragraph, Image
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from django.urls import reverse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from io import BytesIO
from reportlab.pdfgen import canvas
from .models import Myprofile, companyprofile, Punch
from django.db.models import Q, F
from django.shortcuts import render
from django.db import transaction
from django.db.models import Case, IntegerField, When
import calendar
from calendar import month_abbr, monthcalendar, monthrange
from app1 import views
from django.db.models import Sum
from dateutil.relativedelta import relativedelta
from django.core.serializers import serialize
from django.contrib.auth import get_user
from datetime import time
import datetime
from itertools import groupby
from operator import itemgetter
import io
import json
import math
import random
from collections import defaultdict, OrderedDict
from datetime import date, datetime, timedelta
from time import gmtime, strftime
from django.db.models.functions import TruncDate

from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth import login
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.core.mail import (EmailMessage, EmailMultiAlternatives,
                              get_connection, send_mail)
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Max, OuterRef, Q, Subquery
from django.http import FileResponse, HttpResponse, HttpResponseNotFound, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.html import strip_tags
# from django_mail_admin import mail
from django.views.decorators.cache import cache_control
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view
from rest_framework.response import Response
from sqlalchemy import over

from app1.decorators import allowed_users
from app1.serializer import LogSerializer, UserSerializer
from app1.utils import shopUserHome
from decimal import Decimal, getcontext
from .models import *
from .forms import *
from itertools import zip_longest

import pandas as pd
import os
from django.core.files.base import ContentFile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pathlib import Path
from .utils import parse_and_format_date, calculate_distance, is_within_geofence
from math import radians, cos, sin, asin, sqrt
from django.utils.dateparse import parse_date
from .models import companyprofile, Myprofile, User
from django.contrib.auth.decorators import login_required
from .decorators import allowed_users
from django.core.cache import cache
from dateutil import parser
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.views.decorators.http import require_http_methods
from django.http import HttpResponseForbidden

BASE_DIR = Path(__file__).resolve().parent.parent
import logging
logger = logging.getLogger("HRMS")



def adminlogin(request):

    if request.method == 'POST':
        print("test")
        name = request.POST.get('email')
        password = request.POST.get('password')
        print(name, password)
        user = authenticate(email=name, password=password)
        if user:
            if not user.is_superuser and user.status in ["Active", "Onboarding"]:
                login(request, user)
                return redirect('empdash')

            elif user.is_superuser:
                login(request, user)
                return redirect('dashboard')
        else:
            messages.warning(request, 'Invalid Email /Password')
            return redirect('/')
    return render(request, 'login.html')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_empdash(request):
    datas = User.objects.all()

    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=user_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    attendance_rule = AssignAttendanceRule.objects.filter(
        user_id__id=user_id).first()
    leave_notification = LeaveNotification.objects.filter(
        Q(user=user_id) | Q(user_id__admin_id=user_id))
    notifications = []
    button_flag = True

    now = timezone.now()

    if request.user.status not in ["Active", "active"]:
        button_flag = False

    current_date = timezone.now()

    for i in leave_notification:
        if i.admin_id == 0:
            profile = Myprofile.objects.filter(myuser=i.user).first()
        else:
            profile = Myprofile.objects.filter(myuser__id=i.admin_id).first()

        notifications.append(
            {
                "message": i.message,
                "image_url": profile.image.url
                if profile and profile.image else "/static/logo/userlogo.png",
                "notification_id": i.id,
                "user": i.user,
                'admin_id': i.admin_id,
                'is_approved': i.is_approved,
                "readadmin": i.readadmin,
                "events": i.events,
            }
        )

    currentdate = datetime.now()
    year = current_date.year
    month = current_date.strftime("%B")

    users = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id))
    birthday_users = [user for user in users if user.is_birthday_today()]
    work_ann = [user for user in users if user.is_workanniversary_today()]

    print("bool(work_ann) :", bool(work_ann))

    users_count = User.objects.filter(
        Q(id=user_id) | Q(admin_id=user_id)).count()
    print("users_count :", users_count)
    punchedcount = Punch.objects.filter(Q(date__date=current_date) & (Q(is_first_clocked_in=True) | Q(
        is_second_clocked_in=True)) & (Q(user__id=user_id) | Q(user__admin_id=user_id))).count()
    print("punched_count :", punchedcount)
    punched_count = 0
    if users_count > 0:
        punched_count = (punchedcount * 100) // users_count
    print("Percentage of punched count relative to users count:", punched_count)

    pending_leavecount = Leave.objects.filter(Q(Appliedon__month=currentdate.month, Appliedon__year=currentdate.year) & Q(
        status="Applied") & (Q(applicant_email=user_id) | Q(applicant_email__admin_id=user_id))).count()

    admin_id = request.user.id
    admin_wrklcn_id = request.user.wrklcn.id if request.user.wrklcn else None
    holiday_count = HolidayLocationList.objects.filter(
        Holiday_List__Myuser_13=user_id, HolidayLocation__id=admin_wrklcn_id, Holiday_List__HolidayDate__icontains=currentdate.strftime("%B %Y")).count()

    punchrequest_count = Punch.objects.filter(Q(date__month=currentdate.month, date__year=currentdate.year) & Q(
        is_requested=True) & Q(is_approved=False) & Q(is_rejected=False) & (Q(user__id=user_id) | Q(user__admin_id=user_id))).count()

    users = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id))

    today = datetime.now().date()
    current_month = today.month

    birthday_users = [user for user in users if user.is_birthday_today(
    ) and datetime.strptime(user.dob, '%d %B %Y').date().month == current_month]

    work_anniversary_users = [user for user in users if user.is_workanniversary_today(
    ) and datetime.strptime(user.datejoin, '%d %B %Y').date().month == current_month]

    birthday_count = len(birthday_users)
    work_anniversary_count = len(work_anniversary_users)

    punch_object = Punch.objects.filter(
        user__id=user_id, date__date=current_date.date()).last()

    clock_in_type = 2
    if punch_object:
        clock_in_type = punch_object.last_punch_type
    if not attendance_rule:
        messages.info(request, 'Attendance Rules not assinged')

    # display current work duration
    in_time = time(hour=0, minute=0, second=0)
    try:
        punch_data = Punch.objects.get(
            user=request.user, date__date=now.date())
        in_time = punch_data.first_clock_in_time
    except Punch.DoesNotExist:
        pass
    print("In time: ", in_time)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "notifications": notifications,
        "punched_count": punched_count,
        "pending_leavecount": pending_leavecount,
        "holiday_count": holiday_count,
        "punchrequest_count": punchrequest_count,
        "clock_in_type": clock_in_type,
        "button_flag": button_flag,
        "in_time": in_time,
    }
    return render(
        request,
        "employee/inactive_dashboard.html",
        {
            "datas": datas,
            "leave_notification": leave_notification,
            "users": users,
            "birthday_users": birthday_users,
            "show_birthday_image": bool(birthday_users),
            "work_ann": work_ann,
            "show_workann_image": bool(work_ann),
            "month": month,
            "birthday_count": birthday_count,
            "work_anniversary_count": work_anniversary_count,
            **x,
        },
    )


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_overview(request):
    user_id = request.user.id 
    admin_id = User.objects.get(id=user_id).admin_id    
    datas = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id).first()    
    user_obj = User.objects.filter(id=request.user.id).first()
    company_details = None 
    if user_obj:
            
        if user_obj.company_type:  
            print(f"User's Company Type: {user_obj.company_type}")  
            company_type = user_obj.company_type.type_of_company  

            if company_type == 'Main Company':
                company_details = datas.filter(type_of_company='Main Company').first()
                print(f"Main Company Details: {company_details}")  
            elif company_type == 'Sub Company':
                company_details = datas.filter(type_of_company='Sub Company').first()
                print(f"Sub Company Details: {company_details}")  
    context = {
            "k": k,
            "datas": datas.first() if datas.exists() else None,  
            "company_details": company_details,  
            "user_obj": user_obj  
        }

    return render(request, "Employee/inactive_overview.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_address(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    k = corporateaddress.objects.filter(admin_id=admin_id)
    data = registeredaddress.objects.filter(admin_id=admin_id)

    y = companyprofile.objects.filter(admin_id=admin_id)
    w = Myprofile.objects.filter(myuser__id=request.user.id)
    x = {
        "y": y[0] if y.exists() else y,
        "w": w[0] if w.exists() else w,
    }

    return render(request, "Employee/inactive_address.html", {'data': data, 'k': k, **x})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_personalinfo(request):
    user = request.user
    admin_id = User.objects.get(id=user.id).admin_id
    print("user :", user, admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        myprofile = None
    context = {
        'myprofile': myprofile if myprofile else '',
        'c': c[0] if c.exists() else c
    }

    return render(request, "Employee/inactive_personalinfo.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_work(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    k = Myprofile.objects.filter(myuser__id=user_id)
    c = companyprofile.objects.filter(admin_id=admin_id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "Employee/inactive_workmypro.html", x)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_team(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    x = Directreports.objects.filter(myuser_3__id=request.user.id)

    l = Reportingmanager.objects.filter(userid=user_id)

    rpt_users = Reportingmanager.objects.filter(myuser_2=user_id)
    directreport_users = []
    for rpt_user in rpt_users:
        user = User.objects.filter(id=rpt_user.userid).first()
        if user:
            directreport_users.append(user)

    y = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'l': l,
        'x': x,
        'directreport_users': directreport_users
    }
    return render(request, "Employee/inactive_team.html", y)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_education(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Educationalinfo.objects.filter(myuser_4__id=request.user.id)
    y = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "l": l,
    }
    print("EMPLOYEE")
    return render(request, "Employee/inactive_education.html", y)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_family(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Familymembers.objects.filter(myuser_5__id=request.user.id)
    x = Emergencycontact.objects.filter(myuser_6__id=request.user.id)

    y = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "x": x,
    }

    return render(request, "Employee/inactive_family.html", y)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_document(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    uploads = Uploadeddocs.objects.filter(myuser__id=request.user.id)
    proofs_object = Proof.objects.all()
    proofs = {
        proof.id: proof.proof_name for uploaded_doc in uploads for proof in uploaded_doc.proof.all()}

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "uploads": uploads,
        "data": data[0] if data.exists() else data,
        "proofs": proofs,
        "proofs_object": proofs_object
    }

    return render(request, "Employee/inactive_document.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_certifications(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    certificates = Certifications.objects.filter(myuser_8__id=request.user.id)

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "certificates": certificates,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "Employee/inactive_certifications.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_docwork(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id

    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    work = Work.objects.filter(myuser_9__id=request.user.id)

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "work": work,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "Employee/inactive_docwork.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_workweek(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)
    work_week = AssignWorkWeek.objects.filter(user_id=user_id)
    for i in work_week:
        print("Work week : ",
              work_week, i.rules_applied.half_day)
    context = {
        'work_week': work_week,
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "Employee/inactive_workweek.html", context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_e_exit(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)
    resignation = ResignationForm.objects.filter(user=request.user.id)
    context = {
        'k': k[0] if k.exists() else k,
        'c': c[0] if c.exists() else c,
    }
    return render(request, 'Employee/inactive_e_exit.html', {'resignation': resignation, **context})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_payroll(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)

    regaddress = registeredaddress.objects.filter(admin_id=admin_id)
    print("regofficeaddress :", regaddress)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
    print("selected_year :", selected_year, selected_month)

    bank_details = Bank_account.objects.filter(myuser_11=user_id)
    print("bank_details :", bank_details)

    assign_salarystructure = AssignSalaryStructure.objects.filter(
        user_id=user_id, effective_date__month=selected_month, effective_date__year=selected_year).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)

    assign_data = []

    ctc_amount = 0

    selected_date = datetime(selected_year, selected_month, 1)
    print("selected_date :", selected_date)

    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(
            effective_date__lte=selected_date, user_id=user_id).order_by('-effective_date').first()

        if nearest_date:
            assign_salarystructure = nearest_date

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        names = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure)
        amounts = AssignSalaryStructureAmount.objects.filter(
            salaryname__in=names)
        print("names ; amounts 554 :", names, amounts)

        ctc_amount += sum(amount.amount for amount in amounts)
        zipped_data = zip_longest(names, amounts)

        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })
    print("ctc_amount :", ctc_amount)
    print("assign_data :", assign_data)

    adhoc_data = Adhoc.objects.filter(user_id=user_id, createddate__year=selected_year,
                                      createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
    print("adhoc_data : ", adhoc_data)

    earning_amount = 0
    deduction_amount = 0
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    total_earnings = ctc_amount + earning_amount
    print("CCCCCCCCCCCCCCCCCC :", total_earnings)

    num_days = calendar.monthrange(selected_year, selected_month)[1]

    payregister = PayRegister.objects.filter(createddate__month=selected_month, createddate__year=selected_year, user_id=user_id, status__in=[
                                             "Payslip Generated", "Payslip Downloaded"])
    print("payregister :", payregister)

    lop_data = Runpayroll_lop.objects.filter(
        lop_date__month=selected_month, lop_date__year=selected_year, user_id=user_id)
    lopcount = 0
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    absent_AN_count = Punch.objects.filter(
        user__id=user_id, date__year=selected_year, date__month=selected_month, status__in=['A', 'AN']).count()
    punchcount = Punch.objects.filter(
        user__id=user_id, date__year=selected_year, date__month=selected_month).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count)
    working_days = punchcount - absent_AN_count
    print("working_days :", working_days)
    total_lop = absent_AN_count + missing_date_count + lopcount

    per_day_amount = ctc_amount / num_days
    print("per_day_amount :", per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount)
    lopamount = round(lop_amount)

    total_deductions = deduction_amount + lop_amount
    totaldeductions = round(total_deductions)
    net_amount = round(total_earnings - total_deductions)
    print("net_amount :", net_amount)

    net_amount_words = num2words(net_amount, lang='en_IN')

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "regaddress": regaddress,
        "payregister": payregister,
        "num_days": num_days,
        "bank_details": bank_details,
        "assign_salarystructure": assign_salarystructure,
        "assign_data": assign_data,
        "ctc_amount": ctc_amount,
        "adhoc_data": adhoc_data,
        "total_earnings": total_earnings,
        "totaldeductions": totaldeductions,
        "total_lop": total_lop,
        "lopamount": lopamount,
        "net_amount": net_amount,
        "net_amount_words": net_amount_words,
        "month_str": month_str,
        "selected_year": selected_year,
    }

    return render(request, "Employee/inactive_payroll.html", x)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_salarystructure(request):
    user_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)

    assign = AssignSalaryStructure.objects.filter(user_id=user_id).last()

    assign_data = []

    ctc_amount = 0
    annualctc_amount = 0

    if assign:
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign)
        amounts = AssignSalaryStructureAmount.objects.filter(
            salaryname__in=names)

        annual_amount = [amount.amount * 12 for amount in amounts]

        annualctc_amount += sum(annual_amount)

        ctc_amount += sum(amount.amount for amount in amounts)
        zipped_data = zip_longest(names, amounts, annual_amount)

        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, 'Employee/inactive_salarystructure.html',
                  {'assign': assign, 'assign_data': assign_data,
                   'ctc_amount': ctc_amount, 'annualctc_amount': annualctc_amount, **x})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_bankaccount(request):
    user = request.user
    admin_id = User.objects.get(id=request.user.id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    datas = Bank_account.objects.filter(myuser_11=request.user.id)

    try:
        bankdatas = Bank_account.objects.get(myuser_11=user)
    except Bank_account.DoesNotExist:
        bankdatas = None

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'bankdatas': bankdatas,
    }

    return render(request, 'Employee/inactive_bankaccount.html', {'datas': datas, **x})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Inactive'])
def inactive_reimbursements(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    p = Myprofile.objects.filter(myuser__id=request.user.id)
    k = reimbursement.objects.filter(myuser_11=user_id)
    x = {
        "p": p[0] if p.exists() else p,
        "c": c[0] if c.exists() else c,
    }
    return render(request, 'Employee/inactive_reimbursements.html', {'k': k, **x})


def sign_up(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phoneno = request.POST.get('phoneno')
        status = 'Active'
        role = 'Admin'
        otp = random.randint(000000, 999999)
        empid = 10000
        admin_id = request.user.id

        if User.objects.filter(email=email).exists():
            msg1 = 'Email already taken try another one'
            return render(request, "sign_up.html", {'name': name, 'email': email, 'phoneno': phoneno, 'msg1': msg1, })
        if User.objects.filter(phone=phoneno).exists():
            msg2 = 'Phone number already taken try another one'
            return render(request, "sign_up.html", {'name': name, 'email': email, 'phoneno': phoneno, 'msg2': msg2, })

        user = User.objects.create(empid=empid, username=name, email=email,
                                   phone=phoneno, status=status, role=role, otp=otp, password=otp, is_superuser=1,
                                   is_staff=1)

        if not Proof.objects.exists():
            proof_names = ["photo_id", "date_of_birth",
                           "current_address", "permanent_address"]
            for name in proof_names:
                Proof.objects.create(proof_name=name)

        to = [user.email]
        subject = 'OTP For Login'
        html_body = render_to_string(
            'index/emp_otpemail.html', {'user': user, 'otp': otp})

        msg = EmailMultiAlternatives(
            subject=subject, from_email=settings.EMAIL_HOST_USER, to=to)
        msg.attach_alternative(html_body, "text/html")
        msg.send()
        messages.success(request, 'Created Successfully!',extra_tags='bg-success text-white')

        return redirect('login')
    return render(request, "sign_up.html")


def loginf(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        u = User.objects.filter(email=email, otp=password).first()
        if u is not None:
            r = User.objects.filter(email=email, otp=password)
            status = r[0].status
            role = r[0].role
            last_login = r[0].last_login
            if last_login:
                if remember_me:
                    # print("Remember Me is enabled", request.session)
                    request.session.set_expiry(1209600) # 2 weeks
                else:
                    # print("Remember Me is disabled", request.session)
                    request.session.set_expiry(0) 
                if role == "Employee" and status in ["Active", "Onboarding"]:
                    login(request, u)
                    return redirect('empdash')
                if role == "Employee" and status == "Inactive":
                    login(request, u)
                    return redirect('inactive_overview')
                elif role == "Admin":
                    login(request, u)
                    return redirect('dashboard')
            elif not last_login:
                return redirect('resetpassword', id=r[0].id)
            else:
                messages.info(request, " Invalid Email or Password")
                return render(request, 'login.html')
        else:
            messages.info(request, " Invalid Email or Password")
            return render(request, 'login.html')

    return shopUserHome(request)


def resetpassword(request, id):
    u = User.objects.filter(id=id).exists
    if u:
        user = User.objects.get(id=id)

        if request.method == 'POST':
            number = request.POST.get('number')

            existing_password = request.POST.get('password')
            newpass = request.POST.get('newpassword')
            confirmpassword = request.POST.get('confirmpassword')
            u = User.objects.filter(id=id, otp=existing_password).exists()
            if u:
                if newpass == confirmpassword:

                    user.otp = newpass
                    user.password = make_password(newpass)
                    user.save()
                    # Login in logout is doing to update last_login date, so on next time login rest option won't ask
                    login(request, user)
                    logout(request)
                    messages.success(
                        request, "Password reset successfully, please login again",extra_tags='bg-success text-white')
                    return render(request, 'login.html')
                else:
                    messages.info(
                        request, "New and confirm password not matching")

            else:
                messages.info(request, "Existing password not matching")

        return render(request, 'resetpassword.html', {'number': user.email, 'id': user.id})
    else:
        messages.info(request, "User not exit")
        return render(request, 'login.html')


def generateOTP():
    digits = "0123456789"
    OTP = ""
    for i in range(6):
        OTP += digits[math.floor(random.random() * 10)]
    return OTP


def send_otp(request):
    email = request.POST.get("email")
    print(email)
    if User.objects.filter(email=email):
        r = get_object_or_404(User, email=email)
        print('name:', r.username)
        k = generateOTP()
        print(k)
        subject = 'Password Reset Request'
        htmlgen = k
        html_body = render_to_string(
            'index/reset_otp.html', {'r': r.username, 'k': k})
        send_mail(subject, k, '<cydezt@gmail.com>',
                  [email], fail_silently=False, html_message=html_body)
        return HttpResponse(k)


def change_password(request):
    if request.method == 'POST':
        email = request.POST.get("email")
        pwd = request.POST.get("Npassword")
        pwd1 = request.POST.get("Cpassword")
        ca = User.objects.filter(email=email).first()
        if pwd == pwd1:
            ca.set_password(pwd)
            ca.otp = pwd
            ca.save()
            login(request, ca)
            logout(request)
            messages.success(
                request, "Password reset successfully, please login again",extra_tags='bg-success text-white')
            return render(request, 'login.html')
        else:
            messages.warning(request, 'Password must be same')
    return render(request, "forgotpassword.html")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def logout_user(request):
    response = HttpResponseRedirect('/')
    for cookie in request.COOKIES:
        response.delete_cookie(cookie)
    logout(request) 
    return response


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
def clock_in_out(request):
    user = request.user
    current_date = timezone.now().date()

    try:
        attendancerule = AssignAttendanceRule.objects.get(user_id=user.id)
    except AssignAttendanceRule.DoesNotExist:
        messages.info(request, "Rule not assigned, try to contact Admin")
        return redirect("dashboard" if user.role.lower() == 'admin' else "empdash")

    try:
        punch_object = Punch.objects.get(user=user, date__date=current_date)
    except Punch.DoesNotExist:
        punch_object = None

    if request.method == 'POST':
        clock_in_type = request.POST.get("clock_in_type", None)
        clock_out_type = request.POST.get("clock_out_type", None)
        location = request.POST.get("location", None)

        try:
            lat = float(request.POST.get('lat'))
            lon = float(request.POST.get('lon'))
        except (ValueError, TypeError):
            messages.error(request, "Invalid location coordinates.")
            return redirect("empdash")

        print("🧪 Raw POST data:", request.POST)
        print("📥 Raw lat:", lat, "Raw lon:", lon)

        fence_lat = fence_lon = fence_radius = None

        if location == "WFH":
            try:
                fence = EmployeeGeoFence.objects.get(user=request.user)
            except EmployeeGeoFence.DoesNotExist:

                return redirect("set_home_location")

            if fence.home_lat and fence.home_lon and fence.home_radius:
                fence_lat = fence.home_lat
                fence_lon = fence.home_lon
                fence_radius = fence.home_radius
            else:
                messages.error(request, "Location radius not set. Kindly contact HR for quick resolution.")
                return redirect("empdash")

        elif location == "WFO":
            if hasattr(user, "employeeprofile") and user.employeeprofile.branch_location:
                branch = user.employeeprofile.branch_location
                fence_lat = branch.lat
                fence_lon = branch.lon
                fence_radius = branch.radius
            else:
                messages.error(request, "Branch location not assigned. Kindly contact HR for quick resolution.")
                return redirect("empdash")
        else:
            messages.error(request, "Invalid work location selected.")
            return redirect("empdash")
        

        if not all([fence_lat, fence_lon, fence_radius]):
            messages.error(request, f"{location} geofence not set. Kindly contact HR for quick resolution.")
            return redirect("empdash")

        distance = calculate_distance(lat, lon, fence_lat, fence_lon)

        print(f"📍 Employee Location: ({lat}, {lon})")
        print(f"🎯 Target Location: ({fence_lat}, {fence_lon})")
        print(f"📏 Distance: {distance:.2f}m")
        print(f"📐 Allowed Radius: {fence_radius}m")
        print(f"🔹 Inside Circle: {distance <= fence_radius}")

        if distance > fence_radius:
            messages.error(request, f"Clock-in denied: You are outside the allowed {fence_radius:.2f}m Location Radius.")
            return redirect("empdash")

        if clock_in_type:
            result = clocked_in(request, user, punch_object, attendancerule, location)
            if result is True:
                request.session["clocked_in"] = True
                messages.success(request, "Clocked in successfully.")
            elif result is False:
                messages.error(request, "Clock-in failed. Kindly contact HR for quick resolution.")
            return redirect("dashboard" if user.role.lower() == 'admin' else "empdash")

        elif clock_out_type:
            result = clocked_out(user, punch_object, attendancerule)
            if result is True:
                request.session["clocked_in"] = False
                messages.success(request, "Clocked out successfully.")
            elif result is False:
                messages.error(request, "Clock-out failed. Kindly contact HR for quick resolution.")
            return redirect("dashboard" if user.role.lower() == 'admin' else "empdash")

    return redirect("empdash")

    

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def dashboard(request):
    auto_approve_leaves_once_per_day()
    print('Auto-approved leave request ')
    approve_resignation() 
    print('Auto-approved resignation ')

    user_id = request.user.id
    my_profile = Myprofile.objects.filter(myuser_id=user_id).first()
    cache_key = f"dashboard_data_{user_id}"
    today = datetime.today().date()
    next_30_days = [(today + timedelta(days=i)).strftime("%d-%m") for i in range(30)]
    cached_data = cache.get(cache_key)
    if cached_data:
        print("cache hit")
        return render(request, "index/dashboard.html", cached_data)

    current_date = timezone.now()
    today = current_date.date()

    # Calculate this month and last month ranges
    
    this_month_start = today.replace(day=1)
    this_month_end = today.replace(day=monthrange(today.year, today.month)[1])
    last_month = this_month_start - timedelta(days=1)
    last_month_start = last_month.replace(day=1)
    last_month_end = last_month

    # users = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id)).select_related("reptmgr")
    users = User.objects.filter(
        Q(id=user_id) | Q(admin_id=user_id)
    ).exclude(
        Q(resignationform__status='Approved') &
        Q(resignationform__actual_last_working_day__lt=date.today())
    ).select_related("reptmgr")

    # Function to parse `datejoin` safely
    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            try:
                return datetime.strptime(date_str, "%d %B %Y").date()
            except:
                return None

    # Joined last month
    last_month_joined = sum(
        1 for u in users
        if u.status in ["Active", "Onboarding"] and (d := parse_date(u.datejoin)) and last_month_start <= d <= last_month_end
    )

    # Joined this month
    this_month_joined = sum(
        1 for u in users
        if (d := parse_date(u.datejoin)) and this_month_start <= d <= this_month_end
    )
    
    # Resigned last month

    
    last_month_resigned = ResignationForm.objects.filter(
        last_workingday__range=(last_month_start, last_month_end)
    ).count()

    # Also count resignations submitted last month (not just last working day)
    submitted_last_month = ResignationForm.objects.filter(
        resignation_date__range=(last_month_start, last_month_end),
        user__in=users
    ).count()

    # Count resignations with last working day in the future (pending exit)
    pending_exit = ResignationForm.objects.filter(
        last_workingday__gt=today,
        user__in=users
    ).count()

    # Resigned this month
    this_month_resigned = ResignationForm.objects.filter(
        user__in=users,
        last_workingday__range=(this_month_start, this_month_end)
    ).count()

    seven_days_ago = current_date - timedelta(days=7)
    # Fetch profiles in one query
    profiles = Myprofile.objects.filter(myuser__id__in=users.values_list("id", flat=True)).select_related("myuser")
    profile_dict = {profile.myuser_id: profile for profile in profiles}

    # Fetch leave notifications
    leave_notification = LeaveNotification.objects.filter(
        Q(user=user_id) | Q(user_id__admin_id=user_id) & Q(timestamp__gte=seven_days_ago)
    ) 

    print("Leave Notification: ", leave_notification)
    notifications = [
        
        {
            "message": i.message,
            "image_url": (
                profile_dict.get(i.user_id, {}).image.url
                if profile_dict.get(i.user_id) and profile_dict[i.user_id].image
                else "/static/logo/userlogo.png"
            ),
            "notification_id": i.id,
            "user": i.user,
            "admin_id": i.admin_id,
            "is_approved": i.is_approved,
            "readadmin": i.readadmin,
            "events": i.events,
        }
        for i in leave_notification
    ]

    logger.info(notifications)

    # Only consider "Active" status employees for attendance calculations
    attendance_rule = AssignAttendanceRule.objects.filter(user_id=user_id).first()

    # Filter users with status "Active"
    active_users = users.filter(
        status__iexact="Active"
    ).exclude(
        Q(resignationform__status='Approved') &
        Q(resignationform__actual_last_working_day__lt=date.today())
    ).exclude(
        is_superuser=True
    )
    active_user_ids = active_users.values_list("id", flat=True)

    

    # Count punches for "Active" users only
    punched_count = Punch.objects.filter(
        date__date=current_date.date(),
        user__id__in=active_user_ids,
        is_first_clocked_in=True,
    ).count()

    total_active_users = active_users.count()

    male_count = active_users.filter(gender__iexact='Male').count()
    female_count = active_users.filter(gender__iexact='Female').count()
    gender_ratio = f"M {male_count} / F {female_count}"


    total_active_users = active_users.count()
    
    # Define male and female user querysets
    male_users = active_users.filter(gender__iexact='Male')
    female_users = active_users.filter(gender__iexact='Female')

    male_count = male_users.count()
    female_count = female_users.count()
    # Avoid divide by zero
    male_percent = round((male_count / total_active_users) * 100, 2) if total_active_users else 0
    female_percent = round((female_count / total_active_users) * 100, 2) if total_active_users else 0


    punched_percentage = (punched_count * 100) // total_active_users if total_active_users else 0

    today = datetime.now().date()
    today_mmdd = today.strftime("%m-%d")
    next_30_days_dates = [(today + timedelta(days=i)).strftime("%d-%m") for i in range(1, 31)] 

    # First get all active users with birthdays
    birthday_users = [
        user for user in users
        if user.status.lower() == "active" and user.dob
    ]

    # Now filter for today's birthdays only
    today_birthday_users = []
    for user in birthday_users:
        try:
            # Handle both date objects and string formats
            if isinstance(user.dob, (date, datetime)):
                dob_mmdd = user.dob.strftime("%m-%d")
            else:
                # Try parsing different string formats
                try:
                    dob_date = datetime.strptime(user.dob, "%Y-%m-%d")
                except ValueError:
                    try:
                        dob_date = datetime.strptime(user.dob, "%d %B %Y")
                    except ValueError:
                        continue
                dob_mmdd = dob_date.strftime("%m-%d")
            
            if dob_mmdd == today_mmdd:
                today_birthday_users.append(user)
        except Exception as e:
            print(f"Error processing birthday for user {user.id}: {e}")
            continue

    birthday_count = len(today_birthday_users)

    # Now calculate upcoming birthdays (excluding today's)
    upcoming_birthday_users = [
        user for user in users
        if user.status.lower() == "active"
        and user.dob is not None
        and user not in today_birthday_users  # Use today_birthday_users instead of birthday_users
        and (
            (user.dob.strftime("%d-%m") if isinstance(user.dob, datetime) 
            else datetime.strptime(user.dob, "%d %B %Y").strftime("%d-%m"))
        ) in next_30_days_dates
    ]
    upcoming_birthday_count = len(upcoming_birthday_users)
    show_birthday_image = bool(today_birthday_users)

    # Calculate work anniversaries within the next 30 days
    work_anniversaries = []
    today = timezone.now().date()  # Use timezone-aware date

    for user in users.filter(status__iexact="active"):  # More efficient filtering
        if not user.datejoin:  # Skip if no join date
            continue
            
        try:
            # Parse datejoin to date object (handling multiple formats)
            if isinstance(user.datejoin, (date, datetime)):
                doj = user.datejoin.date() if isinstance(user.datejoin, datetime) else user.datejoin
            else:
                try:
                    doj = datetime.strptime(str(user.datejoin), "%Y-%m-%d").date()
                except ValueError:
                    doj = datetime.strptime(str(user.datejoin), "%d %B %Y").date()
            
            # Calculate this year's anniversary
            anniv_this_year = date(today.year, doj.month, doj.day)
            
            # If already passed this year, use next year
            if anniv_this_year < today:
                anniv_this_year = date(today.year + 1, doj.month, doj.day)
            
            # Check if today is the anniversary
            if anniv_this_year == today:
                work_anniversaries.append(user)
                
        except Exception as e:
            print(f"Error processing work anniversary for user {user.id} ({user.username}): {str(e)}")
            continue

    work_anniversary_count = len(work_anniversaries)

    # # Group birthdays by date
    # birthdays_by_date = defaultdict(list)
    # for user in birthday_users:
    #     dob_date = user.dob if isinstance(user.dob, datetime) else datetime.strptime(user.dob, "%d %B %Y").date()
    #     birthdays_by_date[dob_date.strftime("%d-%m")].append(user)

    # Fetch pending leaves, punch requests, and holidays
    pending_leave_count = Leave.objects.filter(
        Appliedon__month=current_date.month,
        Appliedon__year=current_date.year,
        status="Pending",
    ).filter(
        Q(applicant_email__admin_id=user_id) | Q(applicant_email=user_id)
    ).count()

    punch_request_count = Punch.objects.filter(
        date__month=current_date.month, date__year=current_date.year,
        is_requested=True, is_approved=False, is_rejected=False,
        user__id__in=users.values_list("id", flat=True),
    ).count()

    # Fetch work location-based holiday count
    admin_wrklcn_id = request.user.wrklcn.id if request.user.wrklcn else None

    # Count holidays for the user's work location for the whole month
    holiday_count = HolidayLocationList.objects.filter(
        Holiday_List__Myuser_13=user_id,
        HolidayLocation__id=admin_wrklcn_id,
        Holiday_List__HolidayDate__icontains=current_date.strftime("%B %Y"),
    ).count()

    # Get last punch object & clock-in type
    punch_object = Punch.objects.filter(user__id=user_id, date__date=current_date.date()).last()
    clock_in_type = punch_object.last_punch_type if punch_object else 2

    # Attendance rule check
    if not attendance_rule:
        messages.info(request, "⚠️ Attendance Rules not assigned.")

    # Punch-in time
    in_time = Punch.objects.filter(user=request.user, date__date=current_date.date()).first()
    in_time = in_time.first_clock_in_time if in_time else ""

    dashboard_data = {
        "my_profile": my_profile,
        "notifications": notifications,
        "punched_count": punched_percentage,
        "active_user_punched_count": f"{punched_count} / {total_active_users}",
        "pending_leavecount": pending_leave_count,
        "holiday_count": holiday_count,
        "punchrequest_count": punch_request_count,
        "clock_in_type": clock_in_type,
        "button_flag": request.user.status.lower() == "active",
        "in_time": in_time,
        "month": current_date.strftime("%B"),
        "upcoming_birthday_count": upcoming_birthday_count,
        "birthday_users": today_birthday_users,
        "birthday_count": birthday_count,
        # "show_birthday_image": show_birthday_image,
        "work_anniversary_count": len(work_anniversaries),
        "show_birthday_image": bool(birthday_users),
        "show_workann_image": bool(work_anniversaries),
        "last_month_joined": last_month_joined,
        "this_month_joined": this_month_joined,
        "last_month_resigned": last_month_resigned,
        "this_month_resigned": this_month_resigned,
        "gender_ratio": gender_ratio,
        "male_users": male_users,
        "female_users": female_users,
        "male_percent": male_percent,
        "female_percent": female_percent,
        "total_active_users": total_active_users,
    }

    unread_count = LeaveNotification.objects.filter(readadmin=False, admin_id=user_id).count()
    dashboard_data["unread_count"] = unread_count

    cache.set(cache_key, dashboard_data, timeout=300)
    return render(request, "index/dashboard.html", dashboard_data)
    

def view_emp(request, id):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.all()
    datas = User.objects.all()
    data = get_object_or_404(User, pk=id)
    data1 = get_object_or_404(Myprofile, pk=id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/empdetails.html", {'data': data, 'datas': datas, 'data1': data1, **x})


def dashboard_search(request, id):
    data = get_object_or_404(User, pk=id)
    return render(request, 'index/dashboard.html', {'data': data})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def view_employee(request, id):
    admin_id = User.objects.get(id=id).admin_id
    print("ADMINID", admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    datas = User.objects.all()
    data = get_object_or_404(User, pk=id)
    k = Myprofile.objects.filter(myuser_id=data.id)
    p = Myprofile.objects.filter(myuser_id=request.user.id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "p": p[0] if p.exists() else p,
    }
    return render(request, "Employee/empdetails.html", {'data': data, 'datas': datas, **x})


def reverse_geocode(lat, lon):
    geolocator = Nominatim(user_agent="geoapi")
    location = geolocator.reverse(f"{lat}, {lon}")
    return location.address if location else "Unknown Location"


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_personalinfo(request, id):
    datas = User.objects.all()
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    cmpprofile = companyprofile.objects.filter(admin_id=admin_id)
    myprofile = Myprofile.objects.filter(myuser=data.id)
    k = Myprofile.objects.filter(myuser=request.user.id)
    try:
        employee_profile = EmployeeProfile.objects.get(user=data)
    except EmployeeProfile.DoesNotExist:
        employee_profile = None

    try:
        wfh_geofence = EmployeeGeoFence.objects.get(user=data)
        wfh_location_name = reverse_geocode(wfh_geofence.home_lat, wfh_geofence.home_lon)
    except EmployeeGeoFence.DoesNotExist:
        wfh_geofence = None
        wfh_location_name = "Not Set"
    x = {
        "data": data,
        "datas": datas,
        "cmpprofile": cmpprofile[0] if cmpprofile.exists() else cmpprofile,
        "myprofile": myprofile[0] if myprofile.exists() else myprofile,
        "k": k[0] if k.exists() else k,
        "employee_profile": employee_profile,
        "wfh_geofence": wfh_geofence,
        "wfh_location_name": wfh_location_name,
    }
    return render(request, "index/emppersonaldetails.html", {'data': data, 'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_create_personalinfo(request, id):
    user = User.objects.get(id=id)
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        myprofile = Myprofile(myuser=user)
    if request.method == "POST":
        username = request.POST['name']
        email = request.POST['email']
        officialemail = request.POST['Officialemail']
        dob = request.POST['dob']
        gender = request.POST['gender']
        phone = request.POST['phone']
        alternativephone = request.POST['alternatephone']
        bldgrp = request.POST['bldgrp']
        marital = request.POST['marital']
        address = request.POST['address']
        personaladdress = request.POST['permanentaddress']
        housetype = request.POST['housetype']
        crntresidencedate = request.POST['currentresidancedate']
        crntcitydate = request.POST['currentcitydate']

        user.username = username
        user.email = email
        user.dob = dob
        user.gender = gender
        user.phone = phone
        user.save()

        myprofile.offemail = officialemail
        myprofile.altphone = alternativephone
        myprofile.bldgrp = bldgrp
        myprofile.marital = marital
        myprofile.address = address
        myprofile.peraddress = personaladdress
        myprofile.housetype = housetype
        myprofile.crntresidencedate = crntresidencedate
        myprofile.crntcitydate = crntcitydate
        img_file = request.FILES.get('img', None)
        if img_file is not None:
            f = FileSystemStorage()
            f1 = f.save(img_file.name, img_file)
            myprofile.image = f1

        myprofile.save()
        return redirect('view_personalinfo', id)
    return render(request, "index/emppersonaldetails.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_work(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id

    # 🧠 Step 1: Get all direct report relationships (Primary only)
    direct_report_entries = Reportingmanager.objects.filter(myuser_2=data.id, type='Primary')

    # 🧠 Step 2: Filter users that are still active and are currently reporting to this user
    active_direct_reports = User.objects.filter(
        status='Active',
        id__in=direct_report_entries.values_list('userid', flat=True)
    )

    direct_reports_count = active_direct_reports.count()
    has_direct_reports = direct_reports_count > 0
    show_warning = False

    # 🧠 Step 3: Handle status change
    if request.method == 'POST':
        new_status = request.POST.get('Empstatus')

        if new_status == 'Inactive' and has_direct_reports:
            show_warning = True
            messages.error(
                request,
                f"The user '{data.username}' has {direct_reports_count} active team member(s). "
                "Please reassign or remove them before marking the user as Inactive.",
                extra_tags="bg-danger text-white"
            )
        else:
            # Safe to update status
            data.status = new_status
            data.save()
            messages.success(request, f"Status updated to {new_status}.", extra_tags="bg-success text-white")
            return redirect('view_work', id=data.id)

    context = {
        'data': data,
        'datas': User.objects.all(),
        'dsn': Designation.objects.filter(admin_id=admin_id),
        'dpt': Department.objects.filter(admin_id=admin_id),
        'sd': Subdepartment.objects.filter(admin_id=admin_id),
        'jttl': Job.objects.filter(admin_id=admin_id),
        'wr': Worklocation.objects.filter(admin_id=admin_id),
        'cmpprofile': companyprofile.objects.filter(admin_id=admin_id).first(),
        'myprofile': Myprofile.objects.filter(myuser=admin_id).first(),
        'direct_reports_count': direct_reports_count,
        'has_direct_reports': has_direct_reports,
        'show_warning': show_warning,
    }

    return render(request, "index/empworkdetails.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_create_empwork(request, id):
    m = User.objects.filter(id=id)
    if request.method == "POST":
        empid = request.POST.get("Empid")
        joindate = request.POST.get("Joindate")
        proper = request.POST.get("Probationperiod")
        etyp = request.POST.get("Emptype")
        empstatus = request.POST.get("Empstatus")
        wrkex = request.POST.get("Workexp")

        designation_id = request.POST.get('designation')
        designation = Designation.objects.get(id=designation_id)

        department_id = request.POST.get('department')
        department = Department.objects.get(id=department_id)

        subdepartment_id = request.POST.get('subdepartment')
        subdepartment = Subdepartment.objects.get(id=subdepartment_id)

        jobtitle_id = request.POST.get('jobtitle')
        jobtitle = Job.objects.get(id=jobtitle_id)

        wrklcn_id = request.POST.get('wrklcn')
        wrklcn = Worklocation.objects.get(id=wrklcn_id)
        m.update(empid=empid, datejoin=joindate, probperiod=proper, emptype=etyp, wrklcn=wrklcn, status=empstatus, wrkexp=wrkex,
                 designation=designation, jobtitle=jobtitle, department=department, subdepartment=subdepartment)
        return redirect('view_work', id)

    return render(request, 'index/empworkdetails.html')



@require_http_methods(["GET", "POST"])
def view_team(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id

    # ----- DELETE Direct Report -----
    delete_id = request.GET.get('delete')
    if delete_id:
        try:
            report_relation = Reportingmanager.objects.get(userid=delete_id, myuser_2=data.id)
            report_relation.delete()
            messages.success(request, "Employee removed successfully.",extra_tags='bg-success text-white')
            return redirect('view_team', id=id)
        except Reportingmanager.DoesNotExist:
            messages.error(request, "Employee not found.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    if request.method == 'POST':
        selected_ids = request.POST.getlist('direct_report_user[]')
        override = request.POST.get('override', 'false') == 'true'
        conflicted_users = []

        for dr_id in selected_ids:
            existing_relations = Reportingmanager.objects.filter(userid=dr_id).exclude(myuser_2=data.id)
            if existing_relations.exists():
                user = User.objects.filter(id=dr_id).first()
                if user:
                    conflicted_users.append(user.username)

        if conflicted_users and not override:
            return JsonResponse({
                'conflict': True,
                'conflicted_users': conflicted_users
            })

        success_count = 0
        for dr_id in selected_ids:
            try:
                if override:
                    Reportingmanager.objects.filter(userid=dr_id).exclude(myuser_2=data.id).delete()

                existing_primary = Reportingmanager.objects.filter(userid=dr_id, type="Primary").exists()
                assigned_type = "Direct"
                if not existing_primary:
                    assigned_type = "Primary"

                if not Reportingmanager.objects.filter(userid=dr_id, myuser_2=data.id).exists():
                    new_report = Reportingmanager.objects.create(userid=dr_id, type=assigned_type)
                    new_report.myuser_2.add(data)
                    success_count += 1
            except Exception as e:
                continue

        return JsonResponse({
            'success': True,
            'added': success_count
        })

    # Profile + Team Data
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=request.user.id)
    user_id = request.user.id
    users = User.objects.filter(status='Active').exclude(id=data.id)
    reporting_managers = Reportingmanager.objects.filter(userid=data.id)

    reporting_manager_user = None
    primary_manager_id = None
    primary_manager_user_id = None
    secondary_manager_id = None
    secondary_manager_user_id = None
    primary_manager_user = None
    secondary_manager_user = None 
    primary_manager_type = None
    primary_manager_department = None
    primary_manager_designation = None
    secondary_manager_type = None
    secondary_manager_department = None
    secondary_manager_designation = None

    for reporting_manager in reporting_managers:
        if reporting_manager.type == 'Primary':
            primary_manager_user = reporting_manager.myuser_2.first()
            primary_manager_id = reporting_manager.id
            primary_manager_user_id = primary_manager_user.id if primary_manager_user else None

            if primary_manager_user:
                primary_manager_department = primary_manager_user.department.name if primary_manager_user.department else ""
                primary_manager_designation = primary_manager_user.designation.name if primary_manager_user.designation else ""
                primary_manager_type = reporting_manager.type

        elif reporting_manager.type == 'Secondary':
            secondary_manager_user = reporting_manager.myuser_2.first()
            secondary_manager_id = reporting_manager.id
            secondary_manager_user_id = secondary_manager_user.id if secondary_manager_user else None

            if secondary_manager_user:
                secondary_manager_department = secondary_manager_user.department.name if secondary_manager_user.department else ""
                secondary_manager_designation = secondary_manager_user.designation.name if secondary_manager_user.designation else ""
                secondary_manager_type = reporting_manager.type

    rpt_users = Reportingmanager.objects.filter(myuser_2=data.id)
    directreport_users = [
        User.objects.get(id=r.userid) 
        for r in rpt_users 
        if User.objects.filter(id=r.userid, status="Active").exists()
    ]
    reporting_manager_ids = Reportingmanager.objects.filter(userid=data.id).values_list('myuser_2__id', flat=True)
    directreport_users = [u for u in directreport_users if u.id not in reporting_manager_ids]

        # Build a mapping of report.id -> manager type (Primary/Secondary)
    report_manager_types = {}
    for report in directreport_users:
        relation = Reportingmanager.objects.filter(userid=report.id, myuser_2=data.id).first()
        if relation:
            report_manager_types[report.id] = relation.type


    already_ids = [u.id for u in directreport_users]
    eligible_users = User.objects.exclude(id__in=already_ids + list(reporting_manager_ids) + [data.id]).filter(status="Active")

    # Check if trying to inactivate a user with direct reports
    show_inactive_warning = False
    if request.method == 'POST' and request.POST.get('Empstatus') == 'Inactive':
        primary_directs = Reportingmanager.objects.filter(myuser_2=data.id, type='Primary')
        if primary_directs.exists():
            show_inactive_warning = True
            messages.warning(request, f"The user '{data.username}' has {primary_directs.count()} employee(s) as their direct reports. Please reassign or remove them before marking as Inactive.")

    context = {
        "data": data,
        "users": users,
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "reporting_managers": reporting_managers,
        "primary_manager": primary_manager_user.username if primary_manager_user else "",
        "primary_manager_id": primary_manager_id,
        "secondary_manager_id": secondary_manager_id,
        "primary_manager_type": primary_manager_type,
        "primary_manager_department": primary_manager_department,
        "primary_manager_designation": primary_manager_designation,
        "secondary_manager": secondary_manager_user.username if secondary_manager_user else "",
        "secondary_manager_type": secondary_manager_type,
        "secondary_manager_department": secondary_manager_department,
        "secondary_manager_designation": secondary_manager_designation,
        "primary_manager_user_id": primary_manager_user_id,
        "secondary_manager_user_id": secondary_manager_user_id,
        "directreport_users": directreport_users,
        "eligible_users": eligible_users,
        "show_inactive_warning": show_inactive_warning,
        "report_manager_types": report_manager_types,
    }
    return render(request, "index/empteamdetails.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active']) 
def adminadd_reportingmanager(request, id):
    data = get_object_or_404(User, pk=id)
    direct_reports = Reportingmanager.objects.filter(myuser_2=data).values_list('userid', flat=True)
    existing_rm_ids = Reportingmanager.objects.filter(userid=data.id).values_list('myuser_2__id', flat=True)

    users = User.objects.exclude(id=data.id).exclude(id__in=direct_reports).exclude(id__in=existing_rm_ids).distinct()

    if request.method == 'POST':
        name_id = request.POST.get('Name1')
        emptype = request.POST.get('Type')

        if name_id and emptype:
            user = User.objects.get(id=name_id)

            existing_primary = Reportingmanager.objects.filter(myuser_2=user, type='Primary', userid=data.id).exists()
            existing_secondary = Reportingmanager.objects.filter(myuser_2=user, type='Secondary', userid=data.id).exists()
            existing_type = Reportingmanager.objects.filter(userid=data.id, type=emptype).exists()

            if existing_type:
                messages.warning(request, f"A {emptype} reporting manager is already assigned.") 
            else:
                if emptype == 'Primary':
                    if existing_primary:
                        messages.warning(request, f"{user.username} is already added as a Primary reporting manager.")
                    elif existing_secondary:
                        messages.warning(request, f"{user.username} is already assigned as a Secondary reporting manager.") 
                    else:
                        reporting_manager = Reportingmanager.objects.create(userid=data.id, type=emptype)
                        reporting_manager.myuser_2.add(user) 
                        reporting_manager.save()
                        messages.success(request, f"{user.username} added as Primary reporting manager.",extra_tags='bg-success text-white') 

                elif emptype == 'Secondary':
                    if existing_secondary:
                        messages.warning(request, f"{user.username} is already added as a Secondary reporting manager.") 
                    elif existing_primary:
                        messages.warning(request, f"{user.username} is already assigned as a Primary reporting manager.") 
                    else:
                        reporting_manager = Reportingmanager.objects.create(userid=data.id, type=emptype) 
                        reporting_manager.myuser_2.add(user)
                        reporting_manager.save()
                        messages.success(request, f"{user.username} added as Secondary reporting manager.",extra_tags='bg-success text-white') 
                        
        return redirect('view_team', id=data.id)

    existing_reports = Reportingmanager.objects.filter(userid=data.id)

    context = {
        'users': users,
        'data': data,
        
        'l': existing_reports,
        'messages': messages.get_messages(request), 
    }

    return render(request, "index/empteamdetails.html", context)


def admindelete_reportingmanager(request, id):
    reporting_manager = get_object_or_404(Reportingmanager, id=id)
    reporting_manager.delete()

    return redirect('view_team', id=reporting_manager.userid)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_education(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    l = Educationalinfo.objects.filter(myuser_4__id=data.id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "is_view_education": True
    }
    return render(request, "index/eduinfomypro.html", {'data': data, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_add_empeducation(request, uid_4):
    if request.method == "POST":
        quali = request.POST.get("Qualification")
        course = request.POST.get("Course")
        insti = request.POST.get("Institutename")
        psout = request.POST.get("Passout")
        perc = request.POST.get("Percentage")
        u_4 = User.objects.get(id=uid_4)
        Educationalinfo.objects.create(
            qualification=quali, course=course, institute=insti, passout=psout, percent=perc, myuser_4=u_4)
    return redirect('view_education', id=uid_4)


def admin_edit_empeducation(request, id):
    data = get_object_or_404(User, pk=id)
    if request.method == "POST":
        qualifcn = request.POST.get("Qualification1")
        cours = request.POST.get("Course1")
        institn = request.POST.get("Institutename1")
        psouty = request.POST.get("Passout1")
        perctge = request.POST.get("Percentage1")
        edu_id = request.POST.get('edu_id')
        x = Educationalinfo.objects.filter(id=edu_id)
        x.update(qualification=qualifcn, course=cours,
                 institute=institn, passout=psouty, percent=perctge)
    return redirect('view_education', data.id)


def admin_delete_empeducation(request, id):
    edu = get_object_or_404(Educationalinfo, id=id)
    user_id = edu.myuser_4.id
    edu.delete()
    return redirect('view_education', id=user_id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_family(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    l = Familymembers.objects.filter(myuser_5=data.id)
    x = Emergencycontact.objects.filter(myuser_6=data.id)
    y = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "x": x,
        "is_view_family": True
    }
    return render(request, "index/familymypro.html", {'data': data, **y})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_add_empfamily(request, uid_5):
    if request.method == "POST":
        nm3 = request.POST.get("Name3")
        reln = request.POST.get("Relationship")
        dob1 = request.POST.get("DOB1")
        dpndnt = request.POST.get("Dependant")
        u_5 = User.objects.get(id=uid_5)
        Familymembers.objects.create(
            name3=nm3, relation=reln, dob1=dob1, dependant=dpndnt, myuser_5=u_5)
    return redirect('view_family', id=uid_5)


def admin_edit_empfamily(request, id):
    data = get_object_or_404(User, pk=id)
    if request.method == "POST":
        nm4 = request.POST.get("Name6")
        reln1 = request.POST.get("Relationship1")
        dob2 = request.POST.get("DOB2")
        dpndnt1 = request.POST.get("Dependant1")
        fam_member_id = request.POST.get('fam_memberid')

        d = Familymembers.objects.filter(id=fam_member_id)
        d.update(name3=nm4, relation=reln1, dob1=dob2, dependant=dpndnt1)
    return redirect('view_family', data.id)


def admin_delete_empfamily(request, id):
    fam = get_object_or_404(Familymembers, id=id)
    user_id = fam.myuser_5.id
    fam.delete()
    return redirect('view_family', id=user_id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_add_emergencycontact(request, uid_6):
    if request.method == "POST":
        nm4 = request.POST.get("Name4")
        reln1 = request.POST.get("Relationship1")
        phn1 = request.POST.get("Phone1")
        u_6 = User.objects.get(id=uid_6)
        Emergencycontact.objects.create(
            name4=nm4, relation1=reln1, phone1=phn1, myuser_6=u_6)
    return redirect('view_family', id=uid_6)


def admin_edit_emergencycontact(request, id):
    data = get_object_or_404(User, pk=id)
    if request.method == "POST":
        nm5 = request.POST.get("Name5")
        reln2 = request.POST.get("Relationship2")
        phn2 = request.POST.get("Phone2")

        emergency_id = request.POST.get('emergencyid')
        b = Emergencycontact.objects.filter(id=emergency_id)
        b.update(name4=nm5, relation1=reln2, phone1=phn2)
    return redirect('view_family', data.id)


def admin_delete_emergencycontact(request, id):
    emg = get_object_or_404(Emergencycontact, id=id)
    user_id = emg.myuser_6.id
    emg.delete()
    return redirect('view_family', id=user_id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_document(request, id):
    datas = User.objects.all()
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    user_id = request.user.id

    uploads = Uploadeddocs.objects.filter(myuser__id=id)
    proofs_object = Proof.objects.all()
    proofs = {
        proof.id: proof.proof_name for uploaded_doc in uploads for proof in uploaded_doc.proof.all()}

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "uploads": uploads,
        "proofs": proofs,
        "proofs_object": proofs_object
    }
    return render(request, "index/empdocumentdetails.html", {'data': data, 'datas': datas, **x})

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_deletedocument(request, id):
    document = get_object_or_404(Uploadeddocs, id=id)
    if request.user == document.myuser or request.user.is_staff:
        document.delete() 
    return redirect('view_document', id=document.myuser.id)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def doc_verification(request, id):
    if request.method == 'POST':
        document_id = request.POST.get('docid')
        document = Uploadeddocs.objects.get(id=document_id)
        document.verificationstatus = 'Verified'
        document.save()
        return redirect('view_document', id)
    return render(request, "index/empdocumentdetails.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_certification(request, id):
    datas = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    company = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser=admin_id)
    certificates = Certifications.objects.filter(myuser_8__id=id)
    x = {
        "company": company[0] if company.exists() else company,
        "profile": profile[0] if profile.exists() else profile,
        "certificates": certificates,
        "is_view_certification": True
    }
    return render(request, "index/certifications.html", {'datas': datas, **x})

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def delete_certification(request, cert_id):
    certification = get_object_or_404(Certifications, pk=cert_id)
    if request.user == Certifications.myuser_8 or request.user.is_staff:  
        certification.delete()
    

    return redirect('view_certification',id=certification.myuser_8.id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])   
def cert_verification(request, id):
    data = get_object_or_404(User, pk=id)
    if request.method == 'POST':
        cert_id = request.POST.get('certid')
        cert = Certifications.objects.get(id=cert_id)
        cert.verification = 'Verified'
        cert.save()
        return redirect('view_certification', data.id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_docwork(request, id):
    datas = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    company = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser=admin_id)
    work = Work.objects.filter(myuser_9__id=id)

    for item in work:
        if item.uploadedon:
            try:
                uploadedon_date = datetime.strptime(item.uploadedon, '%d-%m-%Y')
                item.uploadedon_formatted = uploadedon_date.strftime('%d %B %Y')
            except ValueError:
                item.uploadedon_formatted = None
        else:
            item.uploadedon_formatted = None
    x = {
        "company": company[0] if company.exists() else company,
        "profile": profile[0] if profile.exists() else profile,
        "work": work,
        "is_view_docwork": True
    }
    return render(request, "index/doc_work.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_workweek(request, id):
    datas = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    work_week = AssignWorkWeek.objects.filter(user_id=id)
    x = {
        "data": data[0] if data.exists() else data,
        "k": k[0] if k.exists() else k,
        "work_week": work_week,
        "is_view_workweek": True
    }
    return render(request, "index/workweekmypro.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_payroll(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    regaddress = registeredaddress.objects.filter(admin_id=admin_id)
    print("regofficeaddress ################### :", regaddress)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
    print("selected_year :", selected_year, selected_month)

    bank_details = Bank_account.objects.filter(myuser_11=id)
    print("bank_details :", bank_details)

    assign_salarystructure = AssignSalaryStructure.objects.filter(
        user_id=id, effective_date__month=selected_month, effective_date__year=selected_year).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)

    assign_data = []
    assigndata = []
    ctc_assigndata = []
    gross_salary_amount = 0
    work_from_office_allowance_amount = 0
    total_net_salary = 0
    total_ctc_salary = 0
    wfocount = 0
    leave_count = 0
    total_gross_salary = 0

    selected_date = datetime(selected_year, selected_month, 1)
    print("selected_date :", selected_date)

    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(
            effective_date__lte=selected_date, user_id=id).order_by('-effective_date').first()

        if nearest_date:
            assign_salarystructure = nearest_date

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
        work_from_office_component = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance", Parentcomponentname__componentname__iexact="Gross Salary").first()
        print("gross_salary_component ; work_from_office_component : ", gross_salary_component, work_from_office_component)
        net_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="Net Salary")
        print("net_salary_component : ", net_salary_component)
        ctc_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="CTC")
        print("ctc_salary_component : ", ctc_salary_component)

        name = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
        amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=name)
        print("name ; amount 1st :", name, amount)
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname=gross_salary_component)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        print("names ; amounts  :", names, amounts)

        net_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="Net Salary")
        net_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=net_names)
        print("names ; amounts  :", net_names, net_amounts)

        ctc_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="CTC").exclude(salarycomponent__componentname__icontains="professional tax")
        ctc_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=ctc_names)
        print("ctc_names ; ctc_amounts  :", ctc_names, ctc_amounts)

        if gross_salary_component:
            gross_amount = amount.filter(salaryname__salarycomponent=gross_salary_component).first()
            gross_salary_amount = gross_amount.amount if gross_amount else 0            
            
        if work_from_office_component:
            work_amount = amounts.filter(salaryname__salarycomponent=work_from_office_component).first()
            work_from_office_allowance_amount = work_amount.amount if work_amount else 0
        
        for netsalry in net_salary_component:
            net_salary = amount.filter(salaryname__salarycomponent=netsalry)
            total_net_salary += net_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total Net Salary:", total_net_salary)

        for ctcsalry in ctc_salary_component:
            if ctcsalry.componentname.lower() != 'professional tax':
                ctc_salary = amount.filter(salaryname__salarycomponent=ctcsalry)
                total_ctc_salary += ctc_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total CTC Salary:", total_ctc_salary)

        total_gross_salary = gross_salary_amount - work_from_office_allowance_amount
        print("total_gross_salary : ", total_gross_salary)

        zipped_data = zip_longest(names, amounts)
        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })

        zippeddata = zip_longest(net_names, net_amounts)
        assigndata.append({
            'rule': rule,
            'zippeddata': zippeddata,
        })

        ctc_zippeddata = zip_longest(ctc_names, ctc_amounts)
        ctc_assigndata.append({
            'rule': rule,
            'ctc_zippeddata': ctc_zippeddata,
        })

    punch_obj = Punch.objects.filter(user__id=id,date__year=selected_year,date__month=selected_month)
    print("punch_obj : ", punch_obj)

    for punch in punch_obj:
        if punch.WfhOrWfo == "WFO":
            wfocount += 1
                    
        leave_data = Leave.objects.filter(
                applicant_email=id,  
                strtDate=punch.date,    
                status="Approved"      
            ).first()       
        print(f"Leave data for {id} on {punch.date}: ", leave_data)
        
        if punch.status == "H":
            leave_count += 1
        elif punch.status == "L":
            if leave_data:
                if leave_data.leavetyp != "Loss Of Pay":
                    leave_count += 1
        elif punch.status == "HL":
            if leave_data:
                print("ccccccccccccccccccccccccccccccccccccc  ")
                if leave_data.leavetyp == "Loss Of Pay":
                    leave_count -= 0.5
                    print("KKKKKKKKKKKKKKKKKKKKK", leave_count)
                
    wfo_count = wfocount + leave_count
    print("wfo_count :", wfo_count, "month_numeric , selected_year:" , selected_month, selected_year)

    year_select = int(selected_year)
    num_days = calendar.monthrange(year_select, selected_month)[1]
    first_day_of_month = datetime(year_select, selected_month, 1)
    if selected_month == 12: 
        next_month = datetime(year_select + 1, 1, 1)
        print("next_month 1 :", next_month, year_select)
    else:
        next_month = datetime(year_select, selected_month + 1, 1)
        print("next_month 2 :", next_month)

    day_count = 0
    current_day = first_day_of_month
    while current_day < next_month:
        if current_day.weekday() != 6: 
            day_count += 1
        current_day += timedelta(days=1)
    count_sundays = num_days - day_count
    print("day_count ############ :", day_count, num_days, count_sundays)

    print("work_from_office_allowance_amount:", work_from_office_allowance_amount)
    perday_WFOamount = work_from_office_allowance_amount / day_count
    total_WFOamount = perday_WFOamount * wfo_count
    print("total_WFOamount : ", wfo_count, total_WFOamount)
    
    WFOamount = round(work_from_office_allowance_amount - total_WFOamount) #This amount add to the deduction
    print("WFOamount :", WFOamount)

    adhoc_data = Adhoc.objects.filter(user_id=id, createddate__year=selected_year,
                                      createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
    print("adhoc_data : ", adhoc_data)

    earning_amount = 0
    deduction_amount = 0
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    total_earnings = gross_salary_amount + earning_amount
    print("total_earnings ; gross_salary_amount ; earning_amount :", total_earnings, gross_salary_amount, earning_amount)

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    total_anomaly_count = 0
    attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=id)
    print("attendance_rule :", attendance_rule)
    for att_rule in attendance_rule:
        rule_type = att_rule.rules_applied
        print("rule_type :", rule_type, )
        if rule_type:
            full_day_hours = rule_type.fullhours
            full_day_minutes = rule_type.fullminutes
            full_time = timedelta(hours=full_day_hours,
                                  minutes=full_day_minutes)
            half_day_hours = rule_type.halfhours
            half_day_minutes = rule_type.halfminutes
            half_time = timedelta(hours=half_day_hours,
                                  minutes=half_day_minutes)
            print("Full Day Hours:", full_day_hours,
                  full_day_minutes, full_time)
            print("Half Day Hours:", half_day_hours,
                  half_day_minutes, half_time)
            in_grace_period = rule_type.inGracePeriod
            out_grace_period = rule_type.outGracePeriod
            print("Grace period:", in_grace_period, out_grace_period)
            in_grace_timedelta = timedelta(
                hours=in_grace_period.hour, minutes=in_grace_period.minute)
            out_grace_timedelta = timedelta(
                hours=out_grace_period.hour, minutes=out_grace_period.minute)

            total_grace_period = in_grace_timedelta + out_grace_timedelta
            print("Total Grace period:", total_grace_period)
            total_fullday_time = full_time + total_grace_period
            print("Total Time:", total_fullday_time)
            total_halfday_time = half_time + total_grace_period
            print("total_halfday_time :", total_halfday_time)

    num_days = calendar.monthrange(selected_year, selected_month)[1]

    payregister = PayRegister.objects.filter(createddate__month=selected_month, createddate__year=selected_year, user_id=id, status__in=[
                                             "Payslip Generated", "Payslip Downloaded"])
    print("payregister :", payregister)

    lop_data = Runpayroll_lop.objects.filter(
        lop_date__month=selected_month, lop_date__year=selected_year, user_id=id)
    lopcount = 0
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    punches = Punch.objects.filter(
        user__id=id,
        date__year=selected_year,
        date__month=selected_month,
        status="AN", is_penalty_reverted=False
    )
    print("Punch Object :", punches)

    for punch in punches:
        total_work_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)
            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

        elif punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration
        if total_work_duration > total_fullday_time:
            AN_count = 0.5
        elif total_work_duration < total_halfday_time:
            AN_count = 1.0
        else:
            AN_count = 0.5

        total_anomaly_count += AN_count
        print("total_anomaly_count :", total_anomaly_count)

    print("total_anomaly_count 2:", total_anomaly_count)
    absent_count = Punch.objects.filter(user__id=id, date__year=selected_year,
                                        date__month=selected_month, status='A', is_penalty_reverted=False).count()

    absent_AN_count = absent_count + total_anomaly_count
    print("absent_AN_count : ", absent_count, absent_AN_count)

    punchcount = Punch.objects.filter(user__id=id, date__year=selected_year, date__month=selected_month).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count, absent_AN_count, lopcount)
    total_lop = absent_AN_count + missing_date_count + lopcount

    per_day_amount = total_gross_salary / num_days
    print("per_day_amount :", per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount)
    lopamount = round(lop_amount)

    total_deductions = deduction_amount + lop_amount + total_net_salary + WFOamount
    totaldeductions = round(total_deductions)
    net_amount = round(total_earnings - total_deductions)
    print("net_amount :", net_amount)

    net_amount_words = num2words(net_amount, lang='en_IN')

    print("ctc_assigndata : ", ctc_assigndata)

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "regaddress": regaddress,
        "payregister": payregister,
        "num_days": num_days,
        "bank_details": bank_details,
        "assign_salarystructure": assign_salarystructure,
        "assign_data": assign_data,
        "assigndata": assigndata,
        "ctc_assigndata": ctc_assigndata,
        "adhoc_data": adhoc_data,
        "total_earnings": total_earnings,
        "total_ctc_salary": total_ctc_salary,
        "totaldeductions": totaldeductions,
        "total_lop": total_lop,
        "lopamount": lopamount,
        "wfo_count": wfo_count, 
        "WFOamount": WFOamount,
        "net_amount": net_amount,
        "net_amount_words": net_amount_words,
        "month_str": month_str,
        "selected_year": selected_year,
        "is_view_payroll": True
    }
    return render(request, "index/pay_slip.html", {'data': data, **context})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_salarystructure(request, id):
    user_id = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    assign = AssignSalaryStructure.objects.filter(user_id=user_id).last()

    assign = AssignSalaryStructure.objects.filter(user_id=user_id).last()

    hierarchy = defaultdict(list)  

    if assign:
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        amount_map = {amount.salaryname.salarycomponent.first().id: amount.amount for amount in amounts}
        print("amount_map : ", amount_map)
        for name in names:
            for component in name.salarycomponent.all():
                parent_name = component.Parentcomponentname.componentname if component.Parentcomponentname else None
                monthly_amount = amount_map.get(component.id, 0)
                component_data = {
                    'name': component.componentname,
                    'monthly_amount': monthly_amount,
                    'annual_amount': monthly_amount * 12,
                    'is_parent': parent_name is None
                }
                print("parent_name : ", parent_name)
                print("component_data : ", component_data)
                if parent_name:
                    hierarchy[parent_name].append(component_data)
                    print("hierarchy[parent_name] : ",  hierarchy[parent_name])
                else:
                    hierarchy[component.componentname].append(component_data)
                    print("hierarchy[component.componentname] : ",  hierarchy[component.componentname])

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "is_view_salarystructure": True,

    }

    return render(
        request,
        'index/salarystructure.html',
        {
            'user_id': user_id,
            'assign': assign,
            'hierarchy': dict(hierarchy),
            **x
        }
    )


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_declaration(request, id):
    datas = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "is_view_declaration": True
    }
    return render(request, "index/declaration.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_bankaccount(request, id):
    datas = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    # bankdatas = Bank_account.objects.filter(myuser_11__id=datas.id)
    try:
        bankdatas = Bank_account.objects.get(myuser_11=datas)
    except Bank_account.DoesNotExist:
        return redirect('admin_update_bankaccount', id=id)

    x = {
        "data": data[0] if data.exists() else data,
        "k": k[0] if k.exists() else k,
        'bankdatas': bankdatas,
    }
    return render(request, "index/empbankaccount.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_update_bankaccount(request, id):
    datas = User.objects.get(id=id)
    data = companyprofile.objects.filter(admin_id=request.user.id)
    k = Myprofile.objects.filter(myuser=request.user.id)
    try:
        bankdatas = Bank_account.objects.get(myuser_11=datas)
    except Bank_account.DoesNotExist:
        bankdatas = Bank_account(myuser_11=datas)

    if request.method == 'POST':
        acc_holder = request.POST['name']
        accnt_no = request.POST['acctno']
        bnk_name = request.POST['bankname']
        branch = request.POST['branchname']
        ifsc = request.POST['ifsccode']
        cty = request.POST['city']
        esa = request.POST['esa']
        pfnum = request.POST['pfnum'] 

        bankdatas.account_holder_name = acc_holder
        bankdatas.account_number = accnt_no
        bankdatas.bank_name = bnk_name
        bankdatas.branch_name = branch
        bankdatas.IFSC_code = ifsc
        bankdatas.city = cty
        bankdatas.esa = esa
        bankdatas.pfnum = pfnum
        bankdatas.save()
        return redirect('view_bankaccount', id=id)

    x = {
        "data": data[0] if data.exists() else data,
        "k": k[0] if k.exists() else k,
    }
    return render(request, "index/empbankaccount.html", {'datas': datas, **x})


def view_e_exit(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    resignation = ResignationForm.objects.filter(user=data.id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
    }
    return render(request, "Employee/e_exit.html", {'data': data, 'resignation': resignation, **x})


def view_filemanager(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "is_view_filemanager": True
    }
    return render(request, "Employee/filemanagermypro.html", {'data': data, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_daily_log(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)

    leave_notification = LeaveNotification.objects.filter(
        Q(user=admin_id) | Q(user_id__admin_id=admin_id))
    notifications = []

    for i in leave_notification:
        if i.admin_id == 0:
            profile = Myprofile.objects.filter(myuser=i.user).first()
        else:
            profile = Myprofile.objects.filter(myuser__id=i.admin_id).first()

        notifications.append(
            {
                "message": i.message,
                "image_url": profile.image.url
                if profile and profile.image
                else "/static/logo/userlogo.png",
                "notification_id": i.id,
            }
        )

    attendance_rule_obj = (
        AssignAttendanceRule.objects.filter(user_id=id)
        .values_list("rules_applied__inTime", "rules_applied__outTime")
        .first()
    )

    today = timezone.now().strftime("%d")
    selected_date = request.GET.get("selected_date", None)

    if attendance_rule_obj:
        rule_in_time, rule_out_time = attendance_rule_obj
        rule_in_time = datetime.combine(datetime.now().date(), rule_in_time)
        rule_out_time = datetime.combine(datetime.now().date(), rule_out_time)

        if selected_date is None:
            selected_date = timezone.now().date()
        else:
            selected_date = datetime.strptime(
                selected_date, '%d %B %Y').date()

        print('Selected Date: ', selected_date)
        regular_time_duration = rule_out_time - rule_in_time
        punch_datas = Punch.objects.filter(
            Q(user__id=id) & Q(date__date=selected_date)
        )
        print('Punch datas: ', punch_datas)
        punch_list = []
        json_data = {
            "org_in_time": rule_in_time,
            "org_out_time": rule_out_time,
            "custom_punch_logs": [],
            "break_duration": "--",
            "work_duration": "--",
            "overtime_duration": "--",
            "status": "",
        }

        anomaly_type = []
        work_duration = timedelta()
        break_duration = timedelta()

        for punch in punch_datas:
            anomaly_type.append('Clock In' if punch.in_time_anomaly else None)
            anomaly_type.append(
                'Clock Out' if punch.out_time_anomaly else None)
            anomaly_type.append(
                'Work Duration' if punch.work_duration_anomaly else None)

            for suffix in ["first", "second"]:
                in_time_key = f"{suffix}_clock_in_time"
                out_time_key = f"{suffix}_clock_out_time"
                is_clocked_in_key = f"is_{suffix}_clocked_in"
                is_clocked_out_key = f"is_{suffix}_clocked_out"

                in_time_value = (
                    getattr(punch, in_time_key, "--")
                    if getattr(punch, is_clocked_in_key, False)
                    else "--"
                )
                out_time_value = (
                    getattr(punch, out_time_key, "--")
                    if getattr(punch, is_clocked_out_key, False)
                    else "--"
                )

                if in_time_value != "--":
                    punch_list.append(
                        {
                            "in_time": in_time_value,
                            "in_type": "In Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "in_time": in_time_value,
                        "in_type": "In Time",
                        "ip_address": punch.ip_address,
                    })

                if out_time_value != "--":
                    punch_list.append(
                        {
                            "out_time": out_time_value,
                            "out_type": "Out Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "out_time": out_time_value,
                        "out_type": "Out Time",
                        "ip_address": punch.ip_address,
                    })

            # print(punch.is_first_clocked_in ,punch.is_first_clocked_out ,punch.is_second_clocked_in , sep='\n')
            if punch.is_first_clocked_in and punch.is_second_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_duration = work_end_time - work_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                work_duration = work_end_time - work_start_time
            if punch.is_first_clocked_out and punch.is_second_clocked_in and punch.is_second_clocked_out:
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                break_duration = break_end_time - break_start_time

                if work_duration >= break_duration:

                    work_duration = work_duration - break_duration
                else:

                    work_duration = break_duration-work_duration

            json_data['status'] = punch.status

        work_duration_str = str(work_duration).split(".")[0]
        try:
            work_duration = (
                datetime.strptime(work_duration_str, "%H:%M:%S")
                if work_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            work_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")

        break_duration_str = str(break_duration)
        try:
            break_duration = (
                datetime.strptime(
                    str(break_duration_str).split(".")[0], "%H:%M:%S")
                if break_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            break_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")
        work_duration_minutes = work_duration.hour * 60 + work_duration.minute
        regular_time_minutes = regular_time_duration.seconds // 60
        overtime_duration = max(
            work_duration_minutes - regular_time_minutes, 0)
        over_time_duration = timedelta(minutes=overtime_duration)

        if overtime_duration != "00:00:00:" or overtime_duration != None:
            over_time_duration = datetime.strptime(
                str(over_time_duration).split(".")[0], "%H:%M:%S")

        json_data['break_duration'] = break_duration
        json_data['work_duration'] = work_duration
        json_data['overtime_duration'] = over_time_duration

        for punch in punch_datas:
            punch.work_duration = work_duration
            punch.break_duration = break_duration
            punch.overtime = over_time_duration
            punch.save()

        json_data['org_in_time'] = json_data['org_in_time'].isoformat()
        json_data['org_out_time'] = json_data['org_out_time'].isoformat(

        )
        json_data['custom_punch_logs'] = [
            {
                'time_type': 'in',
                'time': log['in_time'].isoformat() if log.get('in_time') is not None else None,
                'type': log.get('in_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'in_time' in log and log.get('in_time') is not None
        ] + [
            {
                'time_type': 'out',
                'time': log['out_time'].isoformat() if log.get('out_time') is not None else None,
                'type': log.get('out_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'out_time' in log and log.get('out_time') is not None
        ]

        json_data['break_duration'] = json_data['break_duration'].isoformat()
        json_data['work_duration'] = json_data['work_duration'].isoformat()
        json_data['overtime_duration'] = str(json_data['overtime_duration'])

        # Serialize the data
        serialized_json_data = json.dumps(json_data)
        # serialized_json_data = serialize('json',json_data , use_natural_keys=True)
        context = {
            "punch_list": punch_list,
            "work_duration": work_duration,
            "break_duration": break_duration,
            "overtime_duration": over_time_duration,
            "json_data": serialized_json_data,
            "anomaly_type": anomaly_type,
            "selected_date": selected_date,
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            "data": data,
        }
        return render(request, "index/empdailylogdetails.html", context)

    else:
        return render(request, "index/empdailylogdetails.html", {"notifications": notifications, 'data': data})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_monthly_log(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=admin_id)

    page = request.GET.get('page', 1)
    # Fetch attendance rule

    assignattendancerule = AssignAttendanceRule.objects.filter(user_id=id).values_list(
        "rules_applied__inTime", "rules_applied__outTime").first()

    selected_month = request.POST.get('month_range', None)
    if assignattendancerule:
        rule_in_time, rule_out_time = assignattendancerule
        rule_in_time = datetime.combine(datetime.min, rule_in_time)
        rule_out_time = datetime.combine(datetime.min, rule_out_time)
        regular_time_duration = rule_out_time - rule_in_time

        current_date = timezone.now()
        current_year = current_date.year
        start_date = current_date - timedelta(45)

        if selected_month:
            current_date = datetime.strptime(selected_month.split(
                ' - ')[1], '%d/%m/%Y').strftime('%Y-%m-%d')
            start_date = datetime.strptime(selected_month.split(
                ' - ')[0], '%d/%m/%Y').strftime('%Y-%m-%d')

        punch_datas = Punch.objects.filter(
            Q(user__id=id) & Q(
                date__date__range=(start_date, current_date))
        ).order_by('-date', '-first_clock_in_time')

        punch_collections = []
        for punch in punch_datas:
            work_duration = timedelta()
            break_duration = timedelta()

            if punch.is_first_clocked_in and punch.is_second_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)

                work_duration = work_end_time - work_start_time
                break_duration = break_end_time - break_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                work_duration = work_end_time - work_start_time

            try:
                if work_duration != '00:00:00':
                    work_duration = datetime.strptime(
                        str(work_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                work_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")
            try:
                if break_duration != "00:00:00":
                    break_duration = datetime.strptime(
                        str(break_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                break_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")

            work_duration_minutes = work_duration.hour * 60 + work_duration.minute
            regular_time_minutes = regular_time_duration.seconds // 60
            overtime_duration = max(
                work_duration_minutes - regular_time_minutes, 0)
            over_time_duration = timedelta(minutes=overtime_duration)
            if over_time_duration != '00:00:00':
                over_time_duration = datetime.strptime(
                    str(over_time_duration).split(".")[0], "%H:%M:%S")

            punch_records = {
                "id": punch.id,
                "date": punch.date,
                "status": punch.status,
                "in_time": punch.first_clock_in_time if punch.first_clock_in_time else '--',
                "out_time": punch.second_clock_out_time if punch.second_clock_out_time else (punch.first_clock_out_time or '--'),
                "work_duration": work_duration,
                "break_duration": break_duration,
                "overtime_duration": over_time_duration,
                "breaks": punch.break_count,
            }
            punch_collections.append(punch_records)

        paginator = Paginator(punch_collections, 10)
        try:
            punch_collections = paginator.page(page)
        except PageNotAnInteger:
            punch_collections = paginator.page(1)
        except EmptyPage:
            punch_collections = paginator.page(paginator.num_pages)

        context = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            'punch_in_month': punch_collections,
            "data": data,
        }
        return render(request, "index/empmonthlylog.html", context)
    else:
        return render(request, "index/empmonthlylog.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_automation_log(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect', None)
    page = request.GET.get('page', 1)
    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    penalty_log_data = PenaltyLogs.objects.filter(
        user=id, punch_data__date__month=selected_month).order_by("-punch_data__date")

    paginator = Paginator(penalty_log_data, 10)
    try:
        penalty_log_data = paginator.page(page)
    except PageNotAnInteger:
        penalty_log_data = paginator.page(1)
    except EmptyPage:
        penalty_log_data = paginator.page(paginator.num_pages)

    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        'automation_data': penalty_log_data,
        # "selected_date": selected_date,
        "is_view_automation_log": True
    }
    return render(request, "index/automation logs.html", {'data': data, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_comp_off(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    page = request.GET.get('page', 1)

    comp_off_data = CompOff.objects.filter(user__id=data.id)
    paginator = Paginator(comp_off_data, 10)
    print('Comp off data: ', comp_off_data)
    try:
        comp_off_data = paginator.page(page)
    except PageNotAnInteger:
        comp_off_data = paginator.page(1)
    except EmptyPage:
        comp_off_data = paginator.page(paginator.num_pages)

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "comp_off_data": comp_off_data,
        "is_view_comp_off": True
    }
    return render(request, "index/CompOff.html", {'data': data, **context})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_attendance_rule(request, id):
    data = get_object_or_404(User, pk=id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser=admin_id)
    rule = AssignAttendanceRule.objects.filter(user_id=data.id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
    }
    return render(request, "index/empattendancerule.html", {'data': data, 'rule': rule, **x})


def view_empwork(request, id):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.all()
    data = User.objects.all()
    datas = get_object_or_404(User, pk=id)
    img = get_object_or_404(Myprofile, pk=id)
    # datas = Reportingmanager.objects.all()
    # datas1 = Directreports.objects.all()

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/empwork_details.html", {'img': img, 'data': data, 'datas': datas, **x})


def view_empedu_details(request, id):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.all()
    datas = User.objects.all()
    datas1 = get_object_or_404(User, pk=id)
    data = get_object_or_404(Educationalinfo, pk=id)
    img = get_object_or_404(Myprofile, pk=id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/empedu_details.html",
                  {'img': img, 'data': data, 'datas': datas, 'datas1': datas1, **x})


def view_empfam_details(request, id):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.all()
    datas = User.objects.all()
    datas1 = get_object_or_404(User, pk=id)
    data = get_object_or_404(Familymembers, pk=id)
    # datas1 = Emergencycontact.objects.all()
    data1 = get_object_or_404(Emergencycontact, pk=id)
    # img=get_object_or_404(Myprofile,pk=id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/empfam_details.html",
                  {'data': data, 'data1': data1, 'datas': datas, 'datas1': datas1, **x})


def view_empdoc_details(request, id):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.all()
    datas = User.objects.all()
    datas1 = get_object_or_404(User, pk=id)
    data = get_object_or_404(Uploadeddocs, pk=id)
    datas1 = Certifications.objects.all()
    data1 = get_object_or_404(Certifications, pk=id)
    datas2 = Work.objects.all()
    data2 = get_object_or_404(Work, pk=id)
    img = get_object_or_404(Myprofile, pk=id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/empdoc_details.html",
                  {'img': img, 'data': data, 'data1': data1, 'data2': data2, 'datas': datas, **x})


def employee_search(request, id):
    # datas=User.objects.filter(id=id)
    data = get_object_or_404(User, pk=id)
    return render(request, 'Employee/empdash.html', {'data': data})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_employee(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=admin_id)
    print("Admin Id : ", admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    print(f"Company profiles: {c}") 
    dn = Designation.objects.filter(admin_id=admin_id)
    dp = Department.objects.filter(admin_id=admin_id)
    sd = Subdepartment.objects.filter(admin_id=admin_id)
    jb = Job.objects.filter(admin_id=admin_id)
    wr = Worklocation.objects.filter(admin_id=admin_id)
    rp = User.objects.filter(Q(id=request.user.id) |
                             Q(admin_id=request.user.id))
    empid = 10001 if User.objects.filter(admin_id=request.user.id).count() == 0 else \
        User.objects.filter(admin_id=request.user.id).aggregate(
            max=Max('empid'))["max"] + 1

    if request.method == 'POST':

        name = request.POST.get('username')
        em = request.POST.get('email')
        ph = request.POST.get('phone')
        gen = request.POST.get('gender')
        dob = request.POST.get('dob')
        empt = request.POST.get('emptype')
        pd = request.POST.get('probperiod')
        Joindate = request.POST.get('Joindate')
        status = 'Onboarding'
        role = 'Employee'
        otp = random.randint(000000, 999999)

        designation = request.POST.get('designation')
        des = Designation.objects.get(id=designation)
        print('designation', designation)

        department = request.POST.get('department')
        dep = Department.objects.get(id=department)
        print('department', department)

        jobtitle = request.POST.get('jobtitle')
        job = Job.objects.get(id=jobtitle)
        print('jobtitle', jobtitle)

        wrklcn = request.POST.get('wrklcn')
        wrk = Worklocation.objects.get(id=wrklcn)
        print('wrklcn', wrklcn)

        reptmgr = request.POST.get('reptmgr')

        subdepartment = request.POST.get('subdepartment')

        company_type_id = request.POST.get('company_type')
        ct = companyprofile.objects.get(id=company_type_id)

        
        

        user = None

        if User.objects.filter(email=em).exists():
            msg1 = 'Email already taken try another one'
            return render(request, "index/admin.html",
                          {'name': name, 'em': em, 'ph': ph, 'ph': ph, 'dob': dob, 'gen': gen, 'empt': empt, 'pd': pd,
                           'msg1': msg1, 'dn': dn, 'dp': dp, 'sd': sd, "jb": jb, "wr": wr, 'rp': rp, 'status': status})

        if User.objects.filter(phone=ph).exists():
            msg2 = 'Phone number already taken try another one'
            return render(request, "index/admin.html",
                          {'name': name, 'em': em, 'ph': ph, 'ph': ph, 'dob': dob, 'gen': gen, 'empt': empt, 'pd': pd,
                           'msg2': msg2, 'dn': dn, 'dp': dp, 'sd': sd, "jb": jb, "wr": wr, 'rp': rp, 'status': status})

        if reptmgr == '' and subdepartment != '':
            sub = Subdepartment.objects.get(id=subdepartment)
            user = User.objects.create(empid=empid, username=name, email=em, phone=ph, gender=gen, dob=dob,
                                       emptype=empt, status=status, role=role, otp=otp, password=otp,
                                       probperiod=pd, datejoin=Joindate, designation=des, department=dep,
                                       subdepartment=sub, jobtitle=job, wrklcn=wrk, admin_id=admin_id,company_type=ct)
        elif subdepartment == '' and reptmgr != '':
            rep = User.objects.get(id=reptmgr)
            user = User.objects.create(empid=empid, username=name, email=em, phone=ph, gender=gen, dob=dob,
                                       emptype=empt, status=status, role=role, otp=otp, password=otp, reptmgr=rep,
                                       probperiod=pd, datejoin=Joindate, designation=des, department=dep, jobtitle=job,
                                       wrklcn=wrk, admin_id=admin_id,company_type=ct)
        elif reptmgr == '' and subdepartment == '':
            user = User.objects.create(empid=empid, username=name, email=em, phone=ph, gender=gen, dob=dob,
                                       emptype=empt, status=status, role=role, otp=otp, password=otp,
                                       probperiod=pd, datejoin=Joindate, designation=des, department=dep, jobtitle=job,
                                       wrklcn=wrk, admin_id=admin_id,company_type=ct)
        elif reptmgr != '' and subdepartment != '':
            sub = Subdepartment.objects.get(id=subdepartment)
            rep = User.objects.get(id=reptmgr)
            user = User.objects.create(empid=empid, username=name, email=em, phone=ph, gender=gen, dob=dob,
                                       emptype=empt, status=status, role=role, otp=otp, password=otp, reptmgr=rep,
                                       probperiod=pd, datejoin=Joindate, designation=des, department=dep,
                                       subdepartment=sub, jobtitle=job, wrklcn=wrk, admin_id=admin_id,company_type=ct)
        
        if reptmgr:
            rep = User.objects.get(id=reptmgr)
            reporting_manager = Reportingmanager.objects.create(
                userid=user.id, type="Primary")
            reporting_manager.myuser_2.add(rep)

        today = datetime.now().strftime("%d %B %Y")
        print("today: ", today, user.id)
        attrule = AttendanceRule.objects.filter(user_id=admin_id).first()
        if attrule:
            AssignAttendanceRule.objects.create(
                user_id=user, rules_applied=attrule, effective_date=today)

        workweek = Workweek.objects.filter(admin_id=admin_id).first()
        if workweek:
            AssignWorkWeek.objects.create(
                user_id=user, rules_applied=workweek, effective_date=today)

        to = [user.email]

        subject = 'OTP For Login'
        html_body = render_to_string(
            'index/emp_otpemail.html', {'data': emailconfig, 'user': user, 'otp': otp})
        # ) # render with dynamic value
        # text_content = strip_tags(html_content) # Strip the html tag. So people can see the pure text at least.

        msg = EmailMultiAlternatives(
            subject=subject, from_email=settings.EMAIL_HOST_USER, to=to)
        msg.attach_alternative(html_body, "text/html")
        msg.send()
        messages.success(request, 'Created Successfully!',extra_tags='bg-success text-white')

        return redirect('directory')
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "company_profiles": c
    }

    return render(request, "index/admin.html", {'dn': dn, 'dp': dp, 'sd': sd, "jb": jb, "wr": wr, 'rp': rp, **x})


def update_employee(request, userid):
    # Retrieve the employee object based on the employee_id
    user_id = request.user.id
    employee = User.objects.get(id=userid)

    m = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=user_id)

    # Retrieve the required data for the form
    dn = Designation.objects.filter(admin_id=user_id)
    dp = Department.objects.filter(admin_id=user_id)
    sd = Subdepartment.objects.filter(admin_id=user_id)
    jb = Job.objects.filter(admin_id=user_id)
    wr = Worklocation.objects.filter(admin_id=user_id)
    
    
    all_employees = User.objects.filter(status__in=['Active', 'Onboarding']).exclude(id=employee.id)

    current_reptmgr = Reportingmanager.objects.filter(userid=employee.id,type='Primary').first()
    current_reporting_manager = None
    reporting_managers = Reportingmanager.objects.filter(userid=employee.id)
   
    x = {
        "m": m[0] if m.exists() else m,
        "c": c[0] if c.exists() else c,
    }

    if request.method == 'POST':
        # Retrieve the updated data from the form
        name = request.POST.get('username')
        em = request.POST.get('email')
        ph = request.POST.get('phone')
        gen = request.POST.get('gender')
        dob = request.POST.get('dob')
        status = request.POST.get('status')
        emptype = request.POST.get('emptype')
        pd = request.POST.get('probperiod')
        reptmgr_id = request.POST.get('reptmgr')
        designation = request.POST.get('designation')
        des = Designation.objects.get(id=designation)
        department = request.POST.get('department')
        dep = Department.objects.get(id=department)
        jobtitle = request.POST.get('jobtitle')
        job = Job.objects.get(id=jobtitle)
        wrklcn = request.POST.get('wrklcn')
        wrk = Worklocation.objects.get(id=wrklcn)
        reptmgr = request.POST.get('reptmgr')
        subdepartment = request.POST.get('subdepartment')
        Joindate = request.POST.get('Joindate')
        company_type_id = request.POST.get('company_type')
        if company_type_id:
            company_type = companyprofile.objects.get(id=company_type_id)
            employee.company_type = company_type  
        
       
        # Update the employee object with the new data
        employee.username = name
        employee.email = em
        employee.phone = ph
        employee.gender = gen
        employee.status = status
        employee.emptype = emptype
        employee.dob = dob
        employee.probperiod = pd
        employee.designation = des
        employee.department = dep
        employee.jobtitle = job
        employee.wrklcn = wrk
        employee.datejoin = Joindate
        

        if reptmgr_id:
            employee.reptmgr_id = reptmgr_id
            reporting_manager, created = Reportingmanager.objects.update_or_create(
                userid=employee.id,type='Primary', defaults={}
            )
            reporting_manager.myuser_2.clear()
            reporting_manager.myuser_2.add(User.objects.get(id=reptmgr_id))
        

        
            rep = User.objects.get(id=reptmgr)
            if rep.email != employee.reptmgr.email:
                to = [employee.email]
                subject = "Changed Reporting Manager"
                text_content = f"Dear {employee.username},\n\nYour reporting manager has been changed. Your new reporting manager is {rep.username}.\n\nBest regards,\nYour Company"
                html_content = f"""
                <p>Dear {employee.username},</p>
                <p>Your reporting manager has been changed. Your new reporting manager is <strong>{rep.username} </strong>.</p>
                <p>Best regards,<br>Cydez Technologies</p>
                """
                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=text_content,
                    from_email=settings.EMAIL_HOST_USER,
                    to=to
                )
                msg.attach_alternative(html_content, "text/html")
                msg.send()
        else:
            employee.reptmgr = None
            Reportingmanager.objects.filter(userid=employee.id, type='Primary').delete()
        employee.save()
        # Check for duplicate email and phone number
        if User.objects.filter(email=em).exclude(id=employee.id).exists():
            msg1 = 'Email already taken. Please try another one.'
            return render(request, "index/editemployee.html",
                          {'employee': employee, 'msg1': msg1, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr,
                           'rp': all_employees,'current_reptmgr': current_reptmgr, **x})

        if User.objects.filter(phone=ph).exclude(id=employee.id).exists():
            msg2 = 'Phone number already taken. Please try another one.'
            return render(request, "index/editemployee.html",
                          {'employee': employee, 'msg2': msg2, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr,
                           'rp': all_employees,'current_reptmgr': current_reptmgr, **x})

        if reptmgr == '' and subdepartment != '':
            sub = Subdepartment.objects.get(id=subdepartment)
            employee.subdepartment = sub
        elif subdepartment == '' and reptmgr != '':
            rep = User.objects.get(id=reptmgr)
            employee.reptmgr = rep
        elif reptmgr != '' and subdepartment != '':
            rep = User.objects.get(id=reptmgr)
            sub = Subdepartment.objects.get(id=subdepartment)
            employee.reptmgr = rep
            employee.subdepartment = sub
        elif reptmgr == '' and subdepartment == '':
            employee.username = name
            employee.email = em
            employee.phone = ph
            employee.gender = gen
            employee.dob = dob
            employee.status = status
            employee.emptype = emptype
            employee.probperiod = pd
            employee.designation = des
            employee.department = dep
            employee.jobtitle = job
            employee.wrklcn = wrk
            employee.datejoin = Joindate
            employee.save()
            return redirect('directory')
        employee.save()

        return redirect('directory')
   
   

  
    return render(request, "index/editemployee.html",
                  {'employee': employee, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'rp': all_employees,'current_reptmgr': current_reptmgr,'company_profiles': c, **x})


def delete_directory(request, userid):
    a = User.objects.get(id=userid)
    a.delete()
    print(a)
    return redirect('directory')


# EXPORT EMPLOYEE DATA IN DIRECTORY PAGE

def export_empdetails(request):
    if request.method == 'POST':
        selected_employee_ids = request.POST.getlist('selected_employees')
        selected_users = User.objects.filter(id__in=selected_employee_ids)

        selected_fields = [
            'Employee ID', 'Name', 'Department', 'Designation', 'Job Title', 'DOJ', 'Work Location', 'Employee Status',
            'Employee Type', 'Probation Period', 'Contact Number', 'Official Email ID', 'Personal Email ID',
            'Emergency Contact Name', 'Emergency Contact Number', 'Relation', 'Annual CTC',
        ]

        optional_fields = [
            ('subdep', 'Sub Department'),
            ('marital-status', 'Marital Status'),
            ('bloodgp', 'Blood Group'),
            ('address1', 'Current Address'),
            ('address2', 'Permanent Address'),
            ('eduinfo', 'Educational Info'),
            ('familymembers', 'Family Members'),
            ('idproof', 'ID Proof'),
            ('lastworkingday', 'Last Working Day'),
            ('workhistory', 'Work History'),
            ('leaverules', 'Leave Rules'),
            ('attendancerules', 'Attendance Rules'),
            ('workweekrules', 'Workweek Rules'),
            ('gender', 'Gender'),
            ('dob', 'DOB'),
        ]

        for field_key, field_name in optional_fields:
            if request.POST.get(field_key):
                selected_fields.append(field_name)

        data = []

        for user in selected_users:
            myprofile = Myprofile.objects.filter(myuser=user.id).first()

            primary_managers = Reportingmanager.objects.filter(
                userid=user.id, type='Primary')
            secondary_managers = Reportingmanager.objects.filter(
                userid=user.id, type='Secondary')

            primary_manager_usernames = []
            secondary_manager_usernames = []

            for primary_manager in primary_managers:
                primary_managers_users = primary_manager.myuser_2.all()
                for primary_manager_user in primary_managers_users:
                    primary_manager_usernames.append(
                        primary_manager_user.username)

            for secondary_manager in secondary_managers:
                secondary_managers_users = secondary_manager.myuser_2.all()
                for secondary_manager_user in secondary_managers_users:
                    secondary_manager_usernames.append(
                        secondary_manager_user.username)

            emergencycontact = Emergencycontact.objects.filter(
                myuser_6=user.id)
            emgnames = []
            emgphones = []
            emgrelations = []

            for emg in emergencycontact:
                emgnames.append(emg.name4)
                emgphones.append(str(emg.phone1))
                emgrelations.append(emg.relation1)
            emgname = '\n '.join(emgnames) if emgnames else ''
            emgphone = '\n '.join(emgphones) if emgphones else ''
            emgrelation = '\n '.join(emgrelations) if emgrelations else ''

            educationalenfo = Educationalinfo.objects.filter(myuser_4=user.id)
            eduinfo_list = []
            for eduinfo in educationalenfo:
                info = f"Institute Name: {eduinfo.institute} \nQualification: {eduinfo.qualification} \nCourse Name: {eduinfo.course} \nPassout : {eduinfo.passout} \nScore: {eduinfo.percent} \n"
                eduinfo_list.append(info)
            eduinfo_formatted = '\n'.join(eduinfo_list) if eduinfo_list else ''

            workhistory_list = []
            workhist = f"Department: {user.department.name if user.department else ''} \nDesignation: {user.designation.name if user.designation else ''} \n Start Date: {user.datejoin} \n"
            workhistory_list.append(workhist)
            workhistory_formated = '\n'.join(
                workhistory_list) if workhistory_list else ''

            family = Familymembers.objects.filter(myuser_5=user.id)
            family_list = []
            for faminfo in family:
                fam = f"Name: {faminfo.name3} \nRelation: {faminfo.relation}\n"
                family_list.append(fam)
            faminfo_formatted = '\n'.join(family_list) if family_list else ''

            idproof = Uploadeddocs.objects.filter(myuser=user.id)
            idproofs = []
            for doc in idproof:
                proofs = f"Type: {doc.type1} \nID NO: {doc.id_no} \n"
                idproofs.append(proofs)
            idproof_formatted = '\n'.join(idproofs) if idproofs else ''

            last_working_day = ResignationForm.objects.filter(user=user.id)
            lastworkingday = ''
            for lwd in last_working_day:
                lastworkingday = lwd.last_workingday or lastworkingday

            leaverules = assignrule.objects.filter(user_id=user.id)
            leave_names = []
            for leaverule in leaverules:
                rulesapplied = leaverule.rules_applied.all()
                if rulesapplied:
                    for rule in rulesapplied:
                        leave_names.append(rule.leavename)

            attendancerule = AssignAttendanceRule.objects.filter(
                user_id=user.id)
            rulename = ''
            for rule in attendancerule:
                rulename = rule.rules_applied.rulename or rulename

            workweekrules = AssignWorkWeek.objects.filter(user_id=user.id)
            workweekname = ''
            for workweekrule in workweekrules:
                workweekname = workweekrule.rules_applied.rule_name or workweekname

            assign_salarystructure = AssignSalaryStructure.objects.filter(
                user_id=user.id)
            user_annual_amount = 0

            for assign in assign_salarystructure:
                names = SalaryStructureName.objects.filter(
                    salaryrule=assign.assign_salary)
                amounts = SalaryStructureAmount.objects.filter(
                    salaryname__in=names)

                annual_amount = sum(amount.amount * 12 for amount in amounts)

                user_annual_amount += annual_amount

            annualctc_amount = user_annual_amount

            data_item = [user.empid, user.username, user.department.name if user.department else '',
                         user.designation.name if user.designation else '', user.jobtitle.name if user.jobtitle else '',
                         user.datejoin if user.datejoin else '', user.wrklcn.location if user.wrklcn else '',
                         user.status, user.emptype, user.probperiod, user.phone,
                         myprofile.offemail if myprofile else '', user.email,
                         emgname, emgphone, emgrelation, annualctc_amount]

            if 'Sub Department' in selected_fields:
                data_item.append(user.subdepartment.subname if user.subdepartment else '')
            if 'Marital Status' in selected_fields:
                data_item.append(myprofile.marital if myprofile else '')
            if 'Blood Group' in selected_fields:
                data_item.append(myprofile.bldgrp if myprofile else '')
            if 'Current Address' in selected_fields:
                data_item.append(myprofile.address if myprofile else '')
            if 'Permanent Address' in selected_fields:
                data_item.append(myprofile.peraddress if myprofile else '')
            if 'Educational Info' in selected_fields:
                data_item.append(eduinfo_formatted)
            if 'Family Members' in selected_fields:
                data_item.append(faminfo_formatted)
            if 'ID Proof' in selected_fields:
                data_item.append(idproof_formatted)
            if 'Last Working Day' in selected_fields:
                data_item.append(lastworkingday)
            if 'Work History' in selected_fields:
                data_item.append(workhistory_formated)
            if 'Leave Rules' in selected_fields:
                data_item.append(', '.join(leave_names))
            if 'Attendance Rules' in selected_fields:
                data_item.append(rulename)
            if 'Workweek Rules' in selected_fields:
                data_item.append(workweekname)
            if 'Gender' in selected_fields:
                data_item.append(user.gender if user.gender else '')
            if 'DOB' in selected_fields:
                data_item.append(user.dob if user.dob else '')

            data.append(data_item)

        # Create DataFrame
        df = pd.DataFrame(data, columns=selected_fields)

        # Prepare file name & path
        timestamp = now().strftime("%Y%m%d_%H%M%S")
        file_name = f"employee_details_{timestamp}.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, 'exports', file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Save to disk
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Employees')
            worksheet = writer.sheets['Employees']
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 5, 50)


        
        # Save the exported file to Filemanager
        with open(file_path, 'rb') as f:
            file_instance = File(f, name=file_name)
            file_size = os.path.getsize(file_path)
            Filemanager.objects.create(
            myuser_10=request.user,
            requesttype="Employee Export",
            frmt="xlsx",
            scheduleon=now(),
            size=str(file_size),
            status="Success",
            saveexcel=file_instance
            )

        # Optionally send email with attachment
        email = EmailMessage(
            subject="Employee Export - Success",
            body="Attached is the exported employee data.",
            to=[request.user.email],
        )
        email.attach_file(file_path)
        email.send(fail_silently=True)

        # Serve the file as HTTP response
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response

    return redirect('directory')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def directory(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)

    datas = User.objects.filter(
        (Q(id=admin_id) | Q(admin_id=admin_id)) & (Q(status='Active') | Q(status='Onboarding'))
    ).order_by('status', 'empid', 'username')

    dn = Designation.objects.filter(admin_id=admin_id)
    dp = Department.objects.filter(admin_id=admin_id)
    sd = Subdepartment.objects.filter(admin_id=admin_id)
    jb = Job.objects.filter(admin_id=admin_id)
    wr = Worklocation.objects.filter(admin_id=admin_id)

    manager_lookup = {
        user.id: {'primary': '', 'secondary': '', 'primary_id': None, 'secondary_id': None} for user in datas
    }

    reporting_data = Reportingmanager.objects.prefetch_related('myuser_2').filter(
        userid__in=manager_lookup.keys()
    )
    
    for entry in reporting_data:
        for mgr in entry.myuser_2.all():
            if entry.type == 'Primary':
                manager_lookup[entry.userid]['primary'] = mgr.username
                manager_lookup[entry.userid]['primary_id'] = mgr.id
            elif entry.type == 'Secondary':
                manager_lookup[entry.userid]['secondary'] = mgr.username
                manager_lookup[entry.userid]['secondary_id'] = mgr.id

    for user in datas:
        user.reptmgr_primary = manager_lookup[user.id].get('primary', '')
        user.reptmgr_secondary = manager_lookup[user.id].get('secondary', '')
        user.reptmgr_primary_id = manager_lookup[user.id].get('primary_id')
        user.reptmgr_secondary_id = manager_lookup[user.id].get('secondary_id')


    query = request.GET.get('search')

    count_user = User.objects.filter(
        (Q(id=request.user.id) | Q(admin_id=request.user.id)) & Q(status='Active')
    ).count()

    active_users = User.objects.filter(
        (Q(id=request.user.id) | Q(admin_id=request.user.id)) & Q(status='Active')
    ).count()

    if query:
        filters = [
            Q(empid__icontains=query), Q(username__icontains=query), Q(department__name__icontains=query),
            Q(designation__name__icontains=query), Q(email__icontains=query), Q(phone__icontains=query),
            Q(status__icontains=query), Q(subdepartment__subname__icontains=query),
            Q(jobtitle__name__icontains=query), Q(wrklcn__location__icontains=query),
            Q(emptype__icontains=query), Q(probperiod__icontains=query)
        ]

        query_filter = Q()
        for f in filters:
            query_filter |= f

        datas = User.objects.filter(
            query_filter & (Q(id=request.user.id) | Q(admin_id=request.user.id))
        ).order_by('status', 'empid', 'username')

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, "index/directory.html", {
        'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'datas': datas,
        'query': query, 'data': data, 'k': k, 'count_user': count_user,
        'active_users': active_users, **x
    })

# Inactive Users
@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def list_inactive_employee(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    datas = User.objects.filter(Q(admin_id=admin_id) & (Q(status='Inactive') )).exclude(id=admin_id).order_by('status', 'empid', 'username')

    dn = Designation.objects.filter(admin_id=admin_id)
    dp = Department.objects.filter(admin_id=admin_id)
    sd = Subdepartment.objects.filter(admin_id=admin_id)
    jb = Job.objects.filter(admin_id=admin_id)
    wr = Worklocation.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')

    count_user = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id) & (Q(status='Inactive') )).exclude(id=admin_id).count()
    inactive_users = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id) & (Q(status='Inactive') )).exclude(id=admin_id).count()


    if query:
        datas_list1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list3 = User.objects.filter(Q(department__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list4 = User.objects.filter(Q(designation__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list5 = User.objects.filter(Q(email__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list6 = User.objects.filter(Q(phone__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list7 = User.objects.filter(Q(status__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list8 = User.objects.filter(Q(subdepartment__subname__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list9 = User.objects.filter(Q(jobtitle__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list10 = User.objects.filter(Q(wrklcn__location__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list11 = User.objects.filter(Q(emptype__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list12 = User.objects.filter(Q(probperiod__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')

        if datas_list1 or datas_list2 or datas_list3 or datas_list4 or datas_list5 or datas_list6 or datas_list7 or datas_list8 or datas_list9 or datas_list10 or datas_list11 or datas_list12:

            datas = datas_list1 | datas_list2 | datas_list3 | datas_list4 | datas_list5 | datas_list6 | datas_list7 | datas_list8 | datas_list9 | datas_list10 | datas_list11 | datas_list12

        else:
            datas = []

    # page = request.GET.get('page', 1)
    # paginator = Paginator(datas, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)
    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, "index/notactive.html",
                  {'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'datas': datas, 'query': query, 'data': data,
                   'k': k, 'count_user': count_user, 'inactive_users': inactive_users, **x})

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def verify(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    datas = User.objects.filter(Q(id=request.user.id) | Q(
        admin_id=request.user.id)).order_by('status', 'empid','username' )
    myprofile = Myprofile.objects.filter(myuser__in=datas)
    educational_info = Educationalinfo.objects.filter(myuser_4__in=datas)
    family = Familymembers.objects.filter(myuser_5__in=datas)

    query = request.GET.get('search')

    count_user = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id)).count()
    active_users = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id) & Q(status='Active')).count()

    if request.method == 'POST':
        selected_checkboxes = []

        for key, value in request.POST.items():
            if key.startswith('user_checkbox_') and value == 'on':
                user_id = key.replace('user_checkbox_', '')
                selected_checkboxes.append(user_id)

        action = request.POST.get('action')

        if action == 'notify':
            for user_id in selected_checkboxes:
                user = User.objects.get(id=user_id)
                to = [user.email]
                subject = 'Notification'
                html_body = render_to_string(
                    'index/notifymsg.html', {'name': user.username})
                msg = EmailMultiAlternatives(
                    subject=subject, from_email=settings.EMAIL_HOST_USER, to=to)
                msg.attach_alternative(html_body, "text/html")
                msg.send()

    if query:
        datas_list1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid','username' )
        datas_list2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid','username' )
        datas_list3 = User.objects.filter(Q(email__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid','username' )

        if datas_list1 or datas_list2 or datas_list3:

            datas = datas_list1 | datas_list2 | datas_list3

        else:
            datas = []
            # messages.info(request, 'No Records Found')

    # page = request.GET.get('page', 1)
    # paginator = Paginator(datas, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)

    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "index/verify.html",
                  {'datas': datas, 'educationalinfo_set': educational_info, 'myprofile_set': myprofile,
                   'familymembers_set': family, 'query': query, 'data': data, 'k': k, 'count_user': count_user, 'active_users': active_users, **x})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empdirectory(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    current_user_id = request.user.id
    admin_id = User.objects.get(id=current_user_id).admin_id
    datas = User.objects.filter(Q(admin_id=admin_id) | Q(id=admin_id)).order_by("status", "empid","username")
    c = companyprofile.objects.filter(admin_id=admin_id)
    dn = Designation.objects.all()
    dp = Department.objects.all()
    sd = Subdepartment.objects.all()
    jb = Job.objects.all()
    wr = Worklocation.objects.all()
    rp = User.objects.all()
    query = request.GET.get('search')
    count_user = datas.count()
    if query:
        datas_list1 = User.objects.filter(
            Q(username__contains=query) & (Q(admin_id=admin_id) | Q(id=admin_id))).order_by("status","empid", "username")
        datas_list2 = User.objects.filter(
            Q(email__contains=query) & (Q(admin_id=admin_id) | Q(id=admin_id))).order_by("status","empid", "username")

        if datas_list1 or datas_list2:

            datas = datas_list1 | datas_list2

        else:
            datas = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(datas, 20)
    try:
        datas = paginator.page(page)
    except PageNotAnInteger:
        datas = paginator.page(1)
    except EmptyPage:
        datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "count_user": count_user,
    }

    return render(request, "Employee/directory.html",
                  {'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'rp': rp, 'datas': datas, 'query': query, **x})


# department
@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
@csrf_exempt
def department(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    department = Department.objects.filter(admin_id=admin_id)
    dep = Department.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')

    department = dep.annotate(
        user_count=Count(
            Case(
                When(user__status__iexact='Active', then=1),
                output_field=IntegerField()
            )
        )
    )

    if query:
        datas_list1 = Department.objects.filter(
            name__contains=query, admin_id=admin_id)
        datas_list2 = Department.objects.filter(
            description__contains=query, admin_id=admin_id)

        if datas_list1 or datas_list2:

            department = datas_list1 | datas_list2

        else:
            department = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(department, 10)
    try:
        department = paginator.page(page)
    except PageNotAnInteger:
        department = paginator.page(1)
    except EmptyPage:
        department = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,
    }

    return render(request, 'index/department.html',
                  {'department': department, 'datas': department, 'query': query, 'c': c, 'dep': dep, **x})



@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_department(request):
    if request.method == "POST":
        admin_id = request.user.id
        n = request.POST.get('departname')
        d = request.POST.get('description')

        Department.objects.create(name=n, description=d, admin_id=admin_id)
        return redirect('department')
    return render(request, "index/department.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_depart(request):
    if request.method == "POST":
        n = request.POST.get('departname')
        d = request.POST.get('description')
        department_id = request.POST.get('depid')
        k = Department.objects.filter(id=department_id)
        k.update(name=n, description=d)
        print(k)
        return redirect('department')
    return render(request, "index/department.html", {"data": k})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_department(request, userid):
    k = Department.objects.get(id=userid)
    k.delete()
    return redirect('department')
    return render(request, "index/department.html", {"data": k})


# designation


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def designation(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    designation = Designation.objects.filter(admin_id=admin_id)
    desig = Designation.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')

    designation = desig.annotate(
        user_count=Count(
            Case(
                When(user__status__iexact='Active', then=1),
                output_field=IntegerField()
            )
        )
    )

    if query:
        datas_list1 = Designation.objects.filter(
            name__contains=query, admin_id=admin_id)
        datas_list2 = Designation.objects.filter(
            description__contains=query, admin_id=admin_id)

        if datas_list1 or datas_list2:

            designation = datas_list1 | datas_list2

        else:
            designation = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(designation, 10)
    try:
        designation = paginator.page(page)
    except PageNotAnInteger:
        designation = paginator.page(1)
    except EmptyPage:
        designation = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,
    }

    return render(request, 'index/designation.html',
                  {'designation': designation, 'datas': designation, 'query': query, 'c': c, 'desig': desig, **x})



@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_designation(request):
    if request.method == "POST":
        admin_id = request.user.id
        na = request.POST.get('name')
        dis = request.POST.get('description')
        print(na, dis)
        Designation.objects.create(name=na, description=dis, admin_id=admin_id)
        return redirect("designation")
    return render(request, "index/designation.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_designation(request):
    if request.method == "POST":
        name = request.POST.get('name')
        description = request.POST.get('description')

        designation_id = request.POST.get('designation_id')

        designation_obj = Designation.objects.filter(id=designation_id)

        designation_obj.update(name=name, description=description)
        return redirect("designation")
    return render(request, "index/designation.html", {"des": designation_obj})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_designation(request, userid):
    a = Designation.objects.get(id=userid)
    a.delete()
    print(a)
    return redirect('designation')


# subdepartment

# subdepartment


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def subdepartment(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    subdep = Subdepartment.objects.filter(admin_id=admin_id)
    datas = Subdepartment.objects.filter(admin_id=admin_id)
    department = Department.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')

    datas = subdep.annotate(
        user_count=Count(
            Case(
                When(user__status__iexact='Active', then=1),
                output_field=IntegerField()
            )
        )
    )

    if query:
        data_list1 = Subdepartment.objects.filter(
            depname__name__contains=query)
        data_list2 = Subdepartment.objects.filter(
            subname__contains=query, admin_id=admin_id)
        if data_list1 or data_list2:
            datas = data_list1 | data_list2
        else:
            datas = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(datas, 20)
    try:
        datas = paginator.page(page)
    except PageNotAnInteger:
        datas = paginator.page(1)
    except EmptyPage:
        datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,
    }

    return render(request, "index/subdepartment.html",
                  {'datas': datas, 'query': query, 'sub': datas, 'c': c, 'subdep': subdep, **x,
                   'department': department})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_subdepartment(request):
    # if request.method == "POST":
    #     search_keyword = request.POST.get('sname')
    #     a = Subdepartment.objects.filter(sname__name=search_keyword)
    #     print(search_keyword)
    department = Department.objects.all()
    if request.method == "POST":
        admin_id = request.user.id
        na = request.POST.get('name')
        dis = request.POST.get('description')
        sl = request.POST.get('depname')
        sdepart = Department.objects.get(id=sl)
        Subdepartment.objects.create(subname=na, description=dis, depname=sdepart, admin_id=admin_id)
        return redirect('subdepartment')

    context = {'department': department}
    return render(request, "index/subdepartment.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_subdepartment(request):
    if request.method == "POST":
        v = request.POST.get('depname')
        n = request.POST.get('subname')
        d = request.POST.get('discription')
        subdepartment_id = request.POST.get('subid')
        a = Subdepartment.objects.filter(id=subdepartment_id)
        a.update(subname=n, description=d, depname=v)
        print(v)
        return redirect('subdepartment')
    return render(request, "index/subdepartment.html", {"subdata": a})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_subdepartment(request, userid):
    k = Subdepartment.objects.get(id=userid)
    k.delete()
    return redirect('subdepartment')


# worklocation
@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def worklocation(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=admin_id)
    data = companyprofile.objects.filter(admin_id=admin_id)

    # Work Location with optional search
    query = request.GET.get('search')
    workloc = Worklocation.objects.filter(admin_id=admin_id).annotate(user_count=Count('user'))
    if query:
        datas_list1 = Worklocation.objects.filter(location__icontains=query, admin_id=admin_id)
        workloc = datas_list1 if datas_list1.exists() else []

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(workloc, 10)
    try:
        worklocation = paginator.page(page)
    except PageNotAnInteger:
        worklocation = paginator.page(1)
    except EmptyPage:
        worklocation = paginator.page(paginator.num_pages)

    # Branch count
    branch_location_grouped = CompanyBranchLocation.objects.annotate(
    total_emp=Count('employees')
)

    # Home location grouping using lat/lon -> location name (via geopy)
    geolocator = Nominatim(user_agent="hrms-app")
    grouped_locations = defaultdict(list)

    fences = EmployeeGeoFence.objects.filter(
        home_lat__isnull=False,
        home_lon__isnull=False
    )

    for fence in fences:
        key = f"{fence.home_lat},{fence.home_lon}"
        grouped_locations[key].append(fence)

    home_location_grouped = []

    for latlon, fence_list in grouped_locations.items():
        try:
            location = geolocator.reverse(latlon, timeout=10)
            location_name = location.address.split(",")[0] if location else "Unknown"
        except Exception:
            location_name = "Unknown"

        home_location_grouped.append({
            'location': location_name,
            'total_emp': len(fence_list)
        })

    context = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "worklocation": worklocation,
        "workloc": workloc,
        "query": query,
        'branch_location_grouped': branch_location_grouped,
        'home_location_grouped': home_location_grouped,
    }

    return render(request, 'index/worklocation.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_worklocation(request):
    if request.method == "POST":
        admin_id = request.user.id
        la = request.POST.get('location')
        Worklocation.objects.create(location=la, admin_id=admin_id)
        return redirect("wrk")
    return render(request, "index/worklocation.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_wrklocation(request):
    if request.method == "POST":
        lo = request.POST.get('location')
        worklocation_id = request.POST.get('wid')
        x = Worklocation.objects.filter(id=worklocation_id)
        x.update(location=lo)
        print(x)
        return redirect("wrk")
    return render(request, "index/worklocation.html", {"worklocation": x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_wrklocation(request, userid):
    k = Worklocation.objects.get(id=userid)
    k.delete()
    return redirect('wrk')


# salary component
@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def salary_component(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    salary_component = SalaryComponent.objects.filter(admin_id=admin_id)
    
    query = request.GET.get('search')
    if query:
        datas_list1 = SalaryComponent.objects.filter(componentname__contains=query, admin_id=admin_id)
        datas_list2 = SalaryComponent.objects.filter(description__contains=query, admin_id=admin_id)

        if datas_list1 or datas_list2:

            salary_component = datas_list1 | datas_list2

        else:
            salary_component = []

    page = request.GET.get('page', 1)

    paginator = Paginator(salary_component,15)

    try:
        salarycomponent = paginator.page(page)
    except PageNotAnInteger:
        salarycomponent = paginator.page(1)
    except EmptyPage:
        salarycomponent = paginator.page(paginator.num_pages)
    
    x = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,
    }

    return render(request, 'index/salary_component.html', {'salarycomponent': salarycomponent, 'component':salary_component, 'query': query, **x})

def add_salary_component(request):
    if request.method == "POST":
        admin_id = request.user.id
        componentname = request.POST.get('componentname')
        parent_id = request.POST.get('parent')
        parent = None  
        if parent_id:
            parent = SalaryComponent.objects.get(id=parent_id) 
        percent = request.POST.get('percent')  
        description = request.POST.get('description')
        
        SalaryComponent.objects.create(
            componentname=componentname,
            Parentcomponentname=parent, 
            percent=percent if percent else '0',
            description=description,
            admin_id=admin_id
        )
        return redirect('salary_component')
    
    return render(request, "index/salary_component.html")

def update_salary_component(request):
    if request.method == "POST":
        componentname = request.POST.get('cmpname')
        description = request.POST.get('description')
        parent_id = request.POST.get('parentcmp')
        parent = None  
        if parent_id:
            parent = SalaryComponent.objects.get(id=parent_id) 
        percent = request.POST.get('percent') 
        cmp_id = request.POST.get('cmp_id')
        salarycmp = SalaryComponent.objects.filter(id=cmp_id)
        salarycmp.update(componentname=componentname,
            Parentcomponentname=parent, 
            percent=percent if percent else '0',
            description=description
            )
        print(salarycmp)
        return redirect('salary_component')
    return render(request, "index/salary_component.html", {"data": salarycmp})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_salary_component(request, userid):
    salarycomponent_id = request.POST.get('userid')
    k = SalaryComponent.objects.get(id=userid)
    k.delete()
    return redirect('salary_component')
    return render(request, "index/salary_component.html", {"data": k})


# JobTitle
@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def job(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    job = Job.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')
    job1 = Job.objects.filter(admin_id=admin_id)

    job = job1.annotate(
        user_count=Count(
            Case(
                When(user__status__iexact='Active', then=1),
                output_field=IntegerField()
            )
        )
    )


    if query:
        datas_list1 = Job.objects.filter(
            name__contains=query, admin_id=admin_id)

        if datas_list1:

            job = datas_list1

        else:
            job = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(job, 10)
    try:
        job = paginator.page(page)
    except PageNotAnInteger:
        job = paginator.page(1)
    except EmptyPage:
        job = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, 'index/job_titles.html',
                  {'job': job, 'datas': job, 'query': query, 'data': data, 'job1': job1, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def addjob(request):
    a = Job.objects.all()

    if request.method == "POST":
        admin_id = request.user.id
        n = request.POST.get('name')
        Job.objects.create(name=n, admin_id=admin_id)
        return redirect('job')
    return render(request, 'index/job_titles.html')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_job(request):
    if request.method == "POST":
        n = request.POST.get('name')
        job_id = request.POST.get('jobid')
        k = Job.objects.filter(id=job_id)
        print(k)
        k.update(name=n)
        return redirect('job')
    return render(request, "index/job_titles.html", {"upjob": k})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_job(request, userid):
    k = Job.objects.get(id=userid)
    k.delete()
    return redirect('job')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def rule(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    rule = CompanyRules.objects.filter(admin_id=admin_id)
    query = request.GET.get('search')

    if query:
        datas_list1 = CompanyRules.objects.filter(
            leavename__contains=query, admin_id=admin_id)
        datas_list2 = CompanyRules.objects.filter(
            description__contains=query, admin_id=admin_id)

        if datas_list1 or datas_list2:

            rule = datas_list1 | datas_list2

        else:
            rule = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(rule, 10)
    try:
        rule = paginator.page(page)
    except PageNotAnInteger:
        rule = paginator.page(1)
    except EmptyPage:
        rule = paginator.page(paginator.num_pages)

    default_rules = [
        {"leavename": "Casual Leave", "days": 0},
        {"leavename": "Loss Of Pay", "days": 0},
        {"leavename": "Sick Leave", "days": 0},
        {"leavename": "Maternity Leave", "days": 0},
        {"leavename": "Event Leave", "days": 0},
        {"leavename": "Paternity Leave", "days": 0},
        {"leavename": "Comp Off", "days": 0},
        {"leavename": "Optional Holiday", "days": 0},
    ]

    for rule_data in default_rules:
        rulename, created = CompanyRules.objects.get_or_create(
            admin_id=admin_id, leavename=rule_data["leavename"], defaults=rule_data)

    x = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,
    }

    return render(request, 'index/companyrules.html', {'rule': rule, 'query': query, 'c': c, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_rule(request):
    if request.method == "POST":
        admin_id = request.user.id
        n = request.POST.get('leavename')
        d = request.POST.get('description')
        a = request.POST.get('days')

        CompanyRules.objects.create(
            leavename=n, description=d, days=a, admin_id=admin_id)
        return redirect('rule')
    return render(request, "index/companyrules.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_rules(request):
    if request.method == "POST":
        # n = request.POST.get('leavename')
        d = request.POST.get('description')
        a = request.POST.get('days')
        rule_id = request.POST.get('ruleid')
        # k=CompanyRules.objects.filter(id=rule_id)
        x = CompanyRules.objects.filter(id=rule_id)
        x.update(description=d, days=a)
        print(x)
        return redirect('rule')
    return render(request, "index/companyrules.html", {"data": x})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_rule(request, userid):
    k = CompanyRules.objects.get(id=userid)
    k.delete()
    return redirect('rule')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empoverview(request):
    user_id = request.user.id 
    admin_id = User.objects.get(id=user_id).admin_id    
    datas = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id).first()    
    user_obj = User.objects.filter(id=request.user.id).first()
    company_details = None 
    if user_obj:
            
        if user_obj.company_type:  
            print(f"User's Company Type: {user_obj.company_type}")  
            company_type = user_obj.company_type.type_of_company  

            if company_type == 'Main Company':
                company_details = datas.filter(type_of_company='Main Company').first()
                print(f"Main Company Details: {company_details}")  
            elif company_type == 'Sub Company':
                company_details = datas.filter(type_of_company='Sub Company').first()
                print(f"Sub Company Details: {company_details}")  
    context = {
            "k": k,
            "datas": datas.first() if datas.exists() else None,  
            "company_details": company_details,  
            "user_obj": user_obj  
        }
    return render(request, "Employee/overview.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def leave(request):
    employee_id = request.user.id
    c = companyprofile.objects.filter(admin_id=employee_id)
    k = Myprofile.objects.filter(myuser__id=employee_id)
    assigned_rules = assignrule.objects.filter(user_id=employee_id)
    leaves = CompanyRules.objects.all()
    today = date.today()

    assignleave_name = assignrule.objects.filter(user_id=employee_id)
    print("assignleave_name : ", assignleave_name)

    if assigned_rules:
        query = request.GET.get('Leavetype', None)

        if query is None:
            assign_rule = assignrule.objects.filter(
                user_id=employee_id).first()
        else:
            assign_rule = assignrule.objects.get(
                Q(user_id=employee_id) & Q(rules_applied__id=query))

        if not assign_rule:
            return {}

        effective_date = datetime.strptime(
            assign_rule.effective_date, "%d %B %Y").date()

        leave_name = assign_rule.rules_applied.all().first().leavename
        print("leave_name :", leave_name)
        total_days = assign_rule.rules_applied.all().first().days
        carryforward = assign_rule.rules_applied.all().first().CarryForwardeEnabled
        accrualfrequency = assign_rule.rules_applied.all().first().AccrualFrequency
        accrualperiod = assign_rule.rules_applied.all().first().AccrualPeriod
        print("accrualfrequency ; AccrualPeriod :",
              accrualfrequency, accrualperiod)

        if leave_name not in ["Maternity Leave", "Optional Holiday"]:

            monthly_metrics = {
                "column_name": ["Details"] + [calendar.month_abbr[i] for i in range(1, 13)],
                "data": [
                    ["Credited Leaves"] + ["0.00"] * 12,
                    ["Applied Leaves"] + ["0.00"] * 12,
                    ["Penalty Deduction"] + ["-"] * 12,
                    ["Closing Balance"] + ["-"] * 12,
                ],
            }

            getcontext().prec = 4

        else:

            monthly_metrics = {
                "column_name": ["Details"] + [calendar.month_abbr[i] for i in range(1, 13)],
                "data": [
                    ["Applied Leaves"] + ["0.00"] * 12,
                    ["Penalty Deduction"] + ["-"] * 12,
                    ["Closing Balance"] + ["-"] * 12,
                ],
            }

            getcontext().prec = 3

        first_day_of_current_month = today.replace(day=1)
        print("first_day_of_current_month :", first_day_of_current_month)
        previous_day = first_day_of_current_month - timedelta(days=1)
        print("Previous day: ", previous_day)

        previous_month_first_day = previous_day.replace(day=1)
        print("Previous month first day: ", previous_month_first_day)
        print("leave_name  :", leave_name)

        total_credited_leave = 0
        total_leave_balance = 0

        if carryforward == "on" and leave_name not in ["Maternity Leave", "Optional Holiday"]:
            current_month = effective_date
            print("current_month 1 :", current_month)
            while current_month <= previous_month_first_day:
                if total_days > 0:
                    last_day_of_month = current_month.replace(
                        day=calendar.monthrange(current_month.year, current_month.month)[1])
                    total_days_in_month = calendar.monthrange(
                        current_month.year, current_month.month)[1]
                    print("last_day_of_month ; total_days_in_month :",
                          last_day_of_month, total_days_in_month)

                    if current_month == effective_date:
                        total_day_in_month = (
                            last_day_of_month - effective_date).days + 1
                        print("total_day_in_month :", total_day_in_month)
                    else:
                        total_day_in_month = total_days_in_month
                        print("total_day_in_month 2:", total_day_in_month)

                    one_month_credited_leave = total_days / 12
                    one_day_credited_leave = one_month_credited_leave / total_days_in_month
                    total_credited_leave_in_effective_date = one_day_credited_leave * total_day_in_month

                    total_credited_leave += total_credited_leave_in_effective_date
                    total_leave_balance += total_credited_leave_in_effective_date

                current_month = current_month.replace(
                    day=1) + relativedelta(months=1)

        print("total_credited_leave ; total_leave_balance :",
              total_credited_leave, total_leave_balance)

        if effective_date.year != today.year:
            effective_date = datetime(today.year, 1, 1).date()
        else:
            effective_date = effective_date

        print("Effective_Date :", effective_date)

        previous_leave_balance = Decimal('0.00')
        previous_credited_leave = 0
        previous_credited_leave += Decimal(total_credited_leave)
        previous_applied_leave = 0
        previous_penalty_count = 0
        rule_days = 1
        print("previous_credited_leave : ", previous_credited_leave)
        current_month = effective_date
        print("current_month 2920 : ", current_month)

        while current_month <= today:

            last_day_of_month = current_month.replace(
                day=calendar.monthrange(current_month.year, current_month.month)[1])

            total_days_in_month = calendar.monthrange(
                current_month.year, current_month.month)[1]

            print("last_day_of_month , total_days_in_month : ",
                  last_day_of_month, total_days_in_month)

            # Calculate the total credited_leave for the current month
            if current_month == effective_date:
                total_day_in_month = (
                    last_day_of_month - effective_date).days + 1
                print("total_day_in_month 1 : ",
                      effective_date, total_day_in_month)
            else:
                total_day_in_month = calendar.monthrange(
                    current_month.year, current_month.month)[1]
                print("total_day_in_month 2 : ", total_day_in_month)

            company_rule = assign_rule.rules_applied.all()

            # Checking which rule is need to show in HTML
            for cRule in company_rule:
                # print('Inside the if condition: ')
                print("CFE :", cRule.CarryForwardeEnabled)
                if cRule.days > 0 and cRule.leavename not in ["Maternity Leave", "Optional Holiday"]:
                    print('in if : ', cRule.leavename)
                    onemonth_credited_leave = Decimal(cRule.days / 12)
                    print("onemonth_credited_leave : ",
                          onemonth_credited_leave)
                    one_day_credited_leave = onemonth_credited_leave / \
                        Decimal(total_days_in_month)
                    print("total_day_in_month , one_month_credited_leave , one_day_credited_leave :",
                          total_day_in_month, onemonth_credited_leave, one_day_credited_leave)

                    total_credited_leave = one_day_credited_leave * total_day_in_month
                    print("total_credited_leave 3 : ", total_credited_leave)

                    canceled_request = 0
                    rejected_request = 0
                    # Query the Leave table for the employee's leaves within the current month
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)
                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)

                        print("DATA : ", data)

                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)

                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):

                        credited_leave = total_credited_leave
                        print("credited_leave 3000 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)

                        if applied_leave == Decimal('0.00'):
                            print("previous_credited_leave 3005 : ",
                                  previous_credited_leave)
                            credited_leave += previous_credited_leave
                            print("credited_leave 3006 : ", credited_leave)
                        else:
                            if rejected_request >= 1:
                                credited_leave += previous_credited_leave
                                total_applied_leave -= days
                                previous_applied_leave -= days
                                print("credited_leave 3011 : ", credited_leave,
                                      total_applied_leave, previous_applied_leave)
                                # previous_applied_leave -= applied_leave
                                # applied_leave -= applied_leave_dict[leave_type]
                                # print("preapp :", previous_applied_leave, applied_leave)

                            else:
                                #     previous_credited_leave -= applied_leave
                                print("applied_leave :", applied_leave)
                                print("previous_credited_leave 3017: ",
                                      previous_credited_leave)
                                credited_leave += previous_credited_leave
                                print("credited_leave 3016 : ", credited_leave)

                        print("credited_leave ; crd : ",credited_leave, total_credited_leave)
                        leave_balance = (Decimal(credited_leave) - Decimal(previous_applied_leave))
                        print("leave_balance 3020 : ", leave_balance)

                        if applied_leave == Decimal('0.00'):
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3023 : ",previous_credited_leave)
                        else:
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3024 : ",previous_credited_leave)

                        print('leave_balance , credited_leave 3027: ',leave_balance, credited_leave)
                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        if accrualfrequency == "Monthly" and accrualperiod == "Start":
                            monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                            monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                            monthly_metrics["data"][2][month_index] = "-"
                            monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                            assign_rule.creditedleaves = credited_leave
                            assign_rule.appliedleaves = previous_applied_leave
                            assign_rule.leavebalance = leave_balance
                            assign_rule.save()
                            rule_days = 1

                            print("current_month bbbbbbbbbbbbbbbb :",current_month)
                            # first_day_next_month = current_month.replace(day=1).replace(month=datetime.now().month + 1)
                            if current_month.month == 12:
                                first_day_next_month = current_month.replace(year=current_month.year + 1, month=1, day=1)
                            else:
                                first_day_next_month = current_month.replace(month=current_month.month + 1, day=1)
                            print("first_day_next_month : ", first_day_next_month)
                            last_day_current_month = first_day_next_month - timedelta(days=1)
                            print("first_day_next_month ; last_day_current_month :",first_day_next_month, last_day_current_month)

                        elif accrualfrequency == "Monthly" and accrualperiod == "End":
                            print("last_day_of_month : ", last_day_of_month)
                            if today >= last_day_of_month:
                                monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                                monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                                monthly_metrics["data"][2][month_index] = "-"
                                monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                                # previous_leave_balance = leave_balance
                                # print("previous_leave_balance : ",previous_leave_balance)
                                # assign_rule.creditedleaves = total_credited_leave
                                assign_rule.creditedleaves = credited_leave
                                assign_rule.appliedleaves = previous_applied_leave
                                assign_rule.leavebalance = leave_balance
                                assign_rule.save()
                                rule_days = 1

                elif cRule.days > 0 and (cRule.leavename == "Maternity Leave" or cRule.leavename == "Optional Holiday"):
                    total_credited_leave = Decimal(total_days)
                    print("totalcredited_leave : ", total_credited_leave)

                    canceled_request = 0
                    rejected_request = 0
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)
                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)
                        print("DATA : ", data)
                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)
                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):
                        print("leave_type", leave_type)
                        credited_leave = total_credited_leave
                        print("credited_leave 4431 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)
                        if rejected_request >= 1:
                            total_applied_leave -= days
                            previous_applied_leave -= days
                            print("credited_leave 4439 : ", credited_leave,
                                  total_applied_leave, previous_applied_leave)
                        print("credited_leave ; crd ; previous_applied_leave : ",
                              credited_leave, total_credited_leave, previous_applied_leave)
                        leave_balance = (
                            Decimal(credited_leave) - Decimal(previous_applied_leave))
                        print("leave_balance 4442 : ", leave_balance)

                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        monthly_metrics["data"][0][month_index] = f"{total_applied_leave:.2f}"
                        monthly_metrics["data"][1][month_index] = "-"
                        monthly_metrics["data"][2][month_index] = f"{leave_balance:.2f}"
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.leavebalance = leave_balance
                        assign_rule.save()
                        rule_days = 1

                elif cRule.days <= 0 and cRule.leavename == "Comp Off":

                    compoff_count = CompOff.objects.filter(user=employee_id, punch_data__date__year=current_month.year, punch_data__date__month=current_month.month,
                                                           punch_data__is_compoff_reverted=False).aggregate(total_cred=Sum('creditedleaves'))['total_cred']

                    print("compoff_count 4138:", employee_id)

                    if compoff_count is None:
                        compoff_count = 0
                    print("compoff_count: ", compoff_count)

                    total_credited_leave = compoff_count

                    canceled_request = 0
                    rejected_request = 0
                    # Query the Leave table for the employee's leaves within the current month
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)

                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)
                        print("DATA : ", data)
                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)

                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):

                        credited_leave = compoff_count
                        print("credited_leave 3000 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)

                        if applied_leave == Decimal('0.00'):
                            print("previous_credited_leave 3005 : ",
                                  previous_credited_leave)
                            credited_leave += previous_credited_leave
                            print("credited_leave 3006 : ", credited_leave)
                        else:
                            if rejected_request >= 1:
                                credited_leave += previous_credited_leave
                                total_applied_leave -= days
                                previous_applied_leave -= days
                                print("credited_leave 3011 : ", credited_leave,
                                      total_applied_leave, previous_applied_leave)
                                # previous_applied_leave -= applied_leave
                                # applied_leave -= applied_leave_dict[leave_type]
                                # print("preapp :", previous_applied_leave, applied_leave)

                            else:
                                #     previous_credited_leave -= applied_leave
                                print("applied_leave :", applied_leave)
                                print("previous_credited_leave 3017: ",
                                      previous_credited_leave)
                                credited_leave += previous_credited_leave
                                print("credited_leave 3016 : ", credited_leave)

                        print("credited_leave ; crd : ",
                              credited_leave, total_credited_leave)
                        leave_balance = (
                            Decimal(credited_leave) - Decimal(previous_applied_leave))
                        print("leave_balance 3020 : ", leave_balance)

                        if applied_leave == Decimal('0.00'):
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3023 : ",
                                  previous_credited_leave)
                        else:
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3024 : ",
                                  previous_credited_leave)

                        print('leave_balance , credited_leave 3027: ',
                              leave_balance, credited_leave)
                        # Find the index for the current month and update the data
                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        # monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                        monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                        monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                        monthly_metrics["data"][2][month_index] = "-"
                        monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                        # previous_leave_balance = leave_balance
                        # print("previous_leave_balance : ",previous_leave_balance)
                        # assign_rule.creditedleaves = total_credited_leave
                        assign_rule.creditedleaves = credited_leave
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.leavebalance = leave_balance
                        assign_rule.save()
                        rule_days = 1

                elif cRule.days <= 0 and cRule.leavename == "Loss Of Pay":
                    if cRule.leavename in assign_rule.rules_applied.all().values_list('leavename', flat=True):
                        leave_data = Leave.objects.filter(
                            leavetyp=cRule.leavename,
                            applicant_email_id=employee_id,
                            strtDate__lte=last_day_of_month,
                            endDate__gte=current_month
                        ).values('leavetyp').annotate(
                            applied_leave=Sum('Days'),
                            canceled_request=Sum(Case(
                                When(cancel_requested=True, then=1), default=0, output_field=IntegerField())),
                            rejected_request=Sum(
                                Case(When(rejected=True, then=1), default=0, output_field=IntegerField())),
                        )

                        print("leave_data 3060 : ", leave_data)

                        applied_leave_extract = leave_data.values_list(
                            'applied_leave', flat=True)
                        print("applied_leave_extract : ",
                              applied_leave_extract)
                        if applied_leave_extract.exists():
                            applied_leave_ = applied_leave_extract[0]
                            print("applied_leave_ 1: ", applied_leave_)
                        else:
                            applied_leave_ = 0
                            print("applied_leave_ 2: ", applied_leave_)

                        previous_applied_leave += applied_leave_
                        print("previous_applied_leave 4133:",
                              previous_applied_leave)

                        penalty_count = PenaltyLogs.objects.filter(user=employee_id, punch_data__date__year=current_month.year, punch_data__date__month=current_month.month,
                                                                   punch_data__is_penalty_reverted=False).aggregate(total_deduction=Sum('deduction'))['total_deduction']

                        print("penalty_count 4138:",
                              employee_id, penalty_count)

                        if penalty_count is None:
                            penalty_count = 0
                        print("penalty_count: ", penalty_count)

                        previous_penalty_count += penalty_count
                        month_index = current_month.month

                        print("month_index : ", month_index)

                        # Update only applied_leaves and penalty_deduction values
                        monthly_metrics["data"][1][month_index] = f"{applied_leave_:.2f}"
                        monthly_metrics["data"][2][month_index] = penalty_count
                        # assign_rule.creditedleaves = Decimal("0.00")
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.penaltydeduction = previous_penalty_count
                        # assign_rule.leavebalance = Decimal("0.00")
                        assign_rule.save()
                        rule_days = 0
                        print("applied_leave_ {}" f"{applied_leave_:.2f}")

            print("CFE :", cRule.CarryForwardeEnabled)
            # Move to the next month
            current_month = current_month.replace(
                day=1) + relativedelta(months=1)
            print("current_month : ", current_month)

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        if rule_days == 0:
            data = monthly_metrics
            column_names = data['column_name']
            data_rows = data['data']
            filtered_data = {'column_name': column_names, 'data': []}
            print("filtered_data : ", filtered_data)

            for row in data_rows:
                if row[0] == 'Applied Leaves' or row[0] == 'Penalty Deduction':
                    filtered_data['data'].append(row)

            monthly_metrics = filtered_data
            print("monthly_metrics : ", monthly_metrics)

        context = {
            'k': k[0] if k.exists() else k,
            'c': c[0] if c.exists() else c,
            'assigned_rules': assigned_rules, 'leave': leaves,
            'months': months, 'monthly_metrics': monthly_metrics, 'query': query, 'rule_days': rule_days}

        return render(request, "index/leave.html", context)
    else:
        context = {
            'k': k[0] if k.exists() else k,
            'c': c[0] if c.exists() else c,
        }
        return render(request, "index/leave.html", context)


def update_leaves(request):
    if request.method == 'POST':
        rule_id = request.POST.get('rule_id')
        credited_leaves = request.POST.get('creditedleaves')
        applied_leaves = request.POST.get('appliedleaves')

        rule = assignrule.objects.get(id=rule_id)
        rule.creditedleaves = credited_leaves
        rule.appliedleaves = applied_leaves
        rule.save()

    return redirect('leave')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empleave(request):
    employee_id = request.user.id
    print("employee_id :", employee_id)
    admin_id = User.objects.get(id=employee_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=employee_id)
    leaves = CompanyRules.objects.all()
    today = date.today()
    
    isReportingManager = Reportingmanager.objects.filter(myuser_2__id=employee_id).exists()

    assigned_rules = assignrule.objects.filter(user_id=employee_id)
    print("assigned_rules : ", assigned_rules)
    if assigned_rules:
        query = request.GET.get('Leavetype', None)

        if query is None:
            assign_rule = assignrule.objects.filter(
                user_id=employee_id).first()
            print("assign_rule 1 : ", assign_rule)
        else:

            assign_rule = assignrule.objects.get(
                Q(user_id=employee_id) & Q(rules_applied__id=query))
            print("assign_rule 2 : ", assign_rule)

        if not assign_rule:
            return {}

        effective_date = datetime.strptime(
            assign_rule.effective_date, "%d %B %Y").date()
        print("Effective Date : ", effective_date)

        leave_name = assign_rule.rules_applied.all().first().leavename
        print("leave_name :", leave_name)
        total_days = assign_rule.rules_applied.all().first().days
        carryforward = assign_rule.rules_applied.all().first().CarryForwardeEnabled
        accrualfrequency = assign_rule.rules_applied.all().first().AccrualFrequency
        accrualperiod = assign_rule.rules_applied.all().first().AccrualPeriod
        print("accrualfrequency ; AccrualPeriod :",
              accrualfrequency, accrualperiod)

        if leave_name not in ["Maternity Leave", "Optional Holiday"]:

            monthly_metrics = {
                "column_name": ["Details"] + [calendar.month_abbr[i] for i in range(1, 13)],
                "data": [
                    ["Credited Leaves"] + ["0.00"] * 12,
                    ["Applied Leaves"] + ["0.00"] * 12,
                    ["Penalty Deduction"] + ["-"] * 12,
                    ["Closing Balance"] + ["-"] * 12,
                ],
            }

            getcontext().prec = 4

        else:

            monthly_metrics = {
                "column_name": ["Details"] + [calendar.month_abbr[i] for i in range(1, 13)],
                "data": [
                    ["Applied Leaves"] + ["0.00"] * 12,
                    ["Penalty Deduction"] + ["-"] * 12,
                    ["Closing Balance"] + ["-"] * 12,
                ],
            }

            getcontext().prec = 3

        first_day_of_current_month = today.replace(day=1)
        previous_day = first_day_of_current_month - timedelta(days=1)
        print("Previous day: ", previous_day)

        previous_month_first_day = previous_day.replace(day=1)
        print("Previous month first day: ", previous_month_first_day)
        print("leave_name  :", leave_name)

        total_credited_leave = 0
        total_leave_balance = 0

        if carryforward == "on" and leave_name not in ["Maternity Leave", "Optional Holiday"]:
            current_month = effective_date
            print("current_month 1 :", current_month)
            while current_month <= previous_month_first_day:
                if total_days > 0:
                    last_day_of_month = current_month.replace(
                        day=calendar.monthrange(current_month.year, current_month.month)[1])
                    total_days_in_month = calendar.monthrange(
                        current_month.year, current_month.month)[1]
                    print("last_day_of_month ; total_days_in_month :",
                          last_day_of_month, total_days_in_month)

                    if current_month == effective_date:
                        total_day_in_month = (
                            last_day_of_month - effective_date).days + 1
                        print("total_day_in_month :", total_day_in_month)
                    else:
                        total_day_in_month = total_days_in_month
                        print("total_day_in_month 2:", total_day_in_month)

                    one_month_credited_leave = total_days / 12
                    one_day_credited_leave = one_month_credited_leave / total_days_in_month
                    total_credited_leave_in_effective_date = one_day_credited_leave * total_day_in_month

                    total_credited_leave += total_credited_leave_in_effective_date
                    total_leave_balance += total_credited_leave_in_effective_date

                current_month = current_month.replace(
                    day=1) + relativedelta(months=1)

        print("total_credited_leave ; total_leave_balance :",
              total_credited_leave, total_leave_balance)

        if effective_date.year != today.year:
            effective_date = datetime(today.year, 1, 1).date()
        else:
            effective_date = effective_date

        print("Effective_Date :", effective_date)

        previous_leave_balance = Decimal('0.00')
        previous_credited_leave = 0
        previous_credited_leave += Decimal(total_credited_leave)
        previous_applied_leave = 0
        previous_penalty_count = 0
        rule_days = 1
        print("previous_credited_leave : ",
              previous_credited_leave, previous_applied_leave)
        current_month = effective_date
        print("current_month 2920 : ", current_month)

        while current_month <= today:

            # Calculate the last day of the current month
            last_day_of_month = current_month.replace(
                day=calendar.monthrange(current_month.year, current_month.month)[1])

            total_days_in_month = calendar.monthrange(
                current_month.year, current_month.month)[1]

            print("last_day_of_month , total_days_in_month : ",
                  last_day_of_month, total_days_in_month)

            # Calculate the total credited_leave for the current month
            if current_month == effective_date:
                total_day_in_month = (
                    last_day_of_month - effective_date).days + 1
                print("total_day_in_month 1 : ",
                      effective_date, total_day_in_month)
            else:
                total_day_in_month = calendar.monthrange(
                    current_month.year, current_month.month)[1]
                print("total_day_in_month 2 : ", total_day_in_month)

            company_rule = assign_rule.rules_applied.all()

            # Checking which rule is need to show in HTML
            for cRule in company_rule:
                # print('Inside the if condition: ')
                print("CFE :", cRule.CarryForwardeEnabled)
                if cRule.days > 0 and cRule.leavename not in ["Maternity Leave", "Optional Holiday"]:
                    print('in if : ', cRule.leavename)
                    onemonth_credited_leave = Decimal(cRule.days / 12)
                    print("onemonth_credited_leave : ",
                          onemonth_credited_leave)
                    one_day_credited_leave = onemonth_credited_leave / \
                        Decimal(total_days_in_month)
                    print("total_day_in_month , one_month_credited_leave , one_day_credited_leave :",
                          total_day_in_month, onemonth_credited_leave, one_day_credited_leave)

                    total_credited_leave = one_day_credited_leave * total_day_in_month
                    print("total_credited_leave 3 : ", total_credited_leave)

                    canceled_request = 0
                    rejected_request = 0
                    # Query the Leave table for the employee's leaves within the current month
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)
                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)

                        print("DATA : ", data)

                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)

                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):

                        credited_leave = total_credited_leave
                        print("credited_leave 3000 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)

                        if applied_leave == Decimal('0.00'):
                            print("previous_credited_leave 3005 : ",
                                  previous_credited_leave)
                            credited_leave += previous_credited_leave
                            print("credited_leave 3006 : ", credited_leave)
                        else:
                            if rejected_request >= 1:
                                credited_leave += previous_credited_leave
                                total_applied_leave -= days
                                previous_applied_leave -= days
                                print("credited_leave 3011 : ", credited_leave,
                                      total_applied_leave, previous_applied_leave)

                            else:
                                print("applied_leave :", applied_leave)
                                print("previous_credited_leave 3017: ",previous_credited_leave)
                                credited_leave += previous_credited_leave
                                print("credited_leave 3016 : ", credited_leave)

                        print("credited_leave ; crd : ",credited_leave, total_credited_leave)
                        leave_balance = (Decimal(credited_leave) - Decimal(previous_applied_leave))

                        if applied_leave == Decimal('0.00'):
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3023 : ",previous_credited_leave)
                        else:
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3024 : ",previous_credited_leave)

                        print('leave_balance , credited_leave 3027: ',
                              leave_balance, credited_leave)
                        # Find the index for the current month and update the data
                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        # monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                        if accrualfrequency == "Monthly" and accrualperiod == "Start":
                            monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                            monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                            monthly_metrics["data"][2][month_index] = "-"
                            monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                            # previous_leave_balance = leave_balance
                            # print("previous_leave_balance : ",previous_leave_balance)
                            # assign_rule.creditedleaves = total_credited_leave
                            assign_rule.creditedleaves = credited_leave
                            assign_rule.appliedleaves = previous_applied_leave
                            assign_rule.leavebalance = leave_balance
                            assign_rule.save()
                            rule_days = 1

                            print("current_month bbbbbbbbbbbbbbbb :",
                                  current_month)
                            # first_day_next_month = current_month.replace(day=1).replace(month=datetime.now().month + 1)
                            if current_month.month == 12:
                                first_day_next_month = current_month.replace(year=current_month.year + 1, month=1, day=1)
                            else:
                                first_day_next_month = current_month.replace(month=current_month.month + 1, day=1)
                            last_day_current_month = first_day_next_month - timedelta(days=1)
                            print("first_day_next_month ; last_day_current_month :",
                                  first_day_next_month, last_day_current_month)

                        elif accrualfrequency == "Monthly" and accrualperiod == "End":
                            print("last_day_of_month : ", last_day_of_month)
                            if today >= last_day_of_month:
                                monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                                monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                                monthly_metrics["data"][2][month_index] = "-"
                                monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                                # previous_leave_balance = leave_balance
                                # print("previous_leave_balance : ",previous_leave_balance)
                                # assign_rule.creditedleaves = total_credited_leave
                                assign_rule.creditedleaves = credited_leave
                                assign_rule.appliedleaves = previous_applied_leave
                                assign_rule.leavebalance = leave_balance
                                assign_rule.save()
                                rule_days = 1

                elif cRule.days > 0 and (cRule.leavename == "Maternity Leave" or cRule.leavename == "Optional Holiday"):
                    total_credited_leave = Decimal(total_days)
                    print("totalcredited_leave : ", total_credited_leave)

                    canceled_request = 0
                    rejected_request = 0
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)
                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)
                        print("DATA : ", data)
                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)
                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):
                        print("leave_type", leave_type)
                        credited_leave = total_credited_leave
                        print("credited_leave 4431 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)
                        if rejected_request >= 1:
                            total_applied_leave -= days
                            previous_applied_leave -= days
                            print("credited_leave 4439 : ", credited_leave,
                                  total_applied_leave, previous_applied_leave)
                        print("credited_leave ; crd ; previous_applied_leave : ",
                              credited_leave, total_credited_leave, previous_applied_leave)
                        leave_balance = (
                            Decimal(credited_leave) - Decimal(previous_applied_leave))
                        print("leave_balance 4442 : ", leave_balance)

                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        monthly_metrics["data"][0][month_index] = f"{total_applied_leave:.2f}"
                        monthly_metrics["data"][1][month_index] = "-"
                        monthly_metrics["data"][2][month_index] = f"{leave_balance:.2f}"
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.leavebalance = leave_balance
                        assign_rule.save()
                        rule_days = 1

                elif cRule.days <= 0 and cRule.leavename == "Comp Off":

                    compoff_count = CompOff.objects.filter(user=employee_id, punch_data__date__year=current_month.year, punch_data__date__month=current_month.month,
                                                           punch_data__is_compoff_reverted=False).aggregate(total_cred=Sum('creditedleaves'))['total_cred']

                    print("compoff_count 4138:", employee_id)

                    if compoff_count is None:
                        compoff_count = 0
                    print("compoff_count: ", compoff_count)

                    total_credited_leave = compoff_count

                    canceled_request = 0
                    rejected_request = 0
                    # Query the Leave table for the employee's leaves within the current month
                    leave_data = Leave.objects.filter(
                        applicant_email_id=employee_id,
                        strtDate__lte=last_day_of_month,
                        endDate__gte=current_month
                    ).values('leavetyp', 'Days', 'strtDate', 'endDate').annotate(
                        applied_leave=Sum('Days'),
                        canceled_request=Sum(
                            Case(When(cancel_requested=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                        rejected_request=Sum(
                            Case(When(rejected=True, then='Days'),
                                 default=0, output_field=IntegerField())
                        ),
                    )

                    applied_leave_dict = {leave_type: Decimal(
                        '0.00') for leave_type in assign_rule.rules_applied.all()}
                    print("applied_leave_dict :", applied_leave_dict)
                    total_applied_leave = Decimal('0.00')
                    days = Decimal('0.00')
                    print("DAYS :", days)

                    for data in leave_data:
                        applied_leavename = data['leavetyp']
                        print("NAME :", applied_leavename, cRule.leavename)
                        print("DATA : ", data)
                        if applied_leavename == cRule.leavename:
                            leave_type = data['leavetyp']
                            applied_leave = Decimal(
                                data['applied_leave'] or '0.00')
                            applied_leave_dict[leave_type] = applied_leave
                            print("leave_type :", leave_type, "applied_leave :", applied_leave,
                                  "applied_leave_dict[leave_type] : ", applied_leave_dict[leave_type])

                            canceled_request += data['canceled_request']
                            rejected_request += data['rejected_request']
                            total_applied_leave += applied_leave
                            print("total_applied_leave :", total_applied_leave)

                            if rejected_request >= 1:
                                days += Decimal(data['Days'])
                                print("days: ", days)

                    for leave_type in assign_rule.rules_applied.all().values_list('leavename', flat=True):

                        credited_leave = compoff_count
                        print("credited_leave 3000 : ", credited_leave)
                        applied_leave = applied_leave_dict.get(
                            leave_type, Decimal('0.00'))
                        print("applied_leave : ", applied_leave)
                        previous_applied_leave += total_applied_leave
                        print("prev_appL : ", previous_applied_leave)

                        if applied_leave == Decimal('0.00'):
                            print("previous_credited_leave 3005 : ",
                                  previous_credited_leave)
                            credited_leave += previous_credited_leave
                            print("credited_leave 3006 : ", credited_leave)
                        else:
                            if rejected_request >= 1:
                                credited_leave += previous_credited_leave
                                total_applied_leave -= days
                                previous_applied_leave -= days
                                print("credited_leave 3011 : ", credited_leave,
                                      total_applied_leave, previous_applied_leave)
                                # previous_applied_leave -= applied_leave
                                # applied_leave -= applied_leave_dict[leave_type]
                                # print("preapp :", previous_applied_leave, applied_leave)

                            else:
                                #     previous_credited_leave -= applied_leave
                                print("applied_leave :", applied_leave)
                                print("previous_credited_leave 3017: ",
                                      previous_credited_leave)
                                credited_leave += previous_credited_leave
                                print("credited_leave 3016 : ", credited_leave)

                        print("credited_leave ; crd : ",
                              credited_leave, total_credited_leave)
                        leave_balance = (
                            Decimal(credited_leave) - Decimal(previous_applied_leave))
                        print("leave_balance 3020 : ", leave_balance)

                        if applied_leave == Decimal('0.00'):
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3023 : ",
                                  previous_credited_leave)
                        else:
                            previous_credited_leave = credited_leave
                            print("previous_credited_leave 3024 : ",
                                  previous_credited_leave)

                        print('leave_balance , credited_leave 3027: ',
                              leave_balance, credited_leave)
                        # Find the index for the current month and update the data
                        month_index = current_month.month
                        print("month_index 1 :", month_index)
                        # monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                        monthly_metrics["data"][0][month_index] = f"{total_credited_leave:.2f}"
                        monthly_metrics["data"][1][month_index] = f"{total_applied_leave:.2f}"
                        monthly_metrics["data"][2][month_index] = "-"
                        monthly_metrics["data"][3][month_index] = f"{leave_balance:.2f}"

                        # previous_leave_balance = leave_balance
                        # print("previous_leave_balance : ",previous_leave_balance)
                        # assign_rule.creditedleaves = total_credited_leave
                        assign_rule.creditedleaves = credited_leave
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.leavebalance = leave_balance
                        assign_rule.save()
                        rule_days = 1

                elif cRule.days <= 0 and cRule.leavename == "Loss Of Pay":

                    if cRule.leavename in assign_rule.rules_applied.all().values_list('leavename', flat=True):
                        leave_data = Leave.objects.filter(
                            leavetyp=cRule.leavename,
                            applicant_email_id=employee_id,
                            strtDate__lte=last_day_of_month,
                            endDate__gte=current_month
                        ).values('leavetyp').annotate(
                            applied_leave=Sum('Days'),
                            canceled_request=Sum(Case(
                                When(cancel_requested=True, then=1), default=0, output_field=IntegerField())),
                            rejected_request=Sum(
                                Case(When(rejected=True, then=1), default=0, output_field=IntegerField())),
                        )

                        print("leave_data 3060 : ", leave_data)

                        applied_leave_extract = leave_data.values_list(
                            'applied_leave', flat=True)
                        print("applied_leave_extract : ",
                              applied_leave_extract)
                        if applied_leave_extract.exists():
                            applied_leave_ = applied_leave_extract[0]
                            print("applied_leave_ 1: ", applied_leave_)
                        else:
                            applied_leave_ = 0
                            print("applied_leave_ 2: ", applied_leave_)

                        previous_applied_leave += applied_leave_
                        print("previous_applied_leave 4133:",
                              previous_applied_leave)

                        penalty_count = PenaltyLogs.objects.filter(user=employee_id, punch_data__date__year=current_month.year, punch_data__date__month=current_month.month,
                                                                   punch_data__is_penalty_reverted=False).aggregate(total_deduction=Sum('deduction'))['total_deduction']

                        print("penalty_count 4138:",
                              employee_id, penalty_count)

                        if penalty_count is None:
                            penalty_count = 0
                        print("penalty_count: ", penalty_count)

                        previous_penalty_count += penalty_count

                        month_index = current_month.month

                        print("month_index previous_applied_leave : ",
                              month_index, previous_applied_leave)

                        # Update only applied_leaves and penalty_deduction values
                        monthly_metrics["data"][1][month_index] = f"{applied_leave_:.2f}"
                        monthly_metrics["data"][2][month_index] = penalty_count
                        # assign_rule.creditedleaves = Decimal("0.00")
                        assign_rule.appliedleaves = previous_applied_leave
                        assign_rule.penaltydeduction = previous_penalty_count
                        assign_rule.save()
                        rule_days = 0
                        print("applied_leave_ {}" f"{applied_leave_:.2f}")

            print("CFE :", cRule.CarryForwardeEnabled)
            # Move to the next month
            current_month = current_month.replace(
                day=1) + relativedelta(months=1)
            print("current_month : ", current_month)

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        if rule_days == 0:
            data = monthly_metrics
            column_names = data['column_name']
            data_rows = data['data']
            filtered_data = {'column_name': column_names, 'data': []}
            print("filtered_data : ", filtered_data)

            for row in data_rows:
                if row[0] == 'Applied Leaves' or row[0] == 'Penalty Deduction':
                    filtered_data['data'].append(row)

            monthly_metrics = filtered_data
            print("monthly_metrics : ", monthly_metrics)

        context = {'assigned_rules': assigned_rules, 'leave': leaves,
                   'months': months, 'monthly_metrics': monthly_metrics, 'query': query, 'rule_days': rule_days,
                   'k': k[0] if k.exists() else k,
                   'c': c[0] if c.exists() else c,
                   "isReportingManager":isReportingManager
                   }

        return render(request, "Employee/leave.html", context)
    else:
        context = {
            'k': k[0] if k.exists() else k,
            'c': c[0] if c.exists() else c,
            "isReportingManager":isReportingManager
        }
        return render(request, "Employee/leave.html", context)


@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def overview(request):
    admin_id = request.user.id
   
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = User.objects.filter(user=request.user.id)

 
    is_admin = request.user.groups.filter(name='Admin').exists()

   
    main_company_data = companyprofile.objects.filter(admin_id=admin_id, type_of_company='Main Company').first()
    # sub_company_data = companyprofile.objects.filter(admin_id=admin_id, type_of_company='Sub Company')
    sub_company_data = companyprofile.objects.filter(admin_id=admin_id, type_of_company="Sub Company").first()
  
    print("Sub Company Data:", sub_company_data)
    if sub_company_data:
        print("Sub Company Brand Name:", sub_company_data.brandname)

    x = {  
        "k": k[0] if k.exists() else k,
        "main_company_data": main_company_data,
        "sub_company_data": sub_company_data,
        "is_admin": is_admin,
        "l": l[0] if l.exists() else l,
    }

    return render(request, "index/overview.html", x)


@login_required(login_url='login')
def create_overview(request, uid):
    if request.method == "POST":
        admin_id = request.user.id
        regcmpname = request.POST.get('regcmpnm')
        brndname = request.POST.get('brndnm')
        web = request.POST.get('website')
        dm = request.POST.get('domain')
        Fb = request.POST.get('facebook')
        Twi = request.POST.get('twitter')
        ln = request.POST.get('linkedin')
        logo = request.FILES['logo']
        inverse_logo = request.FILES['inverse_logo']
        company_type = request.POST.get('type_of_company') 
        f = FileSystemStorage()
        f2 = f.save(logo.name, logo)
        if inverse_logo:
            f3 = f.save(inverse_logo.name, inverse_logo)
        else:
            f3 = None     

        companyprofile.objects.create(registeredcompanyname=regcmpname, brandname=brndname,
                                      website=web, domain=dm, fb=Fb, twitter=Twi, linkedin=ln, logo=f2,inverse_logo = f3,type_of_company=company_type,
                                      admin_id=admin_id)
    return redirect('companyprofile')


@login_required(login_url='login')
def update_overview(request):
    admin_id = request.user.id
    
    k = companyprofile.objects.filter(admin_id=admin_id)
    sub_company_data = companyprofile.objects.filter(admin_id=admin_id, type_of_company="Sub Company").first()
    print("Sub Company Data:", sub_company_data)

    if request.method == "POST":
        regcmpname = request.POST.get('regcmpnm')
        brndname = request.POST.get('brndnm')
        web = request.POST.get('website')
        dm = request.POST.get('domain')
        Fb = request.POST.get('facebook')
        Twi = request.POST.get('twitter')
        ln = request.POST.get('linkedin')
        company_type = request.POST.get('type_of_company')

        com_p = k.filter(type_of_company=company_type).first()

        if not com_p:

            com_p = companyprofile(admin_id=admin_id, type_of_company=company_type)

        f1 = FileSystemStorage()

        if 'logo' in request.FILES:
            logo = request.FILES['logo']
            f1.save(logo.name, logo)
            com_p.logo = logo
        else:
            com_p.logo = com_p.logo if com_p.logo else None  

        if 'inverse_logo' in request.FILES:
            inverse_logo = request.FILES['inverse_logo']
            f1.save(inverse_logo.name, inverse_logo)
            com_p.inverse_logo = inverse_logo
        else:
            com_p.inverse_logo = com_p.inverse_logo if com_p.inverse_logo else None  

       
        com_p.registeredcompanyname = regcmpname    
        com_p.brandname = brndname
        com_p.website = web
        com_p.domain = dm
        com_p.fb = Fb    
        com_p.twitter = Twi
        com_p.linkedin = ln
        com_p.save()
        print("subbb",sub_company_data)
        return redirect('companyprofile')
    return render(request, "index/overview.html", {"data": k,"sub_company_data": sub_company_data})


@login_required(login_url='login')
def address(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    data = corporateaddress.objects.filter(admin_id=admin_id)
    reg = registeredaddress.objects.filter(admin_id=admin_id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/address.html", {'data': reg, 'datas': data, **x})


def empaddress(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    k = corporateaddress.objects.filter(admin_id=admin_id)
    data = registeredaddress.objects.filter(admin_id=admin_id)
    
    y = companyprofile.objects.filter(admin_id=admin_id)
    w = Myprofile.objects.filter(myuser__id=request.user.id)
    x = {
        "y": y[0] if y.exists() else y,
        "w": w[0] if w.exists() else w,
    }

    return render(request, "Employee/address.html", {'data': data, 'k': k, **x})


@login_required(login_url='login')
def create_regoffice(request):
    if request.method == "POST":
        admin_id = request.user.id
        regaddress = request.POST.get('regofficeaddress')
        regpin = int(request.POST.get('regpincode'))
        regdis = request.POST.get('regdistrict')
        regst = request.POST.get('regstate')
        regctr = request.POST.get('regcountry')
        registeredaddress.objects.create(
            regofficeaddress=regaddress, regpincode=regpin, regdistrict=regdis, regstate=regst, regcountry=regctr,
            admin_id=admin_id)
        return redirect('address1')
    return render(request, "index/address.html")


@login_required(login_url='login')
def update_regoffice(request):
    admin_id = registeredaddress.objects.filter(admin_id=request.user.id)
    if request.method == "POST":
        regaddress = request.POST.get('regofficeaddress')
        regpin = int(request.POST.get('regpincode'))
        regdis = request.POST.get('regdistrict')
        regst = request.POST.get('regstate')
        regctr = request.POST.get('regcountry')
        admin_id.update(regofficeaddress=regaddress, regpincode=regpin,
                        regdistrict=regdis, regstate=regst, regcountry=regctr)
        return redirect('address1')
    return render(request, "index/address.html")


@login_required(login_url='login')
def create_corpoffice(request):
    admin_id = request.user.id
    k = corporateaddress.objects.filter(admin_id=admin_id)

    if request.method == "POST":
        corpaddress = request.POST.get('corpofficeaddress')
        corppin = int(request.POST.get('corppincode'))
        corpdis = request.POST.get('corpdistrict')
        corpst = request.POST.get('corpstate')
        corpctr = request.POST.get('corpcountry')
        k.create(corpofficeaddress=corpaddress, corppincode=corppin,
                 corpdistrict=corpdis, corpstate=corpst, corpcountry=corpctr, admin_id=admin_id)
        return redirect('address1')
    return render(request, "index/address.html")


@login_required(login_url='login')
def update_corpoffice(request):
    k = corporateaddress.objects.filter(admin_id=request.user.id)
    if request.method == "POST":
        corpaddress = request.POST.get('corpofficeaddress')
        corppin = int(request.POST.get('corppincode'))
        corpdis = request.POST.get('corpdistrict')
        corpst = request.POST.get('corpstate')
        corpctr = request.POST.get('corpcountry')
        k.update(corpofficeaddress=corpaddress, corppincode=corppin,
                 corpdistrict=corpdis, corpstate=corpst, corpcountry=corpctr)
        return redirect('address1')
    return render(request, "index/address.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emppolicy(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companypolicy.objects.filter(admin_id=admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c
    }

    return render(request, "Employee/policy.html", {'data': data, **x})


@login_required(login_url='login')
def create_companypolicy(request):
    if request.method == "POST":
        admin_id = request.user.id
        cmpol = request.POST.get('cmppolicy')
        cdoc = request.FILES['cmdoc']
        f = FileSystemStorage()
        f2 = f.save(cdoc.name, cdoc)
        companypolicy.objects.create(
            companypolicies=cmpol, policydoc=f2, admin_id=admin_id)
    return redirect('policy1')


@login_required(login_url='login')
def delete_companypolicy(request, cid):
    l = companypolicy.objects.get(id=cid)
    l.delete()
    return redirect('policy1')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def policy(request):
    admin_id = request.user.id
    data = companypolicy.objects.filter(admin_id=admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/policy.html", {'data': data, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def personal_info_nav(request):
    user = request.user
    c = companyprofile.objects.filter(admin_id=request.user.id)
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        return redirect('createpersonalinfo')
    context = {
        'myprofile': myprofile,
        'c': c[0] if c.exists() else c
    }

    return render(request, 'index/personalinfo.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_personal_info(request):
    user = request.user
    c = companyprofile.objects.filter(admin_id=user.id)
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        myprofile = Myprofile(myuser=user)

    if request.method == 'POST':
        username = request.POST['name']
        email = request.POST['email']
        officialemail = request.POST['Officialemail']
        dob = request.POST['dob']
        gender = request.POST['gender']
        phone = request.POST['phone']
        alternativephone = request.POST['alternatephone']
        bldgrp = request.POST['bldgrp']
        marital = request.POST['marital']
        address = request.POST['address']
        personaladdress = request.POST['permanentaddress']
        housetype = request.POST['housetype']
        crntresidencedate = request.POST['currentresidancedate']
        crntcitydate = request.POST['currentcitydate']

        user.username = username
        user.email = email
        user.dob = dob
        user.gender = gender
        user.phone = phone
        user.save()

        myprofile.offemail = officialemail
        myprofile.altphone = alternativephone
        myprofile.bldgrp = bldgrp
        myprofile.marital = marital
        myprofile.address = address
        myprofile.peraddress = personaladdress
        myprofile.housetype = housetype
        myprofile.crntresidencedate = crntresidencedate
        myprofile.crntcitydate = crntcitydate
        img_file = request.FILES.get('img', None)
        if img_file is not None:
            f = FileSystemStorage()
            f1 = f.save(img_file.name, img_file)
            myprofile.image = f1

        myprofile.save()
        return redirect('personalinfonav')

    context = {
        'myprofile': myprofile,
        'c': c[0] if c.exists() else c
    }
    return render(request, "index/personalinfo.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def emppersonal_info_nav(request):
    user = request.user
    admin_id = User.objects.get(id=user.id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        return redirect('empcreatepersonalinfo')
    context = {
        'myprofile': myprofile,
        'c': c[0] if c.exists() else c
    }

    return render(request, "Employee/personalinfo.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empcreate_personal_info(request):
    user = request.user
    user_id = request.user.id
    admin_id = User.objects.get(id=user.id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    leave_notification = LeaveNotification.objects.filter(user=user_id)
    notifications = []
    for i in leave_notification:
        if i.admin_id == 0:
            profile = Myprofile.objects.filter(myuser=i.user).first()
        else:
            profile = Myprofile.objects.filter(myuser__id=i.admin_id).first()

        notifications.append(
            {
                "message": i.message,
                "image_url": profile.image.url
                if profile and profile.image
                else "/static/logo/userlogo.png",
                "notification_id": i.id,
            }
        )
    try:
        myprofile = Myprofile.objects.get(myuser=user)
    except Myprofile.DoesNotExist:
        myprofile = Myprofile(myuser=user)

    if request.method == 'POST':
        username = request.POST['name']
        email = request.POST['email']
        officialemail = request.POST['Officialemail']
        dob = request.POST['dob']
        gender = request.POST['gender']
        phone = request.POST['phone']
        alternativephone = request.POST['alternatephone']
        bldgrp = request.POST['bldgrp']
        marital = request.POST['marital']
        address = request.POST['address']
        personaladdress = request.POST['permanentaddress']
        housetype = request.POST['housetype']
        crntresidencedate = request.POST['currentresidancedate']
        crntcitydate = request.POST['currentcitydate']

        user.username = username
        user.email = email
        user.dob = dob
        user.gender = gender
        user.phone = phone
        user.save()

        myprofile.offemail = officialemail
        myprofile.altphone = alternativephone
        myprofile.bldgrp = bldgrp
        myprofile.marital = marital
        myprofile.address = address
        myprofile.peraddress = personaladdress
        myprofile.housetype = housetype
        myprofile.crntresidencedate = crntresidencedate
        myprofile.crntcitydate = crntcitydate
        img_file = request.FILES.get('img', None)
        if img_file is not None:
            f = FileSystemStorage()
            f1 = f.save(img_file.name, img_file)
            myprofile.image = f1

        myprofile.save()
        return redirect('emppersonalinfonav')

    context = {
        'myprofile': myprofile,
        "notifications": notifications,
        'c': c[0] if c.exists() else c
    }
    return render(request, "Employee/personalinfo.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def work_nav(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    x = Workhistory.objects.filter(myuser_1__id=request.user.id)

    dsn = Designation.objects.filter(admin_id=admin_id)
    dpt = Department.objects.filter(admin_id=admin_id)
    sd = Subdepartment.objects.filter(admin_id=admin_id)

    jttl = Job.objects.filter(admin_id=admin_id)
    wr = Worklocation.objects.filter(admin_id=admin_id)

    v = User.STATUS
    b = User.type

    y = {
        "k": k[0] if k.exists() else k,
        "data": c[0] if c.exists() else c,

    }

    return render(request, "index/workmypro.html",
                  {"x": x, 'dsn': dsn, 'dpt': dpt, 'jttl': jttl, 'status': v, 'types': b, 'sd': sd, **y, 'wr': wr})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empwork_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    k = Myprofile.objects.filter(myuser__id=user_id)
    c = companyprofile.objects.filter(admin_id=admin_id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,

    }

    # return render(request,"Employee/workmypro.html",y)
    return render(request, "Employee/workmypro.html", x)


def create_work(request, userid):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    # v=User.objects.filter(id=uid_11).values()
    m = User.objects.filter(id=userid)
    # u=User.objects.get(id=uid_11)
    # c=companyprofile.objects.all()
    dsn = Designation.objects.all()
    dpt = Department.objects.all()
    jttl = Job.objects.all()
    sdpt = Subdepartment.objects.all()
    wr = Worklocation.objects.all()

    if request.method == "POST":
        joindate = request.POST.get("Joindate")
        proper = request.POST.get("Probationperiod")
        etyp = request.POST.get("Emptype")
        # wrklc=request.POST.get("Workloc")
        estts = request.POST.get("Empstatus")
        wrkex = request.POST.get("Workexp")

        designation_id = request.POST.get('designation')
        designation = Designation.objects.get(id=designation_id)

        department_id = request.POST.get('department')
        department = Department.objects.get(id=department_id)

        subdepartment_id = request.POST.get('subdepartment')
        subdepartment = Subdepartment.objects.get(id=subdepartment_id)

        jobtitle_id = request.POST.get('jobtitle')
        jobtitle = Job.objects.get(id=jobtitle_id)

        wrklcn_id = request.POST.get('wrklcn')
        wrklcn = Worklocation.objects.get(id=wrklcn_id)
        m.update(datejoin=joindate, probperiod=proper, emptype=etyp, wrklcn=wrklcn, status=estts, wrkexp=wrkex,
                 designation=designation, jobtitle=jobtitle, department=department, subdepartment=subdepartment
                 )
        # u_11=User.objects.get(id=uid_11)
        # m.update(designation=designame,jobtitle=jobttle,department=deptname,subdepartment=subdeptname,status=estts,emptype=etyp,probperiod=proper)
        # m.update(datejoin=joindate,probperiod=proper,emptype=etyp,wrkloc=wrklc,empstts=estts,wrkexp=wrkex,design=designame,
        # jobtitle=jobttle,dept=deptname,subdept=subdept)

        return redirect('worknav')

    return render(request, 'index/workmypro.html',
                  {'status': User.STATUS, 'types': User.type, 'k': k, 'dsn': dsn, 'dpt': dpt, 'sdpt': sdpt, 'm': m,
                   'wr': wr})


def empcreate_work(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    # v=User.objects.filter(id=uid_11).values()
    m = User.objects.filter(pk=1)

    dsn = Designation.objects.all()
    dpt = Department.objects.all()
    jttl = Job.objects.all()
    sdpt = Subdepartment.objects.all()
    if request.method == "POST":
        joindate = request.POST.get("Joindate")
        proper = request.POST.get("Probationperiod")
        etyp = request.POST.get("Emptype")
        wrklc = request.POST.get("Workloc")
        estts = request.POST.get("Empstatus")
        wrkex = request.POST.get("Workexp")

        desig = request.POST.get("Designation")
        designame = Designation.objects.get(name=desig)

        job = request.POST.get("Jobtitle")
        jobttle = Job.objects.get(name=job)

        depart = request.POST.get("Department")
        deptname = Department.objects.get(name=depart)

        subdept = request.POST.get("Subdepartment")
        # subdeptname=Subdepartment.objects.get(subname=subdept)
        k.update(datejoin=joindate, probperiod=proper, emptype=etyp, wrkloc=wrklc, empstts=estts, wrkexp=wrkex,
                 design=designame, jobtitle=jobttle, dept=deptname, subdept=subdept
                 )
        # u_11=User.objects.get(id=uid_11)
    # m.update(designation=designame,jobtitle=jobttle,department=deptname,subdepartment=subdeptname,status=estts,emptype=etyp,probperiod=proper)
    # m.update(datejoin=joindate,probperiod=proper,emptype=etyp,wrkloc=wrklc,empstts=estts,wrkexp=wrkex,design=designame,
    # jobtitle=jobttle,dept=deptname,subdept=subdept)
    return redirect('empworknav')
    # return render(request,'Employee/workmypro.html',{'status':User.STATUS,'types':User.type,'k':k,'dsn':dsn,'dpt':dpt,'sdpt':sdpt,'m':m})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_work_history(request, uid_1):
    if request.method == "POST":
        # dept1=Myprofile.objects.get(id=did)
        dept1 = request.POST.get("Department1")
        desig1 = request.POST.get("Designation1")
        frm = request.POST.get("From")
        to = request.POST.get("To")
        u_1 = User.objects.get(id=uid_1)
        Workhistory.objects.create(
            dep1=dept1, design1=desig1, from_date=frm, to_date=to, myuser_1=u_1)
    return redirect('worknav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def team_nav(request):
    user_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    directreports = Directreports.objects.filter(admin_id=user_id)
    c = companyprofile.objects.filter(admin_id=user_id)
    designation = Designation.objects.filter(admin_id=user_id)
    department = Department.objects.filter(admin_id=user_id)
    users = User.objects.filter(
        (Q(id=user_id) | Q(admin_id=user_id)), 
        status="Active"
    ).exclude(id=user_id) 

    l = Reportingmanager.objects.filter(userid=user_id)

    rpt_users = Reportingmanager.objects.filter(myuser_2=user_id)
    directreport_users = []
    for rpt_user in rpt_users:
        user = User.objects.filter(id=rpt_user.userid).first()
        if user:
            directreport_users.append(user)
    
    # print("Users: ", users)
    print("Reporting manager: ", rpt_users)
    # print("directreport_users: ", directreport_users)

    context = {
        "data": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "directreports": directreports,
        # "type_choices": type_choices,
        "users": users,
        "designation": designation,
        "department": department,
        'directreport_users': directreport_users
    }
    return render(request, "index/empteamdetails.html", context)


@login_required(login_url='login')
def add_direct_report(request):
    try:
        user_id = request.user.id  # Current admin user
        direct_report_user_id = request.POST.get("Name")  # Select field name in modal

        if not direct_report_user_id:
            messages.error(request, "No user selected.")
            return redirect('teamnav')

        # Prevent adding yourself
        if str(user_id) == str(direct_report_user_id):
            messages.error(request, "You cannot assign yourself as a direct report.")
            return redirect('teamnav')

        # Check if already assigned
        if Directreports.objects.filter(admin_id=user_id, directreport_id=direct_report_user_id).exists():
            messages.warning(request, "This user is already your direct report.")
            return redirect('teamnav')

        # Save the direct report
        Directreports.objects.create(
            admin_id=user_id,
            directreport_id=direct_report_user_id
        )

        messages.success(request, "Direct report added successfully.", extra_tags='bg-success text-white')
    except Exception as e:
        messages.error(request, f"Error adding direct report: {str(e)}")
    return redirect('teamnav')


@login_required(login_url='login')
def deletedirectreport(request, report_id):
    try:
        report = Directreports.objects.get(id=report_id)
        messages.success(request, "Direct report deleted successfully.", extra_tags='bg-success text-white')
    except Directreports.DoesNotExist:
        messages.error(request, "Direct report not found.")
    return redirect('teamnav')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empteam_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    x = Directreports.objects.filter(myuser_3__id=request.user.id)

    l = Reportingmanager.objects.filter(userid=user_id)

    print("The l value: ", l, user_id)
    rpt_users = Reportingmanager.objects.filter(myuser_2=user_id)
    directreport_users = []
    for rpt_user in rpt_users:
        user = User.objects.filter(id=rpt_user.userid).first()
        print('rpt user: ', rpt_user.myuser_2.all().values_list("username"), user)
        if user:
            directreport_users.append(user)

    print("directreport:", directreport_users)
    y = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'l': l,
        'x': x,
        'directreport_users': directreport_users
    }
    return render(request, "Employee/teammypro.html", y)


def table_reporting_manager(request):
    users = User.objects.all()
    if request.method == 'POST':
        name_id = request.POST.get('Name1')
        emptype = request.POST.get('Type')
        user = User.objects.get(id=name_id)

        if emptype == "Primary" and Reportingmanager.objects.filter(type="Primary").exists():
            messages.info(request, "A Primary reporting manager already exists.")
            return redirect('teamnav')

        if emptype == "Secondary" and Reportingmanager.objects.filter(type="Secondary").exists():
            messages.info(request, "A Secondary reporting manager already exists.")
            return redirect('teamnav')

        reporting_manager = Reportingmanager.objects.create(
            type=emptype, userid=request.user.id)
        reporting_manager.myuser_2.add(user)
        reporting_manager.save()

        return redirect('teamnav')

    context = {'users': users}
    return render(request, "index/teammypro.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_reporting_manager(request, uid_2):
    if request.method == "POST":
        nm1 = request.POST.get("eName1")
        typ = request.POST.get("eType")
        dept2 = request.POST.get("eDepartment2")
        desig2 = request.POST.get("eDesignation2")
        u_2 = User.objects.get(id=uid_2)
        Reportingmanager.objects.create(
            name1=nm1, typ=typ, dept2=dept2, design2=desig2, myuser_2=u_2)
    return redirect('empteamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_reportingmanager(request):
    if request.method == "POST":
        name = request.POST.get('myuser_2')
        # username = User.objects.get(id=name)
        typ1 = request.POST.get("typ")

        # dept4 = request.POST.get("Department4")
        # desig4 = request.POST.get("Designation4")
        print("DDDDDDDDDDDDDDDDDDDDDD :", name, typ1)

        reportingmanager_id = request.POST.get('reportingmanager_id')
        k = Reportingmanager.objects.filter(id=reportingmanager_id)
        k.update(myuser_2=name, type=typ1)

    return redirect('teamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_reportingmanager(request):
    if request.method == "POST":
        nm3 = request.POST.get("eName3")
        typ1 = request.POST.get("eType1")

        dept4 = request.POST.get("eDepartment4")
        desig4 = request.POST.get("eDesignation4")

        reportingmanager_id = request.POST.get('reportingmanagerid')
        n = Reportingmanager.objects.filter(id=reportingmanager_id)
        n.update(name1=nm3, typ=typ1, dept2=dept4, design2=desig4, )
        print(n)

    return redirect('empteamnav')


def delete_reportingmanager(request, r_id):
    k = Reportingmanager.objects.get(id=r_id)
    k.delete()
    print(k)
    return redirect('teamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdelete_reportingmanager(request, r_id):
    k = Reportingmanager.objects.get(id=r_id)
    k.delete()
    print(k)
    return redirect('empteamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_direct_reports(request):
    users = User.objects.all()
    if request.method == 'POST':
        name_id = request.POST.get('Name')
        user = User.objects.get(id=name_id)
        Directreports.objects.create(myuser_3=user, admin_id=request.user.id)
        return redirect('teamnav')

    context = {'users': users}

    return render(request, "index/teammypro.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_direct_reports(request, uid_3):
    if request.method == "POST":
        nm2 = request.POST.get("eName2")

        dept3 = request.POST.get("eDepartment3")
        desig3 = request.POST.get("eDesignation3")
        u_3 = User.objects.get(id=uid_3)
        Directreports.objects.create(
            name2=nm2, dept3=dept3, design3=desig3, myuser_3=u_3)
    return redirect('empteamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_directreports(request):
    designation = Designation.objects.filter(admin_id=request.user.id)
    department = Department.objects.filter(admin_id=request.user.id)
    if request.method == "POST":
        username = request.POST.get('username')
        name = User.objects.get(id=username)
        designation = request.POST.get('designation')
        des = Designation.objects.get(id=designation)
        department = request.POST.get('department')
        dep = Department.objects.get(id=department)
        directreport_id = request.POST.get('directreportid')

        j = Directreports.objects.filter(id=directreport_id)

        j.update(myuser_3=name, design3=des,
                 dept3=dep, admin_id=request.user.id)
        print(j)

    return redirect('teamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_directreports(request):
    if request.method == "POST":
        rnm4 = request.POST.get("eName3")
        rdept5 = request.POST.get("eDepartment4")
        rdesig5 = request.POST.get("eDesignation4")
        directreport_id = request.POST.get('directreportid')

        j = Directreports.objects.filter(id=directreport_id)

        j.update(name2=rnm4, dept3=rdept5, design3=rdesig5)
        print(j)

    return redirect('empteamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_directreport(request, myuser_3__id):
    k = Directreports.objects.get(id=myuser_3__id)
    k.delete()
    print(k)
    return redirect('teamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdelete_directreport(request, myuser_3__id):
    k = Directreports.objects.get(id=myuser_3__id)
    k.delete()
    print(k)
    return redirect('empteamnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def education_nav(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Educationalinfo.objects.filter(myuser_4__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "is_view_education": False
    }
    return render(request, "index/eduinfomypro.html", x)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empeducation_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Educationalinfo.objects.filter(myuser_4__id=request.user.id)
    y = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "l": l,
    }
    print("EMPLOYEE")
    return render(request, "Employee/eduinfomypro.html", y)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_edu_info(request, uid_4):
    if request.method == "POST":
        quali = request.POST.get("Qualification")
        course = request.POST.get("Course")
        insti = request.POST.get("Institutename")
        psout = request.POST.get("Passout")
        perc = request.POST.get("Percentage")
        u_4 = User.objects.get(id=uid_4)
        Educationalinfo.objects.create(
            qualification=quali, course=course, institute=insti, passout=psout, percent=perc, myuser_4=u_4)
    return redirect('educationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_edu_info(request, uid_4):
    if request.method == "POST":
        quali = request.POST.get("eQualification")
        course = request.POST.get("eCourse")
        insti = request.POST.get("eInstitutename")
        psout = request.POST.get("ePassout")
        perc = request.POST.get("ePercentage")
        u_4 = User.objects.get(id=uid_4)
        Educationalinfo.objects.create(
            qualification=quali, course=course, institute=insti, passout=psout, percent=perc, myuser_4=u_4)
    return redirect('empeducationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_eduinfo(request, myuser_4__id):
    h = Educationalinfo.objects.get(id=myuser_4__id)
    h.delete()
    print(h)
    return redirect('educationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdelete_eduinfo(request, myuser_4__id):
    h = Educationalinfo.objects.get(id=myuser_4__id)
    h.delete()
    print(h)
    return redirect('empeducationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_eduinfo(request):
    if request.method == "POST":
        qualifcn = request.POST.get("Qualification1")
        cours = request.POST.get("Course1")
        institn = request.POST.get("Institutename1")
        psouty = request.POST.get("Passout1")
        perctge = request.POST.get("Percentage1")
        edu_id = request.POST.get('edu_id')
        x = Educationalinfo.objects.filter(id=edu_id)
        x.update(qualification=qualifcn, course=cours,
                 institute=institn, passout=psouty, percent=perctge, )

    return redirect('educationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_eduinfo(request):
    if request.method == "POST":
        qualifcn = request.POST.get("eQualification1")
        cours = request.POST.get("eCourse1")
        institn = request.POST.get("eInstitutename1")
        psouty = request.POST.get("ePassout1")
        perctge = request.POST.get("ePercentage1")
        edu_id = request.POST.get('eduid')
        x = Educationalinfo.objects.filter(id=edu_id)
        x.update(qualification=qualifcn, course=cours,
                 institute=institn, passout=psouty, percent=perctge, )

    return redirect('empeducationnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def family_nav(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Familymembers.objects.filter(myuser_5__id=request.user.id)
    x = Emergencycontact.objects.filter(myuser_6__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    y = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "x": x,
        "is_view_family": False
    }
    return render(request, "index/familymypro.html", y)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empfamily_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Familymembers.objects.filter(myuser_5__id=request.user.id)
    x = Emergencycontact.objects.filter(myuser_6__id=request.user.id)

    y = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        "l": l,
        "x": x,
    }

    return render(request, "Employee/familymypro.html", y)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_fam_members(request, uid_5):
    if request.method == "POST":
        nm3 = request.POST.get("Name3")
        reln = request.POST.get("Relationship")
        dob1 = request.POST.get("DOB1")
        dpndnt = request.POST.get("Dependant")
        u_5 = User.objects.get(id=uid_5)
        Familymembers.objects.create(
            name3=nm3, relation=reln, dob1=dob1, dependant=dpndnt, myuser_5=u_5)
    return redirect('familynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_fam_members(request, uid_5):
    if request.method == "POST":
        nm3 = request.POST.get("eName3")
        reln = request.POST.get("eRelationship")
        dob1 = request.POST.get("eDOB1")
        dpndnt = request.POST.get("eDependant")
        u_5 = User.objects.get(id=uid_5)
        Familymembers.objects.create(
            name3=nm3, relation=reln, dob1=dob1, dependant=dpndnt, myuser_5=u_5)
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_fam_members(request):
    if request.method == "POST":
        nm4 = request.POST.get("Name6")
        reln1 = request.POST.get("Relationship1")
        dob2 = request.POST.get("DOB2")
        dpndnt1 = request.POST.get("Dependant1")
        fam_member_id = request.POST.get('fam_memberid')

        d = Familymembers.objects.filter(id=fam_member_id)
        d.update(name3=nm4, relation=reln1, dob1=dob2, dependant=dpndnt1)
    return redirect('familynav')
    # return render(request,"index/familymypro.html",{'data':data})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_fam_members(request):
    if request.method == "POST":
        nm4 = request.POST.get("Name6")
        reln1 = request.POST.get("Relationship1")
        dob2 = request.POST.get("DOB2")
        dpndnt1 = request.POST.get("Dependant1")
        fam_member_id = request.POST.get('fam_memberid')

        d = Familymembers.objects.filter(id=fam_member_id)
        d.update(name3=nm4, relation=reln1, dob1=dob2, dependant=dpndnt1)
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_fam_members(request, myuser_5__id):
    k = Familymembers.objects.get(id=myuser_5__id)
    k.delete()
    print(k)
    return redirect('familynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdelete_fam_members(request, myuser_5__id):
    k = Familymembers.objects.get(id=myuser_5__id)
    k.delete()
    print(k)
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_emergency_contact(request, uid_6):
    if request.method == "POST":
        nm4 = request.POST.get("Name4")
        reln1 = request.POST.get("Relationship1")
        phn1 = request.POST.get("Phone1")
        u_6 = User.objects.get(id=uid_6)
        Emergencycontact.objects.create(
            name4=nm4, relation1=reln1, phone1=phn1, myuser_6=u_6)
    return redirect('familynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_emergency_contact(request, uid_6):
    if request.method == "POST":
        nm4 = request.POST.get("Name4")
        reln1 = request.POST.get("Relationship1")
        phn1 = request.POST.get("Phone1")
        u_6 = User.objects.get(id=uid_6)
        Emergencycontact.objects.create(
            name4=nm4, relation1=reln1, phone1=phn1, myuser_6=u_6)
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_emergency_contact(request):
    if request.method == "POST":
        nm5 = request.POST.get("Name5")
        reln2 = request.POST.get("Relationship2")
        phn2 = request.POST.get("Phone2")

        emergency_id = request.POST.get('emergencyid')
        b = Emergencycontact.objects.filter(id=emergency_id)
        b.update(name4=nm5, relation1=reln2, phone1=phn2)
    return redirect('familynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_emergency_contact(request):
    if request.method == "POST":
        nm5 = request.POST.get("Name5")
        reln2 = request.POST.get("Relationship2")
        phn2 = request.POST.get("Phone2")

        emergency_id = request.POST.get('emergencyid')
        b = Emergencycontact.objects.filter(id=emergency_id)
        b.update(name4=nm5, relation1=reln2, phone1=phn2)
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_emergency_contact(request, myuser_6__id):
    k = Emergencycontact.objects.get(id=myuser_6__id)
    k.delete()
    return redirect('familynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdelete_emergency_contact(request, myuser_6__id):
    k = Emergencycontact.objects.get(id=myuser_6__id)
    k.delete()
    return redirect('empfamilynav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def documents_nav(request):
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    uploads = Uploadeddocs.objects.filter(myuser__id=request.user.id)
    certificates = Certifications.objects.filter(myuser_8__id=request.user.id)
    work = Work.objects.filter(myuser_9__id=request.user.id)
    company = companyprofile.objects.filter(admin_id=request.user.id)
    proofs_object = Proof.objects.all()
    print("proofs_object : ", proofs_object)

    proofs = {
        proof.id: proof.proof_name for uploaded_doc in uploads for proof in uploaded_doc.proof.all()}

    print("FFFFFFFFFFFFFFFF :", proofs)

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "uploads": uploads,
        "certificates": certificates,
        "work": work,
        "company": company[0] if company.exists() else company,
        "proofs": proofs,
        "proofs_object": proofs_object
    }
    return render(request, "index/documentsmypro.html", context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_document(request, doc_id):
    try:

        result = get_object_or_404(Uploadeddocs, id=doc_id)
        result.delete()
        return redirect("documentsnav")
    except Uploadeddocs.DoesNotExist:
        messages.info(request, "No data found")
        return redirect("documentsnav")

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdocuments_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    uploads = Uploadeddocs.objects.filter(myuser__id=request.user.id)
    proofs_object = Proof.objects.all()
    proofs = {
        proof.id: proof.proof_name for uploaded_doc in uploads for proof in uploaded_doc.proof.all()}

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "uploads": uploads,
        "data": data[0] if data.exists() else data,
        "proofs": proofs,
        "proofs_object": proofs_object
    }

    return render(request, "Employee/documentsmypro.html", context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_emp_document(request, doc_id):
    try:

        result = get_object_or_404(Uploadeddocs, id=doc_id)
        result.delete()
        return redirect("empdocumentsnav")
    except Uploadeddocs.DoesNotExist:
        messages.info(request, "No data found")
        return redirect("empdocumentsnav")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_uploaded_docs(request, uid_7):
    if request.method == "POST":
        typ1 = request.POST.get("Type1")
        id_no = request.POST.get("Idno")
        proof = request.POST.getlist("proof", [])

        user_id = User.objects.get(id=uid_7)
        img = request.FILES['img']
        obj = FileSystemStorage()
        img_obj = obj.save(img.name, img)
        uploads = Uploadeddocs.objects.create(
            type1=typ1, id_no=id_no, myuser=user_id, image1=img_obj)

        uploads.proof.add(*proof)
    return redirect('documentsnav')


def update_docs(request):
    if request.method == "POST":
        editdoc_id = request.POST.get("editdoc")
        type1 = request.POST.get("Type")
        IdNo = request.POST.get("IdNo")
        proof = request.POST.getlist("proof", [])

        editdoc = Uploadeddocs.objects.get(id=editdoc_id)

        img = request.FILES.get('docimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editdoc.image1 = img_obj

        editdoc.type1 = type1
        editdoc.id_no = IdNo
        editdoc.save()

        editdoc.proof.clear()

        editdoc.proof.add(*proof)
    return redirect('documentsnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emptable_uploaded_docs(request, uid_7):
    if request.method == "POST":
        typ1 = request.POST.get("Type1")
        id_no = request.POST.get("Idno")
        proof = request.POST.getlist("proof", [])

        user_id = User.objects.get(id=uid_7)
        img = request.FILES['img']
        obj = FileSystemStorage()
        img_obj = obj.save(img.name, img)
        uploads = Uploadeddocs.objects.create(
            type1=typ1, id_no=id_no, myuser=user_id, image1=img_obj)

        uploads.proof.add(*proof)
    return redirect('empdocumentsnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empupdate_docs(request):
    if request.method == "POST":
        editdoc_id = request.POST.get("editdoc")
        type1 = request.POST.get("Type")
        IdNo = request.POST.get("IdNo")
        proof = request.POST.getlist("proof", [])

        editdoc = Uploadeddocs.objects.get(id=editdoc_id)

        img = request.FILES.get('docimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editdoc.image1 = img_obj

        editdoc.type1 = type1
        editdoc.id_no = IdNo
        editdoc.save()

        editdoc.proof.clear()

        editdoc.proof.add(*proof)
    return redirect('empdocumentsnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_uploaded_docs(request):
    if request.method == "POST":
        utyp1 = request.POST.get("uType1")
        uid_no = request.POST.get("uIdno")
        uupldby = request.POST.get("uUploadedby")
        upload_id = request.POST.get('uploadid')
        k = Uploadeddocs.objects.filter(id=upload_id)
        k.update(type1=utyp1, id_no=uid_no, uploadedby=uupldby)
    return redirect('documentsnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def certifications(request):
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    certificates = Certifications.objects.filter(myuser_8__id=request.user.id)
    company = companyprofile.objects.filter(admin_id=request.user.id)

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "certificates": certificates,
        "company": company[0] if company.exists() else company,
        "is_view_certification": False
    }
    return render(request, "index/certifications.html", context)


def table_certifications(request, uid_8):
    if request.method == "POST":
        ctitle = request.POST.get("Coursetitle")
        upldby1 = request.POST.get("Uploadedby1")
        typ2 = request.POST.get("Type2")
        verifi = request.POST.get("Verification")
        img = request.FILES['img']
        f = FileSystemStorage()
        f1 = f.save(img.name, img)
        u_8 = User.objects.get(id=uid_8)
        Certifications.objects.create(
            coursetitle=ctitle, uploadedby1=upldby1, type2=typ2, verification=verifi, myuser_8=u_8, image2=f1)
    return redirect('certifications')


def edit_certifications(request):
    if request.method == "POST":
        editcert_id = request.POST.get("editcert")
        coursetype = request.POST.get("coursetype")
        crttitle = request.POST.get("crttitle")

        editcert = Certifications.objects.get(id=editcert_id)

        img = request.FILES.get('certimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editcert.image2 = img_obj

        editcert.coursetitle = coursetype
        editcert.type2 = crttitle
        editcert.save()

    return redirect('view_certification', id=editcert.myuser_8.id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empcertifications(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    certificates = Certifications.objects.filter(myuser_8__id=request.user.id)

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "certificates": certificates,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "Employee/certifications.html", context)


def edit_empcertifications(request):
    if request.method == "POST":
        editcert_id = request.POST.get("editcert")
        coursetype = request.POST.get("coursetype")
        crttitle = request.POST.get("crttitle")

        editcert = Certifications.objects.get(id=editcert_id)

        img = request.FILES.get('certimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editcert.image2 = img_obj

        editcert.coursetitle = coursetype
        editcert.type2 = crttitle
        editcert.save()

    return redirect('empcertifications')

def delete_empcertifications(request, cert_id):
  
  
    if request.method == 'GET':
        certification = get_object_or_404(Certifications, id=cert_id)  
        certification.delete()  
    return redirect('empcertifications')  


def emptable_certifications(request, uid_8):
    if request.method == "POST":
        ctitle = request.POST.get("Coursetitle")
        upldby1 = request.POST.get("Uploadedby1")
        typ2 = request.POST.get("Type2")
        verifi = request.POST.get("Verification")
        img = request.FILES['img']
        f = FileSystemStorage()
        f1 = f.save(img.name, img)
        u_8 = User.objects.get(id=uid_8)
        Certifications.objects.create(
            coursetitle=ctitle, uploadedby1=upldby1, type2=typ2, verification=verifi, myuser_8=u_8, image2=f1)
    return redirect('empcertifications')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_certifications(request):
    if request.method == "POST":
        uctitle = request.POST.get("uCoursetitle")
        uupldby1 = request.POST.get("uUploadedby1")
        utyp2 = request.POST.get("uType2")
        uverifi = request.POST.get("uVerification")

        certification_id = request.POST.get('certificationid')
        x = Certifications.objects.filter(id=certification_id)
        x.update(coursetitle=uctitle, uploadedby1=uupldby1,
                 type2=utyp2, verification=uverifi)
    return redirect('documentsnav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def doc_work(request):
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    company = companyprofile.objects.filter(admin_id=request.user.id)
    work = Work.objects.filter(myuser_9__id=request.user.id)
   

    for item in work:
        print(f"Processing work item with uploadedon: {item.uploadedon}")
        
        if item.uploadedon:
            try:    
                uploaded_on_date = datetime.strptime(item.uploadedon, '%d-%m-%Y') 
  
                formatted_date = uploaded_on_date.strftime('%d %B %Y')
                item.uploadedon_formatted = formatted_date
                print(f"Formatted uploadedon: {item.uploadedon} -> {formatted_date}")
            except ValueError:
                item.uploadedon_formatted = item.uploadedon  


    context = {
            "profile": profile[0] if profile.exists() else profile,
            "work": work,
            "company": company[0] if company.exists() else company,
            "is_view_docwork": False
        }
    return render(request, "index/doc_work.html", context)

# @login_required(login_url='login')
# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# def deletedoc_work(request, work_id):

#     if request.method == 'GET':
#         work = get_object_or_404(Work, id=work_id)
#         work.delete()
#     return redirect('view_docwork', id=work.myuser_9.id)

@login_required(login_url='login')
def deletedoc_work(request, work_id):
    work = get_object_or_404(Work, id=work_id)

    if request.method == 'GET':

        work.delete()
        related_works = Work.objects.filter(myuser_9=work.myuser_9)
     
    return redirect('view_docwork', id=work.myuser_9.id)


def table_work_docs(request, uid_9):
    l = Work.objects.all()
    if request.method == "POST":
        nm5 = request.POST.get("Name6")
        ds1 = request.POST.get("Description")
        upon = date.today()
        upon1 = upon.strftime("%d-%m-%Y")
        img = request.FILES['img']
        f = FileSystemStorage()
        f1 = f.save(img.name, img)

        u_9 = User.objects.get(id=uid_9)
        Work.objects.create(name5=nm5, description1=ds1,
                            uploadedon=upon1, myuser_9=u_9, image3=f1)

    return redirect('doc_work')


def edit_doc_work(request):
    if request.method == "POST":
        editwork_id = request.POST.get("editwork")
        name = request.POST.get("Name")
        description = request.POST.get("Description")
        todaydate = date.today().strftime('%d-%m-%Y')

        editwork = Work.objects.get(id=editwork_id)

        img = request.FILES.get('workimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editwork.image3 = img_obj

        editwork.name5 = name
        editwork.description1 = description
        editwork.uploadedon = todaydate
        editwork.save()
    
    return redirect('view_docwork', id=editwork.myuser_9.id)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empdoc_work(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id

    data = companyprofile.objects.filter(admin_id=admin_id)
    profile = Myprofile.objects.filter(myuser__id=request.user.id)
    work = Work.objects.filter(myuser_9__id=request.user.id)

    for item in work:
        print(f"Processing work item with uploadedon: {item.uploadedon}")
        
        if item.uploadedon:
            try:    
                uploaded_on_date = datetime.strptime(item.uploadedon, '%d-%m-%Y') 
  
                formatted_date = uploaded_on_date.strftime('%d %B %Y')
                item.uploadedon_formatted = formatted_date
                print(f"Formatted uploadedon: {item.uploadedon} -> {formatted_date}")
            except ValueError:
                item.uploadedon_formatted = item.uploadedon  

    context = {
        "profile": profile[0] if profile.exists() else profile,
        "work": work,
        "data": data[0] if data.exists() else data,
    }
    
    return render(request, "Employee/doc_work.html", context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def deleteempdoc_work(request, work_id):
    
    if request.method == 'GET':
        work = get_object_or_404(Work, id=work_id)  
        work.delete()  
    return redirect('empdoc_work')  #

def emptable_work_docs(request, uid_9):
    l = Work.objects.all()
    if request.method == "POST":
        nm5 = request.POST.get("Name5")
        ds1 = request.POST.get("descrip")
        upon = date.today()
        upon1 = upon.strftime("%d-%m-%Y")
        img = request.FILES['img']
        f = FileSystemStorage()
        f1 = f.save(img.name, img)

        u_9 = User.objects.get(id=uid_9)
        Work.objects.create(name5=nm5, description1=ds1,
                            uploadedon=upon1, myuser_9=u_9, image3=f1)

    return redirect('empdoc_work')


def edit_empdoc_work(request):
    if request.method == "POST":
        editwork_id = request.POST.get("editwork")
        name = request.POST.get("Name")
        description = request.POST.get("Description")
        todaydate = date.today().strftime('%d-%m-%Y')
       

        editwork = Work.objects.get(id=editwork_id)

        img = request.FILES.get('workimg')
        if img:
            obj = FileSystemStorage()
            img_obj = obj.save(img.name, img)
            editwork.image3 = img_obj

        editwork.name5 = name
        editwork.description1 = description
        editwork.uploadedon = todaydate
        editwork.save()

    return redirect('empdoc_work')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_work_docs(request):
    if request.method == "POST":
        unm5 = request.POST.get("uName5")
        uupldby2 = request.POST.get("uUploadedby2")
        uupldon = request.POST.get("uUploadedon")
        work_id = request.POST.get('workid')
        k = Work.objects.filter(id=work_id)
        k.update(name5=unm5, uploadedby2=uupldby2, uploadedon=uupldon, )
    return redirect('documentsnav')


# def work_week(request):
#     k = Myprofile.objects.filter(myuser__id=request.user.id)
#     data = companyprofile.objects.all()
#     work_week = AssignWorkWeek.object.filter(user_id=request.user.id)

#     context = {
#             'work_week': work_week,
#             "k": k[0] if k.exists() else k,
#             "data": data[0] if data.exists() else data,
#         }
#     return render(request, 'index/workweekmypro.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def workweek_nav(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    work_week = AssignWorkWeek.objects.filter(user_id=admin_id)
    context = {
        'work_week': work_week,
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "is_view_workweek": False
    }
    return render(request, 'index/workweekmypro.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def attendance_workweek(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    work_week = Workweek.objects.filter(admin_id=admin_id)
    company_rule = CompanyRules.objects.filter(admin_id=admin_id)
    context = {
        'work_week': work_week,
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "company_rules": company_rule
    }
    return render(request, 'index/add_workweek.html', context)


def add_workweek(request):
    admin_id = request.user.id
    if request.method == 'POST':
        off_day = request.POST.get('off_day')
        description = request.POST.get('description')
        rule_name = request.POST.get('rule_name')
        #    = request.POST.get('  ')
        half_day = request.POST.get('half_day')
        day_1 = request.POST.get('day_1_color')
        day_2 = request.POST.get('day_2_color')
        day_3 = request.POST.get('day_3_color')
        day_4 = request.POST.get('day_4_color')
        day_5 = request.POST.get('day_5_color')
        day_6 = request.POST.get('day_6_color')
        day_7 = request.POST.get('day_7_color')
        day_8 = request.POST.get('day_8_color')
        day_9 = request.POST.get('day_9_color')
        day_10 = request.POST.get('day_10_color')
        day_11 = request.POST.get('day_11_color')
        day_12 = request.POST.get('day_12_color')
        day_13 = request.POST.get('day_13_color')
        day_14 = request.POST.get('day_14_color')
        day_15 = request.POST.get('day_15_color')
        day_16 = request.POST.get('day_16_color')
        day_17 = request.POST.get('day_17_color')
        day_18 = request.POST.get('day_18_color')
        day_19 = request.POST.get('day_19_color')
        day_20 = request.POST.get('day_20_color')
        day_21 = request.POST.get('day_21_color')
        day_22 = request.POST.get('day_22_color')
        day_23 = request.POST.get('day_23_color')
        day_24 = request.POST.get('day_24_color')
        day_25 = request.POST.get('day_25_color')
        day_26 = request.POST.get('day_26_color')
        day_27 = request.POST.get('day_27_color')
        day_28 = request.POST.get('day_28_color')
        day_29 = request.POST.get('day_29_color')
        day_30 = request.POST.get('day_30_color')
        day_31 = request.POST.get('day_31_color')
        day_32 = request.POST.get('day_32_color')
        day_33 = request.POST.get('day_33_color')
        day_34 = request.POST.get('day_34_color')
        day_35 = request.POST.get('day_35_color')

        Workweek.objects.create(off_day=off_day, description=description, rule_name=rule_name, half_day=half_day,
                                day_1=day_1, day_2=day_2, day_3=day_3, day_4=day_4, day_5=day_5, day_6=day_6,
                                day_7=day_7, day_8=day_8, day_9=day_9, day_10=day_10,
                                day_11=day_11, day_12=day_12, day_13=day_13, day_14=day_14, day_15=day_15,
                                day_16=day_16, day_17=day_17, day_18=day_18, day_19=day_19, day_20=day_20,
                                day_21=day_21, day_22=day_22, day_23=day_23, day_24=day_24, day_25=day_25,
                                day_26=day_26, day_27=day_27, day_28=day_28,
                                day_29=day_29, day_30=day_30, day_31=day_31, day_32=day_32, day_33=day_33,
                                day_34=day_34, day_35=day_35, admin_id=admin_id)
        return redirect('attendance_workweek')


def edit_workweek(request):
    admin_id = request.user.id
    if request.method == 'POST':
        workweek_id = request.POST.get('workweekid')
        workweek = Workweek.objects.get(id=workweek_id)
        off_day = request.POST.get('off_day')
        description = request.POST.get('description')
        rule_name = request.POST.get('rule_name')
        half_day = request.POST.get('half_day')
        day_1 = request.POST.get('day_1_color')
        day_2 = request.POST.get('day_2_color')
        day_3 = request.POST.get('day_3_color')
        day_4 = request.POST.get('day_4_color')
        day_5 = request.POST.get('day_5_color')
        day_6 = request.POST.get('day_6_color')
        day_7 = request.POST.get('day_7_color')
        day_8 = request.POST.get('day_8_color')
        day_9 = request.POST.get('day_9_color')
        day_10 = request.POST.get('day_10_color')
        day_11 = request.POST.get('day_11_color')
        day_12 = request.POST.get('day_12_color')
        day_13 = request.POST.get('day_13_color')
        day_14 = request.POST.get('day_14_color')
        day_15 = request.POST.get('day_15_color')
        day_16 = request.POST.get('day_16_color')
        day_17 = request.POST.get('day_17_color')
        day_18 = request.POST.get('day_18_color')
        day_19 = request.POST.get('day_19_color')
        day_20 = request.POST.get('day_20_color')
        day_21 = request.POST.get('day_21_color')
        day_22 = request.POST.get('day_22_color')
        day_23 = request.POST.get('day_23_color')
        day_24 = request.POST.get('day_24_color')
        day_25 = request.POST.get('day_25_color')
        day_26 = request.POST.get('day_26_color')
        day_27 = request.POST.get('day_27_color')
        day_28 = request.POST.get('day_28_color')
        day_29 = request.POST.get('day_29_color')
        day_30 = request.POST.get('day_30_color')
        day_31 = request.POST.get('day_31_color')
        day_32 = request.POST.get('day_32_color')
        day_33 = request.POST.get('day_33_color')
        day_34 = request.POST.get('day_34_color')
        day_35 = request.POST.get('day_35_color')

        workweek.off_day = off_day
        workweek.description = description
        workweek.rule_name = rule_name
        workweek.half_day = half_day
        workweek.day_1 = day_1
        workweek.day_2 = day_2
        workweek.day_3 = day_3
        workweek.day_4 = day_4
        workweek.day_5 = day_5
        workweek.day_6 = day_6
        workweek.day_7 = day_7
        workweek.day_8 = day_8
        workweek.day_9 = day_9
        workweek.day_10 = day_10
        workweek.day_11 = day_11
        workweek.day_12 = day_12
        workweek.day_13 = day_13
        workweek.day_14 = day_14
        workweek.day_15 = day_15
        workweek.day_16 = day_16
        workweek.day_17 = day_17
        workweek.day_18 = day_18
        workweek.day_19 = day_19
        workweek.day_20 = day_20
        workweek.day_21 = day_21
        workweek.day_22 = day_22
        workweek.day_23 = day_23
        workweek.day_24 = day_24
        workweek.day_25 = day_25
        workweek.day_26 = day_26
        workweek.day_27 = day_27
        workweek.day_28 = day_28
        workweek.day_29 = day_29
        workweek.day_30 = day_30
        workweek.day_31 = day_31
        workweek.day_32 = day_32
        workweek.day_33 = day_33
        workweek.day_34 = day_34
        workweek.day_35 = day_35
        workweek.save()
        return redirect('attendance_workweek')
    return render(request, 'index/add_workweek.html', {'workweekid:workweekid'})


def delete_workweek(request, id):
    workweek_id = Workweek.objects.get(id=id)
    workweek_id.delete()
    return redirect('attendance_workweek')


def view_assignworkweek(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    datas = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id))
    dn = Designation.objects.all()
    dp = Department.objects.all()
    sd = Subdepartment.objects.all()
    jb = Job.objects.all()
    wr = Worklocation.objects.all()
    com_rule = Workweek.objects.filter(admin_id=request.user.id)
    assg_rule = AssignWorkWeek.objects.all()
    query = request.GET.get('search')

    count_user = User.objects.count()

    # page = request.GET.get('page', 1)
    # paginator = Paginator(datas, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)
    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/assign_workweek.html",
                  {'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'datas': datas, 'query': query, 'k': k,
                   'count_user': count_user, 'assg_rule': assg_rule, 'com_rule': com_rule, **x})


def assignworkweek(request):
    workweek = Workweek.objects.filter(admin_id=request.user.id)
    if request.method == 'POST':
        selected_rules = request.POST.get('rule')
        selected_rulesid = Workweek.objects.get(id=selected_rules)
        effective_date = request.POST.get('effdate')
        selected_employees = request.POST.getlist('selected_employees')
        for employee_id in selected_employees:
            try:
                assign_workweek = AssignWorkWeek.objects.get(
                    user_id_id=employee_id)
                assign_workweek.effective_date = effective_date
                assign_workweek.rules_applied = selected_rulesid
                assign_workweek.save()
            except AssignWorkWeek.DoesNotExist:
                AssignWorkWeek.objects.create(
                    user_id_id=employee_id, effective_date=effective_date, rules_applied=selected_rulesid)

        return redirect('view_assignworkweek')
    return render(request, "index/assign_workweek.html", {'com_rule': workweek})


def delete_assignworkweek(request, assign_rule_id):
    assign_rule = AssignWorkWeek.objects.get(id=assign_rule_id)
    assign_rule.delete()
    return redirect('view_assignworkweek')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empworkweek_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)
    work_week = AssignWorkWeek.objects.filter(user_id=user_id)
    for i in work_week:
        print("Work week : ",
              work_week, i.rules_applied.half_day)
    context = {
        'work_week': work_week,
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "Employee/workweekmypro.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def payroll_nav(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
    }
    return render(request, "index/payrollmypro.html", x)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def emppayroll_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)

    regaddress = registeredaddress.objects.filter(admin_id=admin_id)
    print("regofficeaddress :", regaddress)
    bank_details = Bank_account.objects.filter(myuser_11=user_id)
    print("bank_details :", bank_details)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
    print("selected_year :", selected_year, selected_month)
    use_if_block = selected_year == 2024 and selected_month < 10
    if use_if_block:
        print("Selected month and year are less than October 2024")
        assign_salarystructure = AssignSalaryStructure.objects.filter(user_id=user_id, effective_date__month=selected_month, effective_date__year=selected_year).order_by('effective_date').first()
        print("assignsalary :", assign_salarystructure)

        assign_data = []
        ctc_amount = 0

        selected_date = datetime(selected_year, selected_month, 1)
        print("selected_date :", selected_date)

        if not assign_salarystructure:
            nearest_date = AssignSalaryStructure.objects.filter(
                effective_date__lte=selected_date, user_id=user_id).order_by('-effective_date').first()

            if nearest_date:
                assign_salarystructure = nearest_date

        if assign_salarystructure:
            print("assign_salarystructure :", assign_salarystructure)
            names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
            amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
            print("names ; amounts 7187:", names, amounts)

            ctc_amount += sum(amount.amount for amount in amounts)
            zipped_data = zip_longest(names, amounts)

            assign_data.append({
                'rule': rule,
                'zipped_data': zipped_data,
            })
        print("ctc_amount :", ctc_amount)
        print("assign_data :", assign_data)

        adhoc_data = Adhoc.objects.filter(user_id=user_id, createddate__year=selected_year,
                                        createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
        print("adhoc_data : ", adhoc_data)

        earning_amount = 0
        deduction_amount = 0
        for adhoc_entry in adhoc_data:
            if adhoc_entry.adhocearning:
                earning_amount += adhoc_entry.amount
            elif adhoc_entry.adhocdeduction:
                deduction_amount += adhoc_entry.amount

        total_earnings = ctc_amount + earning_amount
        print("CCCCCCCCCCCCCCCCCC :", total_earnings)

        num_days = calendar.monthrange(selected_year, selected_month)[1]

        payregister = PayRegister.objects.filter(createddate__month=selected_month, createddate__year=selected_year, user_id=user_id, status__in=[
                                                "Payslip Generated", "Payslip Downloaded"])
        print("payregister :", payregister)

        total_fullday_time = timedelta()
        total_halfday_time = timedelta()
        total_anomaly_count = 0
        attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=user_id)
        print("attendance_rule :", attendance_rule)
        for att_rule in attendance_rule:
            rule_type = att_rule.rules_applied
            print("rule_type :", rule_type, )
            if rule_type:
                full_day_hours = rule_type.fullhours
                full_day_minutes = rule_type.fullminutes
                full_time = timedelta(hours=full_day_hours,
                                    minutes=full_day_minutes)
                half_day_hours = rule_type.halfhours
                half_day_minutes = rule_type.halfminutes
                half_time = timedelta(hours=half_day_hours,
                                    minutes=half_day_minutes)
                print("Full Day Hours:", full_day_hours,
                    full_day_minutes, full_time)
                print("Half Day Hours:", half_day_hours,
                    half_day_minutes, half_time)
                in_grace_period = rule_type.inGracePeriod
                out_grace_period = rule_type.outGracePeriod
                print("Grace period:", in_grace_period, out_grace_period)
                in_grace_timedelta = timedelta(
                    hours=in_grace_period.hour, minutes=in_grace_period.minute)
                out_grace_timedelta = timedelta(
                    hours=out_grace_period.hour, minutes=out_grace_period.minute)

                total_grace_period = in_grace_timedelta + out_grace_timedelta
                print("Total Grace period:", total_grace_period)
                total_fullday_time = full_time + total_grace_period
                print("Total Time:", total_fullday_time)
                total_halfday_time = half_time + total_grace_period
                print("total_halfday_time :", total_halfday_time)

        punches = Punch.objects.filter(
            user__id=user_id,
            date__year=selected_year,
            date__month=selected_month,
            status="AN", is_penalty_reverted=False
        )
        print("Punch Object :", punches)

        for punch in punches:
            total_work_duration = timedelta()

            if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                second_clock_in = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)
                second_clock_out = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)
                first_duration = first_clock_out - first_clock_in
                second_duration = second_clock_out - second_clock_in
                total_work_duration += first_duration + second_duration

            elif punch.first_clock_in_time and punch.first_clock_out_time:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                print("first_clock_in ; first_clock_out : ",
                    first_clock_in, first_clock_out)
                first_duration = first_clock_out - first_clock_in
                print("first_duration : ", first_duration)
                total_work_duration += first_duration
            if total_work_duration > total_fullday_time:
                AN_count = 0.5
            elif total_work_duration < total_halfday_time:
                AN_count = 1.0
            else:
                AN_count = 0.5

            total_anomaly_count += AN_count
            print("total_anomaly_count :", total_anomaly_count)

        lop_data = Runpayroll_lop.objects.filter(
            lop_date__month=selected_month, lop_date__year=selected_year, user_id=user_id)
        lopcount = 0
        for lopdata in lop_data:
            lopcount += lopdata.lop_count
            print("lopcount :", lopcount)

        absent_count = Punch.objects.filter(user__id=user_id, date__year=selected_year,
                                            date__month=selected_month, status='A', is_penalty_reverted=False).count()

        absent_AN_count = absent_count + total_anomaly_count
        print("absent_AN_count : ", absent_count, absent_AN_count)

        punchcount = Punch.objects.filter(
            user__id=user_id, date__year=selected_year, date__month=selected_month).count()
        print("punchcount :", punchcount)
        missing_date_count = num_days - punchcount
        print("missing_date_count :", missing_date_count)
        working_days = punchcount - absent_AN_count
        print("working_days :", working_days)
        total_lop = absent_AN_count + missing_date_count + lopcount

        per_day_amount = ctc_amount / num_days
        print("per_day_amount :", per_day_amount)
        lop_amount = per_day_amount * total_lop
        print("lop_amount :", lop_amount)
        lopamount = round(lop_amount)

        total_deductions = deduction_amount + lop_amount
        totaldeductions = round(total_deductions)
        net_amount = round(total_earnings - total_deductions)
        print("net_amount :", net_amount)

        net_amount_words = num2words(net_amount, lang='en_IN')

        x = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            "regaddress": regaddress,
            "payregister": payregister,
            "num_days": num_days,
            "bank_details": bank_details,
            "assign_salarystructure": assign_salarystructure,
            "assign_data": assign_data,
            "ctc_amount": ctc_amount,
            "adhoc_data": adhoc_data,
            "total_earnings": total_earnings,
            "totaldeductions": totaldeductions,
            "total_lop": total_lop,
            "lopamount": lopamount,
            "net_amount": net_amount,
            "net_amount_words": net_amount_words,
            "month_str": month_str,
            "selected_year": selected_year,
            "use_if_block": use_if_block,
        }

        return render(request, "Employee/payrollmypro.html", x)
    
    else:

        assign_salarystructure = AssignSalaryStructure.objects.filter(user_id=user_id, effective_date__month=selected_month, effective_date__year=selected_year).order_by('effective_date').first()
        print("assignsalary :", assign_salarystructure)
        
        # assign_data = []
        assigndata = []
        ctc_assigndata = []
        gross_salary_amount = 0
        work_from_office_allowance_amount = 0
        total_net_salary = 0
        total_ctc_salary = 0
        wfocount = 0
        leave_count = 0
        total_gross_salary = 0
        net_total = 0
        grossamount = 0

        selected_date = datetime(selected_year, selected_month, 1)
        print("selected_date :", selected_date)

        if not assign_salarystructure:
            nearest_date = AssignSalaryStructure.objects.filter(
                effective_date__lte=selected_date, user_id=user_id).order_by('-effective_date').first()

            if nearest_date:
                assign_salarystructure = nearest_date

        if assign_salarystructure:
            gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
            work_from_office_component = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance", Parentcomponentname__componentname__iexact="Gross Salary").first()
            print("gross_salary_component ; work_from_office_component : ", gross_salary_component, work_from_office_component)
            net_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="Net Salary")
            print("net_salary_component : ", net_salary_component)
            ctc_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="CTC")
            print("ctc_salary_component : ", ctc_salary_component)

            name = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
            amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=name)
            print("name ; amount 1st :", name, amount)
            names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname=gross_salary_component)
            amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
            print("names ; amounts  :", names, amounts)

            net_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="Net Salary")
            net_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=net_names)
            print("names ; amounts  :", net_names, net_amounts)

            ctc_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="CTC").exclude(salarycomponent__componentname__icontains="professional tax")
            ctc_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=ctc_names)
            print("ctc_names ; ctc_amounts  :", ctc_names, ctc_amounts)

            if gross_salary_component:
                gross_amount = amount.filter(salaryname__salarycomponent=gross_salary_component).first()
                gross_salary_amount = gross_amount.amount if gross_amount else 0
            
            if work_from_office_component:
                work_amount = amounts.filter(salaryname__salarycomponent=work_from_office_component).first()
                work_from_office_allowance_amount = work_amount.amount if work_amount else 0
            
            for netsalry in net_salary_component:
                net_salary = amount.filter(salaryname__salarycomponent=netsalry)
                total_net_salary += net_salary.aggregate(total=models.Sum('amount'))['total'] or 0
            print("Total Net Salary:", total_net_salary)

            for ctcsalry in ctc_salary_component:
                print("ctcsalry.componentname : ", ctcsalry.componentname)
                if ctcsalry.componentname.lower() != 'professional tax':
                    ctc_salary = amount.filter(salaryname__salarycomponent=ctcsalry)
                    total_ctc_salary += ctc_salary.aggregate(total=models.Sum('amount'))['total'] or 0
            print("Total CTC Salary:", total_ctc_salary)

            total_gross_salary = gross_salary_amount - work_from_office_allowance_amount
            print("total_gross_salary : ", total_gross_salary)

            # zipped_data = zip_longest(names, amounts)
            # assign_data.append({
            #     'rule': rule,
            #     'zipped_data': zipped_data,
            # })

            zippeddata = zip_longest(net_names, net_amounts)
            assigndata.append({
                'rule': rule,
                'zippeddata': zippeddata,
            })

            ctc_zippeddata = zip_longest(ctc_names, ctc_amounts)
            ctc_assigndata.append({
                'rule': rule,
                'ctc_zippeddata': ctc_zippeddata,
            })
    
        print("ctc_assigndata :", ctc_assigndata)

        punch_obj = Punch.objects.filter(user__id=user_id,date__year=selected_year,date__month=selected_month)
        print("punch_obj : ", punch_obj)

        for punch in punch_obj:
                       
            leave_data = Leave.objects.filter(
                    applicant_email=user_id,  
                    strtDate=punch.date,    
                    status="Approved"      
                ).first()       
            print(f"Leave data for {user_id} on {punch.date}: ", leave_data)
            
            if punch.status == "H":
                leave_count += 1
            elif punch.status == "L":
                if leave_data:
                    if leave_data.leavetyp != "Loss Of Pay":
                        leave_count += 1
            elif punch.status == "HL":
                if leave_data:
                    print("ccccccccccccccccccccccccccccccccccccc  ")
                    if leave_data.leavetyp == "Loss Of Pay":
                        leave_count -= 0.5
                        print("KKKKKKKKKKKKKKKKKKKKK", leave_count)
                    
        wfo_count =  WFOCount.objects.filter(user_id=user_id, wfo_date__year=selected_year, wfo_date__month=selected_month)
        print("wfo_count :", wfo_count, "month_numeric , selected_year:" , selected_month, selected_year)
        for i in wfo_count:
            wfocount = i.wfocount
            print("wfocount : ", wfocount)

        year_select = int(selected_year)
        num_days = calendar.monthrange(year_select, selected_month)[1]
        first_day_of_month = datetime(year_select, selected_month, 1)
        if selected_month == 12: 
            next_month = datetime(year_select + 1, 1, 1)
            print("next_month 1 :", next_month, year_select)
        else:
            next_month = datetime(year_select, selected_month + 1, 1)
            print("next_month 2 :", next_month)

        day_count = 0
        current_day = first_day_of_month
        while current_day < next_month:
            if current_day.weekday() != 6: 
                day_count += 1
            current_day += timedelta(days=1)
        count_sundays = num_days - day_count
        print("day_count ############ :", day_count, num_days, count_sundays)

        print("work_from_office_allowance_amount:", work_from_office_allowance_amount)
        perday_WFOamount = work_from_office_allowance_amount / day_count
        total_WFOamount = perday_WFOamount * wfocount
        print("total_WFOamount : ", wfocount, total_WFOamount)
        
        WFOamount = round(work_from_office_allowance_amount - total_WFOamount) #This amount add to the deduction
        print("WFOamount :", WFOamount)

        adhoc_data = Adhoc.objects.filter(user_id=user_id, createddate__year=selected_year,
                                        createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
        print("adhoc_data : ", adhoc_data)

        earning_amount = 0
        deduction_amount = 0
        for adhoc_entry in adhoc_data:
            if adhoc_entry.adhocearning:
                earning_amount += adhoc_entry.amount
            elif adhoc_entry.adhocdeduction:
                deduction_amount += adhoc_entry.amount

        total_earnings = gross_salary_amount
        print("total_earnings ; gross_salary_amount ; earning_amount :", total_earnings, gross_salary_amount, earning_amount)

        total_fullday_time = timedelta()
        total_halfday_time = timedelta()
        total_anomaly_count = 0
        attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=user_id)
        print("attendance_rule :", attendance_rule)
        for att_rule in attendance_rule:
            rule_type = att_rule.rules_applied
            print("rule_type :", rule_type, )
            if rule_type:
                full_day_hours = rule_type.fullhours
                full_day_minutes = rule_type.fullminutes
                full_time = timedelta(hours=full_day_hours,
                                    minutes=full_day_minutes)
                half_day_hours = rule_type.halfhours
                half_day_minutes = rule_type.halfminutes
                half_time = timedelta(hours=half_day_hours,
                                    minutes=half_day_minutes)
                print("Full Day Hours:", full_day_hours,
                    full_day_minutes, full_time)
                print("Half Day Hours:", half_day_hours,
                    half_day_minutes, half_time)
                in_grace_period = rule_type.inGracePeriod
                out_grace_period = rule_type.outGracePeriod
                print("Grace period:", in_grace_period, out_grace_period)
                in_grace_timedelta = timedelta(
                    hours=in_grace_period.hour, minutes=in_grace_period.minute)
                out_grace_timedelta = timedelta(
                    hours=out_grace_period.hour, minutes=out_grace_period.minute)

                total_grace_period = in_grace_timedelta + out_grace_timedelta
                print("Total Grace period:", total_grace_period)
                total_fullday_time = full_time + total_grace_period
                print("Total Time:", total_fullday_time)
                total_halfday_time = half_time + total_grace_period
                print("total_halfday_time :", total_halfday_time)

        num_days = calendar.monthrange(selected_year, selected_month)[1]

        payregister = PayRegister.objects.filter(createddate__month=selected_month, createddate__year=selected_year, user_id=user_id, status__in=[
                                                "Payslip Generated", "Payslip Downloaded"])
        print("payregister :", payregister)

        lop_data = Runpayroll_lop.objects.filter(
            lop_date__month=selected_month, lop_date__year=selected_year, user_id=user_id)
        lopcount = 0
        for lopdata in lop_data:
            lopcount += lopdata.lop_count
            print("lopcount :", lopcount)

        punches = Punch.objects.filter(
            user__id=user_id,
            date__year=selected_year,
            date__month=selected_month,
            status="AN", is_penalty_reverted=False
        )
        print("Punch Object :", punches)

        for punch in punches:
            total_work_duration = timedelta()

            if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                second_clock_in = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)
                second_clock_out = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)
                first_duration = first_clock_out - first_clock_in
                second_duration = second_clock_out - second_clock_in
                total_work_duration += first_duration + second_duration

            elif punch.first_clock_in_time and punch.first_clock_out_time:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                print("first_clock_in ; first_clock_out : ",
                    first_clock_in, first_clock_out)
                first_duration = first_clock_out - first_clock_in
                print("first_duration : ", first_duration)
                total_work_duration += first_duration
            if total_work_duration > total_fullday_time:
                AN_count = 0.5
            elif total_work_duration < total_halfday_time:
                AN_count = 1.0
            else:
                AN_count = 0.5

            total_anomaly_count += AN_count
            print("total_anomaly_count :", total_anomaly_count)

        print("total_anomaly_count 2:", total_anomaly_count)
        absent_count = Punch.objects.filter(user__id=user_id, date__year=selected_year,
                                            date__month=selected_month, status='A', is_penalty_reverted=False).count()

        absent_AN_count = absent_count + total_anomaly_count
        print("absent_AN_count : ", absent_count, absent_AN_count)

        punchcount = Punch.objects.filter(user__id=user_id, date__year=selected_year, date__month=selected_month).count()
        print("punchcount :", punchcount)
        missing_date_count = num_days - punchcount
        print("missing_date_count :", missing_date_count, absent_AN_count, lopcount)
        total_lop = absent_AN_count + missing_date_count + lopcount

        per_day_amount = total_gross_salary / num_days
        print("per_day_amount :", per_day_amount)
        lop_amount = per_day_amount * total_lop
        print("lop_amount :", lop_amount)
        lopamount = round(lop_amount)
        # total_deductions = deduction_amount + lop_amount + total_net_salary + WFOamount
        total_deductions = lop_amount + WFOamount
        totaldeductions = round(total_deductions)
        grossamount = round(total_earnings - total_deductions)
        print("net_amount :", grossamount)

        componentnames = []
        amounts = []
        adhocnames = []
        adhocamounts = []
        ctc_cmpname = []
        ctc_cmpamount = []

        assigndata = []
        netassigndata = []
        ctcassigndata = []
        work_from_office_allowanceamount = 0
        basicamount = 0
        ctc_total = 0
        toataldeda = 0

        if assign_salarystructure:
            print("assign_salarystructure 2:", assign_salarystructure)

            gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
            
            work_from_officecomponent = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance",Parentcomponentname__componentname__iexact="Gross Salary").first()
            print("gross_salary_component and work_from_office_component:", gross_salary_component, work_from_officecomponent)
            
            names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
            amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)

            if work_from_officecomponent:
                workamount = amount.filter(salaryname__salarycomponent=work_from_office_component).first()
                work_from_office_allowanceamount = workamount.amount if workamount else 0
            
            print("gross_salary_component and work_from_office_component:", gross_salary_component, work_from_office_allowanceamount)

            grossnames = AssignSalaryStructureName.objects.filter(
                salaryrule=assign_salarystructure,
                salarycomponent__Parentcomponentname=gross_salary_component
            ).exclude(salarycomponent__componentname__icontains="work from office allowance")
            print("grossnames:", grossnames)

            gross_amounts = []
            calculatedamount = 0
            for name in grossnames:
                for component in name.salarycomponent.all():
                    print("Gross Salary Component Percent:", name, component.percent, component.componentname)
                    if component.percent and component.componentname.lower() != "other allowance":
                        calculated_amount = round(grossamount * (component.percent / 100.0))
                        calculatedamount += calculated_amount
                        print("calculated_amount :", calculatedamount)
                        gross_amounts.append(calculated_amount)
                        if component.componentname.lower() == "basic salary":
                            basicamount = calculated_amount
                    if component.componentname.lower() == "other allowance":
                        other_allowanceamt = grossamount - calculatedamount
                        print("other_allowanceamt :", other_allowanceamt)
                        gross_amounts.append(other_allowanceamt)

            # Zipping gross names and calculated amounts
            print("grossnames :", grossnames , gross_amounts)
            zipped_gross_data = zip_longest(grossnames, gross_amounts)
            assigndata.append({
                'rule': assign_salarystructure,
                'zipped_gross_data': zipped_gross_data,
            })

            # Process Net Salary components
            net_names = AssignSalaryStructureName.objects.filter(
                salaryrule=assign_salarystructure,
                salarycomponent__Parentcomponentname__componentname__iexact="Net Salary"
            )
            net_amounts = []
            for name in net_names:
                for component in name.salarycomponent.all():
                    print("Net Salary Component Percent:", name, component.percent)
                    if component.componentname.lower() == "epf employee":
                        epf_amount = round((basicamount * component.percent) / 100)
                        net_amounts.append(epf_amount)
                    if component.componentname.lower() == "esi employee":
                        esi_amount = round((grossamount * component.percent) / 100) 
                        net_amounts.append(esi_amount)
                    if component.componentname.lower() == "professional tax":
                        profisional_tax_amount = 167 if grossamount < 22000 else 208
                        print("profisional_tax_amount :", profisional_tax_amount)
                        net_amounts.append(profisional_tax_amount)

                    if component.componentname.lower() == "insurance":
                        insurance_amount = 0 if grossamount <= 25000 else 245
                        print("insurance_amount :", insurance_amount)
                        net_amounts.append(insurance_amount)
            
            net_total = sum(net_amounts)
            # Zipping net names and amounts
            zipped_net_data = zip_longest(net_names, net_amounts)
            netassigndata.append({
                'rule': assign_salarystructure,
                'zipped_net_data': zipped_net_data,
            })

            # Process CTC components, excluding professional tax
            ctcsalarynames = AssignSalaryStructureName.objects.filter(
                salaryrule=assign_salarystructure,
                salarycomponent__Parentcomponentname__componentname__iexact="CTC"
            ).exclude(salarycomponent__componentname__icontains="professional tax")
            ctcamounts = []
            for name in ctcsalarynames:
                for component in name.salarycomponent.all():
                    print("CTC Salary Component Percent:", component.percent, component.componentname)
                    if component.componentname == "EPF Employer":
                        epf_amount = round((basicamount * component.percent) / 100)
                        print("amount : ", epf_amount)
                        ctcamounts.append(epf_amount)
                    if component.componentname == "ESI Employer":
                        esi_amount = round((grossamount * component.percent) / 100) 
                        ctcamounts.append(esi_amount)

            ctc_total = sum(ctcamounts)
            print("ctc_total :", ctc_total)
            # Zipping CTC names and amounts
            ctc_zipped_data = zip_longest(ctcsalarynames, ctcamounts)
            ctcassigndata.append({
                'rule': assign_salarystructure,
                'ctc_zipped_data': ctc_zipped_data,
            })

        totalgrossamount = grossamount + earning_amount

        toataldeda = deduction_amount + net_total

        totalNet_Salary = totalgrossamount - toataldeda
        net_amount_words = num2words(totalNet_Salary, lang='en_IN').title()
        print("totalNet_Salary :", totalNet_Salary)
        net_amount_words_formatted = f"({totalNet_Salary})"

        print("assigndata for else part:", assigndata)

        context = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            "regaddress": regaddress,
            "payregister": payregister,
            "num_days": num_days,
            "bank_details": bank_details,
            "assign_salarystructure": assign_salarystructure,
            "assigndata": assigndata,
            "assign_data": netassigndata,
            "ctc_assigndata": ctcassigndata,
            "adhoc_data": adhoc_data,
            "total_earnings": totalgrossamount,
            "total_ctc_salary": ctc_total,
            "totaldeductions": toataldeda,
            "total_lop": total_lop,
            "lopamount": lopamount,
            "wfo_count": wfocount, 
            "WFOamount": WFOamount,
            "net_amount": totalNet_Salary,
            "net_amount_words": net_amount_words,
            "month_str": month_str,
            "selected_year": selected_year,
            "use_if_block": use_if_block,
        }

        return render(request, "Employee/payrollmypro.html", context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def filemanager_nav(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    datas = Filemanager.objects.filter(
        myuser_10__id=request.user.id).order_by('-scheduleon')
    data = companyprofile.objects.filter(admin_id=request.user.id)

    page = request.GET.get('page', 1)
    paginator = Paginator(datas, 20)
    try:
        datas = paginator.page(page)
    except PageNotAnInteger:
        datas = paginator.page(1)
    except EmptyPage:
        datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, "index/filemanagermypro.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def empfilemanager_nav(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    l = Filemanager.objects.filter(myuser_10__id=request.user.id)

    x = {
        "c": c[0] if c.exists() else c,
        "k": k[0] if k.exists() else k,
        'l': l,
        "is_view_filemanager": False
    }
    return render(request, "Employee/filemanagermypro.html", x)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def table_file_manager(request, uid_10):
    if request.method == "POST":
        reqtyp = request.POST.get("Requesttype")
        frmt = request.POST.get("Format")
        schdon = request.POST.get("Scheduleon")
        size = request.POST.get("Size")
        stts = request.POST.get("Status")
        u_10 = User.objects.get(id=uid_10)
        Filemanager.objects.create(requesttype=reqtyp, frmt=frmt,
                                   scheduleon=schdon, size=size, status=stts, myuser_10=u_10)
    return redirect('filemanagernav')


def empadmin(request):
    return render(request, "Employee/empdash.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def apply_leave(request):
    if request.method == 'POST':
        Lt = request.POST.get('Ltype')
        ln = CompanyRules.objects.get(id=Lt)
        Sd = datetime.strptime(request.POST.get(
            'strtdate'), '%d %B %Y').strftime('%Y-%m-%d')
        s1 = request.POST.get('sh1')
        Ed = datetime.strptime(request.POST.get(
            'enddate'), '%d %B %Y').strftime('%Y-%m-%d')
        s2 = request.POST.get('sh2')
        Rn = request.POST.get('reasn')
        Ap = datetime.today().date()
        # st = request.POST.get('Status')
        useremail = request.user.email
        username = request.user.username
        applicant_id = request.user
        status = "Applied"

        # dt=(datetime.datetime.strptime(Ed,"%Y-%m-%d"))-(datetime.datetime.strptime(Sd,"%Y-%m-%d"))
        if s1 == "first half" and s2 == "first half":
            d1 = datetime.strptime(Sd, "%Y-%m-%d")
            d2 = datetime.strptime(Ed, "%Y-%m-%d")
            delt = d2 - d1 + timedelta(days=0)
            delta = delt.days + 0.5
        elif s1 == "first half" and s2 == "second half":
            d1 = datetime.strptime(Sd, "%Y-%m-%d")
            d2 = datetime.strptime(Ed, "%Y-%m-%d")
            delt = d2 - d1 + timedelta(days=1)
            delta = delt.days
        elif s1 == "second half" and s2 == "first half":
            d1 = datetime.strptime(Sd, "%Y-%m-%d")
            d2 = datetime.strptime(Ed, "%Y-%m-%d")
            delt = d2 - d1 + timedelta(days=0)
            delta = delt.days
        elif s1 == "second half" and s2 == "second half":
            d1 = datetime.strptime(Sd, "%Y-%m-%d")
            d2 = datetime.strptime(Ed, "%Y-%m-%d")
            delt = d2 - d1 + timedelta(days=0)
            delta = delt.days + 0.5

        user_id = request.user.id
        assign_rule = assignrule.objects.filter(
            user_id_id=user_id, rules_applied=ln.id).first()
        print("assign rule", assign_rule.leavebalance)

        if assign_rule.creditedleaves != 0 and assign_rule.leavebalance >= Decimal(delta):
            Leave.objects.create(leavetyp=ln.leavename, strtDate=Sd, Reason=Rn, Selecthalf1=s1, endDate=Ed, admin_id=user_id,
                                 Selecthalf2=s2, status=status, Appliedon=Ap, Days=delta, applicant_email=applicant_id)

            assign_rule.leavebalance -= Decimal(delta)
            assign_rule.appliedleaves += Decimal(delta)
            assign_rule.save()
            print("assign_rule :", assign_rule)
            LeaveNotification.objects.create(
                message=f"{applicant_id.username}, apply leave on {datetime.now().strftime('%d %B %Y')}",
                user=applicant_id)
            return redirect('leave')

        elif assign_rule.creditedleaves == 0 and assign_rule.leavebalance == 0:
            Leave.objects.create(leavetyp=ln.leavename, strtDate=Sd, Reason=Rn, Selecthalf1=s1, endDate=Ed,
                                 Selecthalf2=s2, status=status, Appliedon=Ap, Days=delta, applicant_email=applicant_id)
            # assign_rule.penaltydeduction += Decimal(delta)
            assign_rule.appliedleaves += Decimal(delta)
            assign_rule.save()

            LeaveNotification.objects.create(
                message=f"{applicant_id.username}, apply leave on {datetime.now().strftime('%d %B %Y')}", user=applicant_id)
            return redirect('leave')

        messages.error(request, 'Select another leave type')
        return redirect('leave')


def empapply_leave(request):
    if request.method == 'POST':
        Lt = request.POST.get('Ltype')
        ln = CompanyRules.objects.get(id=Lt)
        Sd = datetime.strptime(request.POST.get('strtdate'), '%d %B %Y').date()
        std = Sd.strftime('%d %B %Y')
        s1 = request.POST.get('sh1')
        Ed = datetime.strptime(request.POST.get('enddate'), '%d %B %Y').date()
        edt = Ed.strftime('%d %B %Y')
        s2 = request.POST.get('sh2')
        Rn = request.POST.get('reasn')
        Ap = datetime.today().date()
        status = "Applied"
        applicant_id = request.user
        current_user_id = request.user.id
        admin_id = User.objects.get(id=current_user_id).admin_id

        # Calculate delta (total leave days)
        if s1 == "first half" and s2 == "first half":
            delta = (Ed - Sd).days + 0.5
        elif s1 == "first half" and s2 == "second half":
            delta = (Ed - Sd).days + 1
        elif s1 == "second half" and s2 == "first half":
            delta = (Ed - Sd).days
        elif s1 == "second half" and s2 == "second half":
            delta = (Ed - Sd).days + 0.5

        # Check daily leave limits
        def check_daily_limits(start_date, end_date):

            date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
            
            for day in date_range:
                # Get all approved leaves for this day and check half-day conflicts
                existing_leaves = Leave.objects.filter(
                    applicant_email=applicant_id,
                    strtDate__lte=day,
                    endDate__gte=day,
                    rejected=False,
                    cancel_requested=False
                )
                
                # Check if applying for same half-day that's already taken
                if start_date == end_date == day:
                    # For single day application
                    for leave in existing_leaves:
                        # If existing leave is full day, no more leaves allowed
                        if leave.Selecthalf1 == "full day" or leave.Selecthalf2 == "full day":
                            return False, f"Cannot apply leave on {day.strftime('%d %B %Y')} (already has full day leave)"
                        
                        # Check if applying for same half that's already taken
                        if (s1 == "first half" and leave.Selecthalf1 == "first half") or \
                           (s1 == "second half" and leave.Selecthalf1 == "second half"):
                            return False, f"Cannot apply {s1} leave on {day.strftime('%d %B %Y')} (already applied)"
                
                else:
                    # For multi-day application, check start and end days separately
                    if day == start_date:
                        for leave in existing_leaves:
                            if leave.strtDate == day and leave.Selecthalf1 == s1:
                                return False, f"Cannot apply {s1} leave on {day.strftime('%d %B %Y')} (already applied)"
                    
                    if day == end_date:
                        for leave in existing_leaves:
                            if leave.endDate == day and leave.Selecthalf2 == s2:
                                return False, f"Cannot apply {s2} leave on {day.strftime('%d %B %Y')} (already applied)"
                
                # Also check total leave doesn't exceed 1.0 per day (existing logic)
                total_existing = existing_leaves.aggregate(total=Sum('Days'))['total'] or 0
                if start_date == end_date == day:
                    day_leave = 0.5 if (s1 == "first half" and s2 == "first half") or (s1 == "second half" and s2 == "second half") else 1.0
                else:
                    if day == start_date:
                        day_leave = 0.5 if s1 in ["first half", "second half"] else 1.0
                    elif day == end_date:
                        day_leave = 0.5 if s2 in ["first half", "second half"] else 1.0
                    else:
                        day_leave = 1.0
                
                if total_existing + day_leave > 1.0:
                    return False, f"Cannot apply more than 1.0 leave on {day.strftime('%d %B %Y')}"
            
            return True, ""

        # Validate daily limits
        is_valid, error_msg = check_daily_limits(Sd, Ed)
        if not is_valid:
            messages.error(request, error_msg)
            return redirect('empleave')

        # Check leave balance
        assign_rule = assignrule.objects.filter(
            user_id_id=current_user_id, rules_applied=ln.id).first()

        if not assign_rule:
            messages.error(request, 'No leave rule assigned for this leave type')
            return redirect('empleave')

        if ln.leavename not in ["Loss Of Pay", "Comp Off"] and assign_rule.leavebalance < Decimal(delta):
            messages.error(request, 'Insufficient leave balance')
            return redirect('empleave')

        # Create leave application
        datas = Leave.objects.create(
            leavetyp=ln.leavename, 
            strtDate=Sd, 
            Reason=Rn, 
            Selecthalf1=s1, 
            endDate=Ed,
            Selecthalf2=s2, 
            status=status, 
            Appliedon=Ap, 
            Days=delta, 
            applicant_email=applicant_id,
            admin_id=admin_id
        )

        # Update leave balance (except for special leave types)
        if ln.leavename not in ["Loss Of Pay", "Comp Off"]:
            assign_rule.leavebalance -= Decimal(delta)
            assign_rule.appliedleaves += Decimal(delta)
            assign_rule.save()

        # Create notification
        LeaveNotification.objects.create(
            message=f"{applicant_id.username}, apply leave on {datetime.now().strftime('%d %B %Y')}",
            user=applicant_id
        )

        # Send email notification
        # Get the primary reporting manager
        reporting_manager = applicant_id.reptmgr  # Assuming this is an EmailField
        reporting_manager_user = User.objects.filter(email=reporting_manager).first()

        # Build email addresses
        to = [reporting_manager] if reporting_manager else []
        cc = ['operations@cydeztechnologies.com']

        # Addressing name (manager name fallback)
        manager_name = reporting_manager_user.username if reporting_manager_user else "Manager"

        # Prepare email content
        subject = f"Leave Application Notification for {applicant_id.username}"
        email_from = settings.EMAIL_HOST_USER

        html_body = render_to_string(
            'Employee/mail.html', 
            {
                'data': datas,
                'to': manager_name,  # Use this in template for personalizing greeting
                'emailconfig': "HRMS",
                'std': std,
                'edt': edt
            }
        )

        # Send email
        msg = EmailMultiAlternatives(
            subject=subject,
            from_email=email_from,
            to=to,
            cc=cc,
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()
        messages.success(request, 'Leave applied successfully' )
        return redirect('empleave')
    

from django.views.decorators.cache import cache_control

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def log(request):
    request_user = request.user

    data = Leave.objects.filter(
        Q(applicant_email=request_user.id) | Q(admin_id=request_user.id)).order_by('-Appliedon', '-strtDate', 'applicant_email__username')

    pending_alerts = []
    today = date.today()

    for leave in data:
        user = leave.applicant_email

        # Message logic for pending leaves
        if leave.status == "Pending" and leave.strtDate and leave.strtDate > today:
            remaining_days = (leave.strtDate - today).days

            if remaining_days <= 2:  # threshold for alert
                is_admin = leave.admin_id == request_user.id
                is_reporting_manager = (
                    hasattr(leave.applicant_email, "reptmgr_id") and
                    leave.applicant_email.reptmgr_id == request_user.id
                )

                if is_admin or is_reporting_manager:
                    pending_alerts.append({
                        "username": leave.applicant_email.username,
                        "remaining_days": remaining_days,
                    })
        primary_relation = Reportingmanager.objects.filter(userid=user.id, type='Primary').first()

        if primary_relation and primary_relation.myuser_2.exists():
            manager_user = primary_relation.myuser_2.first()
            user.primary_manager_name = manager_user.username
        else:
            # Fallback to Secondary Manager
            secondary_relation = Reportingmanager.objects.filter(userid=user.id, type='Secondary').first()
            if secondary_relation and secondary_relation.myuser_2.exists():
                manager_user = secondary_relation.myuser_2.first()
                user.primary_manager_name = manager_user.username
            else:
                user.primary_manager_name = "N/A"


        if leave.status in ['Approved', 'Rejected'] and leave.admin_id:
            approver = User.objects.filter(id=leave.admin_id).first()
            if approver:
                if hasattr(leave.applicant_email, "reptmgr_id") and leave.applicant_email.reptmgr_id == approver.id:
                    leave.action_by_name = f"{approver.username} "
                else:
                    leave.action_by_name = f"{approver.username} "
            else:
                leave.action_by_name = "Auto-approved"
        else:
            leave.action_by_name = ""


    k = Myprofile.objects.filter(myuser__id=request_user.id)
    datas = companyprofile.objects.filter(admin_id=request_user.id)

    context = {
        "k": k.first() if k.exists() else None,
        "datas": datas.first() if datas.exists() else None,
        "data": data,
        "pending_alerts": pending_alerts,
    }
    return render(request, "index/log.html", context)



from django.core.cache import cache
def auto_approve_leaves_once_per_day():
    today = date.today()
    cache_key = f"auto_approve_done_{today}"

    if cache.get(cache_key):
        return  

    # Approve all leaves starting today if status is still Pending
    leaves_to_auto_approve = Leave.objects.filter(
        strtDate=today,
        status='Pending'
    )
    for leave in leaves_to_auto_approve:
        leave.status = 'Approved'
        leave.approved_by = None
        leave.save()

    cache.set(cache_key, True, 60 * 60 * 24)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def reporting_manager_team_log(request):
    request_user = request.user
    k = Myprofile.objects.filter(myuser__id=request_user.id)
    datas = companyprofile.objects.filter(admin_id=request_user.id)
    if request_user.role == "Employee" and Reportingmanager.objects.filter(myuser_2=request_user).exists():
        # Get a flat list of employee user IDs
        direct_report_employees = Reportingmanager.objects.filter(myuser_2=request_user).values_list('userid', flat=True)

        # Query the Leave model for direct reports' leave data
        data = Leave.objects.filter(
            Q(applicant_email_id__in=list(direct_report_employees))
        )

        x = {
            "k": k[0] if k.exists() else k,
            "datas": datas[0] if datas.exists() else datas,
        }
        return render(request, "Employee/reporting_manager_team_leave_log.html", {'data': data, **x})
    else:
        return HttpResponse("<h2>Page Not Found</h1>")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def approve_leave(request):
    # print("Approving leave.....")f
    current_time = timezone.now().time().replace(microsecond=0)
    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        leave = Leave.objects.get(pk=leave_id)
        leave.status = 'Approved'

        startdate = leave.strtDate.strftime('%d %B %Y')
        enddate = leave.endDate.strftime('%d %B %Y')

        employee = leave.applicant_email
        attendance_rule = employee.assignattendancerule_set.first()
        rules_applied = attendance_rule.rules_applied if attendance_rule else None
        if leave.leavetyp == "Loss Of Pay":
            lop_date_year = leave.Appliedon.year
            lop_date_month = leave.Appliedon.month

            existing_lop_record = Runpayroll_lop.objects.filter(
                user_id=leave.applicant_email, lop_date__year=lop_date_year, lop_date__month=lop_date_month).first()

            if existing_lop_record:
                existing_lop_record.lop_count += leave.Days
                existing_lop_record.save()
            else:
                # print("No existing Leave record")
                Runpayroll_lop.objects.create(
                    lop_count=leave.Days, user_id=leave.applicant_email, lop_date=leave.Appliedon)

        if leave and leave.punch_data:
            # print("Leave data: ", leave , leave.punch_data)
            punch = Punch.objects.get(id=leave.punch_data.id)
            # print("approved leave ", punch)
            punch.is_approved = True
            punch.is_penalty_reverted = True
            punch.save()
        else:
            punch_data = Punch.objects.filter(
                user=leave.applicant_email, date__gte=leave.strtDate, date__lte=leave.endDate)
            print(punch_data)
            for punch in punch_data:
                # print("Current punch status: ", punch.status)
                if leave.Days == 0.5:
                    # print("Leave day is 0.5")
                    work_duration = get_work_duration(punch, current_time)
                    half_day_duration = get_half_hour(
                        rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                    if work_duration < half_day_duration:
                        # print("Is lessthan!")
                        punch.status = 'AN'
                    # print("Punch data: ", punch, work_duration)
                elif leave.Days == 1.0:
                    # print("Yes it's a full day leave!")
                    punch.status = 'L'

                else:
                    # print("Leave days grater than 1 and 0.5",)
                    print(leave.strtDate, leave.endDate, leave.Selecthalf1,
                          leave.Selecthalf2, punch.date.date())
                    if leave.strtDate == punch.date.date() and '.5' in str(leave.Days) and leave.Selecthalf1 == leave.Selecthalf2:
                        # print("Welcome mister")
                        work_duration = get_work_duration(punch, current_time)
                        half_day_duration = get_half_hour(
                            rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                        if work_duration < half_day_duration:
                            # print("Is lessthan!")
                            punch.status = 'AN'
                        else:
                            punch.status = 'HL'
                    elif leave.endDate == punch.date.date() and '.5' in str(leave.Days) and leave.Selecthalf1 == leave.Selecthalf2:
                        # print("Welcome mister 2")
                        work_duration = get_work_duration(punch, current_time)
                        half_day_duration = get_half_hour(
                            rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                        if work_duration < half_day_duration:
                            # print("Is lessthan!")
                            punch.status = 'AN'
                        else:
                            punch.status = 'HL'

                    else:
                        # print("Welcome mister 3")
                        punch.status = 'L'

                print("Updated punch status: ", punch.status)

                punch.is_penalty_reverted = True
                punch.save()
        email_id = leave.applicant_email.email
        to = [email_id]
        from_email = ''
        password = ''

        default_name = "HRMS"
        subject = f"Leave Application Accepted"
        email_from = settings.EMAIL_HOST_USER
        html_body = render_to_string(
            'index/approve_email.html', {'data': leave, 'emailconfig': default_name, 'startdate': startdate, 'enddate': enddate})
        msg = EmailMultiAlternatives(
            subject=subject, from_email=email_from, to=to)
        msg.attach_alternative(html_body, "text/html")
        # msg.send()

        leave.save()
        LeaveNotification.objects.create(
            user=leave.applicant_email,
            message=f"{leave.applicant_email.username}'s leave has been approved on {datetime.now().strftime('%d %B %Y')}",
            is_approved=True, admin_id=request.user.id
        )
        if request.user.role == "Employee":
            return redirect("reporting_manager_team_log")
        return redirect('log')
        

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def bulk_approve_leave(request):
    if request.method == 'POST':
        admin_id = request.user.id
        applied_leaves = Leave.objects.filter(Q(status='Applied') & (
            Q(applicant_email=admin_id) | Q(admin_id=admin_id)))
        print("applied_leaves :", applied_leaves)
        current_time = timezone.now().time().replace(microsecond=0)
        for leave in applied_leaves:
            leave.status = 'Approved'
            leave.save()
            employee = leave.applicant_email
            attendance_rule = employee.assignattendancerule_set.first()
            rules_applied = attendance_rule.rules_applied if attendance_rule else None
            if leave and leave.punch_data:
                punch = Punch.objects.get(id=leave.punch_data.id)
                punch.is_approved = True
                punch.save()
            else:
                punch_data = Punch.objects.filter(
                    user=leave.applicant_email, date__gte=leave.strtDate, date__lte=leave.endDate)
                print(punch_data)
                for punch in punch_data:
                    # print("Current punch status: ", punch.status)
                    if leave.Days == 0.5:
                        # print("Leave day is 0.5")
                        work_duration = get_work_duration(punch, current_time)
                        half_day_duration = get_half_hour(
                            rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                        if work_duration < half_day_duration:
                            # print("Is lessthan!")
                            punch.status = 'AN'
                            # punch.save()
                        # print("Punch data: ", punch, work_duration)
                        pass
                    elif leave.Days == 1.0:
                        print("Yes it's a full day leave!")
                        punch.status = 'L'
                        # punch.save()
                    else:
                        # print("Leave days grater than 1 and 0.5",)
                        print(leave.strtDate, leave.endDate, leave.Selecthalf1,
                              leave.Selecthalf2, punch.date.date())
                        if leave.strtDate == punch.date.date() and '.5' in str(leave.Days) and leave.Selecthalf1 == leave.Selecthalf2:
                            # print("Welcome mister")
                            work_duration = get_work_duration(
                                punch, current_time)
                            half_day_duration = get_half_hour(
                                rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                            if work_duration < half_day_duration:
                                # print("Is lessthan!")
                                punch.status = 'AN'
                            else:
                                punch.status = 'HL'
                            # punch.save()
                        elif leave.endDate == punch.date.date() and '.5' in str(leave.Days) and leave.Selecthalf1 == leave.Selecthalf2:
                            # print("Welcom mister 2")
                            work_duration = get_work_duration(
                                punch, current_time)
                            half_day_duration = get_half_hour(
                                rules_applied.halfhours, rules_applied.halfminutes) if rules_applied else None
                            if work_duration < half_day_duration:
                                # print("Is lessthan!")
                                punch.status = 'AN'
                            else:
                                punch.status = 'HL'
                            # punch.save()
                        else:
                            # print("Welcom mister 3")
                            punch.status = 'L'
                            # punch.save()
                    # print("Updated punch status: ", punch.status)
                    punch.save()
            email_id = leave.applicant_email.email
            to = [email_id]
            print("to email id :", to)

            default_name = "HRMS"
            subject = 'Leave Application Status'

            email_from = settings.EMAIL_HOST_USER
            html_body = render_to_string(
                'index/approve_email.html', {'data': leave, 'emailconfig': default_name})
            msg = EmailMultiAlternatives(
                subject=subject, from_email=email_from, to=to)
            msg.attach_alternative(html_body, "text/html")
            msg.send()

            LeaveNotification.objects.create(
                user=leave.applicant_email,
                message=f"{leave.applicant_email.username}'s leave has been approved on {datetime.now().strftime('%d %B %Y')}",
                is_approved=True, admin_id=request.user.id
            )

        return redirect('log')


def reject_leave(request):
    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        rejection_reason = request.POST.get('rejection_reason')
        leave = Leave.objects.get(pk=leave_id)
        leave.status = 'Rejected'
        leave.rejected = True
        leave.rejection_reason = rejection_reason
        email_id = leave.applicant_email.email
        to = [email_id]
        from_email = ''
        password = ''

        default_name = "HRMS"
        subject = 'Leave Application Status'
        if leave and leave.punch_data:
            punch = Punch.objects.get(id=leave.punch_data.id)
            punch.is_rejected = True
            punch.save()
        email_from = settings.EMAIL_HOST_USER
        html_body = render_to_string('index/rejection_email.html', {
            'data': leave, 'rejection_reason': rejection_reason, 'emailconfig': default_name, })
        msg = EmailMultiAlternatives(
            subject=subject, from_email=email_from, to=to)
        msg.attach_alternative(html_body, "text/html")
        msg.send()
        leave.save()
        LeaveNotification.objects.create(
            user=leave.applicant_email,
            message=f"{leave.applicant_email.username}'s leave has been rejected on {datetime.now().strftime('%d %B %Y')}",
            is_approved=False, admin_id=request.user.id
        )
        return redirect('log')


def cancel_leave(request):
    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        leave = Leave.objects.get(pk=leave_id)
        leave.cancel_requested = True
        leave.status = 'Pending'
        leave.rejected = False
        leave.save()
    return redirect('log')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def export_leave_data(request):
    admin_id = request.user.id

    if request.method == 'POST':
        month_year_str = request.POST.get('monthselect')

        # Convert the month_year string to a datetime object
        month_year = datetime.strptime(month_year_str, '%B %Y')
        print("month_year_str ; month_year :", month_year_str, month_year)
        leave_data = Leave.objects.filter(
            Q(Appliedon__month=month_year.month, Appliedon__year=month_year.year, applicant_email__id=admin_id) |
            Q(Appliedon__month=month_year.month,
              Appliedon__year=month_year.year, applicant_email__admin_id=admin_id)
        )
        data = []

        for leave in leave_data:
            user_data = [
                leave.applicant_email.empid,
                leave.applicant_email.username,
                leave.Appliedon.strftime('%Y-%m-%d'),
                leave.status,
                leave.strtDate.strftime('%Y-%m-%d'),
                leave.endDate.strftime('%Y-%m-%d'),
                leave.Reason,
            ]
            data.append(user_data)

        print("DATA :", data)

        df = pd.DataFrame(data, columns=[
                          'Employee ID', 'Name', 'Date', 'Status', 'Start Date', 'End Date', 'Reason',])

        excel_file_name = "leave_data_export.xlsx"
        excel_file_path = os.path.join(
            BASE_DIR, 'media/csv/leave_data_export.xlsx')
        df.to_excel(excel_file_path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        fixed_width = 30

        for column in sheet.columns:
            sheet.column_dimensions[get_column_letter(
                column[0].column)].width = fixed_width

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

        wb.save(excel_file_path)

        today = datetime.now()
        filemanager = Filemanager.objects.create(
            myuser_10=request.user, requesttype="Leave Data Export", frmt="XLSX", scheduleon=today, status="In Queue")

        subject = 'Leave Data Export'
        message = 'Attached is the leave data for the specified date range.'
        to_email = [request.user.email]
        from_email = settings.DEFAULT_FROM_EMAIL

        email = EmailMessage(subject, message, from_email, to=to_email)
        email.attach(excel_file_name, open(excel_file_path,
                     'rb').read(), 'application/vnd.ms-excel')
        email.send()
        print("SEND")
        filemanager.status = "Success"
        with open(excel_file_path, 'rb') as file:
            filemanager.saveexcel.save(
                'leave_data_export.xlsx', ContentFile(file.read()))

    return redirect('filemanagernav')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_log(request):
    user_id = request.user.id
    current_date = datetime.now()
    data = Leave.objects.filter(
        applicant_email=request.user.id).order_by('-Appliedon')
    k = Myprofile.objects.filter(myuser__id=user_id)
    datas = companyprofile.objects.filter(admin_id=user_id)

    x = {
        "k": k[0] if k.exists() else k,
        "datas": datas[0] if datas.exists() else datas,
    }

    return render(request, "index/admin_logs.html", {'data': data, 'current_date': current_date, **x})


def cancel_adminleave(request):
    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        leave = Leave.objects.get(id=leave_id)
        leave.delete()

    return redirect('admin_log')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emplog(request):
    user_id = request.user.id
    current_date = datetime.now()
    formatted_date = current_date.strftime("%B %d, %Y")
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    data = Leave.objects.filter(
        applicant_email=request.user.id).order_by('-Appliedon')
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,

    }

    return render(request, "Employee/log.html", {"data": data, 'current_date': current_date, 'formatted_date': formatted_date, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def cancel_empleave(request):
    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        leave = Leave.objects.get(id=leave_id)
        leave.delete()
    return redirect('emplog')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def Rules(request):
    user_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    datas = companyprofile.objects.filter(admin_id=user_id)
    data = CompanyRules.objects.filter(admin_id=user_id)

    leaves_range = list(range(121))

    x = {
        "k": k[0] if k.exists() else k,
        "datas": datas[0] if datas.exists() else datas,
        'leaves_range': leaves_range

    }

    default_rules = [
        {"leavename": "Casual Leave", "days": 0},
        {"leavename": "Loss Of Pay", "days": 0},
        {"leavename": "Sick Leave", "days": 0},
        {"leavename": "Maternity Leave", "days": 0},
        {"leavename": "Event Leave", "days": 0},
        {"leavename": "Paternity Leave", "days": 0},
        {"leavename": "Comp Off", "days": 0},
        {"leavename": "Optional Holiday", "days": 0},
    ]

    for rule_data in default_rules:
        rule, created = CompanyRules.objects.get_or_create(
            admin_id=user_id, leavename=rule_data["leavename"], defaults=rule_data)

    return render(request, "index/Rules.html", {'data': data, **x})


def generalsettings(request):
    if request.method == 'POST':
        rule_id = request.POST.get('rule_id')
        company_rule = CompanyRules.objects.filter(id=rule_id)
        leavename = request.POST.get('leavename')
        description = request.POST.get('description')
        days = request.POST.get('days')
        WBL = request.POST.get('WBL')
        HBL = request.POST.get('HBL')
        CAB = request.POST.get('CAB')
        AF = request.POST.get('AF')
        AP = request.POST.get('AP')
        AUP = request.POST.get('AUP')
        AUNC = request.POST.get('AUNC')
        LEE = request.POST.get('LEE')
        ALE = request.POST.get('ALE')
        MLE = request.POST.get('MLE')
        CFA = request.POST.get('CFA')

        company_rule.update(description=description, days=days, WeekendsBWLeave=WBL, HolidaysBWLeave=HBL, CreditableAccrual=CAB, AccrualFrequency=AF, AccrualPeriod=AP,
                            AllowedUnderProbation=AUP, AllowedUnderNoticePeriod=AUNC, LeaveEncashEnabled=LEE, AllLeaveEncashable=ALE, MaxLeaveEncashable=MLE, CarryForwardeEnabled=CFA)

    return redirect('Rules')


def advancedsettings(request):
    if request.method == 'POST':
        rule_id = request.POST.get('rule_id')
        company_rule = CompanyRules.objects.filter(id=rule_id)
        MLAM = request.POST.get('MLAM')
        CLA = request.POST.get('CLA')
        NLA = request.POST.get('NLA')
        FLA = request.POST.get('FLA')
        FLAA = request.POST.get('FLAA')
        BLA = request.POST.get('BLA')
        BLAUpTo = request.POST.get('BLAUpTo')
        ALNY = request.POST.get('ALNY')

        company_rule.update(MaxLeavesAllowed=MLAM, ContinuousLeavesAllowed=CLA, NegativeLeavesAllowed=NLA, FutureDatedLeavesAllowed=FLA,
                            FutureDatedLeavesAllowedAfter=FLAA, BackdatedLeavesAllowed=BLA, BackdatedLeavesAllowedUpTo=BLAUpTo, ApplyLeavesforNextYearTill=ALNY)

    return redirect('Rules')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empRules(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    assign = assignrule.objects.filter(user_id=user_id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "Employee/Rules.html", {'assign': assign, **x})


# def advanced(request):
#     return render(request,"index/advanced.html")

# def sam(request):
#     return render(request,"index/sample.html")

def forgot(request):
    return render(request, "forgotpassword.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def daily_log(request):
    user_id = request.user.id
    c = companyprofile.objects.filter(admin_id=user_id)
    k = Myprofile.objects.filter(myuser__id=user_id)

    leave_notification = LeaveNotification.objects.filter(
        Q(user=user_id) | Q(user_id__admin_id=user_id))
    notifications = []

    for i in leave_notification:
        if i.admin_id == 0:
            profile = Myprofile.objects.filter(myuser=i.user).first()
        else:
            profile = Myprofile.objects.filter(myuser__id=i.admin_id).first()

        notifications.append(
            {
                "message": i.message,
                "image_url": profile.image.url
                if profile and profile.image
                else "/static/logo/userlogo.png",
                "notification_id": i.id,
            }
        )

    attendance_rule_obj = (
        AssignAttendanceRule.objects.filter(user_id=user_id)
        .values_list("rules_applied__inTime", "rules_applied__outTime")
        .first()
    )
    selected_date = request.GET.get("selected_date", None)
    if attendance_rule_obj:
        rule_in_time, rule_out_time = attendance_rule_obj
        rule_in_time = datetime.combine(datetime.now().date(), rule_in_time)
        rule_out_time = datetime.combine(datetime.now().date(), rule_out_time)

        if selected_date is None:
            selected_date = timezone.now().date()
        else:
            selected_date = datetime.strptime(
                selected_date, '%d %B %Y').date()

        regular_time_duration = rule_out_time - rule_in_time
        punch_datas = Punch.objects.filter(
            Q(user__id=user_id) & Q(date__date=selected_date)
        )

        punch_list = []
        json_data = {
            "org_in_time": rule_in_time,
            "org_out_time": rule_out_time,
            "custom_punch_logs": [],
            "break_duration": "--",
            "work_duration": "--",
            "overtime_duration": "--",
            "status": "",
        }

        anomaly_type = []
        work_duration = timedelta()
        break_duration = timedelta()

        for punch in punch_datas:
            anomaly_type.append('Clock In' if punch.in_time_anomaly else None)
            anomaly_type.append(
                'Clock Out' if punch.out_time_anomaly else None)
            anomaly_type.append(
                'Work Duration' if punch.work_duration_anomaly else None)

            for suffix in ["first", "second"]:
                in_time_key = f"{suffix}_clock_in_time"
                out_time_key = f"{suffix}_clock_out_time"
                is_clocked_in_key = f"is_{suffix}_clocked_in"
                is_clocked_out_key = f"is_{suffix}_clocked_out"

                in_time_value = (
                    getattr(punch, in_time_key, "--")
                    if getattr(punch, is_clocked_in_key, False)
                    else "--"
                )
                out_time_value = (
                    getattr(punch, out_time_key, "--")
                    if getattr(punch, is_clocked_out_key, False)
                    else "--"
                )

                if in_time_value != "--":
                    punch_list.append(
                        {
                            "in_time": in_time_value,
                            "in_type": "In Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "in_time": in_time_value,
                        "in_type": "In Time",
                        "ip_address": punch.ip_address,
                    })

                if out_time_value != "--":
                    punch_list.append(
                        {
                            "out_time": out_time_value,
                            "out_type": "Out Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "out_time": out_time_value,
                        "out_type": "Out Time",
                        "ip_address": punch.ip_address,
                    })

            # print(punch.is_first_clocked_in ,punch.is_first_clocked_out ,punch.is_second_clocked_in , sep='\n')
            if punch.is_first_clocked_in and punch.is_second_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_duration = work_end_time - work_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                work_duration = work_end_time - work_start_time
            if punch.is_first_clocked_out and punch.is_second_clocked_in and punch.is_second_clocked_out:
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                break_duration = break_end_time - break_start_time

                if work_duration >= break_duration:

                    work_duration = work_duration - break_duration
                else:

                    work_duration = break_duration-work_duration

            json_data['status'] = punch.status

        work_duration_str = str(work_duration).split(".")[0]
        try:
            work_duration = (
                datetime.strptime(work_duration_str, "%H:%M:%S")
                if work_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            work_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")

        break_duration_str = str(break_duration)
        try:
            break_duration = (
                datetime.strptime(
                    str(break_duration_str).split(".")[0], "%H:%M:%S")
                if break_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            break_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")
        work_duration_minutes = work_duration.hour * 60 + work_duration.minute
        regular_time_minutes = regular_time_duration.seconds // 60
        overtime_duration = max(
            work_duration_minutes - regular_time_minutes, 0)
        over_time_duration = timedelta(minutes=overtime_duration)

        if overtime_duration != "00:00:00:" or overtime_duration != None:
            over_time_duration = datetime.strptime(
                str(over_time_duration).split(".")[0], "%H:%M:%S")
        print("Overtime Duration: ", over_time_duration)

        json_data['break_duration'] = break_duration
        json_data['work_duration'] = work_duration
        json_data['overtime_duration'] = over_time_duration

        for punch in punch_datas:
            punch.work_duration = work_duration
            punch.break_duration = break_duration
            punch.overtime = over_time_duration
            punch.save()

        json_data['org_in_time'] = json_data['org_in_time'].isoformat()
        json_data['org_out_time'] = json_data['org_out_time'].isoformat(

        )
        json_data['custom_punch_logs'] = [
            {
                'time_type': 'in',
                'time': log['in_time'].isoformat() if log.get('in_time') is not None else None,
                'type': log.get('in_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'in_time' in log and log.get('in_time') is not None
        ] + [
            {
                'time_type': 'out',
                'time': log['out_time'].isoformat() if log.get('out_time') is not None else None,
                'type': log.get('out_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'out_time' in log and log.get('out_time') is not None
        ]

        json_data['break_duration'] = json_data['break_duration'].isoformat()
        json_data['work_duration'] = json_data['work_duration'].isoformat()
        json_data['overtime_duration'] = str(json_data['overtime_duration'])

        # Serialize the data
        serialized_json_data = json.dumps(json_data)
        # serialized_json_data = serialize('json',json_data , use_natural_keys=True)

        context = {
            "punch_list": punch_list,
            "work_duration": work_duration,
            "break_duration": break_duration,
            "overtime_duration": over_time_duration,
            "json_data": serialized_json_data,
            "anomaly_type": anomaly_type,
            "selected_date": selected_date,
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,

        }
        return render(request, "index/daily log.html", context)
    else:
        return render(request, "index/daily log.html", {"notifications": notifications})
    
from django.db.models import F

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def monthly_log(request):

    user_id = request.user.id
    try:
        user_data = User.objects.filter( Q(id=user_id) & ( Q(resignationform__isnull=True) | Q(resignationform__actual_last_working_day__gte=now().date()))).first()

        c = companyprofile.objects.filter(admin_id=user_id)
        k = Myprofile.objects.filter(myuser__id=user_id)

    except User.DoesNotExist:
        # Handle the case when the user does not exist
        messages.info(request, 'User not found')
        return render(request, "Employee/monthly_log.html")

    page = request.GET.get('page', 1)
    # Fetch attendance rule
    assignattendancerule = AssignAttendanceRule.objects.filter(
        user_id=user_id
    ).values_list("rules_applied__inTime", "rules_applied__outTime").first()
    assigned_rules = assignrule.objects.filter(user_id=user_id)

    selected_month_str = request.GET.get('monthselect', None)
    if assignattendancerule:
        rule_in_time, rule_out_time = assignattendancerule
        rule_in_time = datetime.combine(datetime.min, rule_in_time)
        rule_out_time = datetime.combine(datetime.min, rule_out_time)
        regular_time_duration = rule_out_time - rule_in_time

        current_date = timezone.now()
        current_year = current_date.year
        start_date = current_date - timedelta(45)

        if selected_month_str is not None:
            selected_month = datetime.strptime(
                selected_month_str, '%B %Y').month
        else:
            selected_month = current_date.date().month

        punch_datas = Punch.objects.filter(
            Q(user__id=user_id) &
            Q(date__month=selected_month)
        ).filter(
            Q(user__resignationform__isnull=True) |
            Q(user__resignationform__actual_last_working_day__gte=F('date'))
        ).order_by('-date', '-first_clock_in_time')

        punch_collections = []
        for punch in punch_datas:
            work_duration = timedelta()
            break_duration = timedelta()

            if punch.is_first_clocked_in and punch.is_second_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)

                work_duration = work_end_time - work_start_time
                break_duration = break_end_time - break_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                work_duration = work_end_time - work_start_time

            try:
                if work_duration != '00:00:00':
                    work_duration = datetime.strptime(
                        str(work_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                work_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")
            try:
                if break_duration != "00:00:00":
                    break_duration = datetime.strptime(
                        str(break_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                break_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")

            work_duration_minutes = work_duration.hour * 60 + work_duration.minute
            regular_time_minutes = regular_time_duration.seconds // 60
            overtime_duration = max(
                work_duration_minutes - regular_time_minutes, 0)
            over_time_duration = timedelta(minutes=overtime_duration)
            if over_time_duration != '00:00:00':
                over_time_duration = datetime.strptime(
                    str(over_time_duration).split(".")[0], "%H:%M:%S")

            punch_records = {
                "id": punch.id,
                "date": punch.date,
                "status": punch.status,
                "in_time": punch.first_clock_in_time if punch.first_clock_in_time else '--',
                "out_time": punch.second_clock_out_time if punch.second_clock_out_time else (punch.first_clock_out_time or '--'),
                "work_duration": work_duration,
                "break_duration": break_duration,
                "overtime_duration": over_time_duration,
                "breaks": punch.break_count,
                "is_approved": punch.is_approved,
                "is_requested": punch.is_requested,
                "is_rejected": punch.is_rejected,
            }
            punch_collections.append(punch_records)
        paginator = Paginator(punch_collections, 20)
        try:
            punch_collections = paginator.page(page)
        except PageNotAnInteger:
            punch_collections = paginator.page(1)
        except EmptyPage:
            punch_collections = paginator.page(paginator.num_pages)
        selected_date = datetime.strptime(
            selected_month_str, '%B %Y') if selected_month_str is not None else current_date.date()
        context = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            'punch_in_month': punch_collections,
            "assigned_leave_rules": assigned_rules,
            "selected_date": selected_date,
        }
        return render(request, "index/monthly_log.html", context)
    else:
        return render(request, "index/monthly_log.html")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empdaily_log(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)

    attendance_rule_obj = (
        AssignAttendanceRule.objects.filter(user_id=user_id)
        .values_list("rules_applied__inTime", "rules_applied__outTime")
        .first()
    )

    selected_date = request.GET.get("selected_date", None)
    if attendance_rule_obj:
        current_month = timezone.now().month
        rule_in_time, rule_out_time = attendance_rule_obj
        rule_in_time = datetime.combine(datetime.now().date(), rule_in_time)
        rule_out_time = datetime.combine(datetime.now().date(), rule_out_time)

        if selected_date is None:
            selected_date = timezone.now().date()
        else:
            selected_date = datetime.strptime(
                selected_date, '%d %B %Y').date()

        regular_time_duration = rule_out_time - rule_in_time
        punch_datas = Punch.objects.filter(
            Q(user__id=user_id) & Q(date__date=selected_date)
        )

        punch_list = []
        json_data = {
            "org_in_time": rule_in_time,
            "org_out_time": rule_out_time,
            "custom_punch_logs": [],
            "break_duration": "--",
            "work_duration": "--",
            "overtime_duration": "--",
            "status": "",
        }

        work_duration = timedelta()
        break_duration = timedelta()
        anomaly_type = []
        for punch in punch_datas:
            anomaly_type.append('Clock In' if punch.in_time_anomaly else None)
            anomaly_type.append(
                'Clock Out' if punch.out_time_anomaly else None)
            anomaly_type.append(
                'Work Duration' if punch.work_duration_anomaly else None)

            for suffix in ["first", "second"]:
                in_time_key = f"{suffix}_clock_in_time"
                out_time_key = f"{suffix}_clock_out_time"
                is_clocked_in_key = f"is_{suffix}_clocked_in"
                is_clocked_out_key = f"is_{suffix}_clocked_out"

                in_time_value = (
                    getattr(punch, in_time_key, "--")
                    if getattr(punch, is_clocked_in_key, False)
                    else "--"
                )
                out_time_value = (
                    getattr(punch, out_time_key, "--")
                    if getattr(punch, is_clocked_out_key, False)
                    else "--"
                )

                if in_time_value != "--":
                    punch_list.append(
                        {
                            "in_time": in_time_value,
                            "in_type": "In Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "in_time": in_time_value,
                        "in_type": "In Time",
                        "ip_address": punch.ip_address,
                    })

                if out_time_value != "--":
                    punch_list.append(
                        {
                            "out_time": out_time_value,
                            "out_type": "Out Time",
                            "ip_address": punch.ip_address,
                        }
                    )
                    json_data["custom_punch_logs"].append({
                        "out_time": out_time_value,
                        "out_type": "Out Time",
                        "ip_address": punch.ip_address,
                    })

            # print(punch.is_first_clocked_in ,punch.is_first_clocked_out ,punch.is_second_clocked_in , sep='\n')
            if punch.is_first_clocked_in and punch.is_second_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_duration = work_end_time - work_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                work_duration = work_end_time - work_start_time
            if punch.is_first_clocked_out and punch.is_second_clocked_in and punch.is_second_clocked_out:
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time or datetime.strptime(
                        "00:00:00", "%H:%M:%S").time()
                )

                break_duration = break_end_time - break_start_time

                if work_duration >= break_duration:

                    work_duration = work_duration - break_duration
                else:

                    work_duration = break_duration-work_duration

            json_data['status'] = punch.status

        work_duration_str = str(work_duration).split(".")[0]
        try:
            work_duration = (
                datetime.strptime(work_duration_str, "%H:%M:%S")
                if work_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            work_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")

        break_duration_str = str(break_duration)
        try:
            break_duration = (
                datetime.strptime(
                    str(break_duration_str).split(".")[0], "%H:%M:%S")
                if break_duration_str != "00:00:00"
                else timedelta()
            )
        except ValueError:
            break_duration = datetime.strptime(
                "00:00:00", "%H:%M:%S")

        work_duration_minutes = work_duration.hour * 60 + work_duration.minute
        regular_time_minutes = regular_time_duration.seconds // 60
        overtime_duration = max(
            work_duration_minutes - regular_time_minutes, 0)
        over_time_duration = timedelta(minutes=overtime_duration)

        if overtime_duration != "00:00:00:" or overtime_duration != None:
            over_time_duration = datetime.strptime(
                str(over_time_duration).split(".")[0], "%H:%M:%S")

        json_data['break_duration'] = break_duration
        json_data['work_duration'] = work_duration
        json_data['overtime_duration'] = over_time_duration

        for punch in punch_datas:
            punch.work_duration = work_duration
            punch.break_duration = break_duration
            punch.overtime = over_time_duration
            punch.save()

        json_data['org_in_time'] = json_data['org_in_time'].isoformat()
        json_data['org_out_time'] = json_data['org_out_time'].isoformat(

        )
        json_data['custom_punch_logs'] = [
            {
                'time_type': 'in',
                'time': log['in_time'].isoformat() if log.get('in_time') is not None else None,
                'type': log.get('in_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'in_time' in log and log.get('in_time') is not None
        ] + [
            {
                'time_type': 'out',
                'time': log['out_time'].isoformat() if log.get('out_time') is not None else None,
                'type': log.get('out_type', ''),
                'ip_address': log.get('ip_address', '')
            }
            for log in json_data['custom_punch_logs']
            if 'out_time' in log and log.get('out_time') is not None
        ]

        json_data['break_duration'] = json_data['break_duration'].isoformat()
        json_data['work_duration'] = json_data['work_duration'].isoformat()
        json_data['overtime_duration'] = str(json_data['overtime_duration'])

        # Serialize the data
        serialized_json_data = json.dumps(json_data)

        # serialized_json_data = serialize('json',json_data , use_natural_keys=True)

        # print(anomaly_type)
        context = {
            "punch_list": punch_list,
            "work_duration": work_duration,
            "break_duration": break_duration,
            "overtime_duration": over_time_duration,
            "json_data": serialized_json_data,
            "anomaly_type": anomaly_type,
            "selected_date": selected_date,
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
        }

        return render(request, "Employee/daily log.html", context)
    else:
        x = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
        }
        return render(request, "Employee/daily log.html", x)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def emp_monthly_log(request):
    user_id = request.user.id

    # Fetch user-related data
    try:
        user_data = User.objects.get(id=user_id)
        admin_id = user_data.admin_id
        c = companyprofile.objects.filter(admin_id=admin_id)
        k = Myprofile.objects.filter(myuser__id=user_id)

        resigned = ResignationForm.objects.filter(user=user_data, status='Approved', actual_last_working_day__lt=date.today()).exists()

    except User.DoesNotExist:
        # Handle the case when the user does not exist
        messages.info(request, 'User not found')
        return render(request, "Employee/monthly_log.html")

    page = request.GET.get('page', 1)
    # Fetch attendance rule
    assignattendancerule = AssignAttendanceRule.objects.filter(
        user_id=user_id
    ).values_list("rules_applied__inTime", "rules_applied__outTime").first()

    selected_month_str = request.GET.get('monthselect', None)
    assigned_rules = assignrule.objects.filter(user_id=user_id)
    print("Selected Month: ", selected_month_str)
    if assignattendancerule:
        current_date = timezone.now()
        current_year = current_date.year
        start_date = current_date - timedelta(45)

        rule_in_time, rule_out_time = assignattendancerule
        rule_in_time = datetime.combine(datetime.min, rule_in_time)
        rule_out_time = datetime.combine(datetime.min, rule_out_time)
        regular_time_duration = rule_out_time - rule_in_time

        if selected_month_str is not None:
            selected_month = datetime.strptime(
                selected_month_str, '%B %Y').month
        else:
            selected_month = current_date.date().month

        punch_datas = Punch.objects.filter(
            Q(user__id=user_id) & Q(
                date__month=selected_month)
        ).order_by('-date', '-first_clock_in_time')

        punch_collections = []
        for punch in punch_datas:
            work_duration = timedelta()
            break_duration = timedelta()

            if punch.is_first_clocked_in and punch.is_second_clocked_out and punch.first_clock_in_time and punch.second_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)
                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)

                work_duration = work_end_time - work_start_time
                break_duration = break_end_time - break_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time:
                work_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                work_end_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                work_duration = work_end_time - work_start_time

            try:
                if work_duration != '00:00:00':
                    work_duration = datetime.strptime(
                        str(work_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                work_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")
            try:
                if break_duration != "00:00:00":
                    break_duration = datetime.strptime(
                        str(break_duration).split(".")[0], "%H:%M:%S")
            except ValueError:
                break_duration = datetime.strptime(
                    "00:00:00", "%H:%M:%S")

            work_duration_minutes = work_duration.hour * 60 + work_duration.minute
            regular_time_minutes = regular_time_duration.seconds // 60
            overtime_duration = max(
                work_duration_minutes - regular_time_minutes, 0)
            over_time_duration = timedelta(minutes=overtime_duration)
            if over_time_duration != '00:00:00':
                over_time_duration = datetime.strptime(
                    str(over_time_duration).split(".")[0], "%H:%M:%S")

            punch_records = {
                "id": punch.id,
                "date": punch.date,
                "status": punch.status,
                "in_time": punch.first_clock_in_time if punch.first_clock_in_time else '--',
                "out_time": punch.second_clock_out_time if punch.second_clock_out_time else (punch.first_clock_out_time or '--'),
                "work_duration": work_duration,
                "break_duration": break_duration,
                "overtime_duration": over_time_duration,
                "breaks": punch.break_count,
                "is_approved": punch.is_approved,
                "is_requested": punch.is_requested,
                "is_rejected": punch.is_rejected,
            }
            punch_collections.append(punch_records)

        paginator = Paginator(punch_collections, 20)
        print("paginator: ", paginator)
        try:
            punch_collections = paginator.page(page)
        except PageNotAnInteger:
            punch_collections = paginator.page(1)
        except EmptyPage:
            punch_collections = paginator.page(paginator.num_pages)

        selected_date = datetime.strptime(
            selected_month_str, '%B %Y') if selected_month_str is not None else current_date.date()
        context = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
            'punch_in_month': punch_collections,
            "assigned_leave_rules": assigned_rules,
            "selected_date": selected_date,
        }

        return render(request, "Employee/monthly_log.html", context)
    else:
        context = {
            "k": k[0] if k.exists() else k,
            "c": c[0] if c.exists() else c,
        }
        return render(request, "Employee/monthly_log.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def automation_log(request):

    user = request.user
    today = datetime.now()
    selected_month_str = request.GET.get('monthselect')
    employee_name = request.GET.get("employee")
    page = request.GET.get('page', 1)

    if selected_month_str:
        selected_date = datetime.strptime(selected_month_str, '%B %Y').date()
        selected_month = selected_date.month
    else:
        selected_date = today.date()
        selected_month = today.month

    k = Myprofile.objects.filter(myuser__id=user.id).first()
    c = companyprofile.objects.filter(admin_id=user.id).first()

    employees = User.objects.filter(Q(admin_id=user.id) | Q(id=user.id))

    penalty_log_data = []
    loss_of_pay_updates = {}

    # Get all punches for all employees for the selected month
    all_employees_punches = Punch.objects.filter(
        Q(user__in=employees) &
        Q(status__in=['A', 'AN']) &
        Q(is_penalty_reverted=False) &
        Q(date__month=selected_month)
    )

    for punch in all_employees_punches:
        employee = punch.user
        response = {
            'punch_id': punch.id,
            'employee_id': employee.empid,
            'employee_name': employee.username,
            'anomaly_type': "",
            'penalty_type': "",
            'leave_type': "",
            'deduction': "",
            'month': punch.date,
            'penalty_revert': punch.is_penalty_reverted
        }

        print(response, punch.id)
        if punch.status == 'AN':
            work_duration = get_work_duration(punch, today.time())
            half_duration = get_half_hour(
                employee.assignattendancerule_set.first().rules_applied.halfhours,
                employee.assignattendancerule_set.first().rules_applied.halfminutes) if employee.assignattendancerule_set.exists() else None
            response['anomaly_type'] = "Anomaly"
            if work_duration < half_duration:
                response['deduction'] = 1
                response['penalty_type'] = "Full Day"
                penalty_deduction = 1
            else:
                response['deduction'] = Decimal(0.5)
                response['penalty_type'] = "Half Day"
                penalty_deduction = Decimal(0.5)
        elif punch.status == 'A':
            response['anomaly_type'] = "Absent"
            response['deduction'] = 1
            response['penalty_type'] = "Full Day"
            penalty_deduction = 1

        if employee.id not in loss_of_pay_updates:
            loss_of_pay_updates[employee.id] = 0
        loss_of_pay_updates[employee.id] += penalty_deduction

        existing_penalty_log = PenaltyLogs.objects.filter(
            punch_data=punch).first()

        if existing_penalty_log:
            # Update existing entry
            existing_penalty_log.anomaly_type = "Anomaly" if punch.status == 'AN' else "Absent"
            existing_penalty_log.penalty_type = "Full Day" if punch.status == 'A' else "Half Day"
            existing_penalty_log.deduction = penalty_deduction
            existing_penalty_log.leave_type = "Loss Of Pay"
            existing_penalty_log.month = punch.date
            existing_penalty_log.save()

        else:

            penalty_log_data.append(PenaltyLogs(
                user=employee,
                punch_data=punch,
                anomaly_type=response['anomaly_type'],
                penalty_type=response['penalty_type'],
                deduction=response['deduction'],
                leave_type="Loss Of Pay",
                month=punch.date
            ))

    # Bulk create penalty logs
    PenaltyLogs.objects.bulk_create(penalty_log_data)
    if employee_name:
        penalty_log_data = PenaltyLogs.objects.filter(
            Q(user__username__contains=employee_name) |
            Q(user__email__contains=employee_name) &
            Q(user__status='Active'),
            punch_data__date__month=selected_month
            ).order_by("-punch_data__date")
    else:
        penalty_log_data = PenaltyLogs.objects.filter(
            Q(user=user) |
            Q(user__admin_id=user.id) &
            Q(user__status='Active'),
            punch_data__date__month=selected_month
            ).order_by("-punch_data__date")


    context = {
        "c": c,
        "k": k,
        'automation_data': penalty_log_data,
        "is_view_automation_log": False,
        "selected_date": selected_date,
        "employee_name": employee_name,
        "page": page
    }
    return render(request, "index/automation logs.html", context)




@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def export_automation_log(request):
    if request.method == 'POST':
        month_str = request.POST.get('month')
        month = datetime.strptime(month_str, '%B %Y')
        year_selected = month.year
        month_selected = month.month
        userid = request.user.id
        # absent_anomaly_punches = Punch.objects.filter(
        #     Q(user__id=userid) | Q(user__admin_id=userid),
        #     status__in=['A', 'AN'],
        #     date__date__month=month_selected,
        #     date__date__year=year_selected
        # )

        automation_list = []
        absent_count = {}
        anomaly_count = {}
        penalty_log_data = PenaltyLogs.objects.filter(Q(user__id=userid) | Q(
            user__admin_id=userid), punch_data__date__month=month_selected).order_by("-punch_data__date")
        print(penalty_log_data)

        for penalty in penalty_log_data:
            print(penalty.user.username)
            automation_list.append({
                "Employee Id": penalty.user.empid,
                "Employee Name": penalty.user.username,
                "Anomaly Type": 'Absent' if penalty.punch_data.status == 'A' else 'Anomaly' if penalty.punch_data.status == 'AN' else '',
                "Penalty Type": penalty.penalty_type,
                "Leave Type": penalty.leave_type,
                "Deduction": penalty.deduction,
                "Date": penalty.punch_data.date.date().strftime('%d-%m-%Y'),
            })

            if penalty.punch_data.status == 'A':
                absent_count[penalty.punch_data.user.username] = absent_count.get(
                    penalty.punch_data.user.username, 0) + 1
            elif penalty.punch_data.status == 'AN':
                anomaly_count[penalty.punch_data.user.username] = anomaly_count.get(
                    penalty.punch_data.user.username, 0) + 1

        combined_counts = {employee_name: {'Anomaly Count': anomaly_count.get(employee_name, 0),
                                           'Absent Count': absent_count.get(employee_name, 0)}
                           for employee_name in set(anomaly_count) | set(absent_count)}

        counts_df = pd.DataFrame(combined_counts).T.reset_index()
        counts_df.columns = ['Employee Name', 'Anomaly Count', 'Absent Count']

        counts_df = counts_df.sort_values(
            by='Employee Name').reset_index(drop=True)

        df = pd.DataFrame(automation_list)
        # excel_file_name = 'penalty_log.xlsx'
        path = os.path.join(BASE_DIR, 'media/csv/penalty_log.xlsx')

        df.to_excel(path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(path)

        with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
            writer.book = wb
            counts_df.to_excel(writer, index=False, sheet_name='Sheet2')

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
                for col in sheet.columns:
                    for cell in col:
                        alignment_obj = cell.alignment.copy(
                            horizontal='left', vertical='center')
                        cell.alignment = alignment_obj
        wb.save(path)
        today = datetime.now()
        filemanager = Filemanager.objects.create(
            myuser_10=request.user,
            requesttype="Penalty Log Export",
            frmt="XLSX",
            scheduleon=today,
            status="In Queue",
        )
        filemanager.status = "Success"
        with open(path, 'rb') as file:
            filemanager.saveexcel.save(
                'automation_log.xlsx', ContentFile(file.read()))
    return redirect("filemanagernav")


def empautomation_log(request):
    user = request.user
    admin_id = User.objects.get(id=user.id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    page = request.GET.get('page', 1)

    # automation_data = []
    # employee_punches = Punch.objects.filter(user = user , status = 'A').order_by('-date')
    # for punch in employee_punches:
    #     automation_data.append({'employee_id':user.empid,'employee_name': user.username , 'anomaly_type':'Absent' if punch.status == 'A' else punch.status, 'penalty_type':"Full Day",'leave_type':"Loss Of Pay" ,  "deduction":"1.0",'month':punch.date,'penalty_log': punch.is_penalty_reverted})

    penalty_data = PenaltyLogs.objects.filter(user=user)
    print(penalty_data)
    paginator = Paginator(penalty_data, 10)
    try:
        penalty_data = paginator.page(page)
    except PageNotAnInteger:
        penalty_data = paginator.page(1)
    except EmptyPage:
        penalty_data = paginator.page(paginator.num_pages)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "automation_data": penalty_data,
    }

    return render(request, "Employee/automation_logs.html", x)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empdash(request):
    user_id = request.user.id
    wrklcn = request.user.wrklcn.id if request.user.wrklcn else None
    admin_id = request.user.admin_id
    datas = User.objects.all()
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)
    leave_notification = LeaveNotification.objects.filter(user=user_id)
    button_flag = True
    now = timezone.now()
    if request.user.status not in ["Active", "active"]:
        button_flag = False

    today = datetime.now().date()
    current_day = today.day
    current_month = today.month
    currentdate = datetime.now()
    month_name = currentdate.strftime("%B")

    attendance_count = Punch.objects.filter(
        date__month=currentdate.month, date__year=currentdate.year, user_id=user_id).count()

    # Filter holidays in the current month
    first_day = today.replace(day=1)
    last_day = today.replace(day=monthrange(today.year, today.month)[1])

    # Holiday count for the current month
    holiday_count = HolidayLocationList.objects.filter(
        Holiday_List__Myuser_13=admin_id,
        HolidayLocation__id=wrklcn,
        Holiday_List__HolidayDate__range=(first_day, last_day)
    ).count()
    
    pending_leavecount = Leave.objects.filter(Q(Appliedon__month=currentdate.month, Appliedon__year=currentdate.year) & Q(
        status="Applied" or "Pending") & (Q(applicant_email=user_id))).count()

    AN_count = Punch.objects.filter(
        date__month=currentdate.month, date__year=currentdate.year, status="AN", user_id=user_id).count()

    users = User.objects.filter(Q(admin_id=admin_id) | Q(id=admin_id), status__iexact="Active")
    # print("users 2 : ", users)

    birthday_users_count = [user for user in users if user.dob and datetime.strptime(parse_and_format_date(user.dob), '%d %B %Y').date(
    ).month == current_month and datetime.strptime(parse_and_format_date(user.dob), '%d %B %Y').date().day >= current_day]

    work_anniversary_users_count = [user for user in users if user.datejoin and datetime.strptime(parse_and_format_date(user.datejoin), '%d %B %Y').date().month == current_month and datetime.strptime(
        parse_and_format_date(user.datejoin), '%d %B %Y').date().year != today.year and datetime.strptime(parse_and_format_date(user.datejoin), '%d %B %Y').date().day >= current_day]

    birthday_count = len(birthday_users_count)
    work_anniversary_count = len(work_anniversary_users_count)

    notifications = []

    for i in leave_notification:
        admin_id_to_use = i.admin_id if i.admin_id != 0 else user_id
        profile = Myprofile.objects.filter(myuser__id=admin_id_to_use).first()

        notifications.append({
            "message": i.message,
            "image_url": profile.image.url
            if profile and profile.image else "/static/logo/userlogo.png",
            "notification_id": i.id,
            "user": i.user,
            'admin_id': i.admin_id,
            'is_approved': i.is_approved,
            "readuser": i.readuser,
            "events": i.events,
        })

    users = User.objects.filter(Q(admin_id=admin_id) | Q(id=admin_id), status__iexact="Active")

    
    # print("users :", users)
    today = datetime.today().date()
    next_30_days = [(today + timedelta(days=i)).strftime("%d-%m") for i in range(1, 31)] 

    # Fetch birthdays in next 30 days
    upcoming_count = [
        user for user in users
        if user.status.lower() == "active"
        and user.dob is not None
        and (
            (user.dob.strftime("%d-%m") if isinstance(user.dob, datetime) else datetime.strptime(user.dob, "%d %B %Y").strftime("%d-%m"))
        ) in next_30_days
    ]
    upcoming_birthday = len(upcoming_count)

    # Fetch birthdays on in the current day
    today_str = datetime.today().strftime("%m-%d")
    birthday_users = [
        user for user in users
        if user.status.lower() == "active"
        and user.dob is not None
        and (
            (user.dob.strftime("%m-%d") if isinstance(user.dob, datetime) 
            else datetime.strptime(user.dob, "%d %B %Y").strftime("%m-%d"))
        ) == today_str
    ]
    birthday_count = len(birthday_users)


    # Fetch work anniversaries in next 30 days
    today = datetime.today().date()
    next_30_days = [(today + timedelta(days=i)).strftime("%m-%d") for i in range(30)]

    # Work Anniversaries in Next 30 Days
    upcoming_anniv_users = []
    for user in users:
        if user.status.lower() == "active" and user.datejoin is not None:
            try:
                date_str = (
                    user.datejoin.strftime("%m-%d")
                    if isinstance(user.datejoin, date)
                    else datetime.strptime(user.datejoin, "%d %B %Y").strftime("%m-%d")
                )
                if date_str in next_30_days:
                    upcoming_anniv_users.append(user)
            except (ValueError, TypeError):
                continue
    upcoming_work_anniversary = len(upcoming_anniv_users)
    # Work Anniversaries Today
    today_str = today.strftime("%m-%d")
    today_anniv_users = []
    for user in users:
        if user.status.lower() == "active" and user.datejoin is not None:
            try:
                date_str = (
                    user.datejoin.strftime("%m-%d")
                    if isinstance(user.datejoin, date)
                    else datetime.strptime(user.datejoin, "%d %B %Y").strftime("%m-%d")
                )
                if date_str == today_str:
                    today_anniv_users.append(user)
            except (ValueError, TypeError):
                continue
    today_work_anniversary = len(today_anniv_users)
      
    attendance_rule = AssignAttendanceRule.objects.filter(
        user_id__id=user_id).first()
    current_date = datetime.today()
    punch_object = Punch.objects.filter(
        user__id=user_id, date__date=current_date.date()).last()
    clock_in_type = 2
    if punch_object:
        clock_in_type = punch_object.last_punch_type
    if not attendance_rule:
        messages.info(request, 'Attendance Rules not assigned')
    # display current work duration
    in_time = time(hour=0, minute=0, second=0)
    try:
        punch_data = Punch.objects.get(
            user=request.user, date__date=now.date())
        in_time = punch_data.first_clock_in_time
    except Punch.DoesNotExist:
        pass
    print("In time: ", in_time)


    x = {
        "k": k.first(),
        "data": data.first(),
        "notifications": notifications,
        "attendance_count": attendance_count,
        "holiday_count": holiday_count,
        "month_name": month_name,
        "pending_leavecount": pending_leavecount,
        "AN_count": AN_count,
        "clock_in_type": clock_in_type,
        "button_flag": button_flag,
        "in_time": in_time,
        'upcoming_birthday': upcoming_birthday,
        "birthday_count": birthday_count,
        'upcoming_work_anniversary': upcoming_work_anniversary,
        "today_work_anniversary": today_work_anniversary,
    }
    return render(request, "Employee/empdash.html", {"datas": datas, "leave_notification": leave_notification,  "users": users,
                                                     "birthday_users": birthday_users, "show_birthday_image": bool(birthday_users), "work_ann": today_anniv_users, "show_workann_image": bool(today_anniv_users), "month_name": month_name, **x})

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_emailconfig(request):
    if request.method == "POST":
        em = request.POST.get('Email')
        pw = request.POST.get('password')
        us = request.POST.get('Username')
        de = request.POST.get('Designation')
        cm = request.POST.get('C_name')
        md = request.POST.get('Module')

        Email_config.objects.create(
            email=em, password=pw, username=us, designation=de, c_name=cm, module=md)
        return redirect('email')
    return render(request, "index/emailconfig.html")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def emailconfig(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.all()
    em = Email_config.objects.all()
    query = request.GET.get('search')

    if query:
        datas_list1 = Email_config.objects.filter(email__contains=query)
        # datas_list2 = Designation.objects.filter(username__contains=query)

        if datas_list1:

            em = datas_list1

        else:
            em = []
            # messages.info(request, 'No Records Found')

    page = request.GET.get('page', 1)
    paginator = Paginator(em, 10)
    try:
        em = paginator.page(page)
    except PageNotAnInteger:
        em = paginator.page(1)
    except EmptyPage:
        em = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, 'index/emailconfig.html', {'em': em, 'datas': em, 'query': query, 'data': data, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def update_emailconfig(request):
    if request.method == "POST":
        e = request.POST.get('Email')
        p = request.POST.get('password')
        u = request.POST.get('Username')
        d = request.POST.get('Designation')
        c = request.POST.get('C_name')
        m = request.POST.get('Module')
        eid = request.POST.get('emid')
        k = Email_config.objects.filter(id=eid)
        print(k)
        k.update(email=e, password=p, username=u,
                 designation=d, c_name=c, module=m)
        return redirect('email')
    return render(request, "index/emailconfig.html", {"em": k})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_emailconfig(request, userid):
    k = Email_config.objects.get(id=userid)
    k.delete()
    return redirect('email')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def view_profile(request):
    user_id = request.user.id
    user_profile = Myprofile.objects.select_related('reportingmanager').filter(myuser_id=user_id).first()

    # These may be optional—adapt based on your real structure
    primary = user_profile.primary_reporting_manager if user_profile else None
    secondary = user_profile.secondary_reporting_manager if user_profile else None

    # Extract related user IDs
    primary_user_id = primary.myuser.id if primary and primary.myuser else None
    secondary_user_id = secondary.myuser.id if secondary and secondary.myuser else None

    company = companyprofile.objects.all()
    context = {
        'user': user_profile,
        'company': company,
        'primary_manager': primary.myuser.get_full_name() if primary and primary.myuser else '',
        'primary_manager_type': primary.type if primary else '',
        'primary_manager_department': primary.department if primary else '',
        'primary_manager_designation': primary.designation if primary else '',
        'primary_manager_user_id': primary_user_id,
        'secondary_manager': secondary.myuser.get_full_name() if secondary and secondary.myuser else '',
        'secondary_manager_type': secondary.type if secondary else '',
        'secondary_manager_department': secondary.department if secondary else '',
        'secondary_manager_designation': secondary.designation if secondary else '',
        'secondary_manager_user_id': secondary_user_id,
    }
    return render(request, 'index/profile.html', context)



# @login_required(login_url='login')
# @cache_control(no_cache=True, must_revalidate=True,no_store=True)
# def send_email(request):
#     email=request.user.email
#     send_mail('leave apply','hello',
#     'cydezt@gmail.com',
#     [email],
#     fail_silently=False)
#     return render(request,'index/leave.html')


def help_page(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=request.user.id)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, 'index/help.html', x)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def pay_slip(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    regaddress = registeredaddress.objects.filter(admin_id=admin_id)
    print("regofficeaddress :", regaddress)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
    print("selected_year :", selected_year, selected_month)

    bank_details = Bank_account.objects.filter(myuser_11=admin_id)
    print("bank_details :", bank_details)

    assign_salarystructure = AssignSalaryStructure.objects.filter(
        user_id=admin_id, effective_date__month=selected_month, effective_date__year=selected_year).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)

    assign_data = []
    assigndata = []
    ctc_assigndata = []

    gross_salary_amount = 0
    work_from_office_allowance_amount = 0
    total_net_salary = 0
    total_ctc_salary = 0
    wfocount = 0
    leave_count = 0
    total_gross_salary = 0  # Ensure this is always defined

    selected_date = datetime(selected_year, selected_month, 1)
    print("selected_date :", selected_date)

    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(
            effective_date__lte=selected_date, user_id=admin_id).order_by('-effective_date').first()

        if nearest_date:
            assign_salarystructure = nearest_date

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
        work_from_office_component = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance", Parentcomponentname__componentname__iexact="Gross Salary").first()
        print("gross_salary_component ; work_from_office_component : ", gross_salary_component, work_from_office_component)
        net_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="Net Salary")
        print("net_salary_component : ", net_salary_component)
        ctc_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="CTC")
        print("ctc_salary_component : ", ctc_salary_component)

        name = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
        amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=name)
        print("name ; amount 1st :", name, amount)
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname=gross_salary_component)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        print("names ; amounts  :", names, amounts)

        net_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="Net Salary")
        net_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=net_names)
        print("names ; amounts  :", net_names, net_amounts)

        ctc_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="CTC").exclude(salarycomponent__componentname__icontains="professional tax")
        ctc_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=ctc_names)
        print("ctc_names ; ctc_amounts  :", ctc_names, ctc_amounts)

        if gross_salary_component:
            gross_amount = amount.filter(salaryname__salarycomponent=gross_salary_component).first()
            gross_salary_amount = gross_amount.amount if gross_amount else 0            
            
        if work_from_office_component:
            work_amount = amounts.filter(salaryname__salarycomponent=work_from_office_component).first()
            work_from_office_allowance_amount = work_amount.amount if work_amount else 0
        
        for netsalry in net_salary_component:
            net_salary = amount.filter(salaryname__salarycomponent=netsalry)
            total_net_salary += net_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total Net Salary:", total_net_salary)

        for ctcsalry in ctc_salary_component:
            if ctcsalry.componentname.lower() != 'professional tax':
                ctc_salary = amount.filter(salaryname__salarycomponent=ctcsalry)
                total_ctc_salary += ctc_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total CTC Salary:", total_ctc_salary)

        total_gross_salary = gross_salary_amount - work_from_office_allowance_amount
        print("total_gross_salary : ", total_gross_salary)

        zipped_data = zip_longest(names, amounts)
        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })

        zippeddata = zip_longest(net_names, net_amounts)
        assigndata.append({
            'rule': rule,
            'zippeddata': zippeddata,
        })

        ctc_zippeddata = zip_longest(ctc_names, ctc_amounts)
        ctc_assigndata.append({
            'rule': rule,
            'ctc_zippeddata': ctc_zippeddata,
        })
 
    print("total_gross_salary :", total_gross_salary, "work_from_office_allowance_amount : ", work_from_office_allowance_amount)

    punch_obj = Punch.objects.filter(user__id=admin_id,date__year=selected_year,date__month=selected_month)
    print("punch_obj : ", punch_obj)

    for punch in punch_obj:
        if punch.WfhOrWfo == "WFO":
            wfocount += 1
                    
        leave_data = Leave.objects.filter(
                applicant_email=admin_id,  
                strtDate=punch.date,    
                status="Approved"      
            ).first()       
        print(f"Leave data for {admin_id} on {punch.date}: ", leave_data)
        
        if punch.status == "H":
            leave_count += 1
        elif punch.status == "L":
            if leave_data:
                if leave_data.leavetyp != "Loss Of Pay":
                    leave_count += 1
        elif punch.status == "HL":
            if leave_data:
                print("ccccccccccccccccccccccccccccccccccccc  ")
                if leave_data.leavetyp == "Loss Of Pay":
                    leave_count -= 0.5
                    print("KKKKKKKKKKKKKKKKKKKKK", leave_count)
                
    wfo_count = wfocount + leave_count
    print("wfo_count :", wfo_count, "month_numeric , selected_year:" , selected_month, selected_year)

    year_select = int(selected_year)
    num_days = calendar.monthrange(year_select, selected_month)[1]
    first_day_of_month = datetime(year_select, selected_month, 1)
    if selected_month == 12: 
        next_month = datetime(year_select + 1, 1, 1)
        print("next_month 1 :", next_month, year_select)
    else:
        next_month = datetime(year_select, selected_month + 1, 1)
        print("next_month 2 :", next_month)

    day_count = 0
    current_day = first_day_of_month
    while current_day < next_month:
        if current_day.weekday() != 6: 
            day_count += 1
        current_day += timedelta(days=1)
    count_sundays = num_days - day_count
    print("day_count ############ :", day_count, num_days, count_sundays)

    print("work_from_office_allowance_amount:", work_from_office_allowance_amount)
    perday_WFOamount = work_from_office_allowance_amount / day_count
    total_WFOamount = perday_WFOamount * wfo_count
    print("total_WFOamount : ", wfo_count, total_WFOamount)
    
    WFOamount = round(work_from_office_allowance_amount - total_WFOamount) #This amount add to the deduction
    print("WFOamount :", WFOamount)

    adhoc_data = Adhoc.objects.filter(user_id=admin_id, createddate__year=selected_year,
                                      createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
    print("adhoc_data : ", adhoc_data)

    earning_amount = 0
    deduction_amount = 0
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    total_earnings = gross_salary_amount + earning_amount
    print("total_earnings ; gross_salary_amount ; earning_amount :", total_earnings, gross_salary_amount, earning_amount)

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    total_anomaly_count = 0
    attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=admin_id)
    print("attendance_rule :", attendance_rule)
    for att_rule in attendance_rule:
        rule_type = att_rule.rules_applied
        print("rule_type :", rule_type, )
        if rule_type:
            full_day_hours = rule_type.fullhours
            full_day_minutes = rule_type.fullminutes
            full_time = timedelta(hours=full_day_hours,
                                  minutes=full_day_minutes)
            half_day_hours = rule_type.halfhours
            half_day_minutes = rule_type.halfminutes
            half_time = timedelta(hours=half_day_hours,
                                  minutes=half_day_minutes)
            print("Full Day Hours:", full_day_hours,
                  full_day_minutes, full_time)
            print("Half Day Hours:", half_day_hours,
                  half_day_minutes, half_time)
            in_grace_period = rule_type.inGracePeriod
            out_grace_period = rule_type.outGracePeriod
            print("Grace period:", in_grace_period, out_grace_period)
            in_grace_timedelta = timedelta(
                hours=in_grace_period.hour, minutes=in_grace_period.minute)
            out_grace_timedelta = timedelta(
                hours=out_grace_period.hour, minutes=out_grace_period.minute)

            total_grace_period = in_grace_timedelta + out_grace_timedelta
            print("Total Grace period:", total_grace_period)
            total_fullday_time = full_time + total_grace_period
            print("Total Time:", total_fullday_time)
            total_halfday_time = half_time + total_grace_period
            print("total_halfday_time :", total_halfday_time)

    num_days = calendar.monthrange(selected_year, selected_month)[1]

    payregister = PayRegister.objects.filter(createddate__month=selected_month, createddate__year=selected_year, user_id=admin_id, status__in=[
                                             "Payslip Generated", "Payslip Downloaded"])
    print("payregister :", payregister)

    lop_data = Runpayroll_lop.objects.filter(
        lop_date__month=selected_month, lop_date__year=selected_year, user_id=admin_id)
    lopcount = 0
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    punches = Punch.objects.filter(
        user__id=admin_id,
        date__year=selected_year,
        date__month=selected_month,
        status="AN", is_penalty_reverted=False
    )
    print("Punch Object :", punches)

    for punch in punches:
        total_work_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)
            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

        elif punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration
        if total_work_duration > total_fullday_time:
            AN_count = 0.5
        elif total_work_duration < total_halfday_time:
            AN_count = 1.0
        else:
            AN_count = 0.5

        total_anomaly_count += AN_count
        print("total_anomaly_count :", total_anomaly_count)

    print("total_anomaly_count 2:", total_anomaly_count)
    absent_count = Punch.objects.filter(user__id=admin_id, date__year=selected_year,
                                        date__month=selected_month, status='A', is_penalty_reverted=False).count()

    absent_AN_count = absent_count + total_anomaly_count
    print("absent_AN_count : ", absent_count, absent_AN_count)

    punchcount = Punch.objects.filter(user__id=admin_id, date__year=selected_year, date__month=selected_month).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count, absent_AN_count, lopcount)
    total_lop = absent_AN_count + missing_date_count + lopcount

    per_day_amount = total_gross_salary / num_days
    print("per_day_amount :", per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount)
    lopamount = round(lop_amount)

    total_deductions = deduction_amount + lop_amount + total_net_salary + WFOamount
    totaldeductions = round(total_deductions)
    net_amount = round(total_earnings - total_deductions)
    print("net_amount :", net_amount)

    net_amount_words = num2words(net_amount, lang='en_IN')

    print("ctc_assigndata : ", ctc_assigndata)

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "regaddress": regaddress,
        "payregister": payregister,
        "num_days": num_days,
        "bank_details": bank_details,
        "assign_salarystructure": assign_salarystructure,
        "assign_data": assign_data,
        "assigndata": assigndata,
        "ctc_assigndata": ctc_assigndata,
        "adhoc_data": adhoc_data,
        "total_earnings": total_earnings,
        "total_ctc_salary": total_ctc_salary,
        "totaldeductions": totaldeductions,
        "total_lop": total_lop,
        "lopamount": lopamount,
        "wfo_count": wfo_count, 
        "WFOamount": WFOamount,
        "net_amount": net_amount,
        "net_amount_words": net_amount_words,
        "month_str": month_str,
        "selected_year": selected_year,
        "is_view_payroll": False,
    }

    return render(request, 'index/pay_slip.html', context)

def download_payslip_pdf(request):
    payreg_id = request.POST.get("payreg_id")
    month_selected = request.POST.get("month_str", "")
    datetime_object = datetime.strptime(month_selected, "%B")
    month_numeric = datetime_object.month
    year_selected = request.POST.get("selected_year", "")
    print("payreg_id : ", payreg_id, month_selected,
          month_numeric, year_selected)

    custom_width = 700
    custom_height = 1000

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(custom_width, custom_height),
                            topMargin=35, title=f"Payslip_{month_selected}_{year_selected}")
    elements = []

    font_path = str(Path(settings.BASE_DIR) / 'app1'/'arialfont'/'arial.ttf')
    font_path2 = str(Path(settings.BASE_DIR) /
                     'app1'/'arialfont'/'arialbd.ttf')
    pdfmetrics.registerFont(TTFont('Arial', font_path))
    pdfmetrics.registerFont(TTFont('Arial-Bold', font_path2))
    font_size = 12
    font_style = 'Arial'
    bold_font_style = 'Arial-Bold'

    company_name_style = ParagraphStyle(
        name='CompanyName',
        fontName=bold_font_style,
        fontSize=16,
        alignment=TA_CENTER,
    )

    address_style = ParagraphStyle(
        name='CompanyAddress',
        fontSize=11,
        textColor='black',
        fontName=font_style,
        leading=14,
        alignment=TA_CENTER,
    )

    num_days = calendar.monthrange(int(year_selected), month_numeric)[1]

    payreg = PayRegister.objects.filter(id=payreg_id).first()
    admin_id = payreg.user_id.id
    print("admin_id :", admin_id)
    desg = ''
    dept = ''
    loc = ''
    user_data = User.objects.filter(id=admin_id).first()
    if user_data:
        username = user_data.username
        empid = user_data.empid
        desg = user_data.designation.name
        dept = user_data.department.name
        loc = user_data.wrklcn.location

    bankname = ''
    acno = ''
    pfnum = ''
    esa = ''
    bank_details = Bank_account.objects.filter(myuser_11=admin_id).first()
    if bank_details:
        bankname = bank_details.bank_name
        acno = bank_details.account_number
        pfnum = bank_details.pfnum
        esa = bank_details.esa

    idno = ''
    uploadeddocs = Uploadeddocs.objects.filter(
        myuser=admin_id, type1="PAN Card").first()
    if uploadeddocs:
        idno = uploadeddocs.id_no

    cmp_data = companyprofile.objects.filter(admin_id=admin_id).first()
    if cmp_data:
        cmp_logo = cmp_data.logo
        company_logo = Image(cmp_logo.path, width=2*inch, height=1*inch)
        company_name = cmp_data.registeredcompanyname
        company_name_paragraph = Paragraph(company_name, company_name_style)

    company_address_para = ''
    regaddress = registeredaddress.objects.filter(admin_id=admin_id).first()
    if regaddress:
        company_address_str = f"{regaddress.regofficeaddress}, {regaddress.regdistrict}, {regaddress.regstate}, {regaddress.regcountry} - {regaddress.regpincode}"
        company_address_para = Paragraph(company_address_str, address_style)

    assign_data = []
    ctc_amount = 0

    selected_date = datetime(int(year_selected), month_numeric, 1)
    print("selected_date :", selected_date)

    assign_salarystructure = AssignSalaryStructure.objects.filter(
        user_id=admin_id, effective_date__month=month_numeric, effective_date__year=year_selected).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)

    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(
            effective_date__lte=selected_date, user_id=admin_id).order_by('-effective_date').first()

        if nearest_date:
            assign_salarystructure = nearest_date

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        names = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure)
        amounts = AssignSalaryStructureAmount.objects.filter(
            salaryname__in=names)
        print("names ; amounts 9828 :", names, amounts)

        ctc_amount += sum(amount.amount for amount in amounts)
        zipped_data = zip_longest(names, amounts)

        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })
    print("ctc_amount 9837:", ctc_amount)
    print("assign_data :", assign_data)

    adhoc_data = Adhoc.objects.filter(user_id=admin_id, createddate__year=year_selected,
                                      createddate__month=month_numeric).select_related('adhocearning', 'adhocdeduction')
    print("adhoc_data : ", adhoc_data)

    earning_amount = 0
    deduction_amount = 0
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    total_earnings = ctc_amount + earning_amount

    lop_data = Runpayroll_lop.objects.filter(
        lop_date__month=month_numeric, lop_date__year=year_selected, user_id=admin_id)
    lopcount = 0
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    total_anomaly_count = 0
    attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=admin_id)
    print("attendance_rule :", attendance_rule)
    for att_rule in attendance_rule:
        rule_type = att_rule.rules_applied
        print("rule_type :", rule_type, )
        if rule_type:
            full_day_hours = rule_type.fullhours
            full_day_minutes = rule_type.fullminutes
            full_time = timedelta(hours=full_day_hours,
                                  minutes=full_day_minutes)
            half_day_hours = rule_type.halfhours
            half_day_minutes = rule_type.halfminutes
            half_time = timedelta(hours=half_day_hours,
                                  minutes=half_day_minutes)
            print("Full Day Hours:", full_day_hours,
                  full_day_minutes, full_time)
            print("Half Day Hours:", half_day_hours,
                  half_day_minutes, half_time)
            in_grace_period = rule_type.inGracePeriod
            out_grace_period = rule_type.outGracePeriod
            print("Grace period:", in_grace_period, out_grace_period)
            in_grace_timedelta = timedelta(
                hours=in_grace_period.hour, minutes=in_grace_period.minute)
            out_grace_timedelta = timedelta(
                hours=out_grace_period.hour, minutes=out_grace_period.minute)

            total_grace_period = in_grace_timedelta + out_grace_timedelta
            print("Total Grace period:", total_grace_period)
            total_fullday_time = full_time + total_grace_period
            print("Total Time:", total_fullday_time)
            total_halfday_time = half_time + total_grace_period
            print("total_halfday_time :", total_halfday_time)

    punches = Punch.objects.filter(
        user__id=admin_id,
        date__year=year_selected,
        date__month=month_numeric,
        status="AN", is_penalty_reverted=False
    )
    print("Punch Object :", punches)

    for punch in punches:
        print("1111111111111111111111111111111111")
        total_work_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)
            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

        elif punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration
        if total_work_duration > total_fullday_time:
            AN_count = 0.5
        elif total_work_duration < total_halfday_time:
            AN_count = 1.0
        else:
            AN_count = 0.5

        total_anomaly_count += AN_count
        print("total_anomaly_count :", total_anomaly_count)

    print("total_anomaly_count 2:", total_anomaly_count)
    absent_count = Punch.objects.filter(user__id=admin_id, date__year=year_selected,
                                        date__month=month_numeric, status='A', is_penalty_reverted=False).count()

    absent_AN_count = absent_count + total_anomaly_count
    print("absent_AN_count : ", absent_count, absent_AN_count)

    punchcount = Punch.objects.filter(
        user__id=admin_id, date__year=year_selected, date__month=month_numeric).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count)
    working_days = punchcount - absent_AN_count
    print("working_days :", working_days)
    total_lop = absent_AN_count + missing_date_count + lopcount

    per_day_amount = ctc_amount / num_days
    print("per_day_amount :", per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount)
    lopamount = round(lop_amount)

    total_deductions = deduction_amount + lop_amount
    totaldeductions = round(total_deductions)
    net_amount = round(total_earnings - total_deductions)
    print("net_amount :", net_amount)

    # net_amount_words = num2words(net_amount, lang='en_IN')

    net_amount_words = num2words(net_amount, lang='en_IN').title()
    print("net_amountwords :", net_amount_words)
    net_amount_words_formatted = f"({net_amount_words})"

    data = [
        [company_logo, company_name_paragraph, ''],
        ['', company_address_para, ''],
        ['',
            f'Payslip for the Month of {month_selected}, {year_selected}', ''],

    ]

    colWidths = [150, 400, 100]
    rowHeights = [20, 50, 50]

    table = Table(data, colWidths=colWidths, rowHeights=rowHeights)
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, -1), (-1, -1), 13),
        ('FONTNAME', (0, -1), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        # ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))

    data1 = [
        ["Name:", username, 'Employee ID:', empid],
        ["Designation:", desg, 'Bank Name:', bankname],
        ["Department:", dept, 'Bank Account No.:', acno],
        ["Location:", loc, 'PAN No.:', idno],
        ["Effective Work Days:", num_days, 'PF No.:', pfnum],
        ["LOP:", total_lop, 'ESI No.:', esa],

    ]

    colWidths1 = [120, 200, 120, 150]
    rowHeights1 = [20, 20, 20, 20, 20, 20]

    table1 = Table(data1, colWidths=colWidths1, rowHeights=rowHeights1)
    table1.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        # ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBEFORE', (2, 0), (2, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        # ('TOPPADDING', (0, 0), (-1, -1), 300),
    ]))

    data2 = [
        ["Earnings", "Amount", 'Deductions', "Amount"],
    ]

    componentnames = []
    amounts = []
    adhocnames = []
    adhocamounts = []

    for data_entry in assign_data:
        for name, amount in data_entry['zipped_data']:
            component_name = ', '.join(
                component.componentname for component in name.salarycomponent.all()) if name else ''
            amount_value = amount.amount if amount else ''

            componentnames.append(component_name)
            amounts.append(amount_value)

    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            component_name = adhoc_entry.adhocearning.component_name
            amount = adhoc_entry.amount

            componentnames.append(component_name)
            amounts.append(amount)

    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocdeduction:
            adhoc_name = adhoc_entry.adhocdeduction.component_name
            adhoc_amount = adhoc_entry.amount

            adhocnames.append(adhoc_name)
            adhocamounts.append(adhoc_amount)

    adhocnames.append("LOP Amount")
    adhocamounts.append(lopamount)

    max_length = max(len(componentnames), len(adhocnames))
    print("max_length : 10060 : ", max_length)

    for i in range(max_length):
        component_name = componentnames[i] if i < len(componentnames) else ''
        amount_value = amounts[i] if i < len(amounts) else ''

        adhoc_name = adhocnames[i] if i < len(adhocnames) else ''
        adhoc_amount = adhocamounts[i] if i < len(adhocamounts) else ''

        data2.append([component_name, amount_value, adhoc_name, adhoc_amount])

    data2.extend([
        ['Total Earnings (Rs)', total_earnings,
         'Total Deductions (Rs)', totaldeductions],
        ['Net Pay For The Month:', net_amount, '', ''],
        [net_amount_words_formatted, '', '', ''],
        ['', '', '', ''],
        ['', '', 'This is a system generated payslip and does not require signature.', ''],
    ])

    colWidths2 = [200, 95, 200, 95]
    # rowHeights2 = [20, 20, 20, 20, 20, 30, 20]
    row_height = 20
    rowHeights2 = [row_height] * (6 + max_length)

    rowHeights2[0] = 30
    rowHeights2[-4] = 60
    rowHeights2[-3] = 20
    rowHeights2[-2] = 10
    rowHeights2[-1] = 30

    print("rowHeights2 :", rowHeights2)

    table2 = Table(data2, colWidths=colWidths2, rowHeights=rowHeights2)
    table2.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 0), (-1, 0), bold_font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 0.25, colors.black),

        ('FONTNAME', (0, -5), (-1, -4), bold_font_style),  # Total
        # ('LINEBELOW', (0, -4), (-1, -4), 1, colors.black),
        # ('LINEABOVE', (0, -4), (-1, -4), 1, colors.black),

        ('LINEBEFORE', (2, 0), (2, 1 + max_length), 1, colors.black),
        # ('LINEAFTER', (2, 0), (2, -2 -max_length), 1, colors.black),
        ('LINEABOVE', (0, -5), (-1, -5), 0.25, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 0.25, colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
        ('BOX', (0, 0), (-1, 1 + max_length), 1, colors.black),
    ]))

    spacer_height = 20
    spacer = Spacer(1, spacer_height)

    elements.append(table)
    elements.append(table1)
    elements.append(spacer)
    elements.append(table2)

    doc.build(elements)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payslip_{month_selected}_{year_selected}.pdf"'
    response.write(pdf_data)
    return response

def download_emppayslip_pdf(request):
    payreg_id = request.POST.get("payreg_id")
    month_selected = request.POST.get("month_str", "")
    datetime_object = datetime.strptime(month_selected, "%B")
    month_numeric = datetime_object.month
    year_selected = request.POST.get("selected_year", "")
    print("payreg_id : ", payreg_id, month_selected,
          month_numeric, year_selected)

    custom_width = 700
    custom_height = 1000

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(custom_width, custom_height),
                            topMargin=35, title=f"Payslip_{month_selected}_{year_selected}")
    elements = []

    font_path = str(Path(settings.BASE_DIR) / 'app1'/'arialfont'/'arial.ttf')
    font_path2 = str(Path(settings.BASE_DIR) /
                     'app1'/'arialfont'/'arialbd.ttf')
    pdfmetrics.registerFont(TTFont('Arial', font_path))
    pdfmetrics.registerFont(TTFont('Arial-Bold', font_path2))
    font_style = 'Arial'
    bold_font_style = 'Arial-Bold'

    company_name_style = ParagraphStyle(
        name='CompanyName',
        fontName=bold_font_style,
        fontSize=16,
        alignment=TA_CENTER,
    )

    address_style = ParagraphStyle(
        name='CompanyAddress',
        fontSize=11,
        textColor='black',
        fontName=font_style,
        leading=14,
        alignment=TA_CENTER,
    )

    num_days = calendar.monthrange(int(year_selected), month_numeric)[1]

    payreg = PayRegister.objects.filter(id=payreg_id).first()
    user_id = payreg.user_id.id
    admin_id = payreg.user_id.admin_id
    print("user_id :", user_id, admin_id)

    desg = ''
    dept = ''
    loc = ''
    user_data = User.objects.filter(id=user_id).first()
    if user_data:
        username = user_data.username
        empid = user_data.empid
        desg = user_data.designation.name
        dept = user_data.department.name
        loc = user_data.wrklcn.location

    bankname = ''
    acno = ''
    pfnum = ''
    esa = ''
    bank_details = Bank_account.objects.filter(myuser_11=user_id).first()
    if bank_details:
        bankname = bank_details.bank_name
        acno = bank_details.account_number
        pfnum = bank_details.pfnum
        esa = bank_details.esa

    idno = ''
    uploadeddocs = Uploadeddocs.objects.filter(
        myuser=user_id, type1="PAN Card").first()
    if uploadeddocs:
        idno = uploadeddocs.id_no

    cmp_data = companyprofile.objects.filter(admin_id=admin_id).first()
    if cmp_data:
        cmp_logo = cmp_data.logo if cmp_data.logo else ''
        print("Company Logo: ", cmp_data.logo)
        company_logo = Image(cmp_logo.path, width=2*inch, height=1*inch)
        company_name = cmp_data.registeredcompanyname if cmp_data.registeredcompanyname else ''
        company_name_paragraph = Paragraph(company_name, company_name_style)

    company_address_para = ''
    regaddress = registeredaddress.objects.filter(admin_id=admin_id).first()
    if regaddress:
        company_address_str = f"{regaddress.regofficeaddress}, {regaddress.regdistrict}, {regaddress.regstate}, {regaddress.regcountry} - {regaddress.regpincode}"
        company_address_para = Paragraph(company_address_str, address_style)

    assign_data = []
    ctc_amount = 0

    selected_date = datetime(int(year_selected), month_numeric, 1)
    print("selected_date :", selected_date)

    assign_salarystructure = AssignSalaryStructure.objects.filter(
        user_id=user_id, effective_date__month=month_numeric, effective_date__year=year_selected).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)

    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(
            effective_date__lte=selected_date, user_id=user_id).order_by('-effective_date').first()

        if nearest_date:
            assign_salarystructure = nearest_date

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        names = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure)
        amounts = AssignSalaryStructureAmount.objects.filter(
            salaryname__in=names)
        print("names ; amounts 10242 :", names, amounts)

        ctc_amount += sum(amount.amount for amount in amounts)
        zipped_data = zip_longest(names, amounts)

        assign_data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })
    print("ctc_amount :10251: ", ctc_amount)
    print("assign_data :", assign_data)

    adhoc_data = Adhoc.objects.filter(user_id=user_id, createddate__year=year_selected,
                                      createddate__month=month_numeric).select_related('adhocearning', 'adhocdeduction')
    print("adhoc_data : ", adhoc_data)

    earning_amount = 0
    deduction_amount = 0
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    total_earnings = ctc_amount + earning_amount

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    total_anomaly_count = 0
    attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=user_id)
    print("attendance_rule :", attendance_rule)
    for att_rule in attendance_rule:
        rule_type = att_rule.rules_applied
        print("rule_type :", rule_type, )
        if rule_type:
            full_day_hours = rule_type.fullhours
            full_day_minutes = rule_type.fullminutes
            full_time = timedelta(hours=full_day_hours,
                                  minutes=full_day_minutes)
            half_day_hours = rule_type.halfhours
            half_day_minutes = rule_type.halfminutes
            half_time = timedelta(hours=half_day_hours,
                                  minutes=half_day_minutes)
            print("Full Day Hours:", full_day_hours,
                  full_day_minutes, full_time)
            print("Half Day Hours:", half_day_hours,
                  half_day_minutes, half_time)
            in_grace_period = rule_type.inGracePeriod
            out_grace_period = rule_type.outGracePeriod
            print("Grace period:", in_grace_period, out_grace_period)
            in_grace_timedelta = timedelta(
                hours=in_grace_period.hour, minutes=in_grace_period.minute)
            out_grace_timedelta = timedelta(
                hours=out_grace_period.hour, minutes=out_grace_period.minute)

            total_grace_period = in_grace_timedelta + out_grace_timedelta
            print("Total Grace period:", total_grace_period)
            total_fullday_time = full_time + total_grace_period
            print("Total Time:", total_fullday_time)
            total_halfday_time = half_time + total_grace_period
            print("total_halfday_time :", total_halfday_time)

    punches = Punch.objects.filter(
        user__id=user_id,
        date__year=year_selected,
        date__month=month_numeric,
        status="AN", is_penalty_reverted=False
    )
    print("Punch Object :", punches)

    for punch in punches:
        print("1111111111111111111111111111111111")
        total_work_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)
            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

        elif punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration
        if total_work_duration > total_fullday_time:
            AN_count = 0.5
        elif total_work_duration < total_halfday_time:
            AN_count = 1.0
        else:
            AN_count = 0.5

        total_anomaly_count += AN_count
        print("total_anomaly_count :", total_anomaly_count)

    lop_data = Runpayroll_lop.objects.filter(
        lop_date__month=month_numeric, lop_date__year=year_selected, user_id=user_id)
    lopcount = 0
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    absent_count = Punch.objects.filter(user__id=user_id, date__year=year_selected,
                                        date__month=month_numeric, status='A', is_penalty_reverted=False).count()

    absent_AN_count = absent_count + total_anomaly_count
    print("absent_AN_count : ", absent_count, absent_AN_count)

    punchcount = Punch.objects.filter(
        user__id=user_id, date__year=year_selected, date__month=month_numeric).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count)
    working_days = punchcount - absent_AN_count
    print("working_days :", working_days)
    total_lop = absent_AN_count + missing_date_count + lopcount

    per_day_amount = ctc_amount / num_days
    print("per_day_amount :", per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount)
    lopamount = round(lop_amount)

    total_deductions = deduction_amount + lop_amount
    totaldeductions = round(total_deductions)
    net_amount = round(total_earnings - total_deductions)
    print("net_amount :", net_amount)

    # net_amount_words = num2words(net_amount, lang='en_IN')

    net_amount_words = num2words(net_amount, lang='en_IN').title()
    print("net_amountwords :", net_amount_words)
    net_amount_words_formatted = f"({net_amount_words})"

    data = [
        [company_logo, company_name_paragraph, ''],
        ['', company_address_para, ''],
        ['',
            f'Payslip for the Month of {month_selected}, {year_selected}', ''],

    ]

    colWidths = [150, 400, 100]
    rowHeights = [20, 50, 50]

    table = Table(data, colWidths=colWidths, rowHeights=rowHeights)
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, -1), (-1, -1), 13),
        ('FONTNAME', (0, -1), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        # ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))

    data1 = [
        ["Name:", username, 'Employee ID:', empid],
        ["Designation:", desg, 'Bank Name:', bankname],
        ["Department:", dept, 'Bank Account No.:', acno],
        ["Location:", loc, 'PAN No.:', idno],
        ["Effective Work Days:", num_days, 'PF No.:', pfnum],
        ["LOP:", total_lop, 'ESI No.:', esa],

    ]

    colWidths1 = [120, 200, 120, 150]
    rowHeights1 = [20, 20, 20, 20, 20, 20]

    table1 = Table(data1, colWidths=colWidths1, rowHeights=rowHeights1)
    table1.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        # ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBEFORE', (2, 0), (2, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        # ('TOPPADDING', (0, 0), (-1, -1), 300),
    ]))

    data2 = [
        ["Earnings", "Amount", 'Deductions', "Amount"],
    ]

    componentnames = []
    amounts = []
    adhocnames = []
    adhocamounts = []

    for data_entry in assign_data:
        for name, amount in data_entry['zipped_data']:
            component_name = ', '.join(
                component.componentname for component in name.salarycomponent.all()) if name else ''
            amount_value = amount.amount if amount else ''

            componentnames.append(component_name)
            amounts.append(amount_value)

    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            component_name = adhoc_entry.adhocearning.component_name
            amount = adhoc_entry.amount

            componentnames.append(component_name)
            amounts.append(amount)

    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocdeduction:
            adhoc_name = adhoc_entry.adhocdeduction.component_name
            adhoc_amount = adhoc_entry.amount

            adhocnames.append(adhoc_name)
            adhocamounts.append(adhoc_amount)

    adhocnames.append("LOP Amount")
    adhocamounts.append(lopamount)

    max_length = max(len(componentnames), len(adhocnames))
    print("max_length :", max_length)

    for i in range(max_length):
        component_name = componentnames[i] if i < len(componentnames) else ''
        amount_value = amounts[i] if i < len(amounts) else ''

        adhoc_name = adhocnames[i] if i < len(adhocnames) else ''
        adhoc_amount = adhocamounts[i] if i < len(adhocamounts) else ''

        data2.append([component_name, amount_value, adhoc_name, adhoc_amount])

    data2.extend([
        ['Total Earnings (Rs)', total_earnings,
         'Total Deductions (Rs)', totaldeductions],
        ['Net Pay For The Month:', net_amount, '', ''],
        [net_amount_words_formatted, '', '', ''],
        ['', '', '', ''],
        ['', '', 'This is a system generated payslip and does not require signature.', ''],
    ])

    colWidths2 = [200, 95, 200, 95]
    # rowHeights2 = [20, 20, 20, 20, 20, 30, 20]
    row_height = 20
    rowHeights2 = [row_height] * (6 + max_length)

    rowHeights2[0] = 30
    rowHeights2[-4] = 60
    rowHeights2[-3] = 20
    rowHeights2[-2] = 10
    rowHeights2[-1] = 30

    print("rowHeights2 :", rowHeights2)

    table2 = Table(data2, colWidths=colWidths2, rowHeights=rowHeights2)
    table2.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 0), (-1, 0), bold_font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 0.25, colors.black),

        ('FONTNAME', (0, -5), (-1, -4), bold_font_style),  # Total
        # ('LINEBELOW', (0, -4), (-1, -4), 1, colors.black),
        # ('LINEABOVE', (0, -4), (-1, -4), 1, colors.black),

        ('LINEBEFORE', (2, 0), (2, 1 + max_length), 1, colors.black),
        # ('LINEAFTER', (2, 0), (2, -2 -max_length), 1, colors.black),
        ('LINEABOVE', (0, -5), (-1, -5), 0.25, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 0.25, colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
        ('BOX', (0, 0), (-1, 1 + max_length), 1, colors.black),
    ]))

    spacer_height = 20
    spacer = Spacer(1, spacer_height)

    elements.append(table)
    elements.append(table1)
    elements.append(spacer)
    elements.append(table2)

    doc.build(elements)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payslip_{month_selected}_{year_selected}.pdf"'
    response.write(pdf_data)
    return response

def emppay_slip(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=request.user.id)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }

    return render(request, 'Employee/pay_slip.html', x)


# def view_salary_structure(request):
#     admin_id = request.user.id
#     k = Myprofile.objects.filter(myuser__id=request.user.id)
#     data = companyprofile.objects.filter(admin_id=admin_id)
#     datas = salary_structure.objects.all()
#     s = SalaryComponent.objects.filter(admin_id=admin_id)

#     x = {
#         "k": k[0] if k.exists() else k,
#         "data": data[0] if data.exists() else data,
#         # "is_view_salarystructure": False
#     }
#     return render(request, 'index/salarystructure.html', {'datas': datas, 's': s, **x})


def empview_salary_structure(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    data = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    datas = SalaryStructureRule.objects.all()

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, 'Employee/salary_structure.html', {'datas': datas, **x})


def add_salary_structure(request, uid_12):
    if request.method == 'POST':
        bs = request.POST.get('basic')
        hr = request.POST.get('hra')
        pf = request.POST.get('pf')
        esi = request.POST.get('esi')
        sp = request.POST.get('spa')
        dt = request.POST.get('date')
        ct = int(bs) + int(hr) + int(pf) + int(esi) + int(sp)
        cta = int(ct) * 12
        bsa = int(bs) * 12
        hra = int(hr) * 12
        pfa = int(pf) * 12
        esia = int(esi) * 12
        spa = int(sp) * 12

        u_12 = User.objects.get(id=uid_12)
        salary_structure.objects.create(date=dt, ctc=ct, basic=bs, hra=hr, pf_employer=pf, esi_employer=esi,
                                        special_allowance=sp,
                                        ctc_a=cta, basic_a=bsa, hra_a=hra, pf_employer_a=pfa, esi_employer_a=esia,
                                        special_allowance_a=spa, myuser_12=u_12)
        return redirect('salary_structure')


def update_salary_structure(request):
    if request.method == 'POST':
        # ct=request.POST.get('ctc')
        bs = request.POST.get('basic')
        hr = request.POST.get('hra')
        pf = request.POST.get('pf')
        esi = request.POST.get('esi')
        sp = request.POST.get('spa')
        dt = request.POST.get('date')

        bsa = int(bs) * 12
        hra = int(hr) * 12
        pfa = int(pf) * 12
        esia = int(esi) * 12
        spa = int(sp) * 12
        ct = int(bs) + int(hr) + int(pf) + int(esi) + int(sp)
        cta = int(ct) * 12
        salary_id = request.POST.get('salaryid')
        s = salary_structure.objects.filter(id=salary_id)
        s.update(date=dt, ctc=ct, basic=bs, hra=hr, pf_employer=pf, esi_employer=esi, special_allowance=sp,
                 ctc_a=cta, basic_a=bsa, hra_a=hra, pf_employer_a=pfa, esi_employer_a=esia, special_allowance_a=spa)
    return redirect('salary_structure')


def delete_salary_structure(request, myuser_12__id):
    k = salary_structure.objects.get(id=myuser_12__id)
    k.delete()
    return redirect('salary_structure')


def declaration(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "is_view_declaration": False
    }
    return render(request, 'index/declaration.html', x)


def empdeclaration(request):
    user_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, 'Employee/declaration.html', x)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def bank_account(request):
    user = request.user
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=user.id)
    datas = Bank_account.objects.all()

    try:
        bankdatas = Bank_account.objects.get(myuser_11=user)
    except Bank_account.DoesNotExist:
        return redirect('update_bank_account')

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        'bankdatas': bankdatas,
    }

    return render(request, 'index/bank_account.html', {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def update_bank_account(request):
    user = request.user
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=user.id)
    try:
        bankdatas = Bank_account.objects.get(myuser_11=user)
    except Bank_account.DoesNotExist:
        bankdatas = Bank_account(myuser_11=user)

    if request.method == 'POST':
        acc_holder = request.POST['name']
        accnt_no = request.POST['acctno']
        bnk_name = request.POST['bankname']
        branch = request.POST['branchname']
        ifsc = request.POST['ifsccode']
        cty = request.POST['city']

        bankdatas.account_holder_name = acc_holder
        bankdatas.account_number = accnt_no
        bankdatas.bank_name = bnk_name
        bankdatas.branch_name = branch
        bankdatas.IFSC_code = ifsc
        bankdatas.city = cty
        bankdatas.save()
        return redirect('bank_account')

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        'bankdatas': bankdatas
    }

    return render(request, 'index/bank_account.html', x)

def delete_bank_account(request, myuser_11__id):
    k = Bank_account.objects.get(id=myuser_11__id)
    k.delete()
    return redirect('bank_account')


def add_bank_account(request, uid_11):
    if request.method == 'POST':
        accnt_no = request.POST.get('account_number')
        acc_holder = request.POST.get('acc_holder_name')
        bnk_name = request.POST.get('bank')
        branch = request.POST.get('branch')
        IFSC = request.POST.get('ifsc')
        cty = request.POST.get('city')
        u_11 = User.objects.get(id=uid_11)
        Bank_account.objects.create(account_holder_name=acc_holder, account_number=accnt_no,
                                    bank_name=bnk_name, branch_name=branch, IFSC_code=IFSC, city=cty, myuser_11=u_11)
        return redirect('bank_account')


def empbank_account(request):
    user = request.user
    admin_id = User.objects.get(id=request.user.id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    datas = Bank_account.objects.filter(myuser_11=request.user.id)

    try:
        bankdatas = Bank_account.objects.get(myuser_11=user)
    except Bank_account.DoesNotExist:
        return redirect('empupdate_bank-account')

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'bankdatas': bankdatas,
    }

    return render(request, 'Employee/bank_account.html', {'datas': datas, **x})


def empupdate_bank_account(request):
    user = request.user
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=user.admin_id)
    try:
        bankdatas = Bank_account.objects.get(myuser_11=user)
    except Bank_account.DoesNotExist:
        bankdatas = Bank_account(myuser_11=user)

    if request.method == 'POST':
        acc_holder = request.POST['name']
        accnt_no = request.POST['acctno']
        bnk_name = request.POST['bankname']
        branch = request.POST['branchname']
        ifsc = request.POST['ifsccode']
        cty = request.POST['city']
        esa = request.POST['esa']
        pfnum = request.POST['pfnum']
        print(f"ESI: {esa}, PF Number: {pfnum}") 

        bankdatas.account_holder_name = acc_holder
        bankdatas.account_number = accnt_no
        bankdatas.bank_name = bnk_name
        bankdatas.branch_name = branch
        bankdatas.IFSC_code = ifsc
        bankdatas.city = cty
        bankdatas.esa = esa
        bankdatas.pfnum = pfnum
        bankdatas.save()
        return redirect('empbank_account')

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'bankdatas': bankdatas
    }

    return render(request, 'Employee/bank_account.html', x)


def empadd_bank_account(request, uid_11):
    if request.method == 'POST':
        accnt_no = request.POST.get('account_number')
        acc_holder = request.POST.get('acc_holder_name')
        bnk_name = request.POST.get('bank')
        branch = request.POST.get('branch')
        IFSC = request.POST.get('ifsc')
        cty = request.POST.get('city')
        esa = request.POST.get('esa')
        pfnum = request.POST.get('pfnum')
        u_11 = User.objects.get(id=uid_11)
        Bank_account.objects.create(account_holder_name=acc_holder, account_number=accnt_no,
                                    bank_name=bnk_name, branch_name=branch, IFSC_code=IFSC, city=cty, myuser_11=u_11,esa=esa,pfnum=pfnum)
        return redirect('empbank_account')


def empdelete_bank_account(request, myuser_11__id):
    k = Bank_account.objects.get(id=myuser_11__id)
    k.delete()
    return redirect('empbank_account')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def leavebalance(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    datas = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id)).order_by('empid')
    query = request.GET.get('search')
    if query:
        datas_list1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        if datas_list1 or datas_list2:
            datas = datas_list1 | datas_list2
        else:
            datas = []
            # messages.info(request, 'No Records Found')

    user_leave_names = []
    seen_leave_names = set()

    for leave_name in assignrule.objects.filter(Q(user_id=admin_id) | Q(user_id__admin_id=admin_id)).values_list('rules_applied__leavename', flat=True):
        if leave_name not in seen_leave_names:
            user_leave_names.append(leave_name)
            seen_leave_names.add(leave_name)

    print(user_leave_names)

    leave_balance = assignrule.objects.filter(Q(user_id=admin_id) | Q(
        user_id__admin_id=admin_id)).values('rules_applied__leavename', 'leavebalance', 'user_id__id')
    print(leave_balance)

    # page = request.GET.get('page', 1)
    # paginator = Paginator(datas, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)
    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "datas": datas,
        "user_leave_names": user_leave_names,
        "leave_balance": leave_balance,
        "query": query
    }

    return render(request, "index/leavebalance.html", {'datas': datas, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def assignleaverule(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    datas = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id))
    dn = Designation.objects.all()
    dp = Department.objects.all()
    sd = Subdepartment.objects.all()
    jb = Job.objects.all()
    wr = Worklocation.objects.all()
    com_rule = CompanyRules.objects.filter(admin_id=request.user.id)
    assg_rule = assignrule.objects.all()
    query = request.GET.get('search')

    count_user = User.objects.count()

    if query:
        datas_list1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list3 = User.objects.filter(Q(department__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list4 = User.objects.filter(Q(designation__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list5 = User.objects.filter(Q(email__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list6 = User.objects.filter(Q(phone__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list7 = User.objects.filter(Q(status__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list8 = User.objects.filter(Q(subdepartment__subname__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list9 = User.objects.filter(Q(jobtitle__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list10 = User.objects.filter(Q(wrklcn__location__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list11 = User.objects.filter(Q(emptype__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list12 = User.objects.filter(Q(probperiod__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))

        if datas_list1 or datas_list2 or datas_list3 or datas_list4 or datas_list5 or datas_list6 or datas_list7 or datas_list8 or datas_list9 or datas_list10 or datas_list11 or datas_list12:

            datas = datas_list1 | datas_list2 | datas_list3 | datas_list4 | datas_list5 | datas_list6 | datas_list7 | datas_list8 | datas_list9 | datas_list10 | datas_list11 | datas_list12

        else:
            datas = []
            # messages.info(request, 'No Records Found')

    # page = request.GET.get('page', 1)
    # paginator = Paginator(datas, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)
    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/AssignLeaveRule.html",
                  {'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'datas': datas, 'query': query, 'k': k,
                   'count_user': count_user, 'assg_rule': assg_rule, 'com_rule': com_rule, **x})


def addassignrule(request):
    com_rules = CompanyRules.objects.all()

    if request.method == 'POST':
        selected_rules = request.POST.getlist('rule')
        effective_date_str = request.POST.get('effdate')
        print("effective_date:", effective_date_str)
        selected_employees = request.POST.getlist('selected_employees')

        for employee_id in selected_employees:
            total_credited_leave = 0
            total_leave_balance = 0

            for rule_id in selected_rules:
                if assignrule.objects.filter(rules_applied=rule_id, user_id_id=employee_id, effective_date=effective_date_str).exists():
                    print("Selected rule", rule_id, "already exists for employee",
                          employee_id, "and effective date", effective_date_str)
                else:
                    company_rule = CompanyRules.objects.get(id=rule_id)

                    effective_date = datetime.strptime(
                        effective_date_str, "%d %B %Y").date()

                    if company_rule.days <= 0:
                        assign_rule = assignrule.objects.create(
                            user_id_id=employee_id,
                            effective_date=effective_date_str,
                            creditedleaves=0,
                            appliedleaves=0,
                            penaltydeduction=0,
                            leavebalance=0
                        )
                        assign_rule.rules_applied.add(rule_id)
                    else:

                        today = date.today()
                        print("company_rule.carryforward  : ",
                              company_rule, company_rule.CarryForwardeEnabled)
                        if company_rule.CarryForwardeEnabled == "on" and (company_rule.leavename != "Maternity Leave" or company_rule.leavename != "Optional Holiday"):
                            current_month = effective_date
                            print("current_month 1 :", current_month)
                            while current_month <= today:
                                if company_rule.days > 0:
                                    last_day_of_month = current_month.replace(
                                        day=calendar.monthrange(current_month.year, current_month.month)[1])
                                    total_days_in_month = calendar.monthrange(
                                        current_month.year, current_month.month)[1]
                                    print("last_day_of_month ; total_days_in_month :",
                                          last_day_of_month, total_days_in_month)

                                    if current_month == effective_date:
                                        total_day_in_month = (
                                            last_day_of_month - effective_date).days + 1
                                        print("total_day_in_month :",
                                              total_day_in_month)
                                    else:
                                        total_day_in_month = total_days_in_month
                                        print("total_day_in_month 2:",
                                              total_day_in_month)

                                    one_month_credited_leave = company_rule.days / 12
                                    one_day_credited_leave = one_month_credited_leave / total_days_in_month
                                    total_credited_leave_in_effective_date = one_day_credited_leave * total_day_in_month

                                    total_credited_leave += total_credited_leave_in_effective_date
                                    total_leave_balance += total_credited_leave_in_effective_date

                                current_month = current_month.replace(
                                    day=1) + relativedelta(months=1)

                            print("total_credited_leave ; total_leave_balance if :",
                                  total_credited_leave, total_leave_balance)

                        else:

                            today = date.today()
                            current_date = datetime.now().date()
                            current_year = datetime.now().year
                            print("current_date :", current_date, current_year)

                            if effective_date.year != current_year:
                                effective_date = datetime(
                                    current_year, 1, 1).date()
                            else:
                                effective_date = effective_date

                            current_month = effective_date
                            while current_month <= today:
                                if company_rule.days > 0:
                                    if company_rule.leavename != "Maternity Leave" and company_rule.leavename != "Optional Holiday":

                                        last_day_of_month = current_month.replace(
                                            day=calendar.monthrange(current_month.year, current_month.month)[1])
                                        total_days_in_month = calendar.monthrange(
                                            current_month.year, current_month.month)[1]

                                        if current_month == effective_date:
                                            total_day_in_month = (
                                                last_day_of_month - effective_date).days + 1

                                        else:
                                            total_day_in_month = calendar.monthrange(
                                                current_month.year, current_month.month)[1]

                                        one_month_credited_leave = company_rule.days / 12
                                        one_day_credited_leave = one_month_credited_leave / total_days_in_month
                                        total_credited_leave_in_effective_date = one_day_credited_leave * total_day_in_month
                                        total_credited_leave += total_credited_leave_in_effective_date
                                        total_leave_balance += total_credited_leave_in_effective_date
                                        print("total_credited_leave ; total_leave_balance 1 :",
                                              total_credited_leave, total_leave_balance)

                                    else:

                                        total_credited_leave = Decimal(
                                            company_rule.days)
                                        total_leave_balance = Decimal(
                                            company_rule.days)
                                        print("total_credited_leave ; total_leave_balance :",
                                              total_credited_leave, total_leave_balance)

                                    current_month = current_month.replace(
                                        day=1) + relativedelta(months=1)

                        assign_rule = assignrule.objects.create(user_id_id=employee_id, effective_date=effective_date_str,
                                                                creditedleaves=total_credited_leave, appliedleaves=0, penaltydeduction=0, leavebalance=total_leave_balance)
                        assign_rule.rules_applied.add(rule_id)

        return redirect('assignrule')

    return render(request, "index/AssignLeaveRule.html", {'com_rules': com_rules})


def delete_assign_rule(request, assign_rule_id):
    assign_rule = assignrule.objects.get(id=assign_rule_id)
    assign_rule.delete()
    return redirect('assignrule')


# Attendance/Rules
from django.utils.html import escape


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def attendanceRule(request):
    user = get_object_or_404(User, id=request.user.id)
    k = Myprofile.objects.filter(myuser__id=user.id)
    data = companyprofile.objects.filter(admin_id=user.id)
    leave = CompanyRules.objects.filter(admin_id=user.id)
    attendance_rule = AttendanceRule.objects.filter(user_id=user.id)

    # Convert queryset to JSON-safe format
    rule_data = []
    for i in attendance_rule:
        rule_data.append({
            "id": i.id,
            "rulename": escape(i.rulename),
            "description": escape(i.description),
            "edit": f"""<button class='btn btn-link' data-toggle='modal' data-target='#editrule{i.id}'>
                            <i class='fa fa-edit' style='font-size: 24px; color: grey;'></i>
                        </button>""",
            "delete": f"""<a class='dropdown-item' onclick="return confirm('Are you sure to delete?')" 
                        href='{reverse('delete_attendancerule', args=[i.id])}'>
                        <i class='dw dw-delete-3'></i></a>"""
        })

    x = {
        "k": k[0] if k.exists() else None,
        "data": data[0] if data.exists() else None,
        "attendance_rule": attendance_rule,
        "leave": leave,
        "json_data": json.dumps(rule_data, cls=DjangoJSONEncoder)
    }
    return render(request, 'index/attendanceRule.html', x)


def add_attendanceRule(request):
    admin_id = get_object_or_404(User, id=request.user.id)
    if request.method == 'POST':
        rulename = request.POST.get('rulename')
        description = request.POST.get('description')
        inTime = request.POST.get('inTime')
        outTime = request.POST.get('outTime')
        autoDeductionDate = request.POST.get('autoDeductionDate')
        enable_AD = request.POST.get('enable_AD') == 'on'
        enable_AT = request.POST.get('enable_AT') == 'on'

        inTimeGracePeriodHours = request.POST.get("inGracePeriod-hours")
        inTimeGracePeriodMinutes = request.POST.get("inGracePeriod-minutes")
        outTimeGracePeriodHours = request.POST.get("outGracePeriod-hours")
        outTimeGracePeriodMinutes = request.POST.get("outGracePeriod-minutes")
        inGracePeriod = datetime.strptime(
            f"{inTimeGracePeriodHours}:{inTimeGracePeriodMinutes}", "%H:%M")
        outGracePeriod = datetime.strptime(
            f"{outTimeGracePeriodHours}:{outTimeGracePeriodMinutes}", "%H:%M")

        fullhours = request.POST.get('fullday-hours')
        fullminutes = request.POST.get('fullday-minutes')
        halfhours = request.POST.get('halfday-hours')
        halfminutes = request.POST.get('halfday-minutes')
        maximum_TBD = request.POST.get('maximum_TBD') == 'on'
        maximum_NOB = request.POST.get('maximum_NOB') == 'on'
        auto_CO = request.POST.get('auto_CO') == 'on'
        enable_OT = request.POST.get('enable_OT') == 'on'
        enable_IR = request.POST.get('enable_IR') == 'on'
        enable_CO = request.POST.get('enable_CO') == 'on'
        enable_PR = request.POST.get('enable_PR') == 'on'

        in_Time = request.POST.get('in_Time') == 'on'
        lateComingAllowded = request.POST.get('lateComingAllowded')
        penaltyInterval = request.POST.get('penaltyInterval')
        penalty = request.POST.get('penalty')
        leaveDeduction = request.POST.get('leaveDeduction')

        out_Time = request.POST.get('out_Time') == 'on'
        earlyLeavingAllowded = request.POST.get('earlyLeavingAllowded')
        penaltyInterval1 = request.POST.get('penaltyInterval1')
        penalty1 = request.POST.get('penalty1')
        leaveDeduction1 = request.POST.get('leaveDeduction1')

        work_duration = request.POST.get('work_duration') == 'on'
        ShortfallInWDAllowed = request.POST.get('ShortfallInWDAllowed')
        penaltyInterval2 = request.POST.get('penaltyInterval2')
        penalty2 = request.POST.get('penalty2')
        leaveDeduction2 = request.POST.get('leaveDeduction2')

        AttendanceRule.objects.create(user_id=admin_id, rulename=rulename, description=description, inTime=inTime,
                                      outTime=outTime,
                                      enable_AD=enable_AD, autoDeductionDate=autoDeductionDate, enable_AT=enable_AT,
                                      inGracePeriod=inGracePeriod, outGracePeriod=outGracePeriod, fullhours=fullhours,
                                      fullminutes=fullminutes, halfhours=halfhours, halfminutes=halfminutes,
                                      maximum_TBD=maximum_TBD,
                                      maximum_NOB=maximum_NOB, auto_CO=auto_CO, enable_OT=enable_OT,
                                      enable_IR=enable_IR,
                                      enable_CO=enable_CO, enable_PR=enable_PR, in_Time=in_Time,
                                      lateComingAllowded=lateComingAllowded,
                                      penaltyInterval=penaltyInterval, penalty=penalty, leaveDeduction=leaveDeduction,
                                      out_Time=out_Time,
                                      earlyLeavingAllowded=earlyLeavingAllowded, penaltyInterval1=penaltyInterval1,
                                      penalty1=penalty1,
                                      leaveDeduction1=leaveDeduction1, work_duration=work_duration,
                                      ShortfallInWDAllowed=ShortfallInWDAllowed,
                                      penaltyInterval2=penaltyInterval2, penalty2=penalty2,
                                      leaveDeduction2=leaveDeduction2)
        return redirect('attendanceRule')


def edit_attendanceRule(request):
    if request.method == 'POST':
        attendancerule_id = request.POST.get('attendanceruleid')
        attendance_rule = AttendanceRule.objects.filter(id=attendancerule_id)
        rulename = request.POST.get('rulename')
        description = request.POST.get('description')
        inTime = request.POST.get('inTime')
        outTime = request.POST.get('outTime')
        autoDeductionDate = request.POST.get('autoDeductionDate')
        enable_AD = request.POST.get('enable_AD') == 'on'
        enable_AT = request.POST.get('enable_AT') == 'on'

        inTimeGracePeriodHours = request.POST.get("inGracePeriod-hours")
        inTimeGracePeriodMinutes = request.POST.get("inGracePeriod-minutes")
        outTimeGracePeriodHours = request.POST.get("outGracePeriod-hours")
        outTimeGracePeriodMinutes = request.POST.get("outGracePeriod-minutes")
        inGracePeriod = datetime.strptime(
            f"{inTimeGracePeriodHours}:{inTimeGracePeriodMinutes}", "%H:%M")
        outGracePeriod = datetime.strptime(
            f"{outTimeGracePeriodHours}:{outTimeGracePeriodMinutes}", "%H:%M")

        fullhours = request.POST.get('fullday-hours')
        fullminutes = request.POST.get('fullday-minutes')
        halfhours = request.POST.get('halfday-hours')
        halfminutes = request.POST.get('halfday-minutes')
        maximum_TBD = request.POST.get('maximum_TBD') == 'on'
        maximum_NOB = request.POST.get('maximum_NOB') == 'on'
        auto_CO = request.POST.get('auto_CO') == 'on'
        enable_OT = request.POST.get('enable_OT') == 'on'
        enable_IR = request.POST.get('enable_IR') == 'on'
        enable_CO = request.POST.get('enable_CO') == 'on'
        enable_PR = request.POST.get('enable_PR') == 'on'

        in_Time = request.POST.get('in_Time') == 'on'
        lateComingAllowded = request.POST.get('lateComingAllowded')
        penaltyInterval = request.POST.get('penaltyInterval')
        penalty = request.POST.get('penalty')
        leaveDeduction = request.POST.get('leaveDeduction')

        out_Time = request.POST.get('out_Time') == 'on'
        earlyLeavingAllowded = request.POST.get('earlyLeavingAllowded')
        penaltyInterval1 = request.POST.get('penaltyInterval1')
        penalty1 = request.POST.get('penalty1')
        leaveDeduction1 = request.POST.get('leaveDeduction1')

        work_duration = request.POST.get('work_duration') == 'on'
        ShortfallInWDAllowed = request.POST.get('ShortfallInWDAllowed')
        penaltyInterval2 = request.POST.get('penaltyInterval2')
        penalty2 = request.POST.get('penalty2')
        leaveDeduction2 = request.POST.get('leaveDeduction2')

        attendance_rule.update(rulename=rulename, description=description, inTime=inTime, outTime=outTime,
                               enable_AD=enable_AD, autoDeductionDate=autoDeductionDate, enable_AT=enable_AT,
                               inGracePeriod=inGracePeriod, outGracePeriod=outGracePeriod, fullhours=fullhours,
                               fullminutes=fullminutes, halfhours=halfhours, halfminutes=halfminutes,
                               maximum_TBD=maximum_TBD,
                               maximum_NOB=maximum_NOB, auto_CO=auto_CO, enable_OT=enable_OT, enable_IR=enable_IR,
                               enable_CO=enable_CO, enable_PR=enable_PR, in_Time=in_Time,
                               lateComingAllowded=lateComingAllowded,
                               penaltyInterval=penaltyInterval, penalty=penalty, leaveDeduction=leaveDeduction,
                               out_Time=out_Time,
                               earlyLeavingAllowded=earlyLeavingAllowded, penaltyInterval1=penaltyInterval1,
                               penalty1=penalty1,
                               leaveDeduction1=leaveDeduction1, work_duration=work_duration,
                               ShortfallInWDAllowed=ShortfallInWDAllowed,
                               penaltyInterval2=penaltyInterval2, penalty2=penalty2, leaveDeduction2=leaveDeduction2)
        return redirect('attendanceRule')


def delete_attendancerule(request, id):
    k = AttendanceRule.objects.get(id=id)
    k.delete()
    return redirect('attendanceRule')


def view_assignattendancerule(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    dn = Designation.objects.all()
    dp = Department.objects.all()
    sd = Subdepartment.objects.all()
    jb = Job.objects.all()
    wr = Worklocation.objects.all()
    user_id = request.user.id
    datas = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id) & Q(status= 'Active') ).order_by('status', 'empid', 'username')
    atten_rule = AttendanceRule.objects.filter(user_id=user_id)
    assg_rule = AssignAttendanceRule.objects.all()
    query = request.GET.get('search')

    print('DATAS: ', datas)
    if query:
        datas_list1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list3 = User.objects.filter(Q(department__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list4 = User.objects.filter(Q(designation__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list5 = User.objects.filter(Q(email__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list6 = User.objects.filter(Q(phone__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list7 = User.objects.filter(Q(status__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list8 = User.objects.filter(Q(subdepartment__subname__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list9 = User.objects.filter(Q(jobtitle__name__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list10 = User.objects.filter(Q(wrklcn__location__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id)))
        datas_list11 = User.objects.filter(Q(emptype__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')
        datas_list12 = User.objects.filter(Q(probperiod__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('status', 'empid', 'username')

        if datas_list1 or datas_list2 or datas_list3 or datas_list4 or datas_list5 or datas_list6 or datas_list7 or datas_list8 or datas_list9 or datas_list10 or datas_list11 or datas_list12:

            datas = datas_list1 | datas_list2 | datas_list3 | datas_list4 | datas_list5 | datas_list6 | datas_list7 | datas_list8 | datas_list9 | datas_list10 | datas_list11 | datas_list12

        else:
            datas = []

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/assign_attendancerule.html",
                  {'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'query': query, 'k': k, 'assg_rule': assg_rule,
                   'atten_rule': atten_rule, 'datas': datas, **x})


def assignattendancerule(request):
    if request.method == 'POST':
        selected_rules = request.POST.get('rule')
        selected_rulesid = AttendanceRule.objects.get(id=selected_rules)
        effective_date = request.POST.get('effdate')
        selected_employees = request.POST.getlist('selected_employees')

        for employee_id in selected_employees:
            try:
                assign_workweek = AssignAttendanceRule.objects.get(
                    user_id_id=employee_id)
                assign_workweek.effective_date = effective_date
                assign_workweek.rules_applied = selected_rulesid
                assign_workweek.save()
            except AssignAttendanceRule.DoesNotExist:
                AssignAttendanceRule.objects.create(
                    user_id_id=employee_id, effective_date=effective_date, rules_applied=selected_rulesid)

        return redirect('view_assignattendancerule')

    atten_rule = AttendanceRule.objects.filter(user_id=request.user.id)
    datas = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id))

    return render(request, "index/assign_attendancerule.html", {'atten_rule': atten_rule, 'datas': datas})


def delete_assignattendancerule(request, assign_rule_id):
    assign_rule = AssignAttendanceRule.objects.get(id=assign_rule_id)
    assign_rule.delete()
    return redirect('view_assignattendancerule')


def EmpAttendanceRule(request):
    user = request.user
    admin_id = User.objects.get(id=user.id).admin_id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    rule = AssignAttendanceRule.objects.filter(user_id=user.id)
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,

    }
    return render(request, "Employee/EmpAttendanceRule.html", {'rule': rule, **x})


########################################################################################################################
''' API VIEW '''


@api_view()
def empdaily_log_api(request):
    user = request.user
    today = datetime.now()
    date_ = request.GET.get('date', today.strftime('%Y-%m-%d'))
    attendanceRuleObj = AttendanceRule.objects.all()
    assignattendancerule = AssignAttendanceRule.objects.filter(user_id=user)

    # print('assignattendancerule: ', assignattendancerule)
    in_Time = '00:00'
    out_Time = '00:00'
    base_date = datetime(1970, 1, 1).date()

    if not assignattendancerule:
        in_datetime = datetime.strptime(
            f"{base_date} {in_Time}", "%Y-%m-%d %H:%M")
        out_datetime = datetime.strptime(
            f"{base_date} {out_Time}", "%Y-%m-%d %H:%M")
        in_Time = in_datetime.time()
        out_Time = out_datetime.time()
    else:
        for attendanceRule in assignattendancerule:
            in_Time = attendanceRule.rules_applied.inTime
            out_Time = attendanceRule.rules_applied.outTime

    in_datetime = datetime.combine(base_date, in_Time)
    out_datetime = datetime.combine(base_date, out_Time)

    working_time = out_datetime - in_datetime

    punch_data = Punch.objects.filter(user=user)
    # print('punch_data: ', punch_data[::-1])

    serialized_data = []
    for data in punch_data:
        shift_one_clocked_in, shift_one_clocked_out_one, shift_two_clocked_in, shift_two_clocked_out = None, None, None, None

        work_duration_str = "00:00:00"
        break_duration_str = "00:00:00"
        overtime_duration_str = '00:00:00'

        work_duration = timedelta(hours=0)

        if data.is_shift_one:
            shift_one_clocked_in = data.clock_in_time
            shift_one_clocked_out_one = data.clock_out_time

        if data.is_shift_two:
            shift_two_clocked_in = data.clock_in_time
            shift_two_clocked_out = data.clock_out_time

        if shift_one_clocked_in and shift_one_clocked_out_one:
            work_duration = shift_one_clocked_out_one - shift_one_clocked_in
            if work_duration.total_seconds() < working_time.total_seconds():
                data.status = 'AN'

            work_duration_str = str(work_duration)
            data.work_duration = work_duration_str

        if shift_one_clocked_in and shift_two_clocked_out:
            work_duration = shift_two_clocked_out - shift_one_clocked_in
            work_duration_str = str(work_duration)
            data.work_duration = work_duration_str

        if shift_one_clocked_out_one and shift_two_clocked_in:
            break_duration = shift_two_clocked_in - shift_one_clocked_out_one
            break_duration_str = str(break_duration)
            data.break_duration = break_duration_str

        if data.clock_out_time is not None:
            clock_out_time_ = datetime.combine(data.date, out_Time)
            clocked_out_time = data.clock_out_time

            overtime_duration = clocked_out_time - clock_out_time_

            if work_duration > timedelta(hours=8):
                overtime_duration_str = str(overtime_duration).split(".")[0]
                overtime_duration_str = str(overtime_duration_str)
                data.overtime = overtime_duration_str

        data.save()

        serialized_data.append({
            'work_duration': work_duration_str,
            'break_duration': break_duration_str,
        })

    serializer = LogSerializer(punch_data, many=True)

    return Response(serializer.data)


@api_view()
def admin_employee_detials(request):
    user_admin = request.user.admin_id
    user_id = request.user.id
    users_obj = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id))

    serializer = UserSerializer(users_obj, many=True)

    if serializer.data:
        return Response(serializer.data)

    else:
        return Response('No data found')


@api_view()
def employee_details(request):
    users = User.objects.filter(admin_id=request.user.admin_id)

    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)



from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from django.core.cache import cache

# create one geolocator instance (module‑level or view‑level)
geolocator = Nominatim(user_agent="hrms-attendance")

def latlon_to_address(lat, lon):
    """Return readable address for (lat, lon), cached 24 h."""
    if not lat or not lon:
        return "Unknown Location"

    cache_key = f"addr_{lat}_{lon}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    try:
        location = geolocator.reverse((lat, lon), timeout=5)
        address = location.address if location else "Unknown Location"
    except GeocoderTimedOut:
        address = "Geocoder Timeout"
    except Exception:
        address = "Location Error"

    cache.set(cache_key, address, 86400)          # cache 1 day
    return address


from app1.templatetags.location_tags import get_clocked_in_location
import json

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def attendance_logs(request):
    userid = request.user.id
    c = companyprofile.objects.filter(admin_id=userid)
    k = Myprofile.objects.filter(myuser__id=userid)

    user_data = User.objects.filter(
        (Q(id=userid) | Q(admin_id=userid)) & Q(status='Active')).exclude(
        Q(resignationform__status='Approved') &
        Q(resignationform__actual_last_working_day__lt=date.today())
    )

    if request.method == 'POST':
        punch_ids = request.POST.get('edittime')
        punch_data = Punch.objects.get(id=punch_ids)

        WFH_WFO = request.POST.get('WFH_WFO')

        in_time_str = request.POST.get('inTime')
        out_time_str = request.POST.get('outTime')
        in_time = datetime.strptime(in_time_str, '%H:%M').time()
        out_time = datetime.strptime(out_time_str, '%H:%M').time()
        punchedtime_delta = timedelta(hours=out_time.hour, minutes=out_time.minute) - \
            timedelta(hours=in_time.hour, minutes=in_time.minute)
        sec_punchedtime_delta = timedelta()
        if punch_data.is_second_clocked_in:
            inTime_second_str = request.POST.get('inTime_second')
            outTime_second_str = request.POST.get('outTime_second')
            sec_in_time = datetime.strptime(inTime_second_str, '%H:%M').time()
            sec_out_time = datetime.strptime(
                outTime_second_str, '%H:%M').time()
            sec_punchedtime_delta = timedelta(hours=sec_out_time.hour, minutes=sec_out_time.minute) - timedelta(
                hours=sec_in_time.hour, minutes=sec_in_time.minute)

        total_punchedtime_delta = punchedtime_delta + sec_punchedtime_delta
        total_punchedtime = total_punchedtime_delta

        attendancerule = AssignAttendanceRule.objects.get(
            user_id=punch_data.user_id)

        in_grace_period = getattr(
            attendancerule.rules_applied, "inGracePeriod", time()) if attendancerule else time()
        out_grace_period = getattr(
            attendancerule.rules_applied, "outGracePeriod", time()) if attendancerule else time()
        in_Time = (datetime.combine(datetime.today(), getattr(attendancerule.rules_applied, "inTime", time()))
                   + timedelta(hours=in_grace_period.hour,
                               minutes=in_grace_period.minute, seconds=in_grace_period.second)
                   ) if attendancerule.rules_applied else time()
        out_Time = (datetime.combine(datetime.today(), getattr(attendancerule.rules_applied, "outTime", time()))
                    + timedelta(hours=out_grace_period.hour,
                                minutes=out_grace_period.minute, seconds=out_grace_period.second)
                    ) if attendancerule.rules_applied else time()

        rule_outtime = datetime.combine(datetime.today(), getattr(
            attendancerule.rules_applied, "outTime", time()))

        in_grace_period_delta = timedelta(
            hours=in_grace_period.hour, minutes=in_grace_period.minute)
        out_grace_period_delta = timedelta(
            hours=out_grace_period.hour, minutes=out_grace_period.minute)
        total_grace_period_delta = in_grace_period_delta + out_grace_period_delta
        total_grace_period_time = datetime.min + total_grace_period_delta

        half_hr = attendancerule.rules_applied.halfhours
        half_min = attendancerule.rules_applied.halfminutes
        half_time = time(hour=half_hr, minute=half_min)
        combined_time = datetime.combine(
            datetime.min, half_time) + total_grace_period_delta

        full_hr = attendancerule.rules_applied.fullhours
        full_min = attendancerule.rules_applied.fullminutes
        full_time = time(hour=full_hr, minute=full_min)
        combinedfull_time = datetime.combine(datetime.min, full_time)

        outtime = "22:30"

        leave_query = Leave.objects.filter(
            applicant_email=punch_data.user.id, status="Approved", strtDate__lte=punch_data.date, endDate__gte=punch_data.date)
        if leave_query.exists():
            leave = leave_query.first()
            if (leave.Selecthalf1 == "first half" and leave.Selecthalf2 == "second half"):
                punch_data.status = "L"
                existing_penalty_log = PenaltyLogs.objects.filter(
                    punch_data=punch_data).first()
                if existing_penalty_log:
                    existing_penalty_log.delete()
            elif (leave.Selecthalf1 == "first half" and leave.Selecthalf2 == "first half") or (leave.Selecthalf1 == "second half" and leave.Selecthalf2 == "second half"):
                punch_data.status = "HL"
                existing_penalty_log = PenaltyLogs.objects.filter(
                    punch_data=punch_data).first()
                if existing_penalty_log:
                    existing_penalty_log.delete()
        elif (punch_data.is_second_clocked_in and in_time_str <= in_Time.strftime('%H:%M') and sec_out_time.strftime('%H:%M') >= rule_outtime.strftime('%H:%M') and sec_out_time.strftime('%H:%M') <= outtime):
            punch_data.status = "P"
            existing_penalty_log = PenaltyLogs.objects.filter(
                punch_data=punch_data).first()
            if existing_penalty_log:
                existing_penalty_log.delete()
        elif (in_time_str <= in_Time.strftime('%H:%M') and out_time_str >= rule_outtime.strftime('%H:%M') and out_time_str <= outtime):
            punch_data.status = "P"
            existing_penalty_log = PenaltyLogs.objects.filter(
                punch_data=punch_data).first()
            if existing_penalty_log:
                existing_penalty_log.delete()
        else:
            punch_data.status = "AN"

        total_work_duration = timedelta()
        break_duration = timedelta()
        punch_data.first_clock_in_time = in_time
        punch_data.first_clock_out_time = out_time
        punch_data.is_first_clocked_in = True
        punch_data.is_first_clocked_out = True
        punch_data.punch_in_count = 1
        punch_data.punch_out_count = 1
        punch_data.is_shift_one = True
        punch_data.last_punch_type = 1
        punch_data.WfhOrWfo = WFH_WFO
        punch_data.save()
        if in_time <= out_time:
            if punch_data.is_second_clocked_in:
                punch_data.second_clock_in_time = sec_in_time
                punch_data.second_clock_out_time = sec_out_time
                punch_data.is_second_clocked_in = True
                punch_data.is_second_clocked_out = True
                punch_data.punch_in_count = 2
                punch_data.punch_out_count = 2
                punch_data.is_shift_two = True
                punch_data.last_punch_type = 2

                first_clock_in = datetime.combine(datetime.today(), in_time)
                first_clock_out = datetime.combine(datetime.today(), out_time)
                second_clock_in = datetime.combine(
                    datetime.today(), sec_in_time)
                second_clock_out = datetime.combine(
                    datetime.today(), sec_out_time)

                first_duration = first_clock_out - first_clock_in
                second_duration = second_clock_out - second_clock_in
                total_work_duration += first_duration + second_duration

                break_start_time = datetime.combine(datetime.today(), out_time)
                break_end_time = datetime.combine(
                    datetime.today(), sec_out_time)
                break_duration = break_end_time - break_start_time
                print("break_duration :", break_duration)

                punch_data.work_duration = str(total_work_duration)
                punch_data.break_duration = str(break_duration)
                punch_data.save()

            else:
                print("else in_time , out_time : ", in_time, out_time)
                first_clock_in = datetime.combine(datetime.today(), in_time)
                first_clock_out = datetime.combine(datetime.today(), out_time)
                print("first_clock_in ; first_clock_out : ",
                      first_clock_in, first_clock_out)
                first_duration = first_clock_out - first_clock_in
                print("first_duration : ", first_duration)
                total_work_duration += first_duration
                print("total_work_duration : ", str(total_work_duration))
                punch_data.work_duration = str(total_work_duration)
                punch_data.save()

        print("break_duration : ", break_duration, total_work_duration)

        assignattendancerule = AssignAttendanceRule.objects.filter(
            user_id=punch_data.user_id)
        print("assignattendancerule : ", assignattendancerule)
        total_time = timedelta()
        overtime = timedelta()
        for assignrule in assignattendancerule:
            inTime = assignrule.rules_applied.inTime
            outTime = assignrule.rules_applied.outTime
            print("inTime ; outTime : ", inTime, outTime)
            in_Time = datetime.combine(datetime.today(), inTime)
            out_Time = datetime.combine(datetime.today(), outTime)
            print("intime_datetime ; outTime_datetime : ", in_Time, out_Time)
            total_time += out_Time - in_Time

            print("totaltime :", total_time)

            if total_work_duration > total_time:
                overtime += total_work_duration - total_time
                print("overtime :", overtime)
                punch_data.overtime = str(overtime)

                punch_data.save()

    # /////////////
    selected_date_str = request.GET.get('dateselect', None)
    current_date = timezone.now().date()

    if selected_date_str is None:
        selected_date = timezone.now().date()
    else:
        selected_date = datetime.strptime(selected_date_str, '%d %B %Y').date()

    selectedday = selected_date.day
    selectedmonth = selected_date.strftime("%B")
    selectedyear = selected_date.year
    print("selected_date :", selected_date,
          selectedday, selectedmonth, selectedyear)
          

    punches_in_today = Punch.objects.filter(
        Q(date__date=selected_date) &
        (Q(user__id=userid) | Q(user__admin_id=userid))
    ).filter(
        Q(user__resignationform__isnull=True) | 
        Q(user__resignationform__actual_last_working_day__gte=selected_date)
    )
    for punch in punches_in_today:
        total_work_duration = timedelta()
        break_duration = timedelta()

        try:
            if punch.WfhOrWfo == "WFH":
                geo = EmployeeGeoFence.objects.get(user=punch.user)
                if geo.home_lat and geo.home_lon:
                    # 🔄 convert lat/lon to readable address
                    punch.display_location = latlon_to_address(geo.home_lat, geo.home_lon)
                else:
                    punch.display_location = "Home Location Not Set"

            elif punch.WfhOrWfo == "WFO":
                profile = EmployeeProfile.objects.get(user=punch.user)
                if profile.branch_location:
                    punch.display_location = profile.branch_location.name
                else:
                    punch.display_location = "Office Location Not Assigned"

            else:
                punch.display_location = "-- -- --"

        except Exception:
            punch.display_location = "Location Error"

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)

            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

            break_start_time = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            break_end_time = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            break_duration = break_end_time - break_start_time
            print("break_duration :", break_duration)

        elif punch.is_first_clocked_in and punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration

        print("total_work_duration :", total_work_duration)
        total_hours = int(total_work_duration.total_seconds() // 3600)
        total_minutes = int((total_work_duration.total_seconds() % 3600) // 60)
        total_work_duration_str = f"{total_hours} Hours {total_minutes} Mins"
        if total_work_duration_str == "0 Hours 0 Mins":
            punch.work_duration = "-- -- --"
        else:
            punch.work_duration = total_work_duration_str
        print("total_work_duration_str :", punch.work_duration)

        break_hours = int(break_duration.total_seconds() // 3600)
        break_minutes = int((break_duration.total_seconds() % 3600) // 60)
        total_break_duration_str = f"{break_hours} Hours {break_minutes} Mins"
        if total_break_duration_str == "0 Hours 0 Mins":
            punch.break_duration = "-- -- --"
        else:
            punch.break_duration = total_break_duration_str

        assignattendancerule = AssignAttendanceRule.objects.filter(
            user_id=punch.user_id)
        print("assignattendancerule : ", assignattendancerule)
        total_time = timedelta()
        overtime = timedelta()
        for assignrule in assignattendancerule:
            inTime = assignrule.rules_applied.inTime
            outTime = assignrule.rules_applied.outTime
            print("inTime ; outTime : ", inTime, outTime)
            in_Time = datetime.combine(datetime.today(), inTime)
            out_Time = datetime.combine(datetime.today(), outTime)
            print("intime_datetime ; outTime_datetime : ", in_Time, out_Time)
            total_time += out_Time - in_Time

        print("totaltime :", total_time)

        if total_work_duration > total_time:
            overtime += total_work_duration - total_time
        print("overtime :", overtime)
        overtime_hours = int(overtime.total_seconds() // 3600)
        overtime_minutes = int((overtime.total_seconds() % 3600) // 60)
        total_overtime_str = f"{overtime_hours} Hours {overtime_minutes} Mins"
        if total_overtime_str == "0 Hours 0 Mins":
            punch.overtime = "-- -- --"
        else:
            punch.overtime = total_overtime_str

        print("punch overtime :", punch.overtime)

    punches_in_today_users = set(
        punch_data.user_id for punch_data in punches_in_today)
    print("punches_in_today_users :", punches_in_today_users)
    employees = User.objects.filter(
        (Q(id=userid) | Q(admin_id=userid)) & Q(status='Active')
    )
    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'selected_date': selected_date,
        'current_date': current_date,
        'punches_in_today': punches_in_today,
        'selectedday': selectedday,
        'selectedmonth': selectedmonth,
        'selectedyear': selectedyear,
        'user_data': user_data,
        'punches_in_today_users': punches_in_today_users,
        'employees': employees, 
        'selected_username': request.GET.get('username', '')  
    }

    return render(request, 'index/attendance_logs.html', context)


    

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def new_attendance_logs(request):
    if request.method == 'POST':
        userid = request.POST.get('userid')
        user = User.objects.get(id=userid)
        selected_day = request.POST.get('day')
        selected_month = request.POST.get('month')
        datetime_object = datetime.strptime(selected_month, "%B")
        month_numeric = datetime_object.month
        selected_year = request.POST.get('year')
        selected_date = datetime(
            int(selected_year), month_numeric, int(selected_day))
        
        WFH_WFO = request.POST.get('WFH_WFO')

        in_time_str = request.POST.get('inTime')
        out_time_str = request.POST.get('outTime')
        print("in_time_str , out_time_str :", in_time_str, out_time_str)
        in_time = datetime.strptime(in_time_str, '%H:%M').time()
        out_time = datetime.strptime(out_time_str, '%H:%M').time()
        punchedtime_delta = timedelta(hours=out_time.hour, minutes=out_time.minute) - \
            timedelta(hours=in_time.hour, minutes=in_time.minute)
        print("in_time , out_time : ", in_time, out_time, punchedtime_delta)

        total_punched_time = punchedtime_delta

        attendancerule = AssignAttendanceRule.objects.get(user_id=user)
        in_grace_period = getattr(
            attendancerule.rules_applied, "inGracePeriod", time()) if attendancerule else time()
        out_grace_period = getattr(
            attendancerule.rules_applied, "outGracePeriod", time()) if attendancerule else time()

        rulein_Time = (datetime.combine(datetime.today(), getattr(attendancerule.rules_applied, "inTime", time()))
                       + timedelta(hours=in_grace_period.hour,
                                   minutes=in_grace_period.minute, seconds=in_grace_period.second)
                       ) if attendancerule.rules_applied else time()
        ruleout_Time = (datetime.combine(datetime.today(), getattr(attendancerule.rules_applied, "outTime", time()))
                        + timedelta(hours=out_grace_period.hour,
                                    minutes=out_grace_period.minute, seconds=out_grace_period.second)
                        ) if attendancerule.rules_applied else time()
        rule_outtime = datetime.combine(datetime.today(), getattr(
            attendancerule.rules_applied, "outTime", time()))

        in_grace_period_delta = timedelta(
            hours=in_grace_period.hour, minutes=in_grace_period.minute)
        out_grace_period_delta = timedelta(
            hours=out_grace_period.hour, minutes=out_grace_period.minute)
        total_grace_period_delta = in_grace_period_delta + out_grace_period_delta

        half_hr = attendancerule.rules_applied.halfhours
        half_min = attendancerule.rules_applied.halfminutes
        half_time = time(hour=half_hr, minute=half_min)
        combined_time = datetime.combine(
            datetime.min, half_time) + total_grace_period_delta

        full_hr = attendancerule.rules_applied.fullhours
        full_min = attendancerule.rules_applied.fullminutes
        full_time = time(hour=full_hr, minute=full_min)
        combinedfull_time = datetime.combine(datetime.min, full_time)

        outtime = "22:30"

        print("selected_date :", selected_date)
        leave_query = Leave.objects.filter(
            applicant_email=user, status="Approved", strtDate__lte=selected_date, endDate__gte=selected_date)
        print("leave_query :", leave_query)
        if leave_query.exists():
            leave = leave_query.first()
            if (leave.Selecthalf1 == "first half" and leave.Selecthalf2 == "second half"):
                status = "L"
            elif (leave.Selecthalf1 == "first half" and leave.Selecthalf2 == "first half") or (leave.Selecthalf1 == "second half" and leave.Selecthalf2 == "second half"):
                status = "HL"

        elif (in_time_str <= rulein_Time.strftime('%H:%M') and out_time_str >= rule_outtime.strftime('%H:%M') and out_time_str <= outtime):
            status = "P"
        else:
            status = "AN"

        Punch.objects.create(user=user, status=status, first_clock_in_time=in_time, first_clock_out_time=out_time, date=selected_date, WfhOrWfo = WFH_WFO,
                             is_first_clocked_in=True, is_first_clocked_out=True, last_punch_type=1, punch_in_count=1, punch_out_count=1, is_shift_one=True)

    return redirect(reverse('attendance_logs') + '?dateselect=' + selected_day + '+' + selected_month + '+' + selected_year)


# EXPORT EMPLOYEE PUNCH DATA IN ATTENDANCE LOGs PAGE

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def export_attendance_logs(request):
    admin_id = User.objects.filter(id=request.user.id)
    if request.method == 'POST':
        start_date_str = request.POST.get('startdate')
        end_date_str = request.POST.get('enddate')
        start_date = datetime.strptime(start_date_str, '%d %B %Y')
        end_date = datetime.strptime(
            end_date_str, '%d %B %Y') + timedelta(days=1)
        userid = request.user.id

        punchdata = Punch.objects.filter(Q(date__range=[start_date, end_date], user__id=userid) | Q(
            date__range=[start_date, end_date], user__admin_id=userid))

        data_list = []

        for punch in punchdata:
            total_work_duration = timedelta()
            break_duration = timedelta()

            if punch.is_first_clocked_in and punch.is_second_clocked_out and punch.first_clock_in_time and punch.second_clock_out_time:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                second_clock_in = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)
                second_clock_out = datetime.combine(
                    datetime.today(), punch.second_clock_out_time)

                first_duration = first_clock_out - first_clock_in
                second_duration = second_clock_out - second_clock_in
                total_work_duration += first_duration + second_duration

                break_start_time = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                break_end_time = datetime.combine(
                    datetime.today(), punch.second_clock_in_time)
                break_duration = break_end_time - break_start_time

            elif punch.is_first_clocked_in and punch.is_first_clocked_out and punch.first_clock_in_time and punch.first_clock_out_time:
                first_clock_in = datetime.combine(
                    datetime.today(), punch.first_clock_in_time)
                first_clock_out = datetime.combine(
                    datetime.today(), punch.first_clock_out_time)
                first_duration = first_clock_out - first_clock_in
                total_work_duration += first_duration

            print("total_work_duration :", total_work_duration)
            total_hours = int(total_work_duration.total_seconds() // 3600)
            total_minutes = int(
                (total_work_duration.total_seconds() % 3600) // 60)
            total_work_duration_str = f"{total_hours} Hours {total_minutes} Mins"
            if total_work_duration_str == "0 Hours 0 Mins":
                punch.work_duration = "-- -- --"
            else:
                punch.work_duration = total_work_duration_str
            print("total_work_duration_str :", punch.work_duration)

            break_hours = int(break_duration.total_seconds() // 3600)
            break_minutes = int((break_duration.total_seconds() % 3600) // 60)
            total_break_duration_str = f"{break_hours} Hours {break_minutes} Mins"
            if total_break_duration_str == "0 Hours 0 Mins":
                punch.break_duration = "-- -- --"
            else:
                punch.break_duration = total_break_duration_str

            assignattendancerule = AssignAttendanceRule.objects.filter(
                user_id=punch.user_id)
            print("assignattendancerule : ", assignattendancerule)
            total_time = timedelta()
            overtime = timedelta()
            for assignrule in assignattendancerule:
                inTime = assignrule.rules_applied.inTime
                outTime = assignrule.rules_applied.outTime
                print("inTime ; outTime : ", inTime, outTime)
                in_Time = datetime.combine(datetime.today(), inTime)
                out_Time = datetime.combine(datetime.today(), outTime)
                print("intime_datetime ; outTime_datetime : ", in_Time, out_Time)
                total_time += out_Time - in_Time

            print("totaltime :", total_time)
            if total_work_duration > total_time:
                overtime += total_work_duration - total_time
            print("overtime :", overtime)
            overtime_hours = int(overtime.total_seconds() // 3600)
            overtime_minutes = int((overtime.total_seconds() % 3600) // 60)
            total_overtime_str = f"{overtime_hours} Hours {overtime_minutes} Mins"
            if total_overtime_str == "0 Hours 0 Mins":
                punch.overtime = "-- -- --"
            else:
                punch.overtime = total_overtime_str

            print("punch overtime :", punch.overtime)

            anomaly = ''
            if punch.status == "AN":
                anomaly = 1
            outtime = ""
            if punch.is_first_clocked_out and not punch.is_second_clocked_in:
                outtime = punch.first_clock_out_time.strftime(
                    '%I:%M %p') if punch.first_clock_out_time else '-- -- --'
            elif punch.is_second_clocked_out:
                outtime = punch.second_clock_out_time.strftime(
                    '%I:%M %p') if punch.second_clock_out_time else '-- -- --'

            data_list.append({
                'Employee ID': punch.user.empid,
                'Name': punch.user.username,
                'Date': punch.date.strftime('%d %B %Y'),
                'Status': punch.status,
                'In Time': punch.first_clock_in_time.strftime('%I:%M %p') if punch.first_clock_in_time else '-- -- --',
                'Out Time': outtime,
                'Work Duration': punch.work_duration,
                'Overtime Duration': punch.overtime,
                'Break Duration': punch.break_duration,
                'Break Count': punch.break_count,
                'Outstanding Anomalies': anomaly,
            })

        df = pd.DataFrame(data_list)
        excel_file_name = 'attendance_log.xlsx'
        path = os.path.join(BASE_DIR, 'media/csv/attendance_log.xlsx')

        df.to_excel(path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        fixed_width = 20
        for column in sheet.columns:
            sheet.column_dimensions[get_column_letter(
                column[0].column)].width = fixed_width
        wb.save(path)

        today = datetime.now()
        filemanager = Filemanager.objects.create(myuser_10=request.user, requesttype="Attendance Log Export",
                                                 frmt="XLSX", scheduleon=today, status="In Queue")
        subject = 'Punch Data Export'
        message = 'Attached is the punch data for the specified date range.'
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [adminid.email for adminid in admin_id]
        email = EmailMessage(subject, message, from_email, to_email)
        email.attach(excel_file_name, open(path, 'rb').read(),
                     'application/vnd.ms-excel')
        email.send()

        filemanager.status = "Success"
        with open(path, 'rb') as file:
            filemanager.saveexcel.save(
                'attendance_log.xlsx', ContentFile(file.read()))
    return redirect("attendance_logs")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def e_exit(request):
    user = request.user.id
    k = Myprofile.objects.filter(myuser__id=user)
    c = companyprofile.objects.filter(admin_id=user)
    resignation = ResignationForm.objects.filter(user=request.user.id)
    context = {
        'k': k[0] if k.exists() else k,
        'c': c[0] if c.exists() else c,
    }
    return render(request, 'index/e_exit.html', {'resignation': resignation, **context})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def add_e_exit(request):
    user = request.user
    print("USER ID : ", user)
    if request.method == 'POST':
        user = request.user
        resignation_date = datetime.strptime(request.POST.get(
            'resignation_date'), '%d %B %Y').strftime('%Y-%m-%d')
        reason = request.POST.get('reason')
        last_workingday = datetime.strptime(request.POST.get(
            'last_workingday'), '%d %B %Y').strftime('%Y-%m-%d')
        notice_period = request.POST.get('notice_period')
        Shortfall = request.POST.get('Shortfall')
        resignation_letter = request.POST.get('resignation_letter')

        ResignationForm.objects.create(user=user, resignation_date=resignation_date, reason=reason,
                                       last_workingday=last_workingday,
                                       notice_period=notice_period, Shortfall=Shortfall,
                                       resignation_letter=resignation_letter, status="Pending")
        return redirect('e_exit')
    return render(request, 'index/e_exit.html')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def update_e_exit(request):
    if request.method == "POST":
        resignation_date = datetime.strptime(request.POST.get(
            'resignation_date'), '%d %B %Y').strftime('%Y-%m-%d')
        reason = request.POST.get('reason')
        last_workingday = datetime.strptime(request.POST.get(
            'last_workingday'), '%d %B %Y').strftime('%Y-%m-%d')
        notice_period = request.POST.get('notice_period')
        Shortfall = request.POST.get('Shortfall')
        resignation_letter = request.POST.get('resignation_letter')
        ResignationForm.objects.update(resignation_date=resignation_date, reason=reason,
                                       last_workingday=last_workingday,
                                       notice_period=notice_period, Shortfall=Shortfall,
                                       resignation_letter=resignation_letter)
        return redirect('e_exit')


def delete_resignation(request, id):
    resignation = get_object_or_404(ResignationForm, id=id)

    # Check if the user making the request is the owner of the resignation
    if resignation.user == request.user:
        resignation.delete()
    return redirect('e_exit')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empe_exit(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=user_id)
    
    # Latest resignation
    resignation = ResignationForm.objects.filter(user=request.user).order_by('-id')

    context = {
        'k': k.first() if k.exists() else None,
        'c': c.first() if c.exists() else None,
        'resignation': resignation,
    }
    return render(request, 'Employee/e_exit.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url='login')
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empadd_e_exit(request):
    user = request.user
    if request.method == 'POST':
        resignation_date = datetime.strptime(request.POST.get('resignation_date'), '%Y-%m-%d').date()
        resignation_date = datetime.strptime(request.POST.get('resignation_date'), '%Y-%m-%d').date()
        reason = request.POST.get('reason')

        actual_last_working_day_str = request.POST.get('actual_last_working_day')
        if actual_last_working_day_str:
            actual_last_working_day = datetime.strptime(actual_last_working_day_str, '%Y-%m-%d').date()
            actual_last_working_day = datetime.strptime(actual_last_working_day_str, '%Y-%m-%d').date()
        else:
            actual_last_working_day = None

        last_workingday = datetime.strptime(request.POST.get('last_workingday'), '%Y-%m-%d').date()
        last_workingday = datetime.strptime(request.POST.get('last_workingday'), '%Y-%m-%d').date()
        notice_period = request.POST.get('notice_period')
        Shortfall = request.POST.get('Shortfall')
        resignation_letter = request.POST.get('resignation_letter')

        resignation = ResignationForm.objects.create(
            user=user,
            resignation_date=resignation_date,
            reason=reason,
            actual_last_working_day=actual_last_working_day,
            last_workingday=last_workingday,
            notice_period=notice_period,
            Shortfall=Shortfall,
            resignation_letter=resignation_letter,
            status="Pending"
        )

        # ✅ Send email to admin
        subject = f"Resignation Submitted by {user.username}"
        to = ['operations@cydeztechnologies.com']
        from_email = settings.EMAIL_HOST_USER
        context = {
            'employee': user,
            'resignation': resignation,
        }

        html_content = render_to_string('index/resignation_notification.html', context)
        msg = EmailMultiAlternatives(subject, "", from_email, to)
        msg.attach_alternative(html_content, "text/html")
        msg.send()

        return redirect('empe_exit')
    return render(request, 'Employee/e_exit.html')


@login_required
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def edit_resignation(request, resignation_id):
    resignation = get_object_or_404(ResignationForm, id=resignation_id, user=request.user)

    if request.method == "POST":
        try:
            resignation_date = datetime.strptime(request.POST.get('resignation_date'), '%Y-%m-%d')
            last_workingday = datetime.strptime(request.POST.get('last_workingday'), '%Y-%m-%d')
            actual_last_working_day_str = request.POST.get('actual_last_working_day')

            actual_last_working_day = (
                datetime.strptime(actual_last_working_day_str, '%Y-%m-%d')
                if actual_last_working_day_str else None
            )

            resignation.resignation_date = resignation_date
            resignation.reason = request.POST.get('reason')
            resignation.actual_last_working_day = actual_last_working_day
            resignation.last_workingday = last_workingday
            resignation.notice_period = request.POST.get('notice_period')
            resignation.Shortfall = request.POST.get('Shortfall')
            resignation.resignation_letter = request.POST.get('resignation_letter')

            resignation.allow_edit = False
            resignation.update_message = "Please update your reason or last working day."
            resignation.status = "Pending"
            resignation.save()  
            return redirect('empe_exit')
        except Exception as e:
            print("❌ Edit Error:", e)
            return redirect('empe_exit') 
    return redirect('empe_exit')


def empdelete_resignation(request, id):
    resignation = get_object_or_404(ResignationForm, id=id)
    if resignation.user == request.user:
        resignation.delete()
    return redirect('empe_exit')




@login_required
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def update_resignation_status(request, resignation_id):
    resignation = get_object_or_404(ResignationForm, id=resignation_id)

    if request.method == 'POST':
        resignation.reason = request.POST.get('reason')
        resignation.last_workingday = request.POST.get('last_workingday')
        resignation.save()
        return redirect('resignationlist')  # or the same page

    return redirect('resignationlist') 


def resignationlist(request):
    user = request.user.id
    k = Myprofile.objects.filter(myuser__id=user)
    c = companyprofile.objects.filter(admin_id=user)

    # Get queryset and order by newest first
    resg_qs = ResignationForm.objects.filter(
        Q(user__id=user) | Q(user__admin_id=user)
    ).order_by('-id') 

    # Apply pagination
    paginator = Paginator(resg_qs, 25)  # Show 25 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "k": k.first() if k.exists() else None,
        "c": c.first() if c.exists() else None,
        "resg_list": page_obj,
        "today": date.today(),
    }
    return render(request, 'index/resgnationlist.html', context)


@csrf_exempt 
def toggle_edit_permission(request):
    if request.method == 'POST':
        res_id = request.POST.get('resignation_id')
        try:
            resignation = ResignationForm.objects.get(id=res_id)
            resignation.allow_edit = not resignation.allow_edit
            resignation.save()
            return JsonResponse({'status': 'success', 'allow_edit': resignation.allow_edit})
        except ResignationForm.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Resignation not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


def approve_resignation():
    today = date.today()

    pending_resignations = ResignationForm.objects.filter(
        status='Pending',
        actual_last_working_day=today
    )

    for resg in pending_resignations:
        resg.status = 'Approved'
        resg.save()
        resg.save()

        # Send Email Notification
        # Send Email Notification
        email_id = resg.user.email
        if email_id:
            subject = 'Resignation Application Auto-Approved'
            email_from = settings.EMAIL_HOST_USER
            to = [email_id]
        if email_id:
            subject = 'Resignation Application Auto-Approved'
            email_from = settings.EMAIL_HOST_USER
            to = [email_id]

            html_body = render_to_string(
                'index/resgapprovemail.html',
                {'data': resg, 'email': 'HRMS'}
            )

            msg = EmailMultiAlternatives(subject=subject, from_email=email_from, to=to)
            msg.attach_alternative(html_body, "text/html")
            msg.send()
            html_body = render_to_string(
                'index/resgapprovemail.html',
                {'data': resg, 'email': 'HRMS'}
            )

            msg = EmailMultiAlternatives(subject=subject, from_email=email_from, to=to)
            msg.attach_alternative(html_body, "text/html")
            msg.send()


from django.core.files.storage import default_storage
import os
from django.views.decorators.http import require_POST


@csrf_exempt  
@login_required
@require_POST
def reject_resignation(request):
    resg_id = request.POST.get('resg_id')
    reason = request.POST.get('rejection_reason')

    try:
        resignation = ResignationForm.objects.get(id=resg_id)
        resignation.status = "Rejected"
        resignation.rejection_reason = reason
        resignation.save()
        return JsonResponse({'status': 'success'})
    except ResignationForm.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Resignation not found'}, status=404)


@require_POST
@login_required
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def cancel_resignation(request):
    resg_id = request.POST.get("resg_id")

    try:
        resg = ResignationForm.objects.get(id=resg_id)
        resg.status = "Pending"
        resg.rejection_reason = ""
        resg.save()

        # 🔥 Delete generated certificate if exists
        cert = ExperienceCertificate.objects.filter(resignation=resg).first()
        if cert:
            if cert.certificate_file and default_storage.exists(cert.certificate_file.path):
                os.remove(cert.certificate_file.path)
            cert.delete()

        return JsonResponse({"status": "success"})
    except ResignationForm.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Resignation not found"})


@login_required
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def update_resignation_details(request, resignation_id):
    resignation = get_object_or_404(ResignationForm, id=resignation_id)

    if request.method == 'POST':
        resignation.resignation_date = datetime.strptime(request.POST.get('resignation_date'), '%Y-%m-%d')
        resignation.reason = request.POST.get('reason')
        resignation.actual_last_working_day = datetime.strptime(request.POST.get('actual_last_working_day'), '%Y-%m-%d')
        resignation.last_workingday = datetime.strptime(request.POST.get('last_workingday'), '%Y-%m-%d')
        resignation.notice_period = int(request.POST.get('notice_period'))
        resignation.Shortfall = int(request.POST.get('Shortfall'))
        resignation.resignation_letter = request.POST.get('resignation_letter')
        resignation.update_message = "Please update your reason or last working day."
        resignation.status = "Pending"
        resignation.save()  

        messages.success(request, 'Resignation details updated successfully.')
        return redirect('resignationlist')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def notification(request):
    user_id = request.user.id

    # Get logged-in user's profile and company profile
    k = Myprofile.objects.filter(myuser__id=user_id)
    c = companyprofile.objects.filter(admin_id=user_id)

    # Fetch leave notifications
    leave_notifications = LeaveNotification.objects.filter(
        Q(user=user_id) | Q(user__admin_id=user_id)
    ).order_by('-timestamp')

    # Attach profile image to each notification
    for notification in leave_notifications:
        if notification.admin_id == 0:
            user_profile = Myprofile.objects.filter(myuser=notification.user).first()
        else:
            user_profile = Myprofile.objects.filter(myuser__id=notification.admin_id).first()

        if user_profile and user_profile.image and hasattr(user_profile.image, 'url'):
            notification.user_profile_image = user_profile.image.url
        else:
            notification.user_profile_image = "/static/logo/userlogo.png"

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(leave_notifications, 20)

    try:
        leave_notifications = paginator.page(page)
    except PageNotAnInteger:
        leave_notifications = paginator.page(1)
    except EmptyPage:
        leave_notifications = paginator.page(paginator.num_pages)

    context = {
        "k": k[0] if k.exists() else None,
        "c": c[0] if c.exists() else None,
        "leave_notifications": leave_notifications,
    }

    return render(request, 'index/notification.html', context)


def mark_notification_as_read(request, notification_id):
    try:
        notification = LeaveNotification.objects.get(id=notification_id)
        print("D", notification)

        notification.is_read = True
        notification.save()
        return JsonResponse({'status': 'read'})
    except LeaveNotification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notification not found'})


def mark_notification_as_unread(request, notification_id):
    try:
        notification = LeaveNotification.objects.get(id=notification_id)
        notification.is_read = False
        notification.save()
        return JsonResponse({'status': 'unread'})
    except LeaveNotification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notification not found'})


def delete_notifications(request):
    if request.method == 'POST':
        selected_notification_ids = request.POST.getlist('selected_employees')
        action = request.POST.get('action')

        leave_id = LeaveNotification.objects.filter(
            id__in=selected_notification_ids)

        if action == "1":
            leave_id.update(readadmin=True)
        elif action == "2":
            leave_id.update(readadmin=False)
        elif action == "3":
            leave_id.delete()

        return redirect('notification')




def empnotification(request):
    user_id = request.user.id
    admin_id = request.user.admin_id
    k = Myprofile.objects.filter(myuser__id=user_id)
    c = companyprofile.objects.filter(admin_id=admin_id)

    leave_notifications = LeaveNotification.objects.filter(
        user=user_id).order_by('-timestamp')

    for notification in leave_notifications:
        if notification.admin_id == 0:
            user_profile = Myprofile.objects.filter(myuser=notification.user).first()
        else:
            user_profile = Myprofile.objects.filter(myuser__id=notification.admin_id).first()

        if user_profile and user_profile.image and hasattr(user_profile.image, 'url'):
            notification.user_profile_image = user_profile.image.url
        else:
            notification.user_profile_image = "/static/logo/userlogo.png"

    page = request.GET.get('page', 1)
    paginator = Paginator(leave_notifications, 20)

    try:
        leave_notifications = paginator.page(page)
    except PageNotAnInteger:
        leave_notifications = paginator.page(1)
    except EmptyPage:
        leave_notifications = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, 'Employee/notification.html', {'leave_notifications': leave_notifications, **x})


def delete_empnotifications(request):
    if request.method == 'POST':
        selected_notification_ids = request.POST.getlist('selected_employees')
        action = request.POST.get('action')

        leave_id = LeaveNotification.objects.filter(
            id__in=selected_notification_ids)

        if action == "1":
            leave_id.update(readuser=True)
        elif action == "2":
            leave_id.update(readuser=False)
        elif action == "3":
            leave_id.delete()

        return redirect('empnotification')


def send_bday_mail():
    # Get the current date
    today = datetime.date.today()
    today_formatted = today.strftime("%d %B %Y")

    # Get employees whose birthday is today
    employees = User.objects.filter(role='Employee', dob=today_formatted)
    admin = User.objects.filter(role='Admin')

    # Send emails to the admin for each employee
    if employees:
        subject = 'Birthday Reminder'
        message = 'Today is the birthday of the following employees:\n'
        for employee in employees:
            message += f'- {employee.get_username()}\n'
        admin_emails = [admin.email for admin in admin]

        send_mail(subject, message, settings.EMAIL_HOST_USER,
                  admin_emails, fail_silently=False)
        print('Birthday reminder emails sent successfully!')
    else:
        print('No employee birthdays today. No email sent.')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def createsalary(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    salary = SalaryComponent.objects.filter(admin_id=request.user.id)
    salaryRule = SalaryStructureRule.objects.filter(myuser_12=request.user.id)

    data = []
    for rule in salaryRule:
        names = SalaryStructureName.objects.filter(salaryrule=rule)
        amounts = SalaryStructureAmount.objects.filter(salaryname__in=names)
        zipped_data = zip_longest(names, amounts)

        data.append({
            'rule': rule,
            'zipped_data': zipped_data,
        })

    context = {
        'salary': salary,
        'salaryRule': salaryRule,
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, 'index/createsalarystructure.html', {'data': data, **context})


from math import ceil

def add_createsalarystructure(request):
    if request.method == "POST":
        admin_id = request.user.id
        rulename = request.POST.get('salaryname')
        description = request.POST.get('description')
        compname = request.POST.getlist('structurename', [])
        Amount = request.POST.getlist('csalary', [])

        user_instance = User.objects.get(id=admin_id)
        salary_structure_rule = SalaryStructureRule.objects.create(rule_name=rulename, Description=description, myuser_12=user_instance)

        gross_salary_component = SalaryComponent.objects.filter(componentname="Gross Salary").first()
        net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
        ctc_salary_component = SalaryComponent.objects.filter(componentname="CTC").first()

        gross_salary_amount = 0
        total_parent_component_amount = 0
        total_NSparent_component_amount = 0
        total_CTCparent_component_amount = 0
        basic_salary_amount = 0

        if gross_salary_component and str(gross_salary_component.id) in compname:
            gross_salary_index = compname.index(str(gross_salary_component.id))
            gross_salary_amount = float(Amount[gross_salary_index])

        for i, sname in enumerate(compname):
            component = SalaryComponent.objects.get(id=sname)

            print("component.componentname :", component.componentname)
            
            if component.componentname == "Professional Tax" or component.componentname == "Professional tax":
                print("component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 167 if gross_salary_amount < 22000 else 208

            if component.componentname == "Insurance":
                print("component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 0 if gross_salary_amount <= 25000 else 245

                salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
                salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=component.id))
                salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
                salary_amount.save()
                    
                if component.Parentcomponentname == net_salary_component:
                    total_NSparent_component_amount += calculated_amount
                elif component.Parentcomponentname == ctc_salary_component:
                    total_CTCparent_component_amount += calculated_amount

            if component.componentname in ["Other Allowance", "EPF Employer", "EPF Employee", "Net Salary", "CTC", "Professional Tax", "Professional tax","Insurance"]:
                continue

            print("component.componentname :", component.componentname)
            salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
            salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=sname))

            calculated_amount = float(Amount[i])
            if component.percent:
                calculated_amount = ceil(gross_salary_amount * (component.percent / 100.0))
            
            print("component.componentname :", calculated_amount)

            if component.Parentcomponentname == gross_salary_component:
                total_parent_component_amount += calculated_amount
            elif component.Parentcomponentname == net_salary_component:
                total_NSparent_component_amount += calculated_amount
            elif component.Parentcomponentname == ctc_salary_component:
                total_CTCparent_component_amount += calculated_amount
                print("total_CTCparent_component_amount elif: ", total_CTCparent_component_amount)

            if component.componentname == "Basic Salary":
                basic_salary_amount = calculated_amount

            salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
            salary_amount.save()

        process_other_allowance(compname, salary_structure_rule, gross_salary_amount, total_parent_component_amount)

        total_NSparent_component_amount = process_epfempoyee_component(compname, salary_structure_rule, total_NSparent_component_amount, basic_salary_amount)

        total_CTCparent_component_amount = process_epf_component(compname, salary_structure_rule, total_CTCparent_component_amount, basic_salary_amount)

        process_net_salary(compname, salary_structure_rule, total_NSparent_component_amount, gross_salary_amount)

        process_ctc(compname, salary_structure_rule, total_CTCparent_component_amount, gross_salary_amount)

    return redirect('create_salary')

def process_other_allowance(compname, salary_structure_rule, gross_salary_amount, total_parent_component_amount):
    other_allowance_component = SalaryComponent.objects.filter(componentname="Other Allowance").first()

    if other_allowance_component and str(other_allowance_component.id) in compname:
        salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=other_allowance_component.id))
        
        if gross_salary_amount > total_parent_component_amount:
            other_allowance_amount = ceil(gross_salary_amount - total_parent_component_amount)
        else:
            other_allowance_amount = 0 

        other_allowance_amount = abs(other_allowance_amount)

        salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=other_allowance_amount)
        salary_amount.save()

def process_epfempoyee_component(compname, salary_structure_rule, total_NSparent_component_amount, basic_salary_amount):
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    print("component_name 1 ", compname)

    epfempoyee_component = SalaryComponent.objects.filter(componentname="EPF Employee").first()
    print("epfempoyee_component ", epfempoyee_component, "str(net_salary_component.id) :", str(epfempoyee_component.id))
    print("epf_component.percent :", epfempoyee_component.percent )

    if str(epfempoyee_component.id) in compname:
        print("ghghfghhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh")
        salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=epfempoyee_component.id))
        calculated_amount = ceil(basic_salary_amount * (epfempoyee_component.percent / 100.0))
        print("total_NSparent_component_amount 111 :", total_NSparent_component_amount, calculated_amount)
        total_NSparent_component_amount += calculated_amount
        print("total_NSparent_component_amount :", total_NSparent_component_amount)
        salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
        salary_amount.save()
    return total_NSparent_component_amount

def process_epf_component(compname, salary_structure_rule, total_CTCparent_component_amount, basic_salary_amount):
    print("bvcvbbcbcbvvvvvvvvvvvvvvvvv")
    epf_component = SalaryComponent.objects.filter(componentname="EPF Employer").first()
    if str(epf_component.id) in compname:
        print("epf_component nvnbvnbv : ", epf_component)
        salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=epf_component.id))
        calculated_amount = ceil(basic_salary_amount * (epf_component.percent / 100.0))
        print("total_CTCparent_component_amount 1 :", total_CTCparent_component_amount, calculated_amount)
        total_CTCparent_component_amount += calculated_amount
        print("total_CTCparent_component_amount :", total_CTCparent_component_amount)
        salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
        salary_amount.save()
    return total_CTCparent_component_amount

def process_net_salary(compname, salary_structure_rule, total_NSparent_component_amount, gross_salary_amount):
    net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
    if net_salary_component and str(net_salary_component.id) in compname:
        salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=net_salary_component.id))

        net_salary_amount = gross_salary_amount - total_NSparent_component_amount
        salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=net_salary_amount)
        salary_amount.save()

def process_ctc(compname, salary_structure_rule, total_CTCparent_component_amount, gross_salary_amount):
    ctc_component = SalaryComponent.objects.filter(componentname="CTC").first()
    if ctc_component and str(ctc_component.id) in compname:
        salary_name = SalaryStructureName.objects.create(salaryrule=salary_structure_rule)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=ctc_component.id))

        ctc_amount = gross_salary_amount + total_CTCparent_component_amount
        print("gross_salary_amount : ",gross_salary_amount, "total_CTCparent_component_amount : ", total_CTCparent_component_amount, "ctc_amount : ", ctc_amount)
        salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=ctc_amount)
        salary_amount.save()

def edit_salary_structure(request):
    if request.method == "POST":
        structureid = request.POST.get('structure_id')
        rule_name = request.POST.get('rulename')
        description = request.POST.get('description')

        salarystructure = SalaryStructureRule.objects.get(id=structureid)
        salarystructure.rule_name = rule_name
        salarystructure.Description = description
        salarystructure.save()

        existing_components = request.POST.getlist('componentname_existing', [])
        existing_amounts = request.POST.getlist('amount_existing', [])

        print("existing_components :" , existing_components, existing_amounts)

        new_components = request.POST.getlist('componentname_new', [])
        new_amounts = request.POST.getlist('amount_new', [])
        print("new_components :", new_components)

        all_components = existing_components + new_components
        all_amounts = existing_amounts + new_amounts

        print("all_components :", all_components, all_amounts)

        gross_salary_component = SalaryComponent.objects.filter(componentname="Gross Salary").first()
        net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
        ctc_salary_component = SalaryComponent.objects.filter(componentname="CTC").first()
        print("gross_salary_component :", gross_salary_component)

        gross_salary_amount = 0
        total_parent_component_amount = 0
        total_NSparent_component_amount = 0
        total_CTCparent_component_amount = 0
        basic_salary_amount = 0

        if gross_salary_component and str(gross_salary_component.id) in all_components:
            gross_salary_index = all_components.index(str(gross_salary_component.id))
            gross_salary_amount = float(all_amounts[gross_salary_index])
            print("gross_salary_amount 13070:", gross_salary_amount)
            

        for i, sname in enumerate(existing_components):
            component = SalaryComponent.objects.get(id=sname)
            print("existing_components component.componentname :", component , component.componentname)
            
            if component.componentname == "Professional Tax" or component.componentname == "Professional tax":
                print("existing_components component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 167 if gross_salary_amount < 22000 else 208

            if component.componentname == "Insurance":
                print("existing_components component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 0 if gross_salary_amount <= 25000 else 245

                if component.Parentcomponentname == net_salary_component:
                    total_NSparent_component_amount += calculated_amount
                elif component.Parentcomponentname == ctc_salary_component:
                    total_CTCparent_component_amount += calculated_amount
            if component.componentname in ["Other Allowance", "EPF Employer", "EPF Employee", "Net Salary", "CTC", "Professional Tax", "Professional tax","Insurance"]:
                continue
            print("existing_components component.componentname :", component.componentname)
            calculated_amount = float(existing_amounts[i])
            if component.percent:
                calculated_amount = ceil(gross_salary_amount * (component.percent / 100.0))
            print("existing_components component.componentname :", calculated_amount)
            if component.Parentcomponentname == gross_salary_component:
                total_parent_component_amount += calculated_amount
            elif component.Parentcomponentname == net_salary_component:
                total_NSparent_component_amount += calculated_amount
            elif component.Parentcomponentname == ctc_salary_component:
                total_CTCparent_component_amount += calculated_amount
                print(" existing_components total_CTCparent_component_amount elif: ", total_CTCparent_component_amount)
            if component.componentname == "Basic Salary":
                basic_salary_amount = calculated_amount
        
        for i, sname in enumerate(new_components):
            component = SalaryComponent.objects.get(id=sname)

            print("component.componentname :", component , component.componentname)
               
            if component.componentname == "Professional Tax" or component.componentname == "Professional tax":
                print("component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 167 if gross_salary_amount < 22000 else 208

            if component.componentname == "Insurance":
                print("component.componentname gross_salary_amount if:", component.componentname , gross_salary_amount)
                
                calculated_amount = 0 if gross_salary_amount <= 25000 else 245

                salary_name = SalaryStructureName.objects.create(salaryrule=salarystructure)
                salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=component.id))
                salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
                salary_amount.save()
                    
                if component.Parentcomponentname == net_salary_component:
                    total_NSparent_component_amount += calculated_amount
                elif component.Parentcomponentname == ctc_salary_component:
                    total_CTCparent_component_amount += calculated_amount

            if component.componentname in ["Other Allowance", "EPF Employer", "EPF Employee", "Net Salary", "CTC", "Professional Tax", "Professional tax","Insurance"]:
                continue

            print("component.componentname :", component.componentname)
            salary_name = SalaryStructureName.objects.create(salaryrule=salarystructure)
            salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=sname))

            calculated_amount = float(new_amounts[i])
            if component.percent:
                calculated_amount = ceil(gross_salary_amount * (component.percent / 100.0))

            print("component.componentname :", calculated_amount)

            if component.Parentcomponentname == gross_salary_component:
                total_parent_component_amount += calculated_amount
            elif component.Parentcomponentname == net_salary_component:
                total_NSparent_component_amount += calculated_amount
            elif component.Parentcomponentname == ctc_salary_component:
                total_CTCparent_component_amount += calculated_amount
                print("total_CTCparent_component_amount elif: ", total_CTCparent_component_amount)

            if component.componentname == "Basic Salary":
                basic_salary_amount = calculated_amount
            
            salary_amount = SalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount)
            salary_amount.save()
            
        process_other_allowance(new_components, salarystructure, gross_salary_amount, total_parent_component_amount)

        total_NSparent_component_amount = process_epfempoyee_component(new_components, salarystructure, total_NSparent_component_amount, basic_salary_amount)

        total_CTCparent_component_amount = process_epf_component(new_components, salarystructure, total_CTCparent_component_amount, basic_salary_amount)

        process_net_salary(new_components, salarystructure, total_NSparent_component_amount, gross_salary_amount)

        process_ctc(new_components, salarystructure, total_CTCparent_component_amount, gross_salary_amount)

    return redirect('create_salary')

def delete_salary_structure(request, id):
    rule = get_object_or_404(SalaryStructureRule, pk=id)

    names = SalaryStructureName.objects.filter(salaryrule=rule)
    amounts = SalaryStructureAmount.objects.filter(salaryname__in=names)

    amounts.delete()
    names.delete()
    rule.delete()

    return redirect('create_salary')


def delete_component_amount(request, structure_id, name_id):
    rule = get_object_or_404(SalaryStructureRule, pk=structure_id)
    name = get_object_or_404(SalaryStructureName, pk=name_id)

    rule.salarystructurename_set.remove(name)

    amount = SalaryStructureAmount.objects.get(salaryname=name)
    amount.delete()

    name.delete()

    return redirect('create_salary')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_assign_salarystructure(request):
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=request.user.id)
    component = SalaryComponent.objects.filter(admin_id=request.user.id)
    datas = User.objects.filter(Q(id=request.user.id) | Q(admin_id=request.user.id) & ~Q(status="Inactive"))
    com_rule = SalaryStructureRule.objects.filter(myuser_12=request.user.id)

    assg_rule = []
    for user in datas:
        last_assg_rule = AssignSalaryStructure.objects.filter(user_id__id=user.id).order_by('-effective_date').first()
        assg_rule.append(last_assg_rule)

    query = request.GET.get('search')
    count_user = User.objects.count()

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(request, "index/assign_salarystructure.html",{'datas': datas, 'query': query, 'k': k, 'count_user': count_user, 'assg_rule': assg_rule,
                   'com_rule': com_rule, 'component': component, **x})


# EXPORT SALARY STRUCTURE

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def export_salarystructure(request):
    if request.method == 'POST':

        selected_employee_ids = request.POST.getlist('selected_employees')
        selected_users = User.objects.filter(id__in=selected_employee_ids)
        print("selected_employee_ids :", selected_employee_ids, selected_users)

        data = []
        all_components = set()

        for user in selected_users:
            assign = AssignSalaryStructure.objects.filter(
                user_id=user.id).order_by("-effective_date").first()
            print("assgrulestr :", assign)
            ctc_amount = 0
            annualctc_amount = 0

            user_components = set()

            user_data = [
                user.empid,
                user.username,
                user.wrklcn.location,
                user.department.name,
                user.designation.name,
            ]

            if assign:
                # for assign in assgrulestr:
                user_data.append(assign.assign_salary.rule_name)
                effectivedate = assign.effective_date
                formatted_effective_date = effectivedate.strftime("%d %B %Y")
                user_data.append(formatted_effective_date)
                names = AssignSalaryStructureName.objects.filter(
                    salaryrule=assign)
                amounts = AssignSalaryStructureAmount.objects.filter(
                    salaryname__in=names)
                all_components.update(
                    [comp.componentname for name in names for comp in name.salarycomponent.all()])
                user_components.update(
                    [comp.componentname for name in names for comp in name.salarycomponent.all()])
                if request.POST.get('enable_toggle') == 'on':
                    annual_amount = [amount.amount * 12 for amount in amounts]
                    annualctc_amount += sum(annual_amount)
                    print("annual_amount :", annual_amount)
                else:
                    ctc_amount += sum(amount.amount for amount in amounts)
                    print("ctc_amount :", ctc_amount)
            else:
                user_data.extend(['', '',])

            user_data.append(annualctc_amount if request.POST.get(
                'enable_toggle') == 'on' else ctc_amount)

            user_component_amounts = {comp: 0 for comp in all_components}

            if assign:
                names = AssignSalaryStructureName.objects.filter(
                    salaryrule=assign)
                amounts = AssignSalaryStructureAmount.objects.filter(
                    salaryname__in=names)

                print("All Components:", all_components)
                for comp in user_components:
                    for name in names:
                        for component in name.salarycomponent.all():
                            if comp == component.componentname:
                                for amount in name.assignsalarystructureamount_set.all():
                                    user_component_amounts[comp] += (amount.amount * 12) if request.POST.get(
                                        'enable_toggle') == 'on' else amount.amount

            user_data.extend(list(user_component_amounts.values()))
            data.append(user_data)

        table_columns = [
            'Employee ID', 'Name', 'Location', 'Department', 'Designation', 'Structure Name', 'Structure Effective Date',
            'CTC Amount',
        ]
        if all_components:
            table_columns.extend(all_components)
        df = pd.DataFrame(data, columns=table_columns)

        excel_file_name = "salary_structure_export.xlsx"
        excel_file_path = os.path.join(
            BASE_DIR, 'media/csv/salary_structure_export.xlsx')
        df.to_excel(excel_file_path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        fixed_width = 30

        for column in sheet.columns:
            sheet.column_dimensions[get_column_letter(
                column[0].column)].width = fixed_width

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

        wb.save(excel_file_path)

        subject = "Salary Structure Export"
        message = 'Here is the salary structure details'
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [admin.email for admin in User.objects.filter(
            id=request.user.id)]
        print("to_email :", to_email)
        email = EmailMessage(subject, message, from_email, to_email)
        email.attach(excel_file_name, open(excel_file_path,
                                           'rb').read(), 'application/vnd.ms-excel')
        email.send()

        today = datetime.now()
        filemanager = Filemanager.objects.create(myuser_10=request.user, requesttype="Employee Salary Structure",
                                                 frmt="XLSX", scheduleon=today, status="Success")

        with open(excel_file_path, 'rb') as file:
            filemanager.saveexcel.save(
                excel_file_name, ContentFile(file.read()))

    return redirect('filemanagernav')

def assign_salarystructure(request):
    salarystructure = SalaryStructureRule.objects.filter(myuser_12=request.user.id)
    if request.method == 'POST':
        selected_rules = request.POST.get('rule')
        selected_rulesid = SalaryStructureRule.objects.get(id=selected_rules)
        effective_date = request.POST.get('effdate')
        formateddate = datetime.strptime(effective_date, "%B %Y")
        formatted_date = formateddate.strftime("%Y-%m-%d")

        selected_employees = request.POST.getlist('selected_employees')
        for employee_id in selected_employees:
            assign_salarystr = AssignSalaryStructure.objects.create(user_id_id=employee_id, effective_date=formatted_date, assign_salary=selected_rulesid)
            salary_names = SalaryStructureName.objects.filter(salaryrule=selected_rules)
            for salary_name in salary_names:
                new_assign_name = AssignSalaryStructureName.objects.create(salaryrule=assign_salarystr)
                new_assign_name.salarycomponent.set(salary_name.salarycomponent.all())
                salary_amounts = SalaryStructureAmount.objects.filter(salaryname=salary_name)
                for salary_amount in salary_amounts:
                    AssignSalaryStructureAmount.objects.create(amount=salary_amount.amount, salaryname=new_assign_name)

        return redirect('view_assign_salarystructure')
    return render(request, "index/assign_salarystructure.html", {'com_rule': salarystructure})

def delete_assign_salarystructure(request, assign_rule_id):
    assign_rule = AssignSalaryStructure.objects.get(id=assign_rule_id)
    names = AssignSalaryStructureName.objects.filter(salaryrule=assign_rule)
    amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
    amounts.delete()
    names.delete()
    assign_rule.delete()
    return redirect('view_assign_salarystructure')

def salary_structure(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)

    assign = AssignSalaryStructure.objects.filter(user_id=admin_id).last()

    hierarchy = defaultdict(list)  

    if assign:
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        amount_map = {amount.salaryname.salarycomponent.first().id: amount.amount for amount in amounts}
        print("amount_map : ", amount_map)
        for name in names:
            for component in name.salarycomponent.all():
                parent_name = component.Parentcomponentname.componentname if component.Parentcomponentname else None
                monthly_amount = amount_map.get(component.id, 0)
                component_data = {
                    'name': component.componentname,
                    'monthly_amount': monthly_amount,
                    'annual_amount': monthly_amount * 12,
                    'is_parent': parent_name is None
                }
                print("parent_name : ", parent_name)
                print("component_data : ", component_data)
                if parent_name:
                    hierarchy[parent_name].append(component_data)
                    print("hierarchy[parent_name] : ",  hierarchy[parent_name])
                else:
                    hierarchy[component.componentname].append(component_data)
                    print("hierarchy[component.componentname] : ",  hierarchy[component.componentname])

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "is_view_salarystructure": False,

    }

    return render(
        request,
        'index/salarystructure.html',
        {
            'assign': assign,
            'hierarchy': dict(hierarchy),
            **x
        }
    )

def empsalary_structure(request):
    user_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)

    assign = AssignSalaryStructure.objects.filter(user_id=user_id).last()

    hierarchy = defaultdict(list)  

    if assign:
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        amount_map = {amount.salaryname.salarycomponent.first().id: amount.amount for amount in amounts}
        print("amount_map : ", amount_map)
        for name in names:
            for component in name.salarycomponent.all():
                parent_name = component.Parentcomponentname.componentname if component.Parentcomponentname else None
                monthly_amount = amount_map.get(component.id, 0)
                component_data = {
                    'name': component.componentname,
                    'monthly_amount': monthly_amount,
                    'annual_amount': monthly_amount * 12,
                    'is_parent': parent_name is None
                }
                print("parent_name : ", parent_name)
                print("component_data : ", component_data)
                if parent_name:
                    hierarchy[parent_name].append(component_data)
                    print("hierarchy[parent_name] : ",  hierarchy[parent_name])
                else:
                    hierarchy[component.componentname].append(component_data)
                    print("hierarchy[component.componentname] : ",  hierarchy[component.componentname])

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }

    return render(
        request,
        'Employee/salarystructure.html',
        {
            'assign': assign,
            'hierarchy': dict(hierarchy),
            **x
        }
    )

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_calendar(request):
    admin_id = request.user.id
    admin_wrklcn_id = request.user.wrklcn.id if request.user.wrklcn else None
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=admin_id)

    # print("admin_id ; admin_wrklcn_id : ", admin_id, admin_wrklcn_id)
    Holiday_Location_List = HolidayLocationList.objects.filter(
        Holiday_List__Myuser_13__id=admin_id)
    # print("Holiday_Location_List :", Holiday_Location_List)

    holiday_list = HolidayLocationList.objects.filter(
        Holiday_List__Myuser_13=admin_id, HolidayLocation__id=admin_wrklcn_id)
    # print("holiday_list :", holiday_list)

    holiday_data = []
    for holiday_entry in holiday_list:
        holiday_data.append({
            'date': holiday_entry.Holiday_List.get_formatted_date(),
            'toggle_btn': holiday_entry.HolidayToggleBtn_ON,
            'holiday_name': holiday_entry.Holiday_List.HolidayName,
            'holiday_date': holiday_entry.Holiday_List.HolidayDate
        })

    leaveuser_list = Leave.objects.filter((Q(applicant_email__id=admin_id) | Q(
        applicant_email__admin_id=admin_id)) & Q(status="Approved"))
    print("leaveuser_list :", leaveuser_list)
    leave_counts = {}
    leave_data = []
    for leave_entry in leaveuser_list:
        start_date = leave_entry.strtDate
        end_date = leave_entry.endDate

        # Skip if dates are invalid
        if not isinstance(start_date, (datetime, date)) or not isinstance(end_date, (datetime, date)):
            continue

        try:
            leave_range = [
                (start_date + timedelta(days=x))
                for x in range((end_date - start_date).days + 1)
            ]
        except Exception as e:
            continue  # Skip corrupt ranges

        for leave_date in leave_range:
            leave_date_str = leave_date.strftime("%Y-%m-%d")
            leave_counts[leave_date_str] = leave_counts.get(leave_date_str, 0) + 1

        user_profile = Myprofile.objects.filter(
            myuser=leave_entry.applicant_email
        ).first()

        leave_data.append({
            'date': start_date.strftime("%Y-%m-%d"),
            'enddate': end_date.strftime("%Y-%m-%d"),
            'leave_type': str(leave_entry.leavetyp or '').strip(),
            'applicant_name': leave_entry.applicant_email.username if leave_entry.applicant_email else '',
            'designation': leave_entry.applicant_email.designation.name if leave_entry.applicant_email and leave_entry.applicant_email.designation else '',
            'image': user_profile.image.url if user_profile and user_profile.image else "/static/logo/userlogo.png",
        })

    # print("leave_data :", leave_data)

    user_birthday = []
    user_workann = []
    all_users = User.objects.filter(
        (Q(admin_id=admin_id) | Q(id=admin_id)) & Q(status='Active'))

    for user in all_users:
        birthday = datetime.strptime(parse_and_format_date(user.dob), '%d %B %Y').strftime(
            '%d %m') if user.dob else None
        day, month = birthday.split() if birthday else (None, None)
        print("birthday : ", birthday)

        date_obj = datetime.strptime(
            parse_and_format_date(user.dob), '%d %B %Y') if user.dob else None
        formatted_date = date_obj.strftime('%Y-%m-%d') if date_obj else None

        user_profile = Myprofile.objects.filter(myuser=user).first()
        user_birthday.append({
            'username': user.username,
            'day': day if day is not None else "",
            'month': month if day is not None else "",
            'formatted_date': formatted_date if day is not None else datetime.now().strftime("%Y-%m-%d"),
            'designation': user.designation.name if user.designation else '',
            'image': user_profile.image.url if user_profile and user_profile.image else "/static/logo/userlogo.png",
        })

        workanniversary = datetime.strptime(parse_and_format_date(user.datejoin), '%d %B %Y').strftime(
            '%d %m') if user.datejoin else None
        day, month = workanniversary.split() if workanniversary else (None, None)

        date_obj = datetime.strptime(
            parse_and_format_date(user.datejoin), '%d %B %Y') if user.datejoin else None

        today = datetime.now().date()
        years_of_service = datetime.strptime(
            parse_and_format_date(user.datejoin) if user.datejoin else datetime.now().strftime("%d %B %Y"), '%d %B %Y').year != today.year

        ann_date = date_obj.strftime('%Y-%m-%d') if years_of_service else ''

        userprofile = Myprofile.objects.filter(myuser=user).first()
        user_workann.append({
            'username': user.username,
            'ann_day': day if day else "",
            'ann_month': month if month else "",
            'ann_date': ann_date if ann_date else "",
            'designation': user.designation.name if user.designation else '',
            'user_image': userprofile.image.url if userprofile and userprofile.image else "/static/logo/userlogo.png",
        })

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'holiday_list': holiday_list,
        'holiday_data': holiday_data,
        'leave_data': leave_data,
        'leave_counts': leave_counts,
        'user_birthday': user_birthday,
        'user_workann': user_workann
    }

    return render(request, "index/calendar.html", x)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def view_emp_calendar(request):

    user_id = request.user.id
    user_wrklcn = request.user.wrklcn.id
    k = Myprofile.objects.filter(myuser__id=user_id)
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)

    holiday_list = HolidayLocationList.objects.filter(
    Holiday_List__Myuser_13=admin_id, HolidayLocation__id=user_wrklcn)

    holiday_data = []
    for holiday_entry in holiday_list:
        holiday_data.append({
            'date': holiday_entry.Holiday_List.get_formatted_date(),
            'toggle_btn': holiday_entry.HolidayToggleBtn_ON,
            'holiday_name': holiday_entry.Holiday_List.HolidayName,
            'holiday_date': holiday_entry.Holiday_List.HolidayDate
        })

    leaveuser_list = Leave.objects.filter(
        applicant_email_id=user_id, status='Approved')
    # leavecounts = leaveuser_list.values('strtDate','endDate','applicant_email__username')
    # leave_count = leaveuser_list.values('strtDate','endDate').annotate(count=Count('applicant_email', distinct=True))
    # leave_counts_dict = {(entry['strtDate'].strftime("%Y-%m-%d"),entry['endDate'].strftime("%Y-%m-%d")): entry['count'] for entry in leave_count}
    # leave_counts_dicts = {(entry['strtDate'].strftime("%Y-%m-%d"),entry['endDate'].strftime("%Y-%m-%d")): entry['applicant_email__username']  for entry in leavecounts}

    leave_counts = {}
    leave_data = []
    for leave_entry in leaveuser_list:
        start_date = leave_entry.strtDate
        end_date = leave_entry.endDate
        leave_range = [
            start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]

        for leave_date in leave_range:
            leave_date_str = leave_date.strftime("%Y-%m-%d")
            leave_counts[leave_date_str] = leave_counts.get(
                leave_date_str, 0) + 1

        user_profile = Myprofile.objects.filter(
            myuser=leave_entry.applicant_email).first()
        leave_data.append({
            'date': leave_entry.strtDate.strftime("%Y-%m-%d"),
            'enddate': leave_entry.endDate.strftime("%Y-%m-%d"),
            'leave_type': leave_entry.leavetyp,
            'applicant_name': leave_entry.applicant_email.username,
            'designation': leave_entry.applicant_email.designation.name if leave_entry.applicant_email.designation.name else '',
            'image': user_profile.image.url if user_profile and user_profile.image else "/static/logo/userlogo.png",
        })

    all_users = User.objects.filter(Q(admin_id=admin_id) | Q(id=admin_id), status__iexact='Active')

    user_birthday = []
    user_workann = []

    for user in all_users:
        if not user.dob:
            continue

        try:
            dob_parsed = datetime.strptime(parse_and_format_date(user.dob), '%d %B %Y')
            birthday_day = dob_parsed.strftime('%d')
            birthday_month = dob_parsed.strftime('%m')
            formatted_dob = dob_parsed.strftime('%Y-%m-%d')
        except Exception:
            continue

        user_profile = Myprofile.objects.filter(myuser=user).first()
        user_birthday.append({
            'username': user.username,
            'day': birthday_day,
            'month': birthday_month,
            'formatted_date': formatted_dob,
            'designation': user.designation.name if user.designation else '',
            'image': user_profile.image.url if user_profile and user_profile.image else "/static/logo/userlogo.png",
        })

        if user.datejoin:
            try:
                doj_parsed = datetime.strptime(parse_and_format_date(user.datejoin), '%d %B %Y')
                is_anniversary = doj_parsed.year != datetime.now().year
                if is_anniversary:
                    user_workann.append({
                        'username': user.username,
                        'ann_day': doj_parsed.strftime('%d'),
                        'ann_month': doj_parsed.strftime('%m'),
                        'ann_date': doj_parsed.strftime('%Y-%m-%d'),
                        'designation': user.designation.name if user.designation else '',
                        'user_image': user_profile.image.url if user_profile and user_profile.image else "/static/logo/userlogo.png",
                    })
            except Exception:
                continue

    print("user_workann :", user_workann)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'holiday_list': holiday_list,
        'holiday_data': holiday_data,
        'leave_data': leave_data,
        'leave_counts': leave_counts,
        'user_birthday': user_birthday,
        'user_workann': user_workann
    }

    return render(request, "Employee/calendar.html", x)


def view_holiday(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=admin_id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    work_location = Worklocation.objects.filter(admin_id=admin_id)
    holiday_lists = HolidayList.objects.filter(Myuser_13=admin_id)

    data = []
    for holidaylist in holiday_lists:
        holidaylocation = HolidayLocationList.objects.filter(
            Holiday_List=holidaylist)
        selected_location_ids = list(
            holidaylocation.values_list('HolidayLocation__id', flat=True))

        holiday_location_info = []
        for location in work_location:
            location_info = holidaylocation.filter(
                HolidayLocation=location).first()
            holiday_location_info.append(
                {'location': location, 'info': location_info})

        data.append({
            'holidaylist': holidaylist,
            'holidaylocation': holidaylocation,
            'selected_location_ids': selected_location_ids,
            'location_info': holiday_location_info,
        })

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        'work_location': work_location,
        'data': data
    }

    return render(request, 'index/holidaylist.html', context)


def addholiday(request):
    if request.method == 'POST':
        holiday_name = request.POST.get('holiday_name')
        holiday_date = request.POST.get('holiday_date')
        holiday_location_ids = request.POST.getlist('holiday_location', [])

        holiday_list = HolidayList.objects.create(
            Myuser_13=request.user,
            HolidayName=holiday_name,
            HolidayDate=holiday_date
        )
        for loc_id in holiday_location_ids:
            holiday_location = Worklocation.objects.get(id=loc_id)

            toggle_optional = request.POST.get(
                f'toggle_optional_{loc_id}', 'off')

            holiday_location_instance = HolidayLocationList.objects.create(
                Holiday_List=holiday_list,
                HolidayToggleBtn_ON=toggle_optional
            )
            holiday_location_instance.HolidayLocation.add(holiday_location)
        return redirect('Holidayview')


def editholiday(request):
    holiday_id = request.POST.get('holidaylist_id')
    holiday = HolidayList.objects.get(id=holiday_id)
    holiday.HolidayName = request.POST.get('holiday_name')
    holiday.HolidayDate = request.POST.get('holiday_date')
    holiday.save()
    holiday_location_ids = request.POST.getlist('holiday_location', [])
    HolidayLocationList.objects.filter(Holiday_List=holiday).delete()
    for loc_id in holiday_location_ids:
        holiday_location = Worklocation.objects.get(id=loc_id)
        toggle_optional = request.POST.get(f'toggle_optional_{loc_id}', 'off')
        holiday_location_instance = HolidayLocationList.objects.create(
            Holiday_List=holiday,
            HolidayToggleBtn_ON=toggle_optional
        )
        holiday_location_instance.HolidayLocation.add(holiday_location)
    return redirect('Holidayview')


def deleteholiday_list(request, id):
    holidaylist_id = get_object_or_404(HolidayList, pk=id)
    holidaylocation = HolidayLocationList.objects.filter(
        Holiday_List=holidaylist_id)

    holidaylocation.delete()
    holidaylist_id.delete()
    return redirect('Holidayview')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def payroll_dec(request):
    user_id = request.user.id
    cmp = companyprofile.objects.filter(admin_id=user_id)
    profile = Myprofile.objects.filter(myuser__id=user_id)
    k = reimbursement.objects.filter(
        Q(myuser_11=user_id) | Q(myuser_11__admin_id=user_id)).order_by("-created_date")
    x = {
        "cmp": cmp[0] if cmp.exists() else cmp,
        "profile": profile[0] if profile.exists() else profile,
    }
    return render(request, 'index/reimbursements.html', {'k': k, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def reimbursements(request):
    user_id = request.user.id
    admin_id = User.objects.get(id=user_id).admin_id
    c = companyprofile.objects.filter(admin_id=admin_id)
    p = Myprofile.objects.filter(myuser__id=request.user.id)
    k = reimbursement.objects.filter(myuser_11=user_id)
    x = {
        "p": p[0] if p.exists() else p,
        "c": c[0] if c.exists() else c,
    }
    return render(request, 'Employee/reimbursement.html', {'k': k, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active', 'Onboarding'])
def empadd_reimbursements(request):
    if request.method == 'POST':
        user = request.user
        user_id = User.objects.get(id=user.id)
        comp = request.POST.get('comp')
        dec = request.POST.get('dec')
        img = request.FILES['img']
        obj = FileSystemStorage()
        img_obj = obj.save(img.name, img)
        comm = request.POST.get('comm')
        status = 'Pending'
        reimbursement.objects.create(component=comp, declared_amount=dec, document=img_obj,
                                     comment=comm, status=status, myuser_11=user_id)
        return redirect('reimbursements')


def cancel_reimbursement(request):
    if request.method == 'POST':
        leave_id = request.POST.get('re_id')
        leave = reimbursement.objects.get(id=leave_id)
        leave.cancel_requested = True
        leave.status = 'Pending'
        leave.save()
    return redirect('payroll_dec')


def approve_reimbursement(request):
    if request.method == 'POST':
        leave_id = request.POST.get('re_id')
        leave = reimbursement.objects.get(id=leave_id)
        leave.status = 'Approved'
        leave.save()
        return redirect('payroll_dec')


def reject_reimbursement(request):
    if request.method == 'POST':
        riemb_id = request.POST.get('re_id')
        riembid = reimbursement.objects.get(id=riemb_id)
        riembid.status = 'Rejected'
        riembid.save()
        return redirect('payroll_dec')


def birthdaymail():
    today = date.today()
    formatted_today = today.strftime("%d %B")

    # ✅ Only fetch Active employees
    employees = User.objects.filter(role="Employee", status="Active")

    for employee in employees:
        dateofbirth_str = employee.dob
        formatted_dateofbirth = None

        if dateofbirth_str:
            try:
                dateofbirth = datetime.strptime(dateofbirth_str, "%d %B %Y").date()
                formatted_dateofbirth = dateofbirth.strftime("%d %B")
                print(f"Formatted DOB for {employee.username}: {formatted_dateofbirth}")
            except ValueError:
                continue
        else:
            continue

        if formatted_dateofbirth == formatted_today:
            Age = relativedelta(today, dateofbirth).years
            print(f"🎉 Birthday: {employee.username}, Age: {Age} years")

            admin = User.objects.filter(role="Admin", id=employee.admin_id).first()

            if admin and admin.email:
                to = [admin.email]
                subject = f"🎂 Today is {employee.username}'s Birthday!"
                html_body = render_to_string('index/birthdaymail.html', {
                    'employee': employee,
                    'admin': admin,
                    'Age': Age,
                    'formatted_today': formatted_today
                })

                msg = EmailMultiAlternatives(subject=subject, from_email=settings.EMAIL_HOST_USER, to=to)
                msg.attach_alternative(html_body, "text/html")
                msg.send()
                print(f"✅ Birthday mail sent to admin {admin.username}")
            else:
                print(f"⚠️ Admin email not found for {employee.username}")

    return HttpResponse("SUCCESS")



@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def runpayroll(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    current_date = datetime.now()
    print(current_date)
    current_month = current_date.month
    print(current_month)
    current_year = current_date.year

    selected_month_str = request.GET.get('monthselect', None)
    print(f"Selected Month String: {selected_month_str}")
    if selected_month_str is None:
        selected_month = current_month
        selected_year = current_year
        month_str = current_date.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    selected_date = datetime(selected_year, selected_month, 1)
    print("selected_year :", selected_year, selected_month)

    # //// EDIT LOP COUNT & WFO COUNT////
    if request.method == "POST":
        lopcount = request.POST.get("lopcount")
        lopid = request.POST.get('lopid')
        WFOcount = request.POST.get('WFOcount')
        wfoid = request.POST.get('wfoid')
        lop_instances = Runpayroll_lop.objects.filter(user_id_id=lopid, lop_date__month=selected_month, lop_date__year=selected_year)
        if lop_instances.exists():
            for lop_instance in lop_instances:
                lop_instance.lop_count = lopcount
                lop_instance.save()
        else:
            Runpayroll_lop.objects.create(
                user_id_id=lopid, lop_date=selected_date, lop_count=lopcount)

        WFO_instances = WFOCount.objects.filter(user_id_id=wfoid, wfo_date__month=selected_month, wfo_date__year=selected_year)
        print("WFO_instances :", WFO_instances)
        if WFO_instances.exists():
            for WFO_instance in WFO_instances:
                WFO_instance.wfocount = WFOcount
                WFO_instance.save()
        else:
            WFOCount.objects.create(user_id_id=wfoid, wfo_date=selected_date, wfocount=WFOcount)

        payregister_instance = PayRegister.objects.filter(
            user_id=lopid).first()
        if payregister_instance:
            payregister_instance.status = "Pending Calculation"
            payregister_instance.save()

    datas = User.objects.filter(Q(id=admin_id) | Q(
        admin_id=admin_id)).order_by('empid')
    users = User.objects.filter(Q(id=admin_id) | Q(
        admin_id=admin_id) & Q(status='Active'))
    absent_anomaly_punches = Punch.objects.filter(
        Q(user__id=admin_id) | Q(user__admin_id=admin_id),
        status__in=['A', 'AN', 'HL', 'L'],
        date__date__month=selected_month,
        date__date__year=selected_year
    ).values('user__id', 'status', 'date')
    print(absent_anomaly_punches)

    combine_data = {}
    unique_anomalies = []

    for punch in absent_anomaly_punches:
        user_id = punch['user__id']
        status = punch['status']
        date = punch['date']

        print(f"Checking: user_id={user_id}, status={status}, date={date}")
        if not any(entry['user_id'] == user_id and entry['status'] == status and entry['date'].date() == date.date() for entry in unique_anomalies):
            unique_anomalies.append(
                {'user_id': user_id, 'status': status, 'date': date})

            print(f"Added: user_id={user_id}, status={status}, date={date}")
            print("Unique anomalies:", unique_anomalies)
            if user_id not in combine_data:
                user_details = User.objects.filter(id=user_id).first()
                combine_data[user_id] = {'empid': user_details.empid,
                                         'username': user_details.username, 'absent_dates': [],
                                         'absent_count': 0, 'anomaly_count': 0}

            if status == 'A':
                combine_data[user_id]['absent_dates'].append(date)
                combine_data[user_id]['absent_count'] += 1

            elif status == 'AN':

                combine_data[user_id]['anomaly_count'] += 1

    # Use a set to keep track of unique leave entries for each user

    formatted_data = list(combine_data.values())
    leave_data = Leave.objects.filter(
        Q(applicant_email__id=admin_id) | Q(applicant_email__admin_id=admin_id), status='Approved',
        strtDate__month=selected_month, strtDate__year=selected_year,
        endDate__month=selected_month, endDate__year=selected_year

    ).values('leavetyp', 'applicant_email__id', 'strtDate', 'endDate', 'admin_id', 'Days')
    print(leave_data)
    combine_leave = {}
    for leave in leave_data:
        applicant_email_id = leave['applicant_email__id']
        strtDate = leave['strtDate']
        endDate = leave['endDate']
        print(
            f"Processing leave for user {applicant_email_id} with dates {strtDate} to {endDate}")
        leavetyp = leave['leavetyp']
        Days = leave['Days']
        if applicant_email_id not in combine_leave:
            user_details = User.objects.filter(id=applicant_email_id).first()
            combine_leave[applicant_email_id] = {'empid': user_details.empid,
                                                 'username': user_details.username, 'leave_count': 0, 'leave_dates': []}

        if leavetyp == 'Loss Of Pay':
            combine_leave[applicant_email_id]['leave_dates'].append(strtDate)
            combine_leave[applicant_email_id]['leave_dates'].append(endDate)
            combine_leave[applicant_email_id]['leave_count'] += Days
    print(combine_leave)
    formatted_leave = list(combine_leave.values())

    lop_count = Runpayroll_lop.objects.filter(lop_date__year=selected_year, lop_date__month=selected_month)
    print("lop_count :", lop_count)

    WFO_count = WFOCount.objects.filter(wfo_date__year=selected_year, wfo_date__month=selected_month)
    print("WFO_count :", WFO_count)
    for i in WFO_count:
        print(i.wfocount)

    x = {
        "k": k[0] if k.exists() else None,
        "data": c[0] if c.exists() else None,
        "users": users,
        "datas": datas,
        "formatted_data": formatted_data,
        "absent_anomaly_punches": absent_anomaly_punches,
        "leave_data": leave_data,
        "formatted_leave": formatted_leave,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "lop_count": lop_count,
        "month_str": month_str,
        "WFO_count": WFO_count
    }

    return render(request, "index/runpayroll.html", x)

def runpayroll_v2(request):
    admin_id = request.user.id
    current_date = datetime.now()

    # Get selected month and year
    selected_month_str = request.GET.get('monthselect')
    selected_date = datetime.strptime(selected_month_str, '%B %Y') if selected_month_str else current_date
    selected_year, selected_month = selected_date.year, selected_date.month
    month_str = selected_date.strftime('%B')

    # Fetch profile and company data with select_related (avoids additional queries)
    profile = Myprofile.objects.select_related('myuser').only('myuser__id').filter(myuser_id=admin_id).first()
    company = companyprofile.objects.only('admin_id').filter(admin_id=admin_id).first()

    # Handle LOP and WFO updates
    if request.method == "POST":
        lop_id, wfo_id = request.POST.get('lopid'), request.POST.get('wfoid')
        lop_count, wfo_count = request.POST.get("lopcount"), request.POST.get("WFOcount")

        # Use update_or_create to minimize queries
        Runpayroll_lop.objects.update_or_create(
            user_id_id=lop_id, lop_date__year=selected_year, lop_date__month=selected_month,
            defaults={"lop_count": lop_count}
        )
        WFOCount.objects.update_or_create(
            user_id_id=wfo_id, wfo_date__year=selected_year, wfo_date__month=selected_month,
            defaults={"wfocount": wfo_count}
        )

        # Mark PayRegister as "Pending Calculation"
        PayRegister.objects.filter(user_id=lop_id).update(status="Pending Calculation")

    # Fetch users efficiently
    users = User.objects.filter(Q(id=admin_id) | Q(admin_id=admin_id), status='Active').only('id', 'empid', 'username').order_by("username")
    all_users = users.order_by('username')

    # Fetch and process Punches
    punches = Punch.objects.filter(
        Q(user_id=admin_id) | Q(user__admin_id=admin_id),
        status__in=['A', 'AN', 'HL', 'L'],
        date__year=selected_year,
        date__month=selected_month
    ).values('user_id', 'status', 'date')

    combine_data = {}
    user_details_cache = {user.id: {'empid': user.empid, 'username': user.username} for user in users}

    for punch in punches:
        user_id, status, date = punch['user_id'], punch['status'], punch['date']
        if user_id not in combine_data:
            combine_data[user_id] = {**user_details_cache.get(user_id, {}), 'absent_dates': [], 'absent_count': 0, 'anomaly_count': 0}

        if status == 'A':
            combine_data[user_id]['absent_dates'].append(date)
            combine_data[user_id]['absent_count'] += 1
        elif status == 'AN':
            combine_data[user_id]['anomaly_count'] += 1

    formatted_data = list(combine_data.values())

    # Fetch and process Leave data efficiently
    leave_data = Leave.objects.filter(
        Q(applicant_email_id=admin_id) | Q(applicant_email__admin_id=admin_id),
        status='Approved',
        strtDate__year=selected_year, strtDate__month=selected_month
    ).values('leavetyp', 'applicant_email_id', 'strtDate', 'endDate', 'Days')

    combine_leave = {}
    for leave in leave_data:
        user_id = leave['applicant_email_id']
        if user_id not in combine_leave:
            combine_leave[user_id] = {**user_details_cache.get(user_id, {}), 'leave_count': 0, 'leave_dates': []}

        if leave['leavetyp'] == 'Loss Of Pay':
            combine_leave[user_id]['leave_dates'].extend([leave['strtDate'], leave['endDate']])
            combine_leave[user_id]['leave_count'] += leave['Days']

    formatted_leave = list(combine_leave.values())

    # Fetch LOP & WFO Counts efficiently
    lop_count = Runpayroll_lop.objects.filter(lop_date__year=selected_year, lop_date__month=selected_month).only('lop_count')
    wfo_count = WFOCount.objects.filter(wfo_date__year=selected_year, wfo_date__month=selected_month).only('wfocount')

    context = {
        "k": profile,
        "data": company,
        "users": users,
        "datas": all_users,
        "formatted_data": formatted_data,
        "absent_anomaly_punches": punches,
        "leave_data": leave_data,
        "formatted_leave": formatted_leave,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "lop_count": lop_count,
        "month_str": month_str,
        "WFO_count": wfo_count
    }

    return render(request, "index/runpayroll.html", context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def export_runpayroll(request):
    admin_id = User.objects.filter(id=request.user.id)
    if request.method == 'POST':
        month_str = request.POST.get('month')
        month = datetime.strptime(month_str, '%B %Y')
        year_selected = month.year
        month_selected = month.month
        userid = request.user.id

        absent_anomaly_punches = Punch.objects.filter(Q(user__id=userid) | Q(user__admin_id=userid),
                                                      status__in=['A', 'AN', 'HL', 'L'], date__date__month=month_selected,
                                                      date__date__year=year_selected).values('user__id', 'status', 'date')
        print(absent_anomaly_punches)
        payroll_list = []
        combine_data = {}
        unique_anomalies = []

        for punch in absent_anomaly_punches:
            user_id = punch['user__id']
            status = punch['status']
            date = punch['date']

            print(f"Checking: user_id={user_id}, status={status}, date={date}")
            if not any(entry['user_id'] == user_id and entry['status'] == status and entry['date'].date() == date.date() for entry in unique_anomalies):
                unique_anomalies.append(
                    {'user_id': user_id, 'status': status, 'date': date})

                print(
                    f"Added: user_id={user_id}, status={status}, date={date}")
                print("Unique anomalies:", unique_anomalies)
                if user_id not in combine_data:
                    user_details = User.objects.filter(id=user_id).first()
                    combine_data[user_id] = {'empid': user_details.empid,
                                             'username': user_details.username, 'absent_dates': [], 'anomaly_dates': [],
                                             'absent_count': 0, 'anomaly_count': 0}

                if status == 'A':
                    combine_data[user_id]['absent_dates'].append(date)
                    combine_data[user_id]['absent_count'] += 1

                elif status == 'AN':
                    combine_data[user_id]['anomaly_dates'].append(date)
                    combine_data[user_id]['anomaly_count'] += 1

        formatted_data = list(combine_data.values())
        leave_data = Leave.objects.filter(
            Q(applicant_email__id=userid) | Q(applicant_email__admin_id=userid), status='Approved',
            strtDate__month=month_selected, strtDate__year=year_selected,
            endDate__month=month_selected, endDate__year=year_selected

        ).values('applicant_email__id', 'strtDate', 'endDate', 'leavetyp', 'admin_id', 'Days')
        print(leave_data)

        combine_leave = {}

        for leave in leave_data:
            applicant_email_id = leave['applicant_email__id']
            strtDate = leave['strtDate']
            endDate = leave['endDate']
            leavetyp = leave['leavetyp']
            Days = leave['Days']
            if applicant_email_id not in combine_leave:
                user_details = User.objects.filter(
                    id=applicant_email_id).first()
                combine_leave[applicant_email_id] = {'empid': user_details.empid,
                                                     'username': user_details.username, 'leave_count': 0, 'leave_dates': []}

            if leavetyp == 'Loss Of Pay':
                combine_leave[applicant_email_id]['leave_dates'].append(
                    strtDate)
                combine_leave[applicant_email_id]['leave_dates'].append(
                    endDate)
                combine_leave[applicant_email_id]['leave_count'] += Days
        print(combine_leave)
        formatted_leave = list(combine_leave.values())
        user = User.objects.filter(Q(id=userid) | Q(admin_id=userid))
        all_users = [{'empid': user.empid, 'username': user.username,
                      'user_id': user.id} for user in user]

# Iterate over all_users to append details
        for user_data in all_users:
            print("user_data :", user_data)
            user_id = user_data['user_id']
            user_details = next(
                (data for data in formatted_data if data['empid'] == user_data['empid']), None)
            leave_details = next(
                (leave for leave in formatted_leave if leave['empid'] == user_data['empid']), None)

            lopdata = Runpayroll_lop.objects.filter(
                lop_date__month=month_selected, lop_date__year=year_selected, user_id=user_id)

            if lopdata:
                lop_count_total = lopdata.aggregate(total_lop_count=Sum('lop_count'))[
                    'total_lop_count']
                user_data['LOPCount'] = lop_count_total
            else:
                user_data['LOPCount'] = 0

            if user_details:
                user_data.update({
                    'Date': ', '.join(date.strftime('%Y-%m-%d') for date in user_details['absent_dates']) if user_details['absent_dates'] else '',

                    'AbsentCount': user_details['absent_count'],
                    'AnomalyCount': user_details['anomaly_count'],
                })
            else:
                user_data.update({
                    'Date': '',
                    'AbsentCount': 0,
                    'AnomalyCount': 0,
                })
            if leave_details:
                user_data['LeaveCount'] = leave_details['leave_count']
            else:
                user_data['LeaveCount'] = 0
            payroll_list.append({
                'Employee ID': user_data['empid'],
                'Name': user_data['username'],
                'Date': user_data['Date'],
                'Absent Count': user_data['AbsentCount'],
                'Anomaly Count': user_data['AnomalyCount'],
                'Leave Count': user_data['LOPCount'],
            })
        df = pd.DataFrame(payroll_list)
    excel_file_name = 'attendance_payroll.xlsx'
    path = os.path.join(BASE_DIR, 'media/csv/attendance_payroll.xlsx')

    df.to_excel(path, index=False, sheet_name='Sheet1')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    fixed_width = 20
    for column in sheet.columns:
        sheet.column_dimensions[get_column_letter(
            column[0].column)].width = fixed_width
    wb.save(path)

    today = datetime.now()
    filemanager = Filemanager.objects.create(myuser_10=request.user, requesttype="Attendance payroll Export",
                                             frmt="XLSX", scheduleon=today, status="In Queue")
    subject = 'Punch Data Export'
    message = 'Attached is the punch data for the specified date range.'
    from_email = settings.DEFAULT_FROM_EMAIL
    to_email = [adminid.email for adminid in admin_id]
    email = EmailMessage(subject, message, from_email, to_email)
    email.attach(excel_file_name, open(path, 'rb').read(),
                 'application/vnd.ms-excel')
    email.send()

    filemanager.status = "Success"
    with open(path, 'rb') as file:
        filemanager.saveexcel.save(
            'attendance_payroll.xlsx', ContentFile(file.read()))
    return redirect("runpayroll")


def search_runpayroll(request):
    admin_id = request.user.id
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=admin_id)
    current_date = datetime.now()
    print(current_date)
    current_month = current_date.month
    print(current_month)
    current_year = current_date.year
    selected_month_str = request.GET.get('monthselect', None)
    datas = User.objects.filter(Q(id=admin_id) | Q(
        admin_id=admin_id)).order_by('empid')

    print(f"Selected Month String: {selected_month_str}")
    query = request.GET.get('search')
    print(query)
    count_user = User.objects.filter(
        Q(id=request.user.id) | Q(admin_id=request.user.id)).count()
    print("Count :", count_user)

    if selected_month_str is None:
        selected_month = current_month
        selected_year = current_year
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month

    users = User.objects.filter(Q(id=request.user.id) | Q(
        admin_id=request.user.id)).order_by('username')
    print(users)
    if query:
        users1 = User.objects.filter(Q(empid__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('username')
        users2 = User.objects.filter(Q(username__contains=query) & (
            Q(id=request.user.id) | Q(admin_id=request.user.id))).order_by('username')
        if users1 or users2:
            users = users1 | users2
        else:
            users = []
            # messages.info(request, 'No Records Found')

    absent_anomaly_punches = Punch.objects.filter(
        # Include empid in the search
        (Q(user__username__icontains=query) | Q(user__empid__icontains=query)) &
        (Q(user__id=admin_id) | Q(user__admin_id=admin_id)),
        status__in=['A', 'AN', 'HL', 'L'],
        date__date__month=selected_month,
        date__date__year=selected_year
    ).values('user__id', 'status', 'date')

    print(absent_anomaly_punches)
    combine_data = {}
    unique_anomalies = []

    for punch in absent_anomaly_punches:
        user_id = punch['user__id']
        status = punch['status']
        date = punch['date']

        print(f"Checking: user_id={user_id}, status={status}, date={date}")
        if not any(entry['user_id'] == user_id and entry['status'] == status and entry['date'].date() == date.date() for entry in unique_anomalies):
            unique_anomalies.append(
                {'user_id': user_id, 'status': status, 'date': date})

            print(f"Added: user_id={user_id}, status={status}, date={date}")
            print("Unique anomalies:", unique_anomalies)
            if user_id not in combine_data:
                user_details = User.objects.filter(id=user_id).first()
                combine_data[user_id] = {'empid': user_details.empid,
                                         'username': user_details.username, 'absent_dates': [],
                                         'absent_count': 0, 'anomaly_count': 0}

            if status == 'A':
                combine_data[user_id]['absent_dates'].append(date)
                combine_data[user_id]['absent_count'] += 1

            elif status == 'AN':

                combine_data[user_id]['anomaly_count'] += 1

    formatted_data = list(combine_data.values())
    leave_data = Leave.objects.filter(
        # Include empid in the search
        (Q(applicant_email__username__icontains=query) | Q(applicant_email__empid__icontains=query)) &
        (Q(applicant_email__id=admin_id) | Q(applicant_email__admin_id=admin_id)),
        status='Approved',
        strtDate__month=selected_month,
        strtDate__year=selected_year,
        endDate__month=selected_month,
        endDate__year=selected_year
    ).values('leavetyp', 'applicant_email__id', 'strtDate', 'endDate', 'admin_id', 'Days')
    print(leave_data)

    combine_leave = {}

    for leave in leave_data:
        applicant_email_id = leave['applicant_email__id']
        strtDate = leave['strtDate']
        endDate = leave['endDate']
        print(
            f"Processing leave for user {applicant_email_id} with dates {strtDate} to {endDate}")
        leavetyp = leave['leavetyp']
        Days = leave['Days']
        if applicant_email_id not in combine_leave:
            user_details = User.objects.filter(id=applicant_email_id).first()
            combine_leave[applicant_email_id] = {'empid': user_details.empid,
                                                 'username': user_details.username, 'leave_count': 0, 'leave_dates': []}

        if leavetyp == 'Loss Of Pay':
            combine_leave[applicant_email_id]['leave_dates'].append(strtDate)
            combine_leave[applicant_email_id]['leave_dates'].append(endDate)
            combine_leave[applicant_email_id]['leave_count'] += Days
    print(combine_leave)

    formatted_leave = list(combine_leave.values())
    page = request.GET.get('page', 1)
    paginator = Paginator(datas, 20)
    try:
        datas = paginator.page(page)
    except PageNotAnInteger:
        datas = paginator.page(1)
    except EmptyPage:
        datas = paginator.page(paginator.num_pages)
    x = {
        "k": k[0] if k.exists() else None,
        "data": c[0] if c.exists() else None,
        "users": users,
        "datas": datas,


        "formatted_data": formatted_data,
        "absent_anomaly_punches": absent_anomaly_punches,
        "leave_data": leave_data,
        "formatted_leave": formatted_leave,

        "query": query,
        "selected_month": selected_month,
        "selected_year": selected_year

    }

    return render(request, "index/runpayroll.html", x)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def payroll_leave(request):
    admin_id = request.user.id
    c = companyprofile.objects.filter(admin_id=admin_id)
    k = Myprofile.objects.filter(myuser__id=admin_id)
    today = datetime.now()
    query = request.GET.get('search')

    selected_month_str = request.GET.get('monthselect', None)
    print(f"Selected Month String: {selected_month_str}")
    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    data = Leave.objects.filter(Q(strtDate__year=selected_year, strtDate__month=selected_month)
                                & (Q(applicant_email__id=admin_id) |
                                   Q(applicant_email__admin_id=admin_id)) &
                                Q(applicant_email__status='Acttive'))

    if query:

        data = Leave.objects.filter(Q(strtDate__year=selected_year, strtDate__month=selected_month) &
                                    (Q(applicant_email__username__icontains=query) &
                                    (Q(applicant_email__id=admin_id) |
                                     Q(applicant_email__admin_id=admin_id))) &
                                    Q(applicant_email__status='Active')).order_by('username')
    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
    }
    return render(request, "index/payroll_leave.html", {"datas": data, "query": query, **x})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def calculate_salary(request):
    admin_id = request.user.id
    data = User.objects.filter(Q(admin_id=admin_id) | Q(id=admin_id))
    k = Uploadeddocs.objects.filter(type1="PAN Card")
    r = Leave.objects.filter(leavetyp="Loss Of Pay")
    current_date = datetime.now()
    print(current_date)
    current_month = current_date.month
    print(current_month)
    current_year = current_date.year

    query = request.GET.get('search')
    if query:
        datas_list1 = data.filter(Q(applicant_email__username__icontains=query) | Q(applicant_email__empid__icontains=query) &
                                  (Q(applicant_email__id=admin_id) | Q(applicant_email__admin_id=admin_id)))

        if datas_list1:

            data = datas_list1

        else:
            data = []
            # messages.info(request, 'No Records Found')

    count_user = User.objects.count()

    page = request.GET.get('page', 1)
    paginator = Paginator(data, 20)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    return render(request, "index/calculate_salary.html", {"datas": data, "kl": k, "r": r})


def get_user_by_punch(user_id_list,selected_year,selected_month):

    user_punch_counts = {}

    punchvalues = Punch.objects.filter(
            user_id__in=user_id_list, date__year=selected_year, date__month=selected_month).values("user_id").annotate(count=models.Count('id'))
    for punchdata in punchvalues:
        user_punch_counts[punchdata['user_id']] = punchdata['count']
    return user_punch_counts


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def pay_register(request):
    admin_id = request.user.id
    k = Myprofile.objects.only("id","image").filter(myuser__id=request.user.id)
    data = companyprofile.objects.filter(admin_id=admin_id)
    today = datetime.now()
    todaymonth = datetime.now().month
    todayyear = datetime.now().year

    selected_monthstr = request.GET.get('monthselect', None)
    if selected_monthstr is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(selected_monthstr, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_monthstr, '%B %Y')
        month_str = selected_date.strftime('%B')

    createddate = datetime(selected_year, selected_month, 1)
    cache_key = f'user_list_for_admin_payregister'
    cached_user_list = cache.get(cache_key)
    if cached_user_list is None:
        user_list = User.objects.select_related("wrklcn","department","designation").prefetch_related('uploadeddocs_set','bank_account_set').filter(Q(id=admin_id) | Q(admin_id=admin_id) & Q(status='Active')).only('id', 'empid','username', 'wrklcn__location', 'department__name', 'designation__name', 'datejoin').order_by('username')
        user_list = [
            user for user in user_list
            if datetime.strptime(parse_and_format_date(user.datejoin) if user.datejoin else datetime.now().strftime("%d %B %Y"), "%d %B %Y").replace(day=1) <= datetime(selected_year, selected_month, 1)
        ]
        cache.set(cache_key, user_list, timeout=1800)
    else:
        user_list = cached_user_list
    query = request.GET.get('search')
    if query:
        user_list = User.objects.select_related("wrklcn","department","designation").filter(Q(username__contains=query) & (Q(id=request.user.id) | Q(admin_id=request.user.id))).only('id', 'username', 'datejoin', 'wrklcn__location', 'department__name', 'designation__name').order_by('username')
        user_list = [
            user for user in user_list
            if datetime.strptime(user.datejoin, "%d %B %Y").replace(day=1) <= datetime(selected_year, selected_month, 1)
        ]

    user_id_list = [user.id for user in user_list]


    absent_counts = Punch.objects.filter(user_id__in=user_id_list, date__year=selected_year,
                                            date__month=selected_month, status='A', is_penalty_reverted=False).values("user_id").annotate(count=models.Count('user_id'))
    absent_counts_by_user = {item['user_id']: item['count'] for item in absent_counts}

    cache_key = f'wfo_count_by_user_month_wise{selected_year}{selected_month}'
    cached_wfo_count_by_user_month_wise = cache.get(cache_key)
    if cached_wfo_count_by_user_month_wise is None:
        wfo_counts = WFOCount.objects.filter(
            user_id__in=user_id_list,
            wfo_date__year=selected_year,
            wfo_date__month=selected_month
        ).values('user_id', 'wfocount') 
        wfo_counts_by_user = {item['user_id']: item['wfocount'] for item in wfo_counts}
        cache.set(cache_key, wfo_counts_by_user, timeout=3600)
    else:
        wfo_counts_by_user = cached_wfo_count_by_user_month_wise

    cache_key = f'all_user_punch{selected_year}{selected_month}'
    cached_all_user_punch_month_list = cache.get(cache_key)
    print('cache_key',cache_key)
    if cached_all_user_punch_month_list is None:
        print('cache not  hit')
        user_punch_data = {}
        all_user_punches = Punch.objects.filter(
                user_id__in=user_id_list,
                date__year=selected_year,
                date__month=selected_month
            ).order_by('date').values('status', 'date', 'user_id')
        for punch in all_user_punches:
            user_id = punch['user_id']
            if user_id not in user_punch_data:
                user_punch_data[user_id] = []
            user_punch_data[user_id].append(punch)
        cache.set(cache_key, user_punch_data, timeout=1800)
    else:
        print('cache hit')
        user_punch_data = cached_all_user_punch_month_list

    user_punch_counts = {user_id: len(punches) for user_id, punches in user_punch_data.items()}


    attendance_rules = AssignAttendanceRule.objects.filter(
                user_id_id__in=user_id_list).select_related('rules_applied')
    user_attendance_rules = {}
    for rule in attendance_rules:
        user_id = rule.user_id.id
        if user_id not in user_attendance_rules:
            user_attendance_rules[user_id] = [] 
        user_attendance_rules[user_id].append(rule)

    adhoc_datas = Adhoc.objects.filter(user_id_id__in=user_id_list, createddate__year=selected_year, createddate__month=selected_month).select_related('adhocearning', 'adhocdeduction')
    adhoc_datas_query_by_user = {}
    for adhoc in adhoc_datas:
        user_id = adhoc.user_id_id
        if user_id not in adhoc_datas_query_by_user:
            adhoc_datas_query_by_user[user_id] = [] 
        adhoc_datas_query_by_user[user_id].append(adhoc)

    lop_data = Runpayroll_lop.objects.filter(
            lop_date__year=selected_year,
            lop_date__month=selected_month
        ).values('user_id', 'lop_count')

    user_lop_counts = {}
    for lop in lop_data:
        user_id = lop['user_id']
        lop_count = lop['lop_count']
        if user_id not in user_lop_counts:
            user_lop_counts[user_id] = 0
        user_lop_counts[user_id] += lop_count

    user_data = []
    all_components = set()

    component_names = set()
    aggregated_data = []

    for user in user_list:
        userdata = {
            'user': user,
            'adhoc_data': adhoc_datas_query_by_user.get(user.id,None),
            'component_sum': {},
        }

        if adhoc_datas_query_by_user.get(user.id,None):
            for adhoc_entry in adhoc_datas_query_by_user.get(user.id,None):
                component_name = ''
                amount = 0
                if adhoc_entry.adhocearning:
                    component_name = adhoc_entry.adhocearning.component_name
                    amount += adhoc_entry.amount
                elif adhoc_entry.adhocdeduction:
                    component_name = adhoc_entry.adhocdeduction.component_name
                    amount -= adhoc_entry.amount

                if component_name in userdata['component_sum']:
                    userdata['component_sum'][component_name] += amount
                else:
                    userdata['component_sum'][component_name] = amount

                component_names.add(component_name)

        aggregated_data.append(userdata)

        selected_date = datetime(selected_year, selected_month, 1)
        assign_salarystructure = AssignSalaryStructure.objects.filter(Q(effective_date__year=selected_year, effective_date__month=selected_month) & Q(user_id=user.id)).order_by('effective_date').first()

        user_components = []
        user_amounts = []

        if not assign_salarystructure:
            nearest_date = AssignSalaryStructure.objects.filter(
                effective_date__lte=selected_date, user_id=user.id).order_by('-effective_date').first()

            if nearest_date:
                assign_salarystructure = nearest_date

        if assign_salarystructure:
            names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
            amounts = AssignSalaryStructureAmount.objects.only("amount").filter(salaryname__in=names)
            components = names.values_list('salarycomponent__componentname', flat=True)
            amounts = amounts.values_list('amount', flat=True)

            user_components.extend(components)
            user_amounts.extend(amounts)

            all_components.update(components)
        user_data.append({
            'user': user,
            'data': list(zip(user_components, user_amounts)),
        })

    component_names = sorted(list(component_names))

    unique_components = [x for x in all_components if x is not None]
    unique_components.sort()

    num_days = calendar.monthrange(selected_year, selected_month)[1]

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    existing_pay_registers = PayRegister.objects.filter(
        createddate__year=selected_year,
        createddate__month=selected_month
    ).values_list('user_id', flat=True)

    anomaly_user_punch_data = {}
    anomaly_punches = Punch.objects.only("id","user_id","first_clock_in_time","second_clock_in_time","first_clock_out_time","second_clock_out_time"
    ).filter(
        user_id__in=user_id_list,
        date__year=selected_year,
        date__month=selected_month,
        status="AN", is_penalty_reverted=False
    ).values('id', 'user_id', 'first_clock_out_time', 'second_clock_in_time', 'second_clock_out_time', "is_second_clocked_in")
    for punch in anomaly_punches:
        user_id = punch['user_id']
        if user_id not in anomaly_user_punch_data:
            anomaly_user_punch_data[user_id] = []
        anomaly_user_punch_data[user_id].append(punch)

    for user_info in user_data:
        user = user_info['user']
        total_user_amount = sum(
            [amount for component, amount in user_info['data']])
        per_day_amount = total_user_amount / num_days if num_days > 0 else 0

        absent_count = absent_counts_by_user.get(user.id,0)

        punchcount = user_punch_counts.get(user.id,0)
        missing_date_count = num_days - punchcount

        attendance_rule = user_attendance_rules.get(user.id,None)
        
        
        for att_rule in attendance_rule:
            rule_type = att_rule.rules_applied
            if rule_type:
                full_day_hours = rule_type.fullhours
                full_day_minutes = rule_type.fullminutes
                full_time = timedelta(
                    hours=full_day_hours, minutes=full_day_minutes)

                half_day_hours = rule_type.halfhours
                half_day_minutes = rule_type.halfminutes
                half_time = timedelta(
                    hours=half_day_hours, minutes=half_day_minutes)

                in_grace_period = rule_type.inGracePeriod
                out_grace_period = rule_type.outGracePeriod

                in_grace_timedelta = timedelta(
                    hours=in_grace_period.hour, minutes=in_grace_period.minute)
                out_grace_timedelta = timedelta(
                    hours=out_grace_period.hour, minutes=out_grace_period.minute)

                total_grace_period = in_grace_timedelta + out_grace_timedelta

                total_fullday_time = full_time + total_grace_period

                total_halfday_time = half_time + total_grace_period

        total_anomaly_count = 0
        for punch in anomaly_user_punch_data.get(user.id,[]):

            total_work_duration = timedelta()

            if punch.get("first_clock_in_time", None) and punch.get("first_clock_out_time", None) and punch.get("second_clock_in_time", None) and punch.get("second_clock_out_time", None) and punch.get("is_second_clocked_in", None):
                first_clock_in = datetime.combine(
                    datetime.today(), punch.get("first_clock_in_time", None))
                first_clock_out = datetime.combine(
                    datetime.today(), punch.get("first_clock_out_time", None))
                second_clock_in = datetime.combine(
                    datetime.today(), punch.get("second_clock_in_time", None))
                second_clock_out = datetime.combine(
                    datetime.today(), punch.get("second_clock_out_time", None))

                first_duration = first_clock_out - first_clock_in
                second_duration = second_clock_out - second_clock_in
                total_work_duration += first_duration + second_duration

            elif punch.get("first_clock_in_time", None) and punch.get("first_clock_out_time", None):
                first_clock_in = datetime.combine(
                    datetime.today(), punch.get("first_clock_in_time", None))
                first_clock_out = datetime.combine(
                    datetime.today(), punch.get("first_clock_out_time", None))
              
                first_duration = first_clock_out - first_clock_in
                total_work_duration += first_duration

            if total_work_duration > total_fullday_time:
                AN_count = 0.5
            elif total_work_duration < total_halfday_time:
                AN_count = 1.0
            else:
                AN_count = 0.5

            total_anomaly_count += AN_count

        absent_AN_count = absent_count + total_anomaly_count
        working_days = punchcount
        total_lop = absent_AN_count + missing_date_count

        total_lop += user_lop_counts.get(user.id,0)
        if working_days <= 0:
            totalamount_permonth = 0
        else:
            totalamount = per_day_amount * total_lop
            totalamount_permonth = round(total_user_amount - totalamount)

        leave_count = 0

        punches = user_punch_data.get(user.id)
        if punches is not None:
            for punch in punches:
                
                leave_data = Leave.objects.only("id","leavetyp").filter(
                        applicant_email_id=user.id,  
                        strtDate=punch['date'], 
                        status="Approved"     
                    ).first()       
                
                
                if punch['status'] == "H":
                    leave_count += 1
                elif punch['status'] == "L":
                    if leave_data:
                        if leave_data.leavetyp in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                            leave_count += 1
                elif punch['status'] == "HL":
                    if leave_data:
                        if leave_data.leavetyp not in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                            leave_count -= 0.5
        else:
            print("No punches marked for this user")
        wfocount = 0

        wfocount = wfo_counts_by_user.get(user.id,0)

        user_info['absent_AN_count'] = absent_AN_count
        user_info['total_working_days'] = working_days
        user_info['wfo_count'] = wfocount
        user_info['leave_count'] = leave_count
      
        if user.id not in existing_pay_registers:
            PayRegister.objects.create(user_id=user, netpay=totalamount_permonth, status='Pending', createddate=createddate)

    payreg_data = PayRegister.objects.only('status','user_id').select_related('user_id').filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=admin_id) | Q(user_id__admin_id=admin_id)))
    payout_data = PayoutStatus.objects.only('status','user_id').select_related('user_id').filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=admin_id) | Q(user_id__admin_id=admin_id)))
    salary_hold = PayActionStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=admin_id) | Q(user_id__admin_id=admin_id)))

    payreg_data_ids = list(payreg_data.values_list('user_id__id', flat=True))
    salary_hold_ids = list(salary_hold.values_list('user_id__id', flat=True))
    

    user_statuses = {}

    for user_id in payreg_data_ids:
        if user_id in salary_hold_ids:
            user_status = salary_hold.filter(
                user_id__id=user_id).values_list('actiontype', flat=True)
        else:
            user_status = payreg_data.filter(
                user_id__id=user_id).values_list('status', flat=True)

        user_statuses[user_id] = list(user_status)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        'user_data': user_data,
        'unique_components': unique_components,
        'num_days': num_days,
        'user_lop_counts': user_lop_counts,
        "month_str": month_str,
        "selected_year": selected_year,
        'user_statuses': user_statuses,
    }

    return render(request, "index/pay_register.html", 
                  {'user_list': user_list, 
                   'payreg_data': payreg_data, 
                   'payout_data': payout_data, 
                   'salary_hold': salary_hold, 
                   'query': query, 
                   'aggregated_data': aggregated_data, 
                   'component_names': component_names, **x})


def payregister_statuschange(request):
    today = datetime.now()
    todaymonth = datetime.now().month
    todayyear = datetime.now().year

    user_data = []
    all_components = set()

    component_names = set()
    aggregated_data = []

    if request.method == 'POST':
        monthselect = request.POST.get("monthselect")
        yearselect = request.POST.get("yearselect")
        datetime_object = datetime.strptime(monthselect, "%B")
        month_numeric = datetime_object.month
        createddate = datetime(int(yearselect), month_numeric, 1)
        # print("monthselect , yearselect :",month_numeric, yearselect, createddate)

        selected_employee_ids = request.POST.getlist('selected_employees', [])
        
        selected_users = User.objects.only('id','email','username','admin_id').filter(id__in=selected_employee_ids)
        
        action_type = request.POST.get('actionType')

        query = Q()
        for user in selected_users:
            query |= Q(user_id=user, createddate__year=yearselect, createddate__month=month_numeric)
        # Fetch all PayRegister objects in a single query.
        payregs = PayRegister.objects.only('id','status').filter(query)
        # Create a dictionary for quick lookup
        payreg_dict = {payreg.user_id: payreg for payreg in payregs}

        user_punch_data = {}
        all_user_punches = Punch.objects.filter(
                user_id__in=selected_employee_ids,
                date__year=yearselect,
                date__month=month_numeric
            ).order_by('date').values('status', 'date', 'user_id')
        for punch in all_user_punches:
            user_id = punch['user_id']
            if user_id not in user_punch_data:
                user_punch_data[user_id] = []
            user_punch_data[user_id].append(punch)
        # print("user_punch_data",user_punch_data)
        user_punch_counts = {user_id: len(punches) for user_id, punches in user_punch_data.items()}
        # print("user_punch_counts", user_punch_counts)
        ###for taking count of punches for userwise on key of user_id
        # user_punch_counts = {user_id: len(punches) for user_id, punches in user_punch_data.items()}
        # print("user_punch_counts", user_punch_counts)
        wfo_counts = WFOCount.objects.filter(
            user_id_id__in=selected_employee_ids,
            wfo_date__year=yearselect,
            wfo_date__month=month_numeric,
        ).values('user_id_id').annotate(total_wfocount=Sum('wfocount'))
        wfo_count_dict = {item['user_id_id']: item['total_wfocount'] for item in wfo_counts}

        absent_counts = Punch.objects.filter(
                user_id__in=selected_employee_ids, 
                date__year=yearselect, 
                date__month=month_numeric, 
                status='A', is_penalty_reverted=False).values('user_id').annotate(absent_count=Count('id'))
        absent_count_dict = {item['user_id']: item['absent_count'] for item in absent_counts}
        
        ##attendance rule
        attendance_rules = AssignAttendanceRule.objects.filter(
                user_id_id__in=selected_employee_ids).select_related('rules_applied')
        user_attendance_rules = {}

        for rule in attendance_rules:
            user_id = rule.user_id.id
            if user_id not in user_attendance_rules:
                user_attendance_rules[user_id] = []  # Create the list only once

            user_attendance_rules[user_id].append(rule)

        gross_salary_component_caches = cache.get('gross_salary_component_cache')
        if gross_salary_component_caches is None:
            gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
            cache.set('gross_salary_component_cache', gross_salary_component, timeout=36000)
        else:
            gross_salary_component = gross_salary_component_caches

        work_from_office_components = cache.get('work_from_office_component')
        if work_from_office_components is None:
            work_from_office_component = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance", Parentcomponentname__componentname__iexact="Gross Salary").first()
            cache.set('work_from_office_component', work_from_office_component, timeout=36000)
        else:
            work_from_office_component = work_from_office_components
        # print("gross_salary_component ; work_from_office_component : ", gross_salary_component, work_from_office_component)
        net_salary_components = cache.get('net_salary_component')
        if net_salary_components is None:
            net_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="Net Salary")
            cache.set('net_salary_component', net_salary_component, timeout=36000)
        else:
            net_salary_component = net_salary_components
        

        anomaly_user_punch_data = {}
        anomaly_punches = Punch.objects.only("id","user_id","first_clock_in_time","second_clock_in_time","first_clock_out_time","second_clock_out_time"
        ).filter(
            user_id__in=selected_employee_ids,
            date__year=yearselect,
            date__month=month_numeric,
            status="AN", is_penalty_reverted=False
        ).values('id', 'user_id', 'first_clock_out_time', 'second_clock_in_time', 'second_clock_out_time', "is_second_clocked_in")
        for punch in anomaly_punches:
            user_id = punch['user_id']
            if user_id not in anomaly_user_punch_data:
                anomaly_user_punch_data[user_id] = []
            anomaly_user_punch_data[user_id].append(punch)
        for user in selected_users:
            # payreg = PayRegister.objects.get(user_id=user, createddate__year=yearselect, createddate__month=month_numeric)
            payreg = payreg_dict.get(user,None)

            # print("action_type :", action_type)

            # print("company_details :", user)
            company_details = companyprofile.objects.filter(Q(admin_id=user.admin_id) | Q(admin_id=user.id)).first()
            # reg_address = registeredaddress.objects.filter(Q(admin_id=user.admin_id) | Q(admin_id=user.id))

            wfocount = 0
            leave_count = 0
            total_WFO_amount = 0  
            
            # punches = Punch.objects.filter(
            #         user=user,
            #         date__year=yearselect,
            #         date__month=month_numeric
            #     ).order_by('date')
            # print("punches payreg_sts ## : ", punches)
            punches = user_punch_data.get(user.id,None)
            if punches:
                for punch in user_punch_data.get(user.id,None):
                    # if punch.status in ["H", "L"]:
                    leave_data = Leave.objects.only('id','leavetyp').filter(
                            applicant_email=user,  
                            strtDate=punch.get('date',None),    
                            status="Approved"      
                        ).first()       
                    # print(f"Leave data for {user.email} on {punch.date}: ", leave_data)
                    
                    if punch.get('status',None) == "H":
                        leave_count += 1
                    elif punch.get('status',None) == "L":
                        if leave_data:
                            if leave_data.leavetyp in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                                leave_count += 1
                    elif punch.get('status',None) == "HL":
                        if leave_data:
                            # print("ccccccccccccccccccccccccccccccccccccc  ")
                            if leave_data.leavetyp not in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                                leave_count -= 0.5

            wfocount = wfo_count_dict.get(user.id,0)

            if action_type == 'payslip' and payreg.status == "Completed":
                
                # print("From Email: ", settings.EMAIL_HOST_USER)
                # print("To Email: " , user.email)
                to = [user.email]
                subject = f'Payslip - {monthselect} {yearselect}'
                html_body = render_to_string('index/email_payslip.html', {'user': user, 'company_details': company_details, 'monthselect': monthselect, 'yearselect': yearselect})
                msg = EmailMultiAlternatives(subject=subject, from_email=settings.EMAIL_HOST_USER, to=to)
                msg.attach_alternative(html_body, "text/html")
                msg.send()

                payreg.status = "Payslip Generated"
                payreg.save()

            elif action_type == 'revert':
                payreg.status = "Pending"
                payreg.save()

            elif action_type == 'rupayroll' and payreg.status == "Calculated":
                payreg.status = "Completed"
                payreg.save()
            elif action_type == 'calculate' and (payreg.status == "Pending" or payreg.status == "Calculated" or payreg.status == "Pending Calculation" or payreg.status == "Payslip Downloaded"):
                # print("CALCULATE")

                adhoc_data = Adhoc.objects.filter(user_id=user.id, createddate__year=yearselect, createddate__month=month_numeric).select_related('adhocearning', 'adhocdeduction')
                userdata = {
                    'user': user,
                    'adhoc_data': adhoc_data,
                    'component_sum': {},
                }

                # print("############ adhoc_data ############ :", adhoc_data)
                for adhoc_entry in adhoc_data:
                    component_name = ''
                    amount = 0                    

                    if adhoc_entry.adhocearning and adhoc_entry.adhocearning.component_name != "Work from Office allowance":
                        component_name = adhoc_entry.adhocearning.component_name
                        amount += adhoc_entry.amount
                        # print("amount :", component_name, amount)
                    
                    elif adhoc_entry.adhocdeduction:
                        component_name = adhoc_entry.adhocdeduction.component_name
                        amount -= adhoc_entry.amount
                    
                    if component_name in userdata['component_sum']:
                        userdata['component_sum'][component_name] += amount
                    else:
                        userdata['component_sum'][component_name] = amount
                    component_names.add(component_name)
                aggregated_data.append(userdata)

                selected_date = datetime(int(yearselect), month_numeric, 1)
                # print("selected_date :", selected_date)
                assign_salarystructure = AssignSalaryStructure.objects.only('id').filter(Q(effective_date__year=yearselect, effective_date__month=month_numeric) & Q(user_id=user.id)).order_by('effective_date').first()

                user_components = []
                user_amounts = []

                if not assign_salarystructure:
                    nearest_date = AssignSalaryStructure.objects.only('id').filter(effective_date__lte=selected_date, user_id=user.id).order_by('-effective_date').first()

                    if nearest_date:
                        assign_salarystructure = nearest_date
                
                gross_salary_amount = 0
                work_from_office_allowance_amount = 0
                total_net_salary = 0

                if assign_salarystructure:
                    # print("assign_salarystructure :", assign_salarystructure)
                    names = AssignSalaryStructureName.objects.only('id','salarycomponent').filter(salaryrule=assign_salarystructure)
                    amounts = AssignSalaryStructureAmount.objects.only('id','amount').filter(salaryname__in=names)
                    # print("names ; amounts 14621 :", names, amounts)

                    

                    if gross_salary_component:
                        gross_salary_amount = amounts.filter(salaryname__salarycomponent=gross_salary_component).first()
                    if work_from_office_component:
                        work_from_office_allowance_amount = amounts.filter(salaryname__salarycomponent=work_from_office_component).first()
                    
                    for netsalry in net_salary_component:
                        net_salary = amounts.filter(salaryname__salarycomponent=netsalry)
                        total_net_salary += net_salary.aggregate(total=models.Sum('amount'))['total'] or 0

                    # print("Total Net Salary:", total_net_salary)

                    gross_salary_amount = gross_salary_amount.amount if gross_salary_amount else 0
                    work_from_office_allowance_amount = work_from_office_allowance_amount.amount if work_from_office_allowance_amount else 0
                    # print("Gross Salary Amount:", gross_salary_amount)
                    # print("Work From Home Allowance Amount:", work_from_office_allowance_amount)

                    total_gross_salary = gross_salary_amount - work_from_office_allowance_amount
                    # print("total_gross_salary : ", total_gross_salary)

                    components = names.values_list('salarycomponent__componentname', flat=True)
                    amounts = amounts.values_list('amount', flat=True)
                    percentages = names.values_list('salarycomponent__percent', flat=True)

                    user_components.extend(components)
                    user_amounts.extend(amounts)
                    user_percentages = list(percentages)
                    print("user_percentages",user_percentages)

                    all_components.update(components)
                
                # print("######################## total_WFO_amount :", total_WFO_amount, wfocount)

                user_data.append({
                    'user': user,
                    'data': list(zip(user_components, user_amounts, user_percentages)),
                    'payreg': payreg,
                    'wfo_count': wfocount,
                    'total_WFO_amount': total_WFO_amount,
                    'total_gross_salary': total_gross_salary,
                    'work_from_office_allowance_amount': work_from_office_allowance_amount,
                    'total_net_salary':total_net_salary,
                })

        component_names = sorted(list(component_names))
        # print("component_names:", component_names)
        # print("aggregated_data:", aggregated_data)
        unique_components = [x for x in all_components if x is not None]
        unique_components.sort()

        num_days = calendar.monthrange(int(yearselect), month_numeric)[1]
        # print("working_days_count:", num_days)
        # lopcount = Runpayroll_lop.objects.filter(lop_date__year=yearselect, lop_date__month=month_numeric)
        lop_data = Runpayroll_lop.objects.filter(
            lop_date__year=yearselect,
            lop_date__month=month_numeric
        ).values('user_id', 'lop_count')

        # Create a dictionary to store lop counts for each user.
        user_lop_counts = {}
        for lop in lop_data:
            user_id = lop['user_id']
            lop_count = lop['lop_count']
            if user_id not in user_lop_counts:
                user_lop_counts[user_id] = 0
            user_lop_counts[user_id] += lop_count

        total_fullday_time = timedelta()
        total_halfday_time = timedelta()

        year_select = int(yearselect)
        num_days = calendar.monthrange(year_select, month_numeric)[1]
        first_day_of_month = datetime(year_select, month_numeric, 1)
        if month_numeric == 12: 
            next_month = datetime(year_select + 1, 1, 1)
            # print("next_month 1 :", next_month, year_select)
        else:
            next_month = datetime(year_select, month_numeric + 1, 1)
            # print("next_month 2 :", next_month)

        day_count = 0
        current_day = first_day_of_month
        while current_day < next_month:
            if current_day.weekday() != 6: 
                day_count += 1
            current_day += timedelta(days=1)
        count_sundays = num_days - day_count
        # print("day_count ############ :", day_count, num_days, count_sundays)

        
       
        for user_info in user_data:
        
            basic_salary_percentage = 0
            epf_percentage = 0
            esi_percentage = 0
            basic_salary_amount = 0
            epf_amount = 0
            esi_amount = 0
            profisional_tax_amount = 0
            insurance_amount = 0
            totalnetsalary = 0

            user = user_info['user']
            payreg = user_info['payreg']
            total_user_amount = user_info['total_gross_salary']
            print("user :", user, payreg, total_user_amount)
            # total_user_amount = sum([amount for component, amount in user_info['data']])
            # print("total_user_amount : ", total_user_amount)
            per_day_amount = total_user_amount / num_days if num_days > 0 else 0
            # print("per_day_amount :", per_day_amount)
            user_component_sum = {}
            for aggregated_user_data in aggregated_data:
                if aggregated_user_data['user'] == user:
                    user_component_sum = aggregated_user_data['component_sum']
                    break
            # print("sum(user_component_sum.values()):",sum(user_component_sum.values()))

            # lopcount_user = lopcount.filter(user_id=user.id)

            # absent_count = Punch.objects.filter(
            #     user__id=user.id, date__year=yearselect, date__month=month_numeric, status='A', is_penalty_reverted=False).count()
            absent_count = absent_count_dict.get(user.id,0)

            punchcount = user_punch_counts.get(user.id,0)
            # print("punchcount punchcount:", punchcount,user_punch_counts[user.id])

            missing_date_count = num_days - punchcount
            # print("missing_date_count :", missing_date_count)

            # attendance_rule = AssignAttendanceRule.objects.filter(
            #     user_id_id=user.id)
            attendance_rule = user_attendance_rules.get(user.id,None)
            print('attendance_rule',attendance_rule)
            # print('attendance_rule2',user_attendance_rules.get(user.id))
            for att_rule in attendance_rule:
                rule_type = att_rule.rules_applied
                # print("rule_type :", rule_type, )
                if rule_type:
                    full_day_hours = rule_type.fullhours
                    full_day_minutes = rule_type.fullminutes
                    full_time = timedelta(
                        hours=full_day_hours, minutes=full_day_minutes)

                    half_day_hours = rule_type.halfhours
                    half_day_minutes = rule_type.halfminutes
                    half_time = timedelta(
                        hours=half_day_hours, minutes=half_day_minutes)

                    # print("Full Day Hours:", full_day_hours,
                        #   full_day_minutes, full_time)
                    # print("Half Day Hours:", half_day_hours,
                        #   half_day_minutes, half_time)

                    in_grace_period = rule_type.inGracePeriod
                    out_grace_period = rule_type.outGracePeriod
                    # print("Grace period:", in_grace_period, out_grace_period)

                    in_grace_timedelta = timedelta(
                        hours=in_grace_period.hour, minutes=in_grace_period.minute)
                    out_grace_timedelta = timedelta(
                        hours=out_grace_period.hour, minutes=out_grace_period.minute)

                    total_grace_period = in_grace_timedelta + out_grace_timedelta
                    # print("Total Grace period:", total_grace_period)

                    total_fullday_time = full_time + total_grace_period
                    # print("Total Time:", total_fullday_time)

                    total_halfday_time = half_time + total_grace_period
                    # print("total_halfday_time :", total_halfday_time)

            num_days = (createddate.replace(month=createddate.month % 12 + 1, day=1) - timedelta(days=1)).day

            # punches = Punch.objects.filter(
            #     user=user,
            #     date__year=yearselect,
            #     date__month=month_numeric,
            #     status="AN", is_penalty_reverted=False
            # )

            # print("punch onj :", punches)
            print("punch onj 2:", anomaly_user_punch_data.get(user.id),None)
            total_anomaly_count = 0
            GrandTotalAmount = 0
            for punch in anomaly_user_punch_data.get(user.id,[]):
                total_work_duration = timedelta()

                if punch.get('first_clock_in_time',None) and punch.get('first_clock_out_time',None) and punch.get('second_clock_in_time',None) and punch.get('second_clock_out_time',None):
                    first_clock_in = datetime.combine(
                        datetime.today(), punch.get('first_clock_in_time',None))
                    first_clock_out = datetime.combine(
                        datetime.today(), punch.get('first_clock_out_time',None))
                    second_clock_in = datetime.combine(
                        datetime.today(), punch.get('second_clock_in_time',None))
                    second_clock_out = datetime.combine(
                        datetime.today(), punch.get('second_clock_out_time',None))

                    first_duration = first_clock_out - first_clock_in
                    second_duration = second_clock_out - second_clock_in
                    total_work_duration += first_duration + second_duration

                elif punch.get('first_clock_in_time',None) and punch.get('first_clock_out_time',None):
                    first_clock_in = datetime.combine(
                        datetime.today(), punch.get('first_clock_in_time',None))
                    first_clock_out = datetime.combine(
                        datetime.today(), punch.get('first_clock_out_time',None))
                    # print("first_clock_in ; first_clock_out : ",
                        #   first_clock_in, first_clock_out)
                    first_duration = first_clock_out - first_clock_in
                    # print("first_duration : ", first_duration)
                    # total_work_duration += first_duration

                # print("total_work_duration :", total_work_duration)

                if total_work_duration > total_fullday_time:
                    AN_count = 0.5
                elif total_work_duration < total_halfday_time:
                    AN_count = 1.0
                else:
                    AN_count = 0.5

                total_anomaly_count += AN_count
            # print("Total AN count for the selected month:", total_anomaly_count)

            absent_AN_count = absent_count + total_anomaly_count
            # print("absent_AN_count : ", absent_AN_count)
            working_days = punchcount
            # print("working_days :", working_days)
            print("absent_AN_count : missing_date_count",absent_AN_count,missing_date_count,absent_count,total_anomaly_count)
            total_lop = absent_AN_count + missing_date_count
            
            # for lop_instance in lopcount_user:
                # print("lop_instance :", lop_instance.lop_count)
                # total_lop += lop_instance.lop_count
            total_lop += user_lop_counts.get(user.id,0)
            # print("total_lop",total_lop,user_lop_counts.get(user.id,0))
            if working_days <= 0:
                total_amount = 0
            else:
                totalamount = per_day_amount * total_lop
                
                totalamount_permonth = total_user_amount - totalamount
                total_amount = round(totalamount_permonth)
                
                # total_amount = round(totalamount_permonth + sum(user_component_sum.values()))
      
            total_WFO_amount = user_info['work_from_office_allowance_amount']
            # total_netsalary = user_info['total_net_salary']
            total_wfocount = user_info["wfo_count"]
            # print("total_amount:", total_amount, total_netsalary, total_wfocount)
            perday_WFOamount = total_WFO_amount / day_count
            # print("perday_WFOamount:", perday_WFOamount)
            total_WFOamount = perday_WFOamount * total_wfocount
            # print("total_WFOamount:", total_WFOamount)
            grand_total_amount = total_amount + total_WFOamount

            # print(f"For user {user.email}: total_WFO_amount = {total_WFO_amount}, grand_total_amount = {grand_total_amount}")

            for component_name, amount, percentage in user_info['data']:
                # print(f"Component Name: {component_name}, Percentage: {percentage}%")
                if component_name.lower() == "basic salary":
                    basic_salary_percentage = percentage
                if component_name.lower() == "epf employee":
                    epf_percentage = percentage
                if component_name.lower() == "esi employee":
                    esi_percentage = percentage
                if component_name.lower() == "professional tax":
                    profisional_tax_amount = 167 if grand_total_amount < 22000 else 208
                    # print("profisional_tax_amount :", profisional_tax_amount)

                if component_name.lower() == "insurance":
                    insurance_amount = 0 if grand_total_amount <= 25000 else 245
                    
            if basic_salary_percentage:
                basic_salary_amount = (grand_total_amount * basic_salary_percentage) / 100
                # print("Basic Salary Amount :", basic_salary_amount, basic_salary_percentage)
            if epf_percentage:
                epf_amount = round((basic_salary_amount * epf_percentage) / 100)
                # print("EPF Amount :", epf_amount, epf_percentage)
            if esi_percentage:
                esi_amount = round((grand_total_amount * esi_percentage) / 100)
                # print("ESI Amount :", esi_amount, esi_percentage)

            # print("total_WFOamount : ", total_wfocount, total_WFOamount,"grand_total_amount : ", grand_total_amount)

            totalnetsalary = profisional_tax_amount + epf_amount + esi_amount + insurance_amount
            
            if user.id == 53 or user.id == 8 or user.id == 82 or user.id == 124 or user.id == 24:
                print("username:",user.username)
                print("profisional_tax_amount + epf_amount + esi_amount + insurance_amount: ", profisional_tax_amount, epf_amount, esi_amount, insurance_amount, "totalnetsalary :", totalnetsalary)

            GrandTotalAmount = grand_total_amount + sum(user_component_sum.values())
            if user.id == 53 or user.id == 8 or user.id == 82 or user.id == 124 or user.id == 24:
                print("GrandTotalAmount",GrandTotalAmount)
            grandtotalamount = round(GrandTotalAmount - totalnetsalary)
            if user.id == 53 or user.id == 8 or user.id == 82 or user.id == 124 or user.id == 24:
                    print("grandtotalamount : ", grandtotalamount, "GrandTotalAmount : ", GrandTotalAmount)
                    print(user_info['data'])
            
            user_info['total_amount'] = grandtotalamount
            # print("user_info['total_amount']:", user_info['total_amount'])

            payreg.status = 'Calculated'
            payreg.netpay = grandtotalamount
            payreg.save()

    return redirect(reverse("pay_register") + '?monthselect=' + monthselect + '+' + yearselect)

def download_payregister(request):
    today = datetime.now()
    todaymonth = datetime.now().month
    todayyear = datetime.now().year
    print("todaymonth; todayyear:", todaymonth, todayyear)

    payslip_pdfs = []
    user_infos = []

    if request.method == 'POST':
        monthselect = request.POST.get("monthselect")
        yearselect = request.POST.get("yearselect")
        datetime_object = datetime.strptime(monthselect, "%B")
        month_numeric = datetime_object.month
        createddate = datetime(int(yearselect), month_numeric, 1)
        print("monthselect ; yearselect :",month_numeric, yearselect, createddate)

        selected_employee_ids = request.POST.getlist('selected_employees')
        selected_users = User.objects.filter(id__in=selected_employee_ids)
        print("selected_users :", selected_users)
        action_type = request.POST.get('actionType')
        for user in selected_users:
            payreg = PayRegister.objects.get(user_id=user, createddate__year=yearselect, createddate__month=month_numeric)
            print("payreg:", payreg.status)

            print("action_type :", action_type)

            print("company_details :", user.admin_id, user)

            reg_address = registeredaddress.objects.filter(Q(admin_id=user.admin_id) | Q(admin_id=user.id))
            
            print("Reg address: ", reg_address)

            if action_type == 'download' and (payreg.status == "Payslip Generated" or payreg.status == "Payslip Downloaded"):

                user_info = {'username': user.username,'empid': user.empid, 'netpay': payreg.netpay}
                print("user_info 15151:", user_info)

                payslip_pdf = generate_payslip_pdf(monthselect, yearselect, reg_address, payreg)

                # send_payslip_email(user.email, pdf_buffer)
                print("payslip_pdf",payslip_pdf)
                payslip_pdfs.append(payslip_pdf)
                user_infos.append(user_info)

                payreg.status = "Payslip Downloaded"
                payreg.save()

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for user_info, pdf_data in zip(user_infos, payslip_pdfs):
                # Create a folder for payslips
                payslip_folder = f'pay_slip_{monthselect}_{yearselect}'
                zip_file.writestr(f'{payslip_folder}/{user_info["username"]}_{user_info["empid"]}.pdf', pdf_data)
        # Prepare the HTTP response
        response = HttpResponse(zip_buffer.getvalue(),content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="payslips.zip"'
        return response

    return redirect(reverse("pay_register") + '?monthselect=' + monthselect + '+' + yearselect)

def generate_payslip_pdf(monthselect, yearselect, reg_address, payreg):

    print("Called generate_payslip_pdf....................")
    buffer = BytesIO()
    custom_width = 700
    custom_height = 1000
    doc = SimpleDocTemplate(buffer, pagesize=(custom_width, custom_height), topMargin=35, title="Payslip")
    elements = []

    font_path = str(Path(settings.BASE_DIR) / 'app1'/'arialfont'/'arial.ttf')
    font_path2 = str(Path(settings.BASE_DIR) / 'app1'/'arialfont'/'arialbd.ttf')
    pdfmetrics.registerFont(TTFont('Arial', font_path))
    pdfmetrics.registerFont(TTFont('Arial-Bold', font_path2))
    font_style = 'Arial'
    bold_font_style = 'Arial-Bold'

    company_name_style = ParagraphStyle(
        name='CompanyName',
        fontName=bold_font_style,
        fontSize=16,
        alignment=TA_CENTER,
    )

    address_style = ParagraphStyle(
        name='CompanyAddress',
        fontSize=11,
        textColor='black',
        fontName=font_style,
        leading=14,
        alignment=TA_CENTER,
    )

    cmp_logo = ''
    company_name = ''
    company_name_paragraph = None
    company_logo = None
    company_address_para = None

    if payreg.user_id.company_type:
        cmp_logo = payreg.user_id.company_type.logo or ''
        company_logo = Image(cmp_logo, width=2*inch, height=1*inch)
        company_name = payreg.user_id.company_type.registeredcompanyname
        print("company_name : ", company_name)
    company_name_paragraph = Paragraph(company_name, company_name_style)

    company_address_str = ''
    
    print("reg_address: ", reg_address)
    for regaddress in reg_address:
        if regaddress:
            company_address_str = f"{regaddress.regofficeaddress}, {regaddress.regdistrict}, {regaddress.regstate}, {regaddress.regcountry} - {regaddress.regpincode}"
            company_address_para = Paragraph(
                company_address_str, address_style)

    data = [
        [company_logo, company_name_paragraph, ''],
        ['', company_address_para, ''],
        ['', f'Payslip for the Month of {monthselect}, {yearselect}', ''],

    ]

    colWidths = [150, 400, 100]
    rowHeights = [20, 50, 50]

    table = Table(data, colWidths=colWidths, rowHeights=rowHeights)
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, -1), (-1, -1), 13),
        ('FONTNAME', (0, -1), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        # ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))

    desg = ''
    dept = ''
    loc = ''
    bankname = ''
    esa = ''
    pfnum = ''
    acno = ''
    idno = ''
    earning_amount = 0
    deduction_amount = 0
    lopcount = 0

    datetime_object = datetime.strptime(monthselect, "%B")
    month_numeric = datetime_object.month
    num_days = calendar.monthrange(int(yearselect), month_numeric)[1]

    print("payreg :", payreg)

    user_id = payreg.user_id.id
    username = payreg.user_id.username
    user_email = payreg.user_id.email
    empid = payreg.user_id.empid
    desg = payreg.user_id.designation.name
    dept = payreg.user_id.department.name
    loc = payreg.user_id.wrklcn.location

    bank_details = Bank_account.objects.filter(myuser_11=user_id).first()
    if bank_details:
        bankname = bank_details.bank_name
        acno = bank_details.account_number
        pfnum = bank_details.pfnum
        esa = bank_details.esa
    uploadeddocs = Uploadeddocs.objects.filter(myuser=user_id, type1="PAN Card").first()
    if uploadeddocs:
        idno = uploadeddocs.id_no
    
    selected_date = datetime(int(yearselect), month_numeric, 1)
    print("selected_date :", selected_date)
    
    assign_salarystructure = AssignSalaryStructure.objects.filter(user_id=user_id, effective_date__month=month_numeric, effective_date__year=yearselect).order_by('effective_date').first()
    print("assignsalary :", assign_salarystructure)
    if not assign_salarystructure:
        nearest_date = AssignSalaryStructure.objects.filter(effective_date__lte=selected_date, user_id=user_id).order_by('-effective_date').first()
        if nearest_date:
            assign_salarystructure = nearest_date

    gross_salary_amount = 0
    work_from_office_allowance_amount = 0
    total_net_salary = 0
    total_ctc_salary = 0
    wfocount = 0
    leave_count = 0
    totalNet_Salary = 0
    total_gross_salary = 0
    net_total = 0
    grossamount = 0

    if assign_salarystructure:
        print("assign_salarystructure :", assign_salarystructure)
        
        gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
        work_from_office_component = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance", Parentcomponentname__componentname__iexact="Gross Salary").first()
        print("gross_salary_component ; work_from_office_component : ", gross_salary_component, work_from_office_component)
        net_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="Net Salary")
        print("net_salary_component : ", net_salary_component)
        ctc_salary_component = SalaryComponent.objects.filter(Parentcomponentname__componentname__iexact="CTC")
        print("ctc_salary_component : ", ctc_salary_component)

        name = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
        amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=name)
        print("name ; amount 1st :", name, amount)
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname=gross_salary_component)
        amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)
        print("names ; amounts  :", names, amounts)

        net_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="Net Salary")
        net_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=net_names)
        print("names ; amounts  :", net_names, net_amounts)

        ctc_names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure,salarycomponent__Parentcomponentname__componentname__iexact="CTC").exclude(salarycomponent__componentname__icontains="professional tax")
        ctc_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=ctc_names)
        print("ctc_names ; ctc_amounts  :", ctc_names, ctc_amounts)

        if gross_salary_component:
            gross_amount = amount.filter(salaryname__salarycomponent=gross_salary_component).first()
            gross_salary_amount = gross_amount.amount if gross_amount else 0            
            
        if work_from_office_component:
            work_amount = amounts.filter(salaryname__salarycomponent=work_from_office_component).first()
            work_from_office_allowance_amount = work_amount.amount if work_amount else 0
        
        for netsalry in net_salary_component:
            net_salary = amount.filter(salaryname__salarycomponent=netsalry)
            total_net_salary += net_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total Net Salary:", total_net_salary)

        for ctcsalry in ctc_salary_component:
            if ctcsalry.componentname.lower() != 'professional tax':
                ctc_salary = amount.filter(salaryname__salarycomponent=ctcsalry)
                total_ctc_salary += ctc_salary.aggregate(total=models.Sum('amount'))['total'] or 0
        print("Total CTC Salary:", total_ctc_salary)

        total_gross_salary = gross_salary_amount - work_from_office_allowance_amount
        print("total_gross_salary : ", total_gross_salary)

        # zipped_data = zip_longest(names, amounts)
        # assign_data.append({
        #     'rule': rule,
        #     'zipped_data': zipped_data,
        # })

        # zippeddata = zip_longest(net_names, net_amounts)
        # assigndata.append({
        #     'rule': rule,
        #     'zippeddata': zippeddata,
        # })

        # ctc_zippeddata = zip_longest(ctc_names, ctc_amounts)
        # ctc_assigndata.append({
        #     'rule': rule,
        #     'ctc_zippeddata': ctc_zippeddata,
        # })
 

    punch_obj = Punch.objects.filter(user__id=user_id,date__year=yearselect,date__month=month_numeric)
    print("punch_obj : ", punch_obj)

    for punch in punch_obj:
                    
        leave_data = Leave.objects.filter(
                applicant_email=user_id,  
                strtDate=punch.date,    
                status="Approved"      
            ).first()       
        print(f"Leave data for {user_email} on {punch.date}: ", leave_data)
        
        if punch.status == "H":
            leave_count += 1
        elif punch.status == "L":
            if leave_data:
                if leave_data.leavetyp != "Loss Of Pay":
                    leave_count += 1
        elif punch.status == "HL":
            if leave_data:
                print("ccccccccccccccccccccccccccccccccccccc  ")
                if leave_data.leavetyp == "Loss Of Pay":
                    leave_count -= 0.5
                    print("KKKKKKKKKKKKKKKKKKKKK", leave_count)
                
    wfo_count = WFOCount.objects.filter(user_id=user_id, wfo_date__year=yearselect, wfo_date__month=month_numeric)
    print("wfo_count :", wfo_count, "month_numeric , monthselect:" , month_numeric, monthselect)
    for i in wfo_count:
        wfocount = i.wfocount
        print("wfocount : ", wfocount)
    year_select = int(yearselect)
    num_days = calendar.monthrange(year_select, month_numeric)[1]
    first_day_of_month = datetime(year_select, month_numeric, 1)
    if month_numeric == 12: 
        next_month = datetime(year_select + 1, 1, 1)
        print("next_month 1 :", next_month, year_select)
    else:
        next_month = datetime(year_select, month_numeric + 1, 1)
        print("next_month 2 :", next_month)

    day_count = 0
    current_day = first_day_of_month
    while current_day < next_month:
        if current_day.weekday() != 6: 
            day_count += 1
        current_day += timedelta(days=1)
    count_sundays = num_days - day_count
    print("day_count ############ :", day_count, num_days, count_sundays)

    print("work_from_office_allowance_amount:", work_from_office_allowance_amount)
    perday_WFOamount = work_from_office_allowance_amount / day_count
    total_WFOamount = perday_WFOamount * wfocount
    print("total_WFOamount : ", wfocount, total_WFOamount)
    
    WFOamount = round(work_from_office_allowance_amount - total_WFOamount) #This amount add to the deduction
    print("WFOamount :", WFOamount)

    adhoc_data = Adhoc.objects.filter(user_id=user_id, createddate__year=yearselect,
                                      createddate__month=month_numeric).select_related('adhocearning', 'adhocdeduction')
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            earning_amount += adhoc_entry.amount
        elif adhoc_entry.adhocdeduction:
            deduction_amount += adhoc_entry.amount

    # total_earnings = gross_salary_amount + earning_amount
    total_earnings = gross_salary_amount

    total_fullday_time = timedelta()
    total_halfday_time = timedelta()
    total_anomaly_count = 0
    attendance_rule = AssignAttendanceRule.objects.filter(user_id__id=user_id)
    print("attendance_rule :", attendance_rule)
    for att_rule in attendance_rule:
        rule_type = att_rule.rules_applied
        print("rule_type :", rule_type, )
        if rule_type:
            full_day_hours = rule_type.fullhours
            full_day_minutes = rule_type.fullminutes
            full_time = timedelta(hours=full_day_hours,minutes=full_day_minutes)
            half_day_hours = rule_type.halfhours
            half_day_minutes = rule_type.halfminutes
            half_time = timedelta(hours=half_day_hours,
                                  minutes=half_day_minutes)
            print("Full Day Hours:", full_day_hours,
                  full_day_minutes, full_time)
            print("Half Day Hours:", half_day_hours,
                  half_day_minutes, half_time)
            in_grace_period = rule_type.inGracePeriod
            out_grace_period = rule_type.outGracePeriod
            print("Grace period:", in_grace_period, out_grace_period)
            in_grace_timedelta = timedelta(
                hours=in_grace_period.hour, minutes=in_grace_period.minute)
            out_grace_timedelta = timedelta(
                hours=out_grace_period.hour, minutes=out_grace_period.minute)

            total_grace_period = in_grace_timedelta + out_grace_timedelta
            print("Total Grace period:", total_grace_period)
            total_fullday_time = full_time + total_grace_period
            print("Total Time:", total_fullday_time)
            total_halfday_time = half_time + total_grace_period
            print("total_halfday_time :", total_halfday_time)

    punches = Punch.objects.filter(
        user__id=user_id,
        date__year=yearselect,
        date__month=month_numeric,
        status="AN", is_penalty_reverted=False
    )
    print("Punch Object :", punches)

    for punch in punches:
        print("1111111111111111111111111111111111")
        total_work_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time and punch.is_second_clocked_in:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            second_clock_in = datetime.combine(
                datetime.today(), punch.second_clock_in_time)
            second_clock_out = datetime.combine(
                datetime.today(), punch.second_clock_out_time)
            first_duration = first_clock_out - first_clock_in
            second_duration = second_clock_out - second_clock_in
            total_work_duration += first_duration + second_duration

        elif punch.first_clock_in_time and punch.first_clock_out_time:
            first_clock_in = datetime.combine(
                datetime.today(), punch.first_clock_in_time)
            first_clock_out = datetime.combine(
                datetime.today(), punch.first_clock_out_time)
            print("first_clock_in ; first_clock_out : ",
                  first_clock_in, first_clock_out)
            first_duration = first_clock_out - first_clock_in
            print("first_duration : ", first_duration)
            total_work_duration += first_duration
        if total_work_duration > total_fullday_time:
            AN_count = 0.5
        elif total_work_duration < total_halfday_time:
            AN_count = 1.0
        else:
            AN_count = 0.5

        total_anomaly_count += AN_count
        print("total_anomaly_count :", total_anomaly_count)

    lop_data = Runpayroll_lop.objects.filter(lop_date__month=month_numeric, lop_date__year=yearselect, user_id=user_id)
    for lopdata in lop_data:
        lopcount += lopdata.lop_count
        print("lopcount :", lopcount)

    absent_count = Punch.objects.filter(user__id=user_id, date__year=yearselect, date__month=month_numeric, status='A', is_penalty_reverted=False).count()
    absent_AN_count = absent_count + total_anomaly_count
    print("absent_AN_count : ", absent_count, absent_AN_count)
    punchcount = Punch.objects.filter(user__id=user_id, date__year=yearselect, date__month=month_numeric).count()
    print("punchcount :", punchcount)
    missing_date_count = num_days - punchcount
    print("missing_date_count :", missing_date_count)
    working_days = punchcount - absent_AN_count
    print("working_days :", working_days)
    total_lop = absent_AN_count + missing_date_count + lopcount
    # per_day_amount = ctc_amount / num_days
    per_day_amount = total_gross_salary / num_days
    print("per_day_amount :", total_gross_salary, per_day_amount)
    lop_amount = per_day_amount * total_lop
    print("lop_amount :", lop_amount, "total_net_salary :", total_net_salary)
    lopamount = round(lop_amount)
    # total_deductions = deduction_amount + lop_amount + total_net_salary + WFOamount
    total_deductions =  lop_amount + WFOamount
    totaldeductions = round(total_deductions)
    grossamount = round(total_earnings - total_deductions)
    print("grossamount :", grossamount)
    # net_amount_words = num2words(net_amount, lang='en_IN')
    net_amount_words = num2words(grossamount, lang='en_IN').title()
    print("net_amountwords :", net_amount_words)
    net_amount_words_formatted = f"({net_amount_words})"

    data1 = [
        ["Name:", username, 'Employee ID:', empid],
        ["Designation:", desg, 'Bank Name:', bankname],
        ["Department:", dept, 'Bank Account No.:', acno],
        ["Location:", loc, 'PAN No.:', idno],
        ["Effective Work Days:", num_days, 'PF No.:', pfnum],
        ["LOP:", total_lop, 'ESI No.:', esa],
        ["WFO:", wfocount, '', ''],

    ]

    colWidths1 = [120, 200, 120, 150]
    rowHeights1 = [20, 20, 20, 20, 20, 20, 20]

    table1 = Table(data1, colWidths=colWidths1, rowHeights=rowHeights1)
    table1.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        # ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, -1), font_style),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBEFORE', (2, 0), (2, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        # ('TOPPADDING', (0, 0), (-1, -1), 300),
    ]))

    data2 = [
        ["EARNINGS", "AMOUNT", 'CONTRIBUTIONS', "AMOUNT"],
    ]

    componentnames = []
    amounts = []
    adhocnames = []
    adhocamounts = []
    ctc_cmpname = []
    ctc_cmpamount = []

    assigndata = []
    netassigndata = []
    ctcassigndata = []
    work_from_office_allowanceamount = 0
    basicamount = 0
    ctc_total = 0
    toataldeda = 0

    if assign_salarystructure:
        print("assign_salarystructure 2:", assign_salarystructure)

        gross_salary_component = SalaryComponent.objects.filter(componentname__iexact="Gross Salary").first()
        
        work_from_officecomponent = SalaryComponent.objects.filter(componentname__iexact="Work From Office Allowance",Parentcomponentname__componentname__iexact="Gross Salary").first()
        print("gross_salary_component and work_from_office_component:", gross_salary_component, work_from_officecomponent)
        
        names = AssignSalaryStructureName.objects.filter(salaryrule=assign_salarystructure)
        amount = AssignSalaryStructureAmount.objects.filter(salaryname__in=names)

        if work_from_officecomponent:
            workamount = amount.filter(salaryname__salarycomponent=work_from_office_component).first()
            work_from_office_allowanceamount = workamount.amount if workamount else 0
        
        print("gross_salary_component and work_from_office_component:", gross_salary_component, work_from_office_allowanceamount)

        grossnames = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure,
            salarycomponent__Parentcomponentname=gross_salary_component
        ).exclude(salarycomponent__componentname__icontains="work from office allowance")
        print("grossnames:", grossnames)

        gross_amounts = []
        calculatedamount = 0
        for name in grossnames:
            for component in name.salarycomponent.all():
                print("Gross Salary Component Percent:", name, component.percent, component.componentname)
                if component.percent and component.componentname.lower() != "other allowance":
                    calculated_amount = round(grossamount * (component.percent / 100.0))
                    calculatedamount += calculated_amount
                    print("calculated_amount :", calculatedamount)
                    gross_amounts.append(calculated_amount)
                    if component.componentname.lower() == "basic salary":
                        basicamount = calculated_amount
                if component.componentname.lower() == "other allowance":
                    other_allowanceamt = grossamount - calculatedamount
                    print("other_allowanceamt :", other_allowanceamt)
                    gross_amounts.append(other_allowanceamt)

        # Zipping gross names and calculated amounts
        print("grossnames :", grossnames , gross_amounts)
        zipped_gross_data = zip_longest(grossnames, gross_amounts)
        assigndata.append({
            'rule': assign_salarystructure,
            'zipped_gross_data': zipped_gross_data,
        })

        # Process Net Salary components
        net_names = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure,
            salarycomponent__Parentcomponentname__componentname__iexact="Net Salary"
        )
        net_amounts = []
        for name in net_names:
            for component in name.salarycomponent.all():
                print("Net Salary Component Percent:", name, component.percent)
                if component.componentname.lower() == "epf employee":
                    epf_amount = round((basicamount * component.percent) / 100)
                    net_amounts.append(epf_amount)
                if component.componentname.lower() == "esi employee":
                    esi_amount = round((grossamount * component.percent) / 100) 
                    net_amounts.append(esi_amount)
                if component.componentname.lower() == "professional tax":
                    profisional_tax_amount = 167 if grossamount < 22000 else 208
                    print("profisional_tax_amount :", profisional_tax_amount)
                    net_amounts.append(profisional_tax_amount)

                if component.componentname.lower() == "insurance":
                    insurance_amount = 0 if grossamount <= 25000 else 245
                    print("insurance_amount :", insurance_amount)
                    net_amounts.append(insurance_amount)
        
        net_total = sum(net_amounts)

        # Zipping net names and amounts
        zipped_net_data = zip_longest(net_names, net_amounts)
        netassigndata.append({
            'rule': assign_salarystructure,
            'zipped_net_data': zipped_net_data,
        })

        # Process CTC components, excluding professional tax
        ctcsalarynames = AssignSalaryStructureName.objects.filter(
            salaryrule=assign_salarystructure,
            salarycomponent__Parentcomponentname__componentname__iexact="CTC"
        ).exclude(salarycomponent__componentname__icontains="professional tax")
        ctcamounts = []
        for name in ctcsalarynames:
            for component in name.salarycomponent.all():
                print("CTC Salary Component Percent:", component.percent, component.componentname)
                if component.componentname == "EPF Employer":
                    epf_amount = round((basicamount * component.percent) / 100)
                    print("amount : ", epf_amount)
                    ctcamounts.append(epf_amount)
                if component.componentname == "ESI Employer":
                    esi_amount = round((grossamount * component.percent) / 100) 
                    ctcamounts.append(esi_amount)

        ctc_total = sum(ctcamounts)
        print("ctc_total :", ctc_total)

        # Zipping CTC names and amounts
        ctc_zipped_data = zip_longest(ctcsalarynames, ctcamounts)
        ctcassigndata.append({
            'rule': assign_salarystructure,
            'ctc_zipped_data': ctc_zipped_data,
        })

    for data_entry in assigndata:
        for name, amount in data_entry['zipped_gross_data']:
            component_name = ', '.join(component.componentname for component in name.salarycomponent.all()) if name else ''
            amount_value = amount if amount else ''

            componentnames.append(component_name)
            amounts.append(amount_value)

    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocearning:
            component_name = adhoc_entry.adhocearning.component_name
            amount = adhoc_entry.amount

            componentnames.append(component_name)
            amounts.append(amount)

    for ctc_dataentry in ctcassigndata:
        for ctc_name, ctcamounts in ctc_dataentry['ctc_zipped_data']:
            print("ctc_name :", ctc_name)
            ctccomp_name = ', '.join(component.componentname for component in ctc_name.salarycomponent.all()) if ctc_name else ''
            ctcamount = ctcamounts if ctcamounts else ''
            # totalctcamount += ctcamount
            print("ctcamount 16766 :", ctcamount)

            print("ctccomp_name : ", ctccomp_name, "ctcamount :", ctcamount)

            ctc_cmpname.append(ctccomp_name)
            ctc_cmpamount.append(ctcamount)
    
    for adhoc_entry in adhoc_data:
        if adhoc_entry.adhocdeduction:
            adhoc_name = adhoc_entry.adhocdeduction.component_name
            adhoc_amount = adhoc_entry.amount

            adhocnames.append(adhoc_name)
            adhocamounts.append(adhoc_amount)
    
    for dataentry in netassigndata:
        for netnames, netamounts in dataentry['zipped_net_data']:
            print("netnames :", netnames)
            netcomp_name = ', '.join(component.componentname for component in netnames.salarycomponent.all()) if netnames else ''
            netamount = netamounts if netamounts else ''

            print("netcomp_name : ", netcomp_name, "netamount :", netamount)

            adhocnames.append(netcomp_name)
            adhocamounts.append(netamount)
    
    # adhocnames.append("LOP Amount")
    # adhocamounts.append(lopamount)
    # adhocnames.append("Work From Office Allowance")
    # adhocamounts.append(WFOamount)
    totalgrossamount = grossamount + earning_amount

    toataldeda = deduction_amount + net_total

    totalNet_Salary = totalgrossamount - toataldeda
    net_amount_words = num2words(totalNet_Salary, lang='en_IN').title()
    print("totalNet_Salary :", totalNet_Salary)
    net_amount_words_formatted = f"({totalNet_Salary})"

    max_length = max(len(componentnames), len(ctc_cmpname))
    print("max_length 15119 :", max_length)

    for i in range(max_length):
        component_name = componentnames[i] if i < len(componentnames) else ''
        amount_value = amounts[i] if i < len(amounts) else ''

        ctccmpname = ctc_cmpname[i] if i < len(ctc_cmpname) else ''
        ctccmpamount = ctc_cmpamount[i] if i < len(ctc_cmpamount) else ''

        data2.append([component_name, amount_value, ctccmpname, ctccmpamount])

  
    data2.extend([
        ['Total Earnings (Rs)', totalgrossamount, 'Total Contributions (Rs)', ctc_total],
        ['DEDUCTIONS', 'AMOUNT', '', ''],
    ])

    print("len(adhocnames) :", len(adhocnames))
    for i in range(len(adhocnames)):
        data2.append([adhocnames[i], adhocamounts[i], '', ''])

    data2.extend([
        ['Total Deductions (Rs)', toataldeda, '', ''],
        ['Net Pay For The Month:', totalNet_Salary, '', ''],
        [net_amount_words, '', '', ''],
        ['', '', '', ''],
        ['', '', 'This is a system-generated payslip and does not require a signature.', ''],
    ])

    colWidths2 = [200, 95, 200, 95]
    # rowHeights2 = [20, 20, 20, 20, 20, 30, 20]
    row_height = 20
    rowHeights2 = [row_height] * (8 + max_length + len(adhocnames))

    rowHeights2[0] = 30
    rowHeights2[-4] = 60
    rowHeights2[-3] = 20
    rowHeights2[-2] = 10
    rowHeights2[-1] = 30

    print("rowHeights2 :", rowHeights2)

    table2 = Table(data2, colWidths=colWidths2, rowHeights=rowHeights2)
    table2.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 0), (-1, 0), bold_font_style),
        # ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7cb6c0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 0.25, colors.black),
        ('FONTNAME', (0, 1 + max_length), (-1, 1 + max_length), bold_font_style), # Total 
        ('FONTNAME', (0, 2 + max_length), (-1, 2 + max_length), bold_font_style),
        ('FONTNAME', (0, -5), (-1, -5), bold_font_style),  # Total Deduction
        ('LINEBELOW', (0, 1 + max_length), (-1, 1 + max_length), 0.25, colors.black),
        ('LINEABOVE', (0, 1 + max_length), (-1, 1 + max_length), 0.25, colors.black),
        ('LINEBELOW', (0, 2 + max_length), (-3, 2 + max_length), 0.25, colors.black),

        ('LINEBEFORE', (2, 0), (2, 3 + max_length + len(adhocnames)), 1, colors.black),
        # ('LINEAFTER', (2, 0), (2, -2 -max_length), 1, colors.black),
        ('LINEABOVE', (0, -5), (-3, -5), 0.25, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 0.25, colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
        ('BOX', (0, 0), (-1, 3 + max_length + len(adhocnames)), 1, colors.black),
    ]))

    spacer_height = 20
    spacer = Spacer(1, spacer_height)

    elements.append(table)
    elements.append(table1)
    elements.append(spacer)
    elements.append(table2)

    doc.build(elements)

    pdf_data = buffer.getvalue()
    buffer.close()

    return pdf_data

def send_payslip_email(user_email, pdf_buffer):
    subject = 'Payslip'
    message = 'Your payslip is attached.'
    from_email = settings.EMAIL_HOST_USER
    to_email = [user_email]
    
    print("From Email: ", from_email)
    print("To Email: ", to_email)

    email = EmailMessage(subject, message, from_email, to_email)
    email.attach('payslip.pdf', pdf_buffer.read(), 'application/pdf')
    email.send()
    print("SEND")

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def adhoc_component(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    adhocearning = adhoc_earning.objects.filter(user_id=userid)
    adhocdeduction = adhoc_deduction.objects.filter(user_id=userid)
    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, 'index/adhoc_component.html', {'adhocearning': adhocearning, 'adhocdeduction': adhocdeduction, **x})


def adhoc_deductions(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    adhocearning = adhoc_earning.objects.filter(user_id=userid)
    adhocdeduction = adhoc_deduction.objects.filter(user_id=userid)
    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
    }
    return render(request, 'index/adhoc_deduction.html', {'adhocearning': adhocearning, 'adhocdeduction': adhocdeduction, **x})


def add_adhocearningcomp(request):
    if request.method == "POST":
        userid = request.user.id
        earningname = request.POST.get("earningname")
        adhoc_earning.objects.create(
            user_id_id=userid, component_name=earningname)
    return redirect("adhoc_component")


def edit_adhocearningcomp(request):
    if request.method == "POST":
        cmpid = request.POST.get('cmpid')
        adhocearning = adhoc_earning.objects.get(id=cmpid)
        cmpname = request.POST.get("cmpname")
        adhocearning.component_name = cmpname
        adhocearning.save()
    return redirect("adhoc_component")


def delete_adhocearningcomp(request, id):
    adhocearning = adhoc_earning.objects.get(id=id)
    adhocearning.delete()
    return redirect("adhoc_component")


def add_adhocdeductioncomp(request):
    if request.method == "POST":
        userid = request.user.id
        deductionname = request.POST.get("deductionname")
        adhoc_deduction.objects.create(
            user_id_id=userid, component_name=deductionname)
    return redirect("adhoc_deductions")


def edit_adhocdeductioncomp(request):
    if request.method == "POST":
        cmpid = request.POST.get('cmpid')
        adhocdeduction = adhoc_deduction.objects.get(id=cmpid)
        cmpname = request.POST.get("cmpname")
        adhocdeduction.component_name = cmpname
        adhocdeduction.save()
    return redirect("adhoc_deductions")


def delete_adhocdeductioncomp(request, id):
    adhocdeduction = adhoc_deduction.objects.get(id=id)
    adhocdeduction.delete()
    return redirect("adhoc_deductions")


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def adhoc(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    today = datetime.now()
    adhocearning = adhoc_earning.objects.filter(user_id=userid)
    adhocdeduction = adhoc_deduction.objects.filter(user_id=userid)
    userlist = User.objects.filter(Q(id=userid) | Q(admin_id=userid) & Q(status='Active'))

    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)
    print("Saved Adhoc:", Adhoc.objects.latest('id'))

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
        selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
    print("selected_month :", selected_month)
    adhoclist = Adhoc.objects.filter(
        createddate__year=selected_year,
        createddate__month=selected_month,
        adminid=userid
    ).order_by('createddate')

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "month_str": month_str,
        "selected_year": selected_year,
    }
    return render(request, 'index/adhoc.html', {'adhocearning': adhocearning, 'adhocdeduction': adhocdeduction, 'userlist': userlist, 'adhoclist': adhoclist, **x})


@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def create_adhoc(request):  
    if request.method == "POST":
        try:
            monthselect = request.POST.get("monthselect")
            yearselect = request.POST.get("yearselect")

            # Convert month string to numeric
            datetime_object = datetime.strptime(monthselect, "%B")
            month_numeric = datetime_object.month
            createddate = datetime(int(yearselect), month_numeric, 1)

            adminid = request.user.id
            emp_id = request.POST.get('empname')
            component_type = request.POST.get('componentType')

            component_names = request.POST.getlist('componentName[]')
            amounts = request.POST.getlist('amounts[]')

            if not emp_id or not component_type or not component_names or not amounts:
                messages.error(request, "All fields are required.")
                return redirect(reverse("adhoc") + f"?monthselect={monthselect}+{yearselect}")

            user_instance = User.objects.get(id=emp_id)
            print("Component Names:", component_names)
            print("Amounts:", amounts)

            for component_id, amount in zip(component_names, amounts):
                try:
                    if not amount:
                        continue  # skip if amount is empty

                    amount = float(amount)

                    if component_type == 'Earning':
                        adhoc_earning_instance = adhoc_earning.objects.get(id=component_id)
                        Adhoc.objects.create(
                            user_id=user_instance,
                            adhocearning=adhoc_earning_instance,
                            cmptype='Earning',
                            amount=amount,
                            adminid=adminid,
                            createddate=createddate
                        )
                    elif component_type == 'Deduction':
                        adhoc_deduction_instance = adhoc_deduction.objects.get(id=component_id)
                        Adhoc.objects.create(
                            user_id=user_instance,
                            adhocdeduction=adhoc_deduction_instance,
                            cmptype='Deduction',
                            amount=amount,
                            adminid=adminid,
                            createddate=createddate
                        )
                except Exception as ex:
                    print(f"❌ Error saving adhoc for component_id={component_id}: {ex}")

            # Update PayRegister status
            payregister_instance = PayRegister.objects.filter(
                user_id=user_instance,
                createddate__year=int(yearselect),
                createddate__month=month_numeric
            ).first()

            if payregister_instance:
                payregister_instance.status = "Pending Calculation"
                payregister_instance.save()

            return redirect(reverse("adhoc") + '?monthselect=' + monthselect + '+' + yearselect)

        except Exception as e:
            print("🔥 ERROR IN create_adhoc:", e)
            messages.error(request, "Something went wrong while saving the Adhoc entry.")
            return redirect(reverse("adhoc") + '?monthselect=' + monthselect + '+' + yearselect)

    return redirect("adhoc")


def edit_adhoc(request):
    if request.method == "POST":
        monthselect = request.POST.get("monthselect")
        datetime_object = datetime.strptime(monthselect, "%B")
        month_numeric = datetime_object.month
        yearselect = request.POST.get("yearselect")
        emp_id = request.POST.get('empname')
        user_instance = User.objects.get(id=emp_id)
        component_type = request.POST.get('componenttype')
        component_id = request.POST.get('componentname')
        amount = request.POST.get('amount')
        adhoc_id = request.POST.get('adhocid')
        adhoc_instance = Adhoc.objects.get(id=adhoc_id)

        if component_type == 'Earning':
            adhoc_earning_instance = adhoc_earning.objects.get(id=component_id)
            adhoc_instance.user_id = user_instance
            adhoc_instance.adhocearning = adhoc_earning_instance
            adhoc_instance.adhocdeduction = None
            adhoc_instance.cmptype = component_type
            adhoc_instance.amount = amount
            adhoc_instance.save()
        else:
            adhoc_deduction_instance = adhoc_deduction.objects.get(
                id=component_id)
            adhoc_instance.user_id = user_instance
            adhoc_instance.adhocdeduction = adhoc_deduction_instance
            adhoc_instance.adhocearning = None
            adhoc_instance.cmptype = component_type
            adhoc_instance.amount = amount
            adhoc_instance.save()

        payregister_instance = PayRegister.objects.filter(
            user_id=user_instance, createddate__year=yearselect, createddate__month=month_numeric).first()
        if payregister_instance:
            payregister_instance.status = "Pending Calculation"
            payregister_instance.save()

    return redirect(reverse("adhoc") + '?monthselect=' + monthselect + '+' + yearselect)


def delete_adhoc(request, id):
    dateselect = request.GET.get("monthselect")
    if dateselect:
        date_object = datetime.strptime(dateselect, "%B %Y")
        monthselect = date_object.strftime("%B")
        yearselect = date_object.year
        datetime_object = datetime.strptime(monthselect, "%B")
        month_numeric = datetime_object.month

    print("dateselect ", dateselect, month_numeric, monthselect, yearselect)
    adhoc = Adhoc.objects.get(id=id)
    payregister_instance = PayRegister.objects.filter(
        user_id=adhoc.user_id.id, createddate__year=yearselect, createddate__month=month_numeric).first()
    if payregister_instance:
        payregister_instance.status = "Pending Calculation"
        payregister_instance.save()

    adhoc.delete()
    return redirect(reverse("adhoc") + '?monthselect=' + monthselect + '+' + str(yearselect))


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def salary_onhold(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    today = datetime.now()

    selected_month_str = request.GET.get('monthselect', None)
    print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    user_data = User.objects.filter(
        Q(id=userid) | Q(admin_id=userid) & Q(status='Active'))
    payaction_data = PayActionStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid)))

    # page = request.GET.get('page', 1)
    # paginator = Paginator(payaction_data, 20)
    # try:
    #     payaction_data = paginator.page(page)
    # except PageNotAnInteger:
    #     payaction_data = paginator.page(1)
    # except EmptyPage:
    #     payaction_data = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "month_str": month_str,
        "selected_year": selected_year,
    }
    return render(request, 'index/salary_hold.html', {'payaction_data': payaction_data, 'user_data': user_data, **x})


def addsalary_onhold(request):
    if request.method == "POST":
        monthselect = request.POST.get("monthselect")
        yearselect = request.POST.get("yearselect")
        datetime_object = datetime.strptime(monthselect, "%B")
        month_numeric = datetime_object.month
        createddate = datetime(int(yearselect), month_numeric, 1)

        emp_id = request.POST.get('empname')
        user_instance = User.objects.get(id=emp_id)
        action_type = request.POST.get('actiontype')
        PayActionStatus.objects.create(
            user_id=user_instance, actiontype=action_type, createddate=createddate)
        # return redirect("salary_onhold")
        return redirect(reverse('salary_onhold') + '?monthselect=' + monthselect + '+' + yearselect)


def editsalary_onhold(request):
    if request.method == "POST":
        monthselect = request.POST.get("monthselect")
        yearselect = request.POST.get("yearselect")

        emp_id = request.POST.get('empname')
        user_instance = User.objects.get(id=emp_id)
        action_type = request.POST.get('actionType')
        pay_id = request.POST.get('payid')
        pay_instance = PayActionStatus.objects.get(id=pay_id)
        pay_instance.user_id = user_instance
        pay_instance.actiontype = action_type
        pay_instance.save()
        return redirect(reverse('salary_onhold') + '?monthselect=' + monthselect + '+' + yearselect)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def salary_revision(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    today = datetime.now()

    selected_month_str = request.GET.get('monthselect', None)
    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    user_data = User.objects.filter(Q(id=userid) | Q(admin_id=userid) & Q(status='Active'))
    component = SalaryComponent.objects.filter(admin_id=request.user.id)

    user_records = []
    for user in user_data:
        assg_rule = AssignSalaryStructure.objects.filter(effective_date__year=selected_year, effective_date__month=selected_month, user_id__id=user.id)

        ctc_amount = 0
        old_ctc_amount = 0
        for rule in assg_rule:
            salary_names = AssignSalaryStructureName.objects.filter(salaryrule=rule)
            for name in salary_names:
                amounts = AssignSalaryStructureAmount.objects.filter(salaryname=name)
                for amount in amounts:
                    ctc_amount += amount.amount

            previous_assg_rule = AssignSalaryStructure.objects.filter(
                user_id__id=user.id, effective_date__lt=rule.effective_date).order_by('-effective_date').first()
            if previous_assg_rule:
                prev_salary_names = AssignSalaryStructureName.objects.filter(
                    salaryrule=previous_assg_rule)
                for prev_name in prev_salary_names:
                    prev_amounts = AssignSalaryStructureAmount.objects.filter(
                        salaryname=prev_name)
                    for prev_amount in prev_amounts:
                        old_ctc_amount += prev_amount.amount

        percentage_change = ((ctc_amount - old_ctc_amount) /
                             old_ctc_amount) * 100 if old_ctc_amount != 0 else 0

        formatted_percentage_change = "{:.2f}%".format(percentage_change)

        user_records.append({
            "user": user,
            "ctc_amount": ctc_amount,
            "old_ctc_amount": old_ctc_amount,
            "formatted_percentage_change": formatted_percentage_change,
        })

    assgin_rules = AssignSalaryStructure.objects.filter(
        (Q(user_id__id=userid) | Q(user_id__admin_id=userid)), 
        Q(effective_date__year = selected_year) &
        Q(effective_date__month = selected_month)

    )

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "month_str": month_str,
        "selected_year": selected_year,
    }

    return render(request, 'index/salary_revision.html', {'user_records': user_records, 'user_data': user_data, 'component': component, 'assgin_rules': assgin_rules, **x})


def edit_assignsalary(request):
    if request.method == 'POST':
        salary_id = request.POST.get('salaryid')
        amount = request.POST.getlist('amount', [])
        Amount = request.POST.getlist('Amount', [])
        all_amount = amount + Amount
        
        effective_date = request.POST.get('effdate')
        formateddate = datetime.strptime(effective_date, "%B %Y")
        formatted_date = formateddate.strftime("%Y-%m-%d")
        
        gross_salary_component = SalaryComponent.objects.filter(componentname="Gross Salary").first()
        net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
        ctc_salary_component = SalaryComponent.objects.filter(componentname="CTC").first()

        basic_salary_amount = 0
        gross_salary_amount = 0
        total_parent_component_amount = 0
        total_NSparent_component_amount = 0
        total_CTCparent_component_amount = 0

        previous_salarystructure = AssignSalaryStructure.objects.get(id=salary_id)
        
        existing_assignsalary = AssignSalaryStructure.objects.filter(user_id=previous_salarystructure.user_id, effective_date=formatted_date).first()

        if not existing_assignsalary:
            newassign_salary = AssignSalaryStructure.objects.create(user_id=previous_salarystructure.user_id, assign_salary=previous_salarystructure.assign_salary, effective_date=formatted_date)

            previous_assignsalarynames = AssignSalaryStructureName.objects.filter(salaryrule=previous_salarystructure)
            
            for i, sname in enumerate(previous_assignsalarynames):
                components = sname.salarycomponent.all()
                
                for component in components:
                    if component.componentname == "Gross Salary":
                        try:
                            gross_salary_amount = float(all_amount[i])
                        except IndexError:
                            logger.error(f"Amount for Gross Salary is missing.")
                            continue
                    
                    if component.componentname in ["Professional Tax", "Professional tax"]:
                        calculated_amount = 167 if gross_salary_amount < 22000 else 208

                    if component.componentname in ["Insurance"]:
                        calculated_amount = 0 if gross_salary_amount <= 25000 else 245

                        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
                        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=component.id))
                        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount).save()
                        
                        if component.Parentcomponentname == net_salary_component:
                            total_NSparent_component_amount += calculated_amount
                        elif component.Parentcomponentname == ctc_salary_component:
                            total_CTCparent_component_amount += calculated_amount
                        continue

                    if component.componentname in ["Other Allowance", "EPF Employer", "EPF Employee", "Net Salary", "CTC"]:
                        continue
                    
                    try:
                        calculated_amount = float(all_amount[i])
                    except IndexError:
                        logger.error(f"Amount for {component.componentname} is missing.")
                        continue
                    
                    if component.percent:
                        calculated_amount = round(gross_salary_amount * (component.percent / 100.0))
                    
                    if component.Parentcomponentname == gross_salary_component:
                        total_parent_component_amount += calculated_amount
                    elif component.Parentcomponentname == net_salary_component:
                        total_NSparent_component_amount += calculated_amount
                    elif component.Parentcomponentname == ctc_salary_component:
                        total_CTCparent_component_amount += calculated_amount
                    
                    if component.componentname == "Basic Salary":
                        basic_salary_amount = calculated_amount
                    
                    salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
                    salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=component.id))
                    assignsalaryamount = AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount).save()

                process_otherallowance(components, newassign_salary, gross_salary_amount, total_parent_component_amount)
                total_NSparent_component_amount = process_epfempoyeecomponent(components, newassign_salary, total_NSparent_component_amount, basic_salary_amount)
                total_CTCparent_component_amount = process_epfcomponent(components, newassign_salary, total_CTCparent_component_amount, basic_salary_amount)
                process_netsalary(components, newassign_salary, total_NSparent_component_amount, gross_salary_amount)
                process_ctcamount(components, newassign_salary, total_CTCparent_component_amount, gross_salary_amount)
        
        else:

            existing_assignsalarynames = AssignSalaryStructureName.objects.filter(salaryrule=existing_assignsalary)
            existing_amounts = AssignSalaryStructureAmount.objects.filter(salaryname__in=existing_assignsalarynames)
            existing_amounts.delete()

            for i, sname in enumerate(existing_assignsalarynames):
                components = sname.salarycomponent.all()

                for component in components:
                        
                    if component.componentname == "Gross Salary":
                        try:
                            gross_salary_amount = float(all_amount[i])
                            print(f"Gross Salary Amount: {gross_salary_amount}")
                        except IndexError:
                            logger.error(f"Amount for Gross Salary is missing.")
                            continue

                    if component.componentname in ["Professional Tax", "Professional tax"]:
                        calculated_amount = 167 if gross_salary_amount < 22000 else 208

                    if component.componentname in ["Insurance"]:
                        calculated_amount = 0 if gross_salary_amount <= 25000 else 245

                        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=calculated_amount)
                        salary_amount.save()

                        if component.Parentcomponentname == net_salary_component:
                            total_NSparent_component_amount += calculated_amount
                        elif component.Parentcomponentname == ctc_salary_component:
                            total_CTCparent_component_amount += calculated_amount
                        continue

                    if component.componentname in ["Other Allowance", "EPF Employer", "EPF Employee", "Net Salary", "CTC"]:
                        continue
                    try:
                        calculated_amount = float(all_amount[i])
                    except IndexError:
                        logger.error(f"Amount for {component.componentname} is missing.")
                        continue
                    
                    if component.percent:
                        calculated_amount = round(gross_salary_amount * (component.percent / 100.0))

                    if component.Parentcomponentname == gross_salary_component:
                        total_parent_component_amount += calculated_amount
                    elif component.Parentcomponentname == net_salary_component:
                        total_NSparent_component_amount += calculated_amount
                    elif component.Parentcomponentname == ctc_salary_component:
                        total_CTCparent_component_amount += calculated_amount

                    if component.componentname == "Basic Salary":
                        basic_salary_amount = calculated_amount

                    salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=calculated_amount)
                    salary_amount.save()

                processotherallowance(components, sname, existing_assignsalarynames, gross_salary_amount, total_parent_component_amount)
                total_NSparent_component_amount = processepfempoyeecomponent(components, sname, existing_assignsalarynames, total_NSparent_component_amount, basic_salary_amount)
                total_CTCparent_component_amount = processepfcomponent(components, sname, existing_assignsalarynames, total_CTCparent_component_amount, basic_salary_amount)
                processnetsalary(components, sname, existing_assignsalarynames, total_NSparent_component_amount, gross_salary_amount)
                processctcamount(components, sname, existing_assignsalarynames, total_CTCparent_component_amount, gross_salary_amount)

    return redirect('salary_revision')

def process_otherallowance(components, newassign_salary, gross_salary_amount, total_parent_component_amount):
    other_allowance_component = SalaryComponent.objects.filter(componentname="Other Allowance").first()
    if other_allowance_component and str(other_allowance_component.id) in [str(comp.id) for comp in components]:
        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=other_allowance_component.id))
        if gross_salary_amount > total_parent_component_amount:
            other_allowance_amount = round(gross_salary_amount - total_parent_component_amount)
        else:
            other_allowance_amount = 0 

        other_allowance_amount = abs(other_allowance_amount)
        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=other_allowance_amount).save()

def process_epfempoyeecomponent(components, newassign_salary, total_NSparent_component_amount, basic_salary_amount):
    epfempoyee_component = SalaryComponent.objects.filter(componentname="EPF Employee").first()
    if epfempoyee_component and str(epfempoyee_component.id) in [str(comp.id) for comp in components]:
        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=epfempoyee_component.id))
        calculated_amount = round(basic_salary_amount * (epfempoyee_component.percent / 100.0))
        total_NSparent_component_amount += calculated_amount
        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount).save()
    return total_NSparent_component_amount

def process_epfcomponent(components, newassign_salary, total_CTCparent_component_amount, basic_salary_amount):
    epf_component = SalaryComponent.objects.filter(componentname="EPF Employer").first()
    if epf_component and str(epf_component.id) in [str(comp.id) for comp in components]:
        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=epf_component.id))
        calculated_amount = round(basic_salary_amount * (epf_component.percent / 100.0))
        total_CTCparent_component_amount += calculated_amount
        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=calculated_amount).save()
    return total_CTCparent_component_amount

def process_netsalary(components, newassign_salary, total_NSparent_component_amount, gross_salary_amount):
    net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
    if net_salary_component and str(net_salary_component.id) in [str(comp.id) for comp in components]:
        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=net_salary_component.id))
        net_salary_amount = gross_salary_amount - total_NSparent_component_amount
        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=net_salary_amount).save()

def process_ctcamount(components, newassign_salary, total_CTCparent_component_amount, gross_salary_amount):
    ctc_component = SalaryComponent.objects.filter(componentname="CTC").first()
    if ctc_component and str(ctc_component.id) in [str(comp.id) for comp in components]:
        salary_name = AssignSalaryStructureName.objects.create(salaryrule=newassign_salary)
        salary_name.salarycomponent.set(SalaryComponent.objects.filter(id=ctc_component.id))
        ctc_amount = gross_salary_amount + total_CTCparent_component_amount
        AssignSalaryStructureAmount.objects.create(salaryname=salary_name, amount=ctc_amount).save()

# ELSE PART
def processotherallowance(components, sname, existing_assignsalarynames, gross_salary_amount, total_parent_component_amount):
    other_allowance_component = SalaryComponent.objects.filter(componentname="Other Allowance").first()
    if other_allowance_component and str(other_allowance_component.id) in [str(comp.id) for comp in components]:
        if gross_salary_amount > total_parent_component_amount:
            other_allowance_amount = round(gross_salary_amount - total_parent_component_amount)
        else:
            other_allowance_amount = 0 

        other_allowance_amount = abs(other_allowance_amount)
        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=other_allowance_amount)
        salary_amount.save()


def processepfempoyeecomponent(components, sname, existing_assignsalarynames, total_NSparent_component_amount, basic_salary_amount):
    epfempoyee_component = SalaryComponent.objects.filter(componentname="EPF Employee").first()
    component_ids = [str(comp.id) for comp in components]
    if epfempoyee_component and str(epfempoyee_component.id) in component_ids:
        calculated_amount = round(basic_salary_amount * (epfempoyee_component.percent / 100.0))
        total_NSparent_component_amount += calculated_amount
        print("total_NSparent_component_amount calculated_amount :", total_NSparent_component_amount, calculated_amount)
        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=calculated_amount)
        salary_amount.save()
    return total_NSparent_component_amount

def processepfcomponent(components, sname, existing_assignsalarynames, total_CTCparent_component_amount, basic_salary_amount):
    epf_component = SalaryComponent.objects.filter(componentname="EPF Employer").first()
    if epf_component and str(epf_component.id) in [str(comp.id) for comp in components]:
        calculated_amount = round(basic_salary_amount * (epf_component.percent / 100.0))
        total_CTCparent_component_amount += calculated_amount
        print("total_CTCparent_component_amount calculated_amount :", total_CTCparent_component_amount, calculated_amount)
        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=calculated_amount)
        salary_amount.save()
    return total_CTCparent_component_amount


def processnetsalary(components, sname, existing_assignsalarynames, total_NSparent_component_amount, gross_salary_amount):
    net_salary_component = SalaryComponent.objects.filter(componentname="Net Salary").first()
    print("total_NSparent_component_amount 18037:", total_NSparent_component_amount)
    if net_salary_component and str(net_salary_component.id) in [str(comp.id) for comp in components]:
        net_salary_amount = gross_salary_amount - total_NSparent_component_amount
        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=net_salary_amount)
        salary_amount.save()

def processctcamount(components, sname, existing_assignsalarynames, total_CTCparent_component_amount, gross_salary_amount):
    ctc_component = SalaryComponent.objects.filter(componentname="CTC").first()
    print("total_CTCparent_component_amount 18045:", total_CTCparent_component_amount)
    if ctc_component and str(ctc_component.id) in [str(comp.id) for comp in components]:
        ctc_amount = gross_salary_amount + total_CTCparent_component_amount
        salary_amount = AssignSalaryStructureAmount.objects.create(salaryname=sname, amount=ctc_amount)
        salary_amount.save()

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def runpayroll_overview(request):
        
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    data = companyprofile.objects.filter(admin_id=userid)
    today = datetime.now()

    selected_month_str = request.GET.get('monthselect', None)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
        wfo_date = today.replace(day=1).date()
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')
        wfo_date = selected_date.replace(day=1).date()

    user_list = User.objects.filter(Q(id=userid) | Q(admin_id=userid) & Q(status='Active'))

    for user in user_list:
        wfocount = 0
        leave_count = 0

        punches = Punch.objects.filter(
                user=user,
                date__year=selected_year,
                date__month=selected_month
            ).order_by('date')

        for punch in punches:
            if punch.WfhOrWfo == "WFO":
                wfocount += 1
            
            # if punch.status in ["H", "L"]:
            leave_data = Leave.objects.filter(
                    applicant_email=user,  
                    strtDate=punch.date, 
                    status="Approved"     
                ).first()      
                        
            if punch.status == "H":
                leave_count += 1
            elif punch.status == "L":
                if leave_data:
                    if leave_data.leavetyp in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                        leave_count += 1
            elif punch.status == "HL":
                if leave_data:
                    if leave_data.leavetyp not in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                        leave_count -= 0.5
        wfo_count = wfocount + leave_count
        # print("wfo_count :", wfo_count)
        # print("wfo_date 2 : ", wfo_date)
        WFOlist = WFOCount.objects.filter(user_id=user,wfo_date=wfo_date).first()
        # print("WFOlist :", WFOlist)
        if WFOlist:
            WFOlist.wfo_count = wfo_count
        else:
            WFOCount.objects.create(user_id=user, wfocount=wfo_count, wfo_date=wfo_date)

    user_data = User.objects.filter(Q(id=userid) | Q(admin_id=userid)).count()

    salary_onhold = PayActionStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month, actiontype="On Hold") & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid))).count()

    total_netpay_amount = PayRegister.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid))).aggregate(Sum('netpay'))['netpay__sum']

    payreg_data = PayRegister.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month, status__in=[
                                             "Completed", "Payslip Generated", "Payslip Downloaded"]) & (Q(user_id__id=userid) | Q(user_id__admin_id=userid))).count()

    payslip_data = PayRegister.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month, status__in=[
                                              "Payslip Generated", "Payslip Downloaded"]) & (Q(user_id__id=userid) | Q(user_id__admin_id=userid))).count()

    payout_data = PayoutStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month, status="Paid") & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid))).count()

    x = {
        "k": k[0] if k.exists() else k,
        "data": data[0] if data.exists() else data,
        "month_str": month_str,
        "selected_year": selected_year,
    }
    return render(request, 'index/runpayroll_overview.html', 
                            {
                                'user_data': user_data, 
                            'salary_onhold': salary_onhold,
                            'total_netpay_amount': total_netpay_amount, 
                            'payreg_data': payreg_data, 
                            'payslip_data': payslip_data, 
                            'payout_data': payout_data, 
                            **x
                            }
                )


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def runpayroll_overview_v2(request):
    logger.info("Payroll overview version 2")
    
    userid = request.user.id
    cache_key = f"payroll_overview_{userid}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return render(request, 'index/runpayroll_overview.html', cached_data)

    today = datetime.now()
    selected_month_str = request.GET.get('monthselect')

    if selected_month_str:
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
    else:
        selected_date = today

    selected_month = selected_date.month
    selected_year = selected_date.year
    month_str = selected_date.strftime('%B')
    wfo_date = selected_date.replace(day=1).date()

    # Fetch profile and company data
    myprofile = Myprofile.objects.filter(myuser_id=userid).first()
    company_data = companyprofile.objects.filter(admin_id=userid).first()

    # Get active users
    user_list = User.objects.filter(Q(id=userid) | Q(admin_id=userid), status="Active")

    # Get punch records for all users in one query
    punch_data = Punch.objects.filter(
        user__in=user_list, 
        date__year=selected_year, 
        date__month=selected_month
    ).select_related("user")

    # Get leave data for efficiency
    leave_data = Leave.objects.filter(
        applicant_email__in=user_list, 
        strtDate__year=selected_year, 
        strtDate__month=selected_month,
        status="Approved"
    ).select_related("applicant_email")

    leave_map = {leave.applicant_email_id: leave for leave in leave_data}
    
    wfo_counts = {}
    for punch in punch_data:
        user_id = punch.user_id
        if user_id not in wfo_counts:
            wfo_counts[user_id] = 0
        
        if punch.WfhOrWfo == "WFO":
            wfo_counts[user_id] += 1
        
        leave = leave_map.get(user_id)
        if punch.status == "H":
            wfo_counts[user_id] += 1
        elif punch.status == "L" and leave:
            if leave.leavetyp in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                wfo_counts[user_id] += 1
        elif punch.status == "HL" and leave:
            if leave.leavetyp not in ["Casual Leave", "Comp Off", "Optional Holiday"]:
                wfo_counts[user_id] -= 0.5

    # Bulk update WFO count
    wfo_objects = [
        WFOCount(user_id=user, wfocount=wfo_counts[user.id], wfo_date=wfo_date)
        for user in user_list
    ]
    WFOCount.objects.bulk_create(wfo_objects, ignore_conflicts=True)

    # Payroll data aggregation
    payroll_filters = Q(user_id__in=user_list, createddate__year=selected_year, createddate__month=selected_month)
    
    salary_onhold = PayActionStatus.objects.filter(payroll_filters, actiontype="On Hold").count()
    total_netpay_amount = PayRegister.objects.filter(payroll_filters).aggregate(Sum('netpay'))['netpay__sum'] or 0
    payreg_data = PayRegister.objects.filter(payroll_filters, status__in=["Completed", "Payslip Generated", "Payslip Downloaded"]).count()
    payslip_data = PayRegister.objects.filter(payroll_filters, status__in=["Payslip Generated", "Payslip Downloaded"]).count()
    payout_data = PayoutStatus.objects.filter(payroll_filters, status="Paid").count()
    
    total_user_data_count = User.objects.filter(Q(id=userid) | Q(admin_id=userid)).count()
    active_user_data_count = user_list.count()
    inactive_user_data_count = total_user_data_count - active_user_data_count
    response_data = {
        "total_user_data_count":total_user_data_count,
        "active_user_count": active_user_data_count,
        "inactive_user_data_count":inactive_user_data_count,
        "salary_onhold": salary_onhold,
        "total_netpay_amount": total_netpay_amount,
        "payreg_data": payreg_data,
        "payslip_data": payslip_data,
        "payout_data": payout_data,
        "k": myprofile,
        "data": company_data,
        "month_str": month_str,
        "selected_year": selected_year,
    }

    # Store in cache
    cache.set(cache_key, response_data, timeout=600)  # Cache for 10 minutes

    return render(request, 'index/runpayroll_overview.html', response_data)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def payout(request):
    userid = request.user.id
    k = Myprofile.objects.filter(myuser__id=userid)
    c = companyprofile.objects.filter(admin_id=userid)
    today = datetime.now()
    query = request.GET.get('search')

    selected_month_str = request.GET.get('monthselect', None)
    # print("selected_month_str :", selected_month_str)

    if selected_month_str is None:
        selected_month = today.month
        selected_year = today.year
        month_str = today.strftime('%B')
    else:
        selected_month_now = datetime.strptime(
            selected_month_str, '%B %Y').date()
        selected_year = selected_month_now.year
        selected_month = selected_month_now.month
        selected_date = datetime.strptime(selected_month_str, '%B %Y')
        month_str = selected_date.strftime('%B')

    if request.method == "POST":
        payoutid = request.POST.get("payoutid")
        payout = PayoutStatus.objects.get(id=payoutid)
        status = request.POST.get("status")
        reason = request.POST.get("reason")
        payout.status = status
        payout.reason = reason
        payout.save()

    # print("selected_year, selected_month :", selected_month, selected_year)
    user_data = User.objects.filter(Q(id=userid) | Q(admin_id=userid)).order_by('status', 'empid', 'username')
    for i in user_data:
        payout_entry = PayoutStatus.objects.filter(
            user_id=i.id, createddate__year=selected_year, createddate__month=selected_month)
        # print("payout_entry :", payout_entry)
        if not payout_entry:
            payout_entry = PayoutStatus.objects.create(
                user_id=i, status='Unpaid')
            payout_entry.createddate = datetime(
                selected_year, selected_month, 1)
            payout_entry.save()

    payout_data = PayoutStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid)) & Q(user_id__status='Active'))
    payreg_data = PayRegister.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
        Q(user_id__id=userid) | Q(user_id__admin_id=userid)) & Q(user_id__status='Active'))
    # print("payout_data ; payreg_data :", payout_data, payreg_data)

    if query:
        payout_data = PayoutStatus.objects.filter(Q(createddate__year=selected_year, createddate__month=selected_month) & (
            Q(user_id__username__icontains=query) & (Q(user_id=userid) | Q(user_id__admin_id=userid)))).order_by('user_id__username')

    # page = request.GET.get('page', 1)
    # paginator = Paginator(payout_data, 20)
    # try:
    #     datas = paginator.page(page)
    # except PageNotAnInteger:
    #     datas = paginator.page(1)
    # except EmptyPage:
    #     datas = paginator.page(paginator.num_pages)

    x = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "month_str": month_str,
        "selected_year": selected_year,
        "datas": payout_data,
    }
    return render(request, 'index/payout.html', {'user_data': user_data, 'payout_data': payout_data, 'payreg_data': payreg_data, 'query': query, **x})

# Get request



def request_approval(request):
    user = request.user
    if request.POST:
        # for markaspresent and markexacttime modal
        request_approval = int(request.POST.get("request_approval", '0'))
        # for markasleave and mark as LOP modal
        requestApproval = int(request.POST.get("requestApproval", '0'))
        if request_approval == 1:
            punch_id = request.POST.get('punch_id')
            punch = Punch.objects.get(id=punch_id)
            RequestApproval.objects.create(
                user=user, admin_id=user.admin_id, punch_data=punch, reason=None, request_type=1, is_approved=False,)
            punch.is_requested = True
            punch.save()
        elif request_approval == 3:
            inHour = int(request.POST.get('inHour'))
            inMinutes = int(request.POST.get('inMinutes'))
            inMidDay = request.POST.get('inMidDay')
            outHour = int(request.POST.get('outHour'))
            outMinutes = int(request.POST.get('outMinutes'))
            outMidDay = request.POST.get('outMidDay')
            reason = request.POST.get('reason')
            punch_id = request.POST.get('punch_id')
            punch = Punch.objects.get(id=punch_id)

            # converting 12 hours time to 24 hours
            if inMidDay == 'PM':
                if inHour == 12:
                    inHour = 12
                else:
                    inHour += 12
            if outMidDay == 'PM':
                if outHour == 12:
                    outHour = 12
                else:
                    outHour += 12
            in_time = datetime.strptime(
                "{:02d}:{:02d}:00".format(inHour, inMinutes), '%H:%M:%S')
            out_time = datetime.strptime(
                "{:02d}:{:02d}:00".format(outHour, outMinutes), '%H:%M:%S')
            in_time = in_time.strftime("%H:%M:%S")
            out_time = out_time.strftime("%H:%M:%S")
            org_in_time = punch.first_clock_in_time
            org_out_time = punch.second_clock_out_time if punch.is_second_clocked_out else punch.first_clock_out_time

            RequestApproval.objects.create(
                user=user,
                admin_id=user.admin_id,
                punch_data=punch,
                org_in_time=org_in_time,
                org_out_time=org_out_time,
                in_time=in_time,
                out_time=out_time,
                reason=reason,
                request_type=3,
                is_approved=False,
            )
            punch.is_requested = True
            punch.save()

        elif requestApproval == 2:
            print('Something is working and it ok')
            leave_punch_id = request.POST.get('leave_punch_id')
            punch = Punch.objects.get(id=leave_punch_id)
            first_half = request.POST.get('first_half')
            second_half = request.POST.get('second_half')
            leave_type = request.POST.get('leave_type')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            reason = request.POST.get('reason')

            start_date = datetime.strptime(start_date, '%d %B %Y')
            end_date = datetime.strptime(end_date, '%d %B %Y')
            formatted_start_date = datetime.strftime(start_date, "%Y-%m-%d")
            formatted_end_date = datetime.strftime(end_date, "%Y-%m-%d")

            company = CompanyRules.objects.get(id=leave_type)
            print('Company rule: ', company)
            if first_half == "first half" and second_half == "first half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "first half" and second_half == "second half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "second half" and second_half == "first half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "second half" and second_half == "second half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5

            assign_rule = assignrule.objects.filter(
                user_id_id=user.id, rules_applied=company.id).first()
            print("assinged Leave: ",
                  assign_rule.rules_applied.all().values_list('leavename'))
            # checking if the leavetype have any leavebalance left
            if assign_rule.creditedleaves != 0 and assign_rule.leavebalance >= Decimal(delta):
                Leave.objects.create(leavetyp=company.leavename, strtDate=start_date, Reason=reason, Selecthalf1=first_half, endDate=end_date,
                                     Selecthalf2=second_half, status="Applied", Appliedon=datetime.today().date(), Days=delta, applicant_email=user,
                                     admin_id=user.admin_id, punch_data=punch)
                punch.is_requested = True
                punch.save()
            # checking wather the applying leave is not Loss Of Pay, If it's True then this line will be excuted
            elif assign_rule.creditedleaves == 0 and assign_rule.leavebalance == 0 and assign_rule.rules_applied.all().filter(leavename__contains="Loss Off Pay"):
                Leave.objects.create(leavetyp=company.leavename, strtDate=formatted_start_date, Reason=reason, Selecthalf1=first_half, endDate=formatted_end_date,
                                     Selecthalf2=second_half, status='Applied', Appliedon=datetime.today().date(), Days=delta, applicant_email=user,
                                     admin_id=user.admin_id, punch_data=punch)
                punch.is_requested = True
                punch.save()
            # if the leave type is not Loss Of Pay and have no leave balance or creadited leave then the below code will send a message to the User
            else:
                messages.info(request, 'No Leave balance Left')
                if user.role in ['Admin', 'admin']:
                    return redirect('monthly_log')
                else:
                    return redirect('emp_monthly_log')
        elif requestApproval == 4:
            print("the request is number 4")
            leave_punch_id = request.POST.get('leave_punch_id')
            punch = Punch.objects.get(id=leave_punch_id)
            leave_type = request.POST.get('leave_type')
            start_date = request.POST.get('start_date')
            first_half = request.POST.get('first_half')
            second_half = request.POST.get('second_half')
            end_date = request.POST.get('end_date')
            reason = request.POST.get('reason')
            start_date = datetime.strptime(start_date, '%d %B %Y')
            end_date = datetime.strptime(end_date, '%d %B %Y')
            formatted_start_date = datetime.strftime(start_date, "%Y-%m-%d")
            formatted_end_date = datetime.strftime(end_date, "%Y-%m-%d")

            company = CompanyRules.objects.get(id=leave_type)
            if first_half == "first half" and second_half == "first half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "first half" and second_half == "second half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "second half" and second_half == "first half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            elif first_half == "second half" and second_half == "second half":
                delt = end_date - start_date + timedelta(days=0)
                delta = delt.days + 0.5
            Leave.objects.create(leavetyp=company.leavename, strtDate=formatted_start_date, Reason=reason, Selecthalf1=first_half, endDate=formatted_end_date,
                                 Selecthalf2=second_half, status="Applied", Appliedon=datetime.today().date(), Days=delta, applicant_email=user,
                                 admin_id=user.admin_id, punch_data=punch)
            punch.is_requested = True
            punch.save()

        if user.role in ['admin', 'Admin']:
            return redirect('monthly_log')
        else:
            return redirect('emp_monthly_log')


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def list_requests(request):

    user = request.user
    company_profile = companyprofile.objects.filter(admin_id=user.id).first()
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    today = timezone.now()
    query = request.GET.get('monthselect', None)

    if query is not None:
        query = datetime.strptime(query, '%B %Y').strftime('%Y-%m')
        approvals_data = RequestApproval.objects.filter(
            Q(admin_id=user.id) &
            (Q(is_rejected=False) &
             Q(is_approved=False)) &
            Q(created_at__year=query.split("-")[0]) &
            Q(created_at__month=query.split("-")[1])
        ).order_by('-created_at')
    else:
        approvals_data = RequestApproval.objects.filter(
            Q(user__id=user.id) |
            Q(admin_id=user.id) &
            (Q(is_rejected=False) &
             Q(is_approved=False))).order_by('-created_at')

    # page = request.GET.get('page', 1)

    # paginator = Paginator(approvals_data, 10)
    # try:
    #     approvals_data = paginator.page(page)
    # except PageNotAnInteger:
    #     approvals_data = paginator.page(1)
    # except EmptyPage:
    #     approvals_data = paginator.page(paginator.num_pages)
    context = {
        "approvals_data": approvals_data,
        'company_profile': company_profile,
        "k": k[0] if k.exists() else None,
    }
    return render(request, 'index/Approvals.html', context)


def list_approvals(request):
    print("🔍 list_approvals view called")

    user = request.user
    company_profile = companyprofile.objects.filter(admin_id=user.id).first()
    k = Myprofile.objects.filter(myuser__id=request.user.id)
    query = request.GET.get('monthselect', None)

    approvals_data = RequestApproval.objects.filter(
        admin_id=user.id
    ).filter(Q(is_approved=True) | Q(is_rejected=True))

    if query:
        try:
            query_date = datetime.strptime(query, '%B %Y')
            
            approvals_data = approvals_data.filter(
                created_at__year=query_date.year,
                created_at__month=query_date.month
            )
        except ValueError:
            pass


    approvals_data = approvals_data.order_by('-created_at')

    data_list = []
    for a in approvals_data:
        user_url = f"/view_personalinfo{a.user.id}"
        emp_name = f"<a href='{user_url}'>{a.user.username}</a>"
        emp_id = f"<a href='{user_url}'>{a.user.empid}</a>"
        anomalies = []
        if a.punch_data and a.punch_data.in_time_anomaly:
            anomalies.append("In Time")
        if a.punch_data and a.punch_data.out_time_anomaly:
            anomalies.append("Out Time")
        if a.punch_data and a.punch_data.work_duration_anomaly:
            anomalies.append("Work Duration")

        request_type_text = {
            1: "Mark as Present",
            2: "Mark as Leave",
            3: "Mark Exact Time",
            4: "Mark as LOP"
        }.get(a.request_type, "--")

        action_html = ""
        if not a.is_approved and not a.is_rejected:
            # Approve form
            action_html += f'''
            <form action="/accept-request-approval/" method="post">
                <input type="hidden" name="csrfmiddlewaretoken" value="{request.COOKIES.get('csrftoken')}">
                <input type="hidden" name="punch_id" value="{a.punch_data.id if a.punch_data else ''}">
                <input type="hidden" name="request_type" value="{a.request_type}">
                <input type="hidden" name="request_id" value="{a.id}">
                <button class="revert-button" type="submit">Approve</button>
            </form>
            <form action="/reject-request-approval/" method="post" style="margin-top: 5px;">
                <input type="hidden" name="csrfmiddlewaretoken" value="{request.COOKIES.get('csrftoken')}">
                <input type="hidden" name="punch_id" value="{a.punch_data.id if a.punch_data else ''}">
                <input type="hidden" name="request_type" value="{a.request_type}">
                <input type="hidden" name="request_id" value="{a.id}">
                <button class="reject-button" type="submit">Reject</button>
            </form>
            '''
        else:
            action_html = f"<span class='{'is_approved' if a.is_approved else 'is_rejected'}'>{'Approved' if a.is_approved else 'Rejected'}</span>"


        user_ids = [a.user.id for a in approvals_data]
        manager_lookup = {
            uid: {'primary': '', 'secondary': '', 'primary_id': None, 'secondary_id': None}
            for uid in user_ids
        }

        reporting_data = Reportingmanager.objects.prefetch_related('myuser_2').filter(userid__in=user_ids)

        for entry in reporting_data:
            for mgr in entry.myuser_2.all():
                if entry.type == 'Primary':
                    manager_lookup[entry.userid]['primary'] = mgr.username
                    manager_lookup[entry.userid]['primary_id'] = mgr.id
                elif entry.type == 'Secondary':
                    manager_lookup[entry.userid]['secondary'] = mgr.username
                    manager_lookup[entry.userid]['secondary_id'] = mgr.id


        data_list.append({
                "id": emp_id,
                "emp_name": emp_name,
                "date": a.punch_data.date.strftime("%d %B %Y") if a.punch_data and a.punch_data.date else "--",
                "department": str(a.user.department) if a.user.department else "--",
                "manager": f"{manager_lookup[a.user.id]['primary']}   {manager_lookup[a.user.id]['secondary']}" if a.user.id in manager_lookup else "--",
                "in_time": f"<div>{a.org_in_time.strftime('%I:%M %p') if a.org_in_time else '--'}</div><div class='new-time-data'>{a.in_time.strftime('%I:%M %p') if a.in_time else '--'}</div>",
                "out_time": f"<div>{a.org_out_time.strftime('%I:%M %p') if a.org_out_time else '--'}</div><div class='new-time-data'>{a.out_time.strftime('%I:%M %p') if a.out_time else '--'}</div>",
                "work_duration": a.punch_data.work_duration.strftime('%H:%M:%S') if a.punch_data and a.punch_data.work_duration else "--",
                "reason": a.reason if a.reason else "No reason submitted",
                "anomalies": "<br>".join(anomalies) or "--",
                "request_type": request_type_text,
                "status": (
                    '<span class="is_approved">Approved</span>' if a.is_approved else
                    '<span class="is_rejected">Rejected</span>' if a.is_rejected else
                    "--"
                ),
                "action": action_html
            })

    context = {
        "json_data": json.dumps(data_list, cls=DjangoJSONEncoder),
        "company_profile": company_profile,
        "k": k[0] if k.exists() else None,
    }
    return render(request, 'index/Rejectes.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def accept_request_approval(request):

    if request.POST:

        punch_id = request.POST.get('punch_id')
        request_id = request.POST.get('request_id')
        request_type = request.POST.get('request_type')
        punch_data = Punch.objects.get(id=punch_id)
        request_data = RequestApproval.objects.get(id=request_id)
        assign_attendance_rule = AssignAttendanceRule.objects.filter(
            user_id=punch_data.user).first()
        org_in_time = assign_attendance_rule.rules_applied.inTime
        org_out_time = assign_attendance_rule.rules_applied.outTime
        employee = punch_data.user
        company_penalty_rule = CompanyRules.objects.filter(
            admin_id__in=[employee.id, employee.admin_id], leavename='Loss Of Pay').first()
        employee_loss_of_pay_rules = assignrule.objects.filter(
            user_id=employee, rules_applied__id=company_penalty_rule.id if company_penalty_rule else 0).first()

        if int(request_type) == 1 and employee_loss_of_pay_rules:
            if employee_loss_of_pay_rules.penaltydeduction >= 0:
                if punch_data.status == "A":
                    employee_loss_of_pay_rules.penaltydeduction -= 1
                elif punch_data.status == "AN":
                    employee_loss_of_pay_rules.penaltydeduction -= Decimal(0.5)
            punch_data.is_approved = True
            punch_data.first_clock_in_time = org_in_time
            punch_data.first_clock_out_time = org_out_time
            punch_data.in_time_anomaly = False
            punch_data.status = 'P'
            punch_data.out_time_anomaly = False
            punch_data.work_duration_anomaly = False
            request_data.is_approved = True
            punch_data.is_penalty_reverted = True

            punch_data.save()
            request_data.save()
            employee_loss_of_pay_rules.save()
            return redirect('list-request-approvals')
        elif int(request_type) == 3 and employee_loss_of_pay_rules:
            if employee_loss_of_pay_rules.penaltydeduction >= 0:
                if punch_data.status == "A":
                    employee_loss_of_pay_rules.penaltydeduction -= 1
                elif punch_data.status == "AN":
                    employee_loss_of_pay_rules.penaltydeduction -= Decimal(0.5)
            new_in_time = request.POST.get('new_in_time')
            new_out_time = request.POST.get('new_out_time')
            in_time_obj = datetime.strptime(new_in_time, '%I:%M %p')
            out_time_obj = datetime.strptime(new_out_time, '%I:%M %p')
            in_hour = int(in_time_obj.strftime('%I'))
            in_minute = in_time_obj.minute
            in_am_pm = in_time_obj.strftime('%p')
            out_hour = int(out_time_obj.strftime('%I'))
            out_minute = out_time_obj.minute
            out_am_pm = out_time_obj.strftime('%p')
            if in_am_pm == 'PM':
                if in_hour == 12:
                    in_hour = 12
                else:
                    in_hour += 12
            elif out_am_pm == 'PM':
                if out_hour == 12:
                    out_hour = 12
                else:
                    out_hour += 12

            new_in_time = datetime.strptime(
                str(in_hour)+':' + str(in_minute)+':'+'00', '%H:%M:%S')
            new_out_time = datetime.strptime(
                str(out_hour)+':' + str(out_minute)+':'+'00', '%H:%M:%S')
            work_duration = new_out_time - new_in_time
            work_hours, reminder = divmod(work_duration.seconds, 3600)
            work_minutes,  work_second = divmod(reminder, 60)
            work_duration = "{:02d}:{:02d}:{:02d}".format(
                work_hours, work_minutes, work_second)

            punch_data.is_approved = True
            punch_data.first_clock_in_time = new_in_time
            punch_data.first_clock_out_time = new_out_time
            punch_data.in_time_anomaly = False
            punch_data.out_time_anomaly = False
            punch_data.work_duration_anomaly = False
            punch_data.work_duration = work_duration
            punch_data.status = 'P'
            request_data.is_approved = True
            punch_data.is_penalty_reverted = True

            punch_data.save()
            request_data.save()
            employee_loss_of_pay_rules.save()
            return redirect('list-request-approvals')

    return redirect('list-request-approvals')


def reject_request_approval(request):
    punch_id = request.POST.get('punch_id')
    request_id = request.POST.get('request_id')
    punch_data = Punch.objects.get(id=punch_id)
    request_data = RequestApproval.objects.get(id=request_id)
    punch_data.is_rejected = True
    request_data.is_rejected = True
    punch_data.save()
    request_data.save()
    return redirect('list-request-approvals')

def list_comp_off(request):
    user = request.user
    c = companyprofile.objects.filter(admin_id=user.id)
    k = Myprofile.objects.filter(myuser__id=user.id)
    page = request.GET.get('page', 1)
    comp_off_data = CompOff.objects.filter(
        Q(user__admin_id=user.id) | Q(user__id=user.id) & Q(user__status='Active'))

    # paginator = Paginator(comp_off_data, 10)

    # try:
    #     comp_off_data = paginator.page(page)
    # except PageNotAnInteger:
    #     comp_off_data = paginator.page(1)
    # except EmptyPage:
    #     comp_off_data = paginator.page(paginator.num_pages)

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "comp_off_data": comp_off_data,
    }
    return render(request, "index/CompOff.html", context)


def emp_list_comp_off(request):
    user = request.user
    c = companyprofile.objects.filter(admin_id=user.admin_id)
    k = Myprofile.objects.filter(myuser__id=user.id)
    page = request.GET.get('page', 1)
    comp_off_data = CompOff.objects.filter(user__id=user.id)
    paginator = Paginator(comp_off_data, 10)
    try:
        comp_off_data = paginator.page(page)
    except PageNotAnInteger:
        comp_off_data = paginator.page(1)
    except EmptyPage:
        comp_off_data = paginator.page(paginator.num_pages)

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "comp_off_data": comp_off_data,
    }
    return render(request, "Employee/CompOff.html", context)


def export_approvals_request(request):
    user = request.user
    if request.method == 'POST':
        month_str = request.POST.get('month')
        month = datetime.strptime(month_str, '%B %Y')
        year_selected = month.year
        month_selected = month.month
        request_data = RequestApproval.objects.filter(
            Q(user__id=user.id) | Q(admin_id=user.id) &
            (Q(is_rejected=False) & Q(is_approved=False)) &
            Q(created_at__year=year_selected) &
            Q(created_at__month=month_selected)
        )

        automation_list = []
        for iterator in request_data:

            punch_data = {
                "Employee Id": iterator.user.empid,
                "Employee Name": iterator.user.username,
                "Employee Manager": iterator.user.reptmgr,
                "Orginal In Time": iterator.org_in_time.strftime("%I:%M %p ") if iterator.org_in_time else "00:00:00",
                "Original Out Time": iterator.org_out_time.strftime("%I:%M %p") if iterator.org_out_time else "00:00:00",
                "Requested In Time": iterator.in_time.strftime("%I:%M %p") if iterator.in_time else "00:00:00",
                "Requested Out Time": iterator.out_time.strftime("%I:%M %p ") if iterator.out_time else "00:00:00",
                "Work Duration": iterator.punch_data.work_duration,
                "Reason": iterator.reason if iterator.reason else "--",
                "In Time Anomaly": 'Yes' if iterator.punch_data.in_time_anomaly else 'No',
                "Out Time Anomaly": 'Yes' if iterator.punch_data.out_time_anomaly else 'No',
                "Work Duration Anomaly": 'Yes' if iterator.punch_data.work_duration_anomaly else 'No',
                "Request Type": (
                    'Mark as Present' if iterator.request_type == 1 else
                    ('Mark as Leave' if iterator.request_type == 2 else
                     ('Mark Exact Time' if iterator.request_type == 3 else
                      ('Mark as LOP' if iterator.request_type == 4 else "--")))
                ),
                "Date": iterator.created_at.strftime('%d-%m-%Y'),
            }

            automation_list.append(punch_data)
        df = pd.DataFrame(automation_list)
        excel_file_name = 'Requests.xlsx'
        path = os.path.join(BASE_DIR, 'media/csv/Requests.xlsx')

        df.to_excel(path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        for col in sheet.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='left', vertical='center')
                cell.alignment = alignment_obj

        wb.save(path)

        today = datetime.now()
        filemanager = Filemanager.objects.create(myuser_10=request.user, requesttype="Request Export",
                                                 frmt="XLSX", scheduleon=today, status="In Queue")
        # subject = 'Automation Log '
        # message = 'Please find attached the automation log data for your review and analysis, which will provide valuable insights for optimizing our business processes.'
        # from_email = settings.DEFAULT_FROM_EMAIL
        # to_email = [adminid.email for adminid in admin_id]
        # email = EmailMessage(subject, message, from_email, to_email)
        # email.attach(excel_file_name, open(path, 'rb').read(),
        #                  'application/vnd.ms-excel')
        # email.send()

        filemanager.status = "Success"
        with open(path, 'rb') as file:
            filemanager.saveexcel.save(
                'Requests.xlsx', ContentFile(file.read()))
    return redirect("filemanagernav")


def export_request_approved_or_rejected(request):
    user = request.user
    if request.method == 'POST':
        month_str = request.POST.get('month')
        month = datetime.strptime(month_str, '%B %Y')
        year_selected = month.year
        month_selected = month.month

        request_data = RequestApproval.objects.filter(
            Q(user__id=user.id) | Q(admin_id=user.id) &
            (Q(is_rejected=True) | Q(is_approved=True)) &
            Q(created_at__year=year_selected) &
            Q(created_at__month=month_selected)
        )

        print('Request: ', request_data)
        automation_list = []
        for iterator in request_data:

            punch_data = {
                "Employee Id": iterator.user.empid,
                "Employee Name": iterator.user.username,
                "Employee Manager": iterator.user.reptmgr,
                "Orginal In Time": iterator.org_in_time.strftime("%I:%M %p ") if iterator.org_in_time else "00:00:00",
                "Original Out Time": iterator.org_out_time.strftime("%I:%M %p") if iterator.org_out_time else "00:00:00",
                "Requested In Time": iterator.in_time.strftime("%I:%M %p") if iterator.in_time else "00:00:00",
                "Requested Out Time": iterator.out_time.strftime("%I:%M %p ") if iterator.out_time else "00:00:00",
                "Work Duration": iterator.punch_data.work_duration,
                "Reason": iterator.reason if iterator.reason else "--",
                "In Time Anomaly": 'Yes' if iterator.punch_data.in_time_anomaly else 'No',
                "Out Time Anomaly": 'Yes' if iterator.punch_data.out_time_anomaly else 'No',
                "Work Duration Anomaly": 'Yes' if iterator.punch_data.work_duration_anomaly else 'No',
                "Request Type": (
                    'Mark as Present' if iterator.request_type == 1 else
                    ('Mark as Leave' if iterator.request_type == 2 else
                     ('Mark Exact Time' if iterator.request_type == 3 else
                      ('Mark as LOP' if iterator.request_type == 4 else "--")))
                ),
                "status": ("Accepted" if iterator.is_approved else
                           ("Rejected" if iterator.is_rejected else "--")
                           ),
                "Date": iterator.created_at.strftime('%d-%m-%Y'),
            }

            automation_list.append(punch_data)
        df = pd.DataFrame(automation_list)
        excel_file_name = 'Requests.xlsx'
        path = os.path.join(BASE_DIR, 'media/csv/Requests.xlsx')

        df.to_excel(path, index=False, sheet_name='Sheet1')
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        for col in sheet.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='left', vertical='center')
                cell.alignment = alignment_obj

        wb.save(path)

        today = datetime.now()
        filemanager = Filemanager.objects.create(myuser_10=request.user, requesttype="Request Export",
                                                 frmt="XLSX", scheduleon=today, status="In Queue")
        # subject = 'Automation Log '
        # message = 'Please find attached the automation log data for your review and analysis, which will provide valuable insights for optimizing our business processes.'
        # from_email = settings.DEFAULT_FROM_EMAIL
        # to_email = [adminid.email for adminid in admin_id]
        # email = EmailMessage(subject, message, from_email, to_email)
        # email.attach(excel_file_name, open(path, 'rb').read(),
        #                  'application/vnd.ms-excel')
        # email.send()

        filemanager.status = "Success"
        with open(path, 'rb') as file:
            filemanager.saveexcel.save(
                'Requests.xlsx', ContentFile(file.read()))
    return redirect("filemanagernav")


def penalty_log_revert(request):
    # user = request.user
    if request.method == 'POST':
        punch_id = request.POST.get('punch_id')
        punch_data = None
        monthyearselect = request.POST.get("monthyearselect")
        page = request.POST.get('page', 1)
        empname = request.POST.get('empname')

        try:
            punch_data = Punch.objects.get(id=punch_id)
        except Punch.DoesNotExist:
            messages.info(request, 'Punch data not present')
            return redirect('automationlog')
        if punch_data.status in ['A', 'AN']:
            employee = punch_data.user
            assign_lop = assignrule.objects.filter(
                user_id=employee, rules_applied__leavename='Loss Of Pay').first()
            if assign_lop and punch_data.status == 'A':
                assign_lop.penaltydeduction -= 1
                assign_lop.save()
                punch_data.is_penalty_reverted = True
                punch_data.save()
            elif assign_lop and punch_data.status == 'AN':
                assign_lop.penaltydeduction -= Decimal(0.5)
                assign_lop.save()
                punch_data.is_penalty_reverted = True
                punch_data.save()
            else:
                messages.info(request, 'Leave rule not assigned')
                return redirect('automationlog')
        else:
            messages.info(request, 'something went wrong!')
            return redirect('automationlog')

        if empname:
            return redirect(reverse('automationlog') + '?monthselect=' + monthyearselect + '&employee=' + empname + '&page=' + page)
        else:
            return redirect(reverse('automationlog') + '?monthselect=' + monthyearselect + '&page=' + page)


def comp_off_revert(request):

    if request.method == 'POST':
        punch_id = request.POST.get('punch_id')
        try:
            punch_data = Punch.objects.get(id=punch_id)
        except Punch.DoesNotExist:
            messages.info(request, 'Punch Data not present')
            return redirect('list-comp-off')

        employee = punch_data.user
        assign_comp_off = assignrule.objects.filter(
            user_id=employee, rules_applied__leavename='Comp Off').first()
        if assign_comp_off:
            assign_comp_off.creditedleaves -= Decimal(1)
            assign_comp_off.leavebalance -= Decimal(1)
            assign_comp_off.save()
            punch_data.is_compoff_reverted = True
            punch_data.save()
        else:
            messages.info(request, 'Leave rule not assigned')
            return redirect('list-comp-off')
    return redirect('list-comp-off')


def show_work_duration(request):

    now = timezone.now()
    user = request.user

    in_time = time(hour=0, minute=0, second=0)
    if user.is_authenticated:
        # print("User: ", user)
        try:
            punch_data = Punch.objects.get(user=user, date__date=now.date())
            in_time = punch_data.first_clock_in_time
        except Punch.DoesNotExist:
            pass
    # print("IN Time: ", in_time)
    return JsonResponse({"in_time": in_time, "user": user.username})


import json
from django.core.serializers.json import DjangoJSONEncoder

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def search_punch_logs(request):
    userid = request.user.id
    c = companyprofile.objects.filter(admin_id=userid)
    k = Myprofile.objects.filter(myuser__id=userid)
    user_data = User.objects.filter(Q(id=userid) | Q(admin_id=userid))
    today = datetime.now()
    formated = today.strftime("%B %Y")

    Punch_obj = Punch.objects.filter(user__in=user_data, date__year=today.year, date__month=today.month)

    username = ''
    monthselect = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        monthselect = request.POST.get('monthselect')
        month_year = datetime.strptime(monthselect, "%B %Y")

        if username:
            Punch_obj = Punch.objects.filter(user_id=username, date__year=month_year.year, date__month=month_year.month)
        else:
            Punch_obj = Punch.objects.filter(user__in=user_data, date__year=month_year.year, date__month=month_year.month)

    for punch in Punch_obj:
        total_work_duration = timedelta()
        break_duration = timedelta()

        if punch.first_clock_in_time and punch.first_clock_out_time and punch.second_clock_in_time and punch.second_clock_out_time:
            first_duration = datetime.combine(datetime.today(), punch.first_clock_out_time) - datetime.combine(datetime.today(), punch.first_clock_in_time)
            second_duration = datetime.combine(datetime.today(), punch.second_clock_out_time) - datetime.combine(datetime.today(), punch.second_clock_in_time)
            total_work_duration += first_duration + second_duration
            break_duration = datetime.combine(datetime.today(), punch.second_clock_in_time) - datetime.combine(datetime.today(), punch.first_clock_out_time)

        elif punch.is_first_clocked_in and punch.first_clock_in_time and punch.first_clock_out_time:
            total_work_duration += datetime.combine(datetime.today(), punch.first_clock_out_time) - datetime.combine(datetime.today(), punch.first_clock_in_time)

        total_hours = int(total_work_duration.total_seconds() // 3600)
        total_minutes = int((total_work_duration.total_seconds() % 3600) // 60)
        punch.work_duration = f"{total_hours} Hours {total_minutes} Mins" if total_work_duration.total_seconds() else "-- -- --"

        break_hours = int(break_duration.total_seconds() // 3600)
        break_minutes = int((break_duration.total_seconds() % 3600) // 60)
        punch.break_duration = f"{break_hours} Hours {break_minutes} Mins" if break_duration.total_seconds() else "-- -- --"

        assignattendancerule = AssignAttendanceRule.objects.filter(user_id=punch.user_id)
        total_time = timedelta()
        overtime = timedelta()

        for assignrule in assignattendancerule:
            in_Time = datetime.combine(datetime.today(), assignrule.rules_applied.inTime)
            out_Time = datetime.combine(datetime.today(), assignrule.rules_applied.outTime)
            total_time += out_Time - in_Time

        if total_work_duration > total_time:
            overtime += total_work_duration - total_time

        overtime_hours = int(overtime.total_seconds() // 3600)
        overtime_minutes = int((overtime.total_seconds() % 3600) // 60)
        punch.overtime = f"{overtime_hours} Hours {overtime_minutes} Mins" if overtime.total_seconds() else "-- -- --"

    punch_data = [
        {
            "emp_id": punch.user.empid,
            'user_id': punch.user.id,
            "username": punch.user.username,
            "status": punch.status,
            "date": punch.date.strftime("%d %B %Y") if punch.date else "-- -- --",
            "in_time": punch.first_clock_in_time.strftime('%I:%M %p') if punch.first_clock_in_time else "-- -- --",
            "out_time": punch.first_clock_out_time.strftime('%I:%M %p') if punch.first_clock_out_time else "-- -- --",
            "work_duration": punch.work_duration if punch.work_duration else "-- -- --",
            "overtime": punch.overtime if punch.overtime else "-- -- --",
            "break_duration": punch.break_duration if punch.break_duration else "-- -- --",
            "break_count": punch.break_count if punch.break_count else "-- -- --",
            "location": punch.user.wrklcn.location if hasattr(punch.user, 'wrklcn') and punch.user.wrklcn else "-- -- --",
            "department": punch.user.department.name if hasattr(punch.user, 'department') and punch.user.department else "-- -- --",
            "anomalies": "1" if punch.status == "AN" else "-- -- --",
            "wfhwfo": punch.WfhOrWfo if punch.WfhOrWfo else "-- -- --"
        }
        for punch in Punch_obj
    ]

    context = {
        "k": k[0] if k.exists() else k,
        "c": c[0] if c.exists() else c,
        "user_data": user_data,
        "formated": formated,
        "Punch_obj": Punch_obj,
        "username": username,
        "monthselect": monthselect,
        "punch_data_json": json.dumps(punch_data, cls=DjangoJSONEncoder)
    }

    return render(request, 'index/punch_logs.html', context)

from io import BytesIO
from django.core.files import File
from .models import Filemanager
import os

def export_punch_logs(request):
    if request.method == 'POST':
        starting_date_str = request.POST.get('startingdate')
        ending_date_str = request.POST.get('endingdate')
        username = request.POST.get('username')  
        employee_id = request.POST.get('employee_id')  

        try:
            start_date = datetime.strptime(starting_date_str, '%d/%m/%Y')
            end_date = datetime.strptime(ending_date_str, '%d/%m/%Y') + timedelta(days=1)
        except ValueError as e:
            return HttpResponse("Invalid date format provided.", status=400)

        try:
            employee = User.objects.get(username=username, id=employee_id)
        except User.DoesNotExist:
            return HttpResponse("Employee not found.", status=404)

        punchdata = Punch.objects.filter(
            date__range=[start_date, end_date],
            user=employee
        ).order_by('date')

        data_list = []

        for punch in punchdata:
            total_work_duration = timedelta()
            break_duration = timedelta()

            if punch.first_clock_in_time:
                first_clock_in = datetime.combine(datetime.today(), punch.first_clock_in_time)
                if punch.first_clock_out_time:
                    first_clock_out = datetime.combine(datetime.today(), punch.first_clock_out_time)
                    total_work_duration += first_clock_out - first_clock_in

                if punch.second_clock_in_time:
                    second_clock_in = datetime.combine(datetime.today(), punch.second_clock_in_time)
                    if punch.second_clock_out_time:
                        second_clock_out = datetime.combine(datetime.today(), punch.second_clock_out_time)
                        total_work_duration += second_clock_out - second_clock_in

                        if punch.first_clock_out_time and punch.second_clock_in_time:
                            break_duration = second_clock_in - first_clock_out

            total_hours = int(total_work_duration.total_seconds() // 3600)
            total_minutes = int((total_work_duration.total_seconds() % 3600) // 60)
            punch.work_duration = f"{total_hours} Hours {total_minutes} Mins" if total_hours or total_minutes else "-- -- --"

            break_hours = int(break_duration.total_seconds() // 3600)
            break_minutes = int((break_duration.total_seconds() % 3600) // 60)
            punch.break_duration = f"{break_hours} Hours {break_minutes} Mins" if break_hours or break_minutes else "-- -- --"

            assignattendancerule = AssignAttendanceRule.objects.filter(user_id=punch.user_id)
            total_time = timedelta()
            for assignrule in assignattendancerule:
                inTime = assignrule.rules_applied.inTime
                outTime = assignrule.rules_applied.outTime
                if inTime and outTime:
                    total_time += datetime.combine(datetime.today(), outTime) - datetime.combine(datetime.today(), inTime)

            overtime = max(total_work_duration - total_time, timedelta())
            overtime_hours = int(overtime.total_seconds() // 3600)
            overtime_minutes = int((overtime.total_seconds() % 3600) // 60)
            punch.overtime = f"{overtime_hours} Hours {overtime_minutes} Mins" if overtime_hours or overtime_minutes else "-- -- --"

            anomaly = '1' if punch.status == "AN" else ''
            outtime = (
                punch.first_clock_out_time.strftime('%I:%M %p') if punch.first_clock_out_time and (punch.is_first_clocked_out and not punch.is_second_clocked_in)
                else punch.second_clock_out_time.strftime('%I:%M %p') if punch.second_clock_out_time and punch.is_second_clocked_out
                else '-- -- --'
            )

            data_list.append({
                'Employee ID': punch.user.empid,
                'Name': punch.user.username,
                'Date': punch.date.strftime('%d %B %Y') if punch.date else '-- -- --',
                'Status': punch.status,
                'In Time': punch.first_clock_in_time.strftime('%I:%M %p') if punch.first_clock_in_time else '-- -- --',
                'Out Time': outtime,
                'Work Duration': punch.work_duration,
                'Overtime Duration': punch.overtime,
                'Break Duration': punch.break_duration,
                'Break Count': punch.break_count,
                'Outstanding Anomalies': anomaly,
            })

        df = pd.DataFrame(data_list)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Punch Logs')
            sheet = writer.sheets['Punch Logs']
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 5, 50)

        output.seek(0)

        # Save to disk
        excel_file_name = 'punch_log.xlsx'
        save_path = os.path.join(settings.BASE_DIR, 'media/csv', excel_file_name)
        with open(save_path, 'wb') as f:
            f.write(output.getvalue())

        # Create Filemanager record
        filemanager = Filemanager.objects.create(
            myuser_10=request.user,
            requesttype="Punch Log Export",
            frmt="XLSX",
            scheduleon=now(),
            status="Success"
        )

        # Save the file to FileField
        filemanager.saveexcel.save(excel_file_name, ContentFile(output.getvalue()))
        filemanager.save()

        # Send Email
        subject = 'Punch Data Export'
        message = 'Attached is the punch data for the specified date range.'
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [request.user.email] if request.user.email else []
        if to_email:
            email = EmailMessage(subject, message, from_email, to_email)
            email.attach(excel_file_name, output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()

        # Return Excel file as HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{excel_file_name}"'
        return response

    return redirect("attendance_logs")



def get_usernames(request):
    admin_id = request.user.id
    users = User.objects.filter(Q(id=admin_id) | Q(admin_id=admin_id), Q(status='Active'))  
    data = [{'username': user.username} for user in users]
    return JsonResponse(data, safe=False)


def update_inactive_employee(request, userid):
    
    user_id = request.user.id
    employee = User.objects.get(id=userid)

    m = Myprofile.objects.filter(myuser__id=request.user.id)
    c = companyprofile.objects.filter(admin_id=user_id)

   
    dn = Designation.objects.filter(admin_id=user_id)
    dp = Department.objects.filter(admin_id=user_id)
    sd = Subdepartment.objects.filter(admin_id=user_id)
    jb = Job.objects.filter(admin_id=user_id)
    wr = Worklocation.objects.filter(admin_id=user_id)
    rp = User.objects.filter(Q(id=user_id) | Q(admin_id=user_id))

    x = {
        "m": m[0] if m.exists() else m,
        "c": c[0] if c.exists() else c,
    }

    if request.method == 'POST':
       
        name = request.POST.get('username')
        em = request.POST.get('email')
        ph = request.POST.get('phone')
        gen = request.POST.get('gender')
        dob = request.POST.get('dob')
        status = request.POST.get('status')
        emptype = request.POST.get('emptype')
        pd = request.POST.get('probperiod')
        designation = request.POST.get('designation')
        des = Designation.objects.get(id=designation)
        department = request.POST.get('department')
        dep = Department.objects.get(id=department)
        jobtitle = request.POST.get('jobtitle')
        job = Job.objects.get(id=jobtitle)
        wrklcn = request.POST.get('wrklcn')
        wrk = Worklocation.objects.get(id=wrklcn)
        reptmgr = request.POST.get('reptmgr')
        subdepartment = request.POST.get('subdepartment')
        Joindate = request.POST.get('Joindate')
        company_type_id = request.POST.get('company_type')
        if company_type_id:
            company_type = companyprofile.objects.get(id=company_type_id)
            employee.company_type = company_type 

       
        employee.username = name
        employee.email = em
        employee.phone = ph
        employee.gender = gen
        employee.status = status
        employee.emptype = emptype
        employee.dob = dob
        employee.probperiod = pd
        employee.designation = des
        employee.department = dep
        employee.jobtitle = job
        employee.wrklcn = wrk
        employee.datejoin = Joindate

        if User.objects.filter(id = reptmgr).exists():
            rep = User.objects.get(id=reptmgr)
            if rep.email != employee.reptmgr.email:
                to = [employee.email]
                subject = "Changed Reporting Manager"
                text_content = f"Dear {employee.username},\n\nYour reporting manager has been changed. Your new reporting manager is {rep.username}.\n\nBest regards,\nYour Company"
                html_content = f"""
                <p>Dear {employee.username},</p>
                <p>Your reporting manager has been changed. Your new reporting manager is <strong>{rep.username} </strong>.</p>
                <p>Best regards,<br>Cydez Technologies</p>
                """
                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=text_content,
                    from_email=settings.EMAIL_HOST_USER,
                    to=to
                )
                msg.attach_alternative(html_content, "text/html")
                msg.send()
                
        
        if User.objects.filter(email=em).exclude(id=employee.id).exists():
            msg1 = 'Email already taken. Please try another one.'
            return render(request, "index/edit_inactive_employee.html",
                          {'employee': employee, 'msg1': msg1, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr,
                           'rp': rp, **x})

        if User.objects.filter(phone=ph).exclude(id=employee.id).exists():
            msg2 = 'Phone number already taken. Please try another one.'
            return render(request, "index/edit_inactive_employee.html",
                          {'employee': employee, 'msg2': msg2, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr,
                           'rp': rp, **x})

        if reptmgr == '' and subdepartment != '':
            sub = Subdepartment.objects.get(id=subdepartment)
            employee.subdepartment = sub
        elif subdepartment == '' and reptmgr != '':
            rep = User.objects.get(id=reptmgr)
            employee.reptmgr = rep
        elif reptmgr != '' and subdepartment != '':
            rep = User.objects.get(id=reptmgr)
            sub = Subdepartment.objects.get(id=subdepartment)
            employee.reptmgr = rep
            employee.subdepartment = sub
        elif reptmgr == '' and subdepartment == '':
            employee.username = name
            employee.email = em
            employee.phone = ph
            employee.gender = gen
            employee.dob = dob
            employee.status = status
            employee.emptype = emptype
            employee.probperiod = pd
            employee.designation = des
            employee.department = dep
            employee.jobtitle = job
            employee.wrklcn = wrk
            employee.datejoin = Joindate
            employee.save()
            return redirect('list_inactive_employee')
        employee.save()

        return redirect('list_inactive_employee')

    return render(request, "index/edit_inactive_employee.html",
                  {'employee': employee, 'dn': dn, 'dp': dp, 'sd': sd, 'jb': jb, 'wr': wr, 'rp': rp,'company_profiles': c, **x})


@login_required
@csrf_exempt
def attendance_summary(request):
    if request.method == "POST":
        user_id = request.POST.get('user_id')
        new_status = request.POST.get('new_status')

        if not user_id or not user_id.isdigit():
            return JsonResponse({'status': 'error', 'message': 'Invalid user ID'}, status=400)

        user = get_object_or_404(User, id=int(user_id))
        user.status = new_status
        user.save()

        return JsonResponse({'status': 'success', 'new_status': new_status})

    # Handle selected date
    date_str = request.GET.get('date')
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    except ValueError:
        current_date = date.today()

    # Timezone-aware date range for filtering Punches
    start_of_day = datetime.combine(current_date, time.min)
    end_of_day = datetime.combine(current_date, time.max)

    # ✅ Filter only valid punches with actual clock-in
    valid_punches = Punch.objects.filter(
        date__range=(start_of_day, end_of_day),
        user__status__iexact="Active",
        user__is_superuser=False,
        first_clock_in_time__isnull=False
    ).exclude(
        Q(user__resignationform__status='Approved') &
        Q(user__resignationform__actual_last_working_day__lt=date.today())
    ).select_related('user')

    # ✅ Clocked-in user IDs
    clocked_in_user_ids = [p.user.id for p in valid_punches]

    # Build Clocked In User Data
    clocked_in_users_data = []
    for punch in valid_punches:
        user = punch.user

        # WFH Location
        try:
            geo = EmployeeGeoFence.objects.get(user=user)
            if geo.home_lat and geo.home_lon:
                cache_key = f"location_{geo.home_lat}_{geo.home_lon}"
                home_location = cache.get(cache_key)
                if not home_location:
                    try:
                        geolocator = Nominatim(user_agent="hrms-attendance")
                        location = geolocator.reverse((geo.home_lat, geo.home_lon), timeout=5)
                        home_location = location.address if location else "Unknown Location"
                        cache.set(cache_key, home_location, timeout=86400)
                    except GeocoderTimedOut:
                        home_location = "Timeout"
                    except Exception:
                        home_location = "Error"
            else:
                home_location = "Not Set"
        except EmployeeGeoFence.DoesNotExist:
            home_location = "Not Set"

        # WFO Location
        try:
            profile = EmployeeProfile.objects.get(user=user)
            wfo_location = profile.branch_location.name if profile.branch_location else "Not Assigned"
        except EmployeeProfile.DoesNotExist:
            wfo_location = "Not Assigned"

        clocked_in_users_data.append({
            'user': user,
            'id': user.id,
            'username': user.username,
            'in_time': punch.first_clock_in_time,
            'out_time': punch.first_clock_out_time,
            'home_location': home_location,
            'wfo_location': wfo_location,
        })

    # ✅ Not clocked-in users
    not_clocked_in_users = User.objects.filter(
        status__iexact="Active",
        is_superuser=False
    ).exclude(id__in=clocked_in_user_ids).exclude(
        Q(resignationform__status='Approved') &
        Q(resignationform__actual_last_working_day__lt=date.today())
    )

    # Approved leaves on current date
    approved_leaves_today_qs = Leave.objects.filter(
        status__iexact="Approved",
        strtDate__lte=current_date,
        endDate__gte=current_date
    )

    # Leave mapping by applicant_email (FK to User)
    leave_reason_map = {
        leave.applicant_email.id: f"{leave.leavetyp or ''} - {leave.Reason or ''}".strip(" -")
        for leave in approved_leaves_today_qs
        if leave.applicant_email
    }

    # Build Not Logged In list
    not_logged_with_leave_flag = []
    for user in not_clocked_in_users:
        leave_info = leave_reason_map.get(user.id)
        leave_info = leave_reason_map.get(user.id)

        # WFH Location
        try:
            geo = EmployeeGeoFence.objects.get(user=user)
            if geo.home_lat and geo.home_lon:
                cache_key = f"location_{geo.home_lat}_{geo.home_lon}"
                home_location = cache.get(cache_key)
                if not home_location:
                    try:
                        geolocator = Nominatim(user_agent="hrms-attendance")
                        location = geolocator.reverse((geo.home_lat, geo.home_lon), timeout=5)
                        home_location = location.address if location else "Unknown Location"
                        cache.set(cache_key, home_location, timeout=86400)
                    except GeocoderTimedOut:
                        home_location = "Timeout"
                    except Exception:
                        home_location = "Error"
            else:
                home_location = "Not Set"
        except EmployeeGeoFence.DoesNotExist:
            home_location = "Not Set"

        # WFO Location
        try:
            profile = EmployeeProfile.objects.get(user=user)
            wfo_location = profile.branch_location.name if profile.branch_location else "Not Assigned"
        except EmployeeProfile.DoesNotExist:
            wfo_location = "Not Assigned"

        not_logged_with_leave_flag.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'status': user.status,
            'on_leave': leave_info is not None,
            'leave_reason': leave_info or '',
            'home_location': home_location,
            'wfo_location': wfo_location,
        })

    return render(request, 'index/summary.html', {
        'date': current_date,
        'today': date.today(),
        'clocked_in_users': clocked_in_users_data,
        'not_logged_users_with_flag': not_logged_with_leave_flag,
    })


@login_required(login_url='login')
def monthly_employee_movements(request):
    current_date = timezone.now().date()

    # Calculate this and last month ranges
    this_month_start = current_date.replace(day=1)
    this_month_end = current_date.replace(day=monthrange(current_date.year, current_date.month)[1])
    
    last_month = this_month_start - timedelta(days=1)
    last_month_start = last_month.replace(day=1)
    last_month_end = last_month

    def parse_date_safe(value):
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except:
                try:
                    return datetime.strptime(value, "%d %B %Y").date()
                except:
                    return None
        return value

    # Filter based on user type (optional future enhancement)
    users = User.objects.filter(Q(status__iexact="Active") | Q(status__iexact="Resigned"))

    # JOINED employees
    joined_last_month = [
        u for u in users if (d := parse_date_safe(u.datejoin)) and last_month_start <= d <= last_month_end
    ]
    joined_this_month = [
        u for u in users if (d := parse_date_safe(u.datejoin)) and this_month_start <= d <= this_month_end
    ]

    # RESIGNED employees
    resigned_last_month = ResignationForm.objects.filter(
        last_workingday__range=(last_month_start, last_month_end)
    ).select_related("user")

    resigned_this_month = ResignationForm.objects.filter(
        last_workingday__range=(this_month_start, this_month_end)
    ).select_related("user")

    context = {
        "joined_last_month": joined_last_month,
        "joined_this_month": joined_this_month,
        "resigned_last_month": resigned_last_month,
        "resigned_this_month": resigned_this_month,
        "this_month": this_month_start.strftime("%B"),
        "last_month": last_month_start.strftime("%B"),
    }

    return render(request, "index/monthly_movements.html", context)


# ------------------------------------------------------------------------------

@login_required
@allowed_users(allowed_roles=['Employee'], allowed_statuses=['Active'])
def set_home_location(request):
    if request.method == 'POST':
        lat = request.POST.get('home_lat')
        lon = request.POST.get('home_lon')

        try:
            lat = round(float(lat), 6)
            lon = round(float(lon), 6)
        except (ValueError, TypeError):
            messages.error(request, "Invalid location data. Please try again.")
            return redirect('set_home_location')

        try:
            fence = EmployeeGeoFence.objects.get(user=request.user)
            # If home already set, block update
            if fence.home_lat and fence.home_lon:
                messages.warning(request, "Home location already set. Kindly contact HR for reset the Location.")
                return redirect("empdash")
        except EmployeeGeoFence.DoesNotExist:
            fence = EmployeeGeoFence(user=request.user)

        # Save new home location
        fence.home_lat = lat
        fence.home_lon = lon
        fence.save()

        messages.success(request, f"✅Home location saved successfully. Pending for admin approval. Kindly contact HR for quick resolution.")
        return redirect('empdash')

    return render(request, 'Employee/set_home_location.html')



from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from django.core.cache import cache


@login_required
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def view_all_home_locations(request):
    geolocator = Nominatim(user_agent="geoapi")

    fences = EmployeeGeoFence.objects.select_related('user').all()
    total_set_locations = fences.filter(home_lat__isnull=False, home_lon__isnull=False).count()

    for fence in fences:
        if fence.home_lat and fence.home_lon:
            cache_key = f"location_{fence.home_lat}_{fence.home_lon}"
            location_name = cache.get(cache_key)

            if not location_name:
                try:
                    location = geolocator.reverse((fence.home_lat, fence.home_lon), timeout=5)
                    location_name = location.address if location else "Unknown Location"
                    cache.set(cache_key, location_name, timeout=86400)
                except GeocoderTimedOut:
                    location_name = "Timeout while fetching"
                except Exception:
                    location_name = "Error fetching location"
            fence.location_name = location_name
        else:
            fence.location_name = "Location not set"

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        radius = request.POST.get('radius')

        if 'delete' in request.POST:
            try:
                EmployeeGeoFence.objects.get(user_id=user_id).delete()
                messages.success(request, "🗑️ Home location deleted successfully.")
            except EmployeeGeoFence.DoesNotExist:
                messages.error(request, "❌ User not found.")
        else:
            try:
                fence = EmployeeGeoFence.objects.get(user_id=user_id)
                fence.home_radius = radius
                fence.save()
                messages.success(request, "✅ Home radius updated.")
            except EmployeeGeoFence.DoesNotExist:
                messages.error(request, "❌ User not found.")

        return redirect('view_all_home_locations')  # Redirect after POST

    return render(request, 'index/view_home_locations.html', {
        'fences': fences,
        'total_set_locations': total_set_locations
    })


@login_required
def add_branch_location(request):
    if not request.user.role.lower() == 'admin':
        messages.error(request, "Access denied.")
        return redirect("empdash")

    if request.method == "POST":
        form = BranchLocationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Branch location added successfully.")
            return redirect("view_branch_locations")  # if you have a view for listing
    else:
        form = BranchLocationForm()

    return render(request, "index/add_branch_location.html", {"form": form})



@login_required
def view_branch_locations(request):
    branches = CompanyBranchLocation.objects.all()
    return render(request, 'index/view_branch_locations.html', {'branches': branches})

@login_required
def add_branch_location(request):
    if request.method == 'POST':
        form = BranchLocationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch location added successfully.")
            return redirect('view_branch_locations')
    else:
        form = BranchLocationForm()
    return render(request, 'index/add_branch_location.html', {'form': form})

@login_required
def edit_branch_location(request, pk):
    branch = get_object_or_404(CompanyBranchLocation, pk=pk)
    if request.method == 'POST':
        form = BranchLocationForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch location updated successfully.")
            return redirect('view_branch_locations')
    else:
        form = BranchLocationForm(instance=branch)
    return render(request, 'index/edit_branch_location.html', {'form': form, 'branch': branch})


from django.views.decorators.http import require_POST

@require_POST
@login_required
def delete_branch_location(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        branch_id = request.POST.get("id")
        try:
            branch = CompanyBranchLocation.objects.get(id=branch_id)
            branch.delete()
            return JsonResponse({'success': True, 'message': 'Branch deleted successfully.'})
        except CompanyBranchLocation.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Branch not found.'})
    return JsonResponse({'success': False, 'message': 'Invalid request.'})



@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def create_employee_profile(request):
    if request.method == "POST":
        form = EmployeeProfileCreateForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['user']
            if EmployeeProfile.objects.filter(user=user).exists():
                messages.warning(request, f"⚠️ Employee '{user.username}' already has a profile.")
            else:
                form.save()
                messages.success(request, f"✅ Employee profile created and branch assigned.")
                return redirect('assign-branch-location')
        else:
            messages.error(request, "❌ Invalid data submitted.")
    else:
        form = EmployeeProfileCreateForm()

    return render(request, "index/assign_branch_location.html", {'form': form})


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def list_employee_profiles(request):
    admin_id = request.user.id
    profiles = EmployeeProfile.objects.select_related('user', 'branch_location').filter(
        user__role='Employee',
        user__admin_id=admin_id 
    ).order_by('user__empid', 'user__username')

    return render(request, 'index/employee_profile_list.html', {
        "profiles": profiles
    })

@login_required(login_url='login')
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
@csrf_exempt
def delete_employee_profile(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        try:
            profile = EmployeeProfile.objects.get(user_id=user_id)
            profile.delete()
            return JsonResponse({"success": True, "message": "✅ Profile deleted successfully."})
        except EmployeeProfile.DoesNotExist:
            return JsonResponse({"success": False, "message": "❌ Employee profile not found."})
    return JsonResponse({"success": False, "message": "❌ Invalid request."})


# Experience Certificate

@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
@login_required
def list_generated_certificates(request):
    user = request.user.id
    k = Myprofile.objects.filter(myuser__id=user)
    c = companyprofile.objects.filter(admin_id=user)
    certificates = ExperienceCertificate.objects.select_related('user', 'resignation').order_by('-issued_date')

    return render(request, 'index/generated_certificates_list.html', {
        'k': k[0] if k.exists() else k,
        'c': c[0] if c.exists() else c,
        'certificates': certificates
    })


@login_required
@allowed_users(allowed_roles=['Admin'], allowed_statuses=['Active'])
def admin_view_certificate_pdf(request, employee_id):
    user = get_object_or_404(User, id=employee_id)
    cert = ExperienceCertificate.objects.filter(user=user).first()

    if cert and cert.certificate_file:
        return FileResponse(open(cert.certificate_file.path, 'rb'), content_type='application/pdf')

    return HttpResponseNotFound("Certificate not found for this employee.")

    